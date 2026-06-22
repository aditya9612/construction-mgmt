from decimal import Decimal
from typing import Optional
from openpyxl import Workbook
from fastapi.responses import FileResponse
import tempfile
import csv
from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

from app.cache.redis import (
    bump_cache_version,
    cache_get_json,
    cache_set_json,
    get_cache_version,
)
from app.core.dependencies import (
    get_current_active_user,
    get_request_redis,
    require_roles,
)
from app.db.session import get_db_session
from app.middlewares.rate_limiter import default_rate_limiter_dependency
from app.models.boq import BOQ, BOQAudit, BOQGroup
from app.models.settings import CompanySettings
from app.models.master_data import ActivityType
from app.models.project import Project, Task
from app.models.user import User, UserRole
from app.schemas.base import PaginatedResponse, PaginationMeta
from app.schemas.boq import (
    BOQCreate,
    BOQOut,
    BOQUpdate,
    BOQActualsUpdate,
    BOQBulkCreate,
    BOQImportResponse,
    BOQImportError,
)
from app.utils.helpers import InvalidStateError, NotFoundError, ValidationError
from app.core.logger import logger
from app.models.boq import BOQAudit

import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
from io import BytesIO
from pydantic import ValidationError

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

FONT_NAME = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
if os.path.exists("C:/Windows/Fonts/arial.ttf"):
    pdfmetrics.registerFont(TTFont("Arial", "C:/Windows/Fonts/arial.ttf"))
    FONT_NAME = "Arial"
    FONT_BOLD = "Arial" # Fallback if bold not found
elif os.path.exists("C:/Windows/Fonts/arialbd.ttf"):
    pdfmetrics.registerFont(TTFont("Arial-Bold", "C:/Windows/Fonts/arialbd.ttf"))
    FONT_BOLD = "Arial-Bold"

from sqlalchemy.orm import selectinload
from app.models.owner import Owner


router = APIRouter(
    prefix="/boq",
    tags=["boq"],
    dependencies=[default_rate_limiter_dependency()],
)


READ_ONLY_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.ACCOUNTANT,
        UserRole.CLIENT,
    ]
]

WRITE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.ACCOUNTANT,
    ]
]

TASK_GENERATION_ROLES = [
    UserRole.ADMIN,
    UserRole.PROJECT_MANAGER,
    UserRole.SITE_ENGINEER
]

VERSION_KEY = "cache_version:boq"

# ------------------ HELPERS ------------------


def calculate_cost(
    quantity: Decimal, unit_cost: Decimal, actual_cost: Decimal = Decimal(0)
):
    total = quantity * unit_cost
    variance = total - actual_cost
    return total, variance


# ------------------ CREATE ------------------


@router.post("", response_model=BOQOut)
async def create_boq(
    payload: BOQCreate,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Creating BOQ project_id={payload.project_id}")

    project = await db.scalar(select(Project).where(Project.id == payload.project_id))
    if not project:
        logger.warning(f"Project not found project_id={payload.project_id}")
        raise NotFoundError("Project not found")

    # =========================
    # MASTER DATA VALIDATION (ADD HERE)
    # =========================
    activity = await db.get(
        ActivityType,
        payload.activity_type_id
    )

    if not activity:
        raise NotFoundError(
            "Invalid activity type"
        )

    unit_name = "unit"

    if activity.default_unit_id:
        from app.models.master_data import Unit

        unit_obj = await db.get(
            Unit,
            activity.default_unit_id
        )

        if unit_obj:
            unit_name = unit_obj.name

    try:

        quantity = Decimal(str(payload.quantity))
        unit_cost = Decimal(str(payload.unit_cost))

        total_cost, variance = calculate_cost(quantity, unit_cost)

        group = BOQGroup(
            project_id=payload.project_id,
            name=payload.item_name,
        )

        db.add(group)

        await db.flush()

        obj = BOQ(
            project_id=payload.project_id,
            boq_group_id=group.id,
            version_no=1,
            is_latest=True,
            item_name=payload.item_name,
            category=activity.category,
            description=payload.description,
            quantity=quantity,
            unit=unit_name,
            unit_cost=unit_cost,
            total_cost=total_cost,
            actual_quantity=Decimal(0),
            actual_cost=Decimal(0),
            variance_cost=variance,
            status=payload.status,
            approval_status="Draft",
            activity_type_id=payload.activity_type_id,
        )

        db.add(obj)

        await db.flush()

        await bump_cache_version(redis, VERSION_KEY)

        logger.info(f"BOQ created id={obj.id} project_id={payload.project_id}")

        return BOQOut.model_validate(obj)

    except Exception:
        logger.exception("BOQ creation failed")
        raise


# ------------------ LIST ------------------


@router.get("", response_model=PaginatedResponse[BOQOut])
async def list_boq(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
    status: Optional[str] = None,
    approval_status: Optional[str] = None,
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    version_no: Optional[int] = None,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    version = await get_cache_version(redis, VERSION_KEY)

    cache_key = (
        f"cache:boq:list:{version}:{limit}:{offset}:{search}:"
        f"{status}:{project_id}:{category}:{version_no}"
    )

    cached = await cache_get_json(redis, cache_key)
    if cached is not None:
        return PaginatedResponse[BOQOut].model_validate(cached)

    if search:
        logger.info(f"BOQ search query={search}")

    # Exclude soft-deleted BOQs by default
    query = select(BOQ).where(BOQ.status != "Deleted")
    count_query = select(func.count()).select_from(BOQ).where(BOQ.status != "Deleted")

    if search:
        like = f"%{search}%"
        query = query.where(BOQ.item_name.ilike(like))
        count_query = count_query.where(BOQ.item_name.ilike(like))

    # Allow filtering by specific status if provided
    # (e.g. status="Approved", status="Pending")
    if status:
        query = query.where(BOQ.status == status)
        count_query = count_query.where(BOQ.status == status)

    if approval_status:
        query = query.where(BOQ.approval_status == approval_status)
        count_query = count_query.where(BOQ.approval_status == approval_status)

    if project_id is not None:
        query = query.where(BOQ.project_id == project_id)
        count_query = count_query.where(BOQ.project_id == project_id)

    if category:
        query = query.where(BOQ.category == category)
        count_query = count_query.where(BOQ.category == category)

    if version_no:
        query = query.where(BOQ.version_no == version_no)
        count_query = count_query.where(BOQ.version_no == version_no)
    else:
        query = query.where(BOQ.is_latest == True)
        count_query = count_query.where(BOQ.is_latest == True)

    query = query.order_by(BOQ.id.desc()).limit(limit).offset(offset)

    total = await db.scalar(count_query)
    rows = (await db.execute(query)).scalars().all()

    items = [BOQOut.model_validate(r).model_dump() for r in rows]
    meta = PaginationMeta(total=int(total or 0), limit=limit, offset=offset)

    result = {
        "items": items,
        "meta": meta.model_dump(),
    }

    await cache_set_json(redis, cache_key, result)

    return PaginatedResponse[BOQOut].model_validate(result)


# ------------------ GET ------------------



# ------------------ TEMPLATE & IMPORT ------------------

@router.get("/template/excel")
async def download_boq_template(
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Entry"

    headers = ["Item Name", "Description", "Quantity", "Unit Cost", "Activity Type"]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25

    # Fetch Activity Types with their Category and Default Unit
    from app.models.master_data import ActivityType
    from sqlalchemy.orm import selectinload
    activities = (await db.execute(
        select(ActivityType)
        .options(selectinload(ActivityType.default_unit))
        .where(ActivityType.is_active == True)
    )).scalars().all()
    
    # Create Reference Data Sheet
    ws_ref = wb.create_sheet(title="Reference Data")
    ref_headers = ["Activity Type", "Category", "Default Unit"]
    ws_ref.append(ref_headers)
    for col_idx in range(1, len(ref_headers) + 1):
        cell = ws_ref.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BDC3C7")
        ws_ref.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    if activities:
        for idx, act in enumerate(activities, start=2):
            unit_name = act.default_unit.name if act.default_unit else ""
            ws_ref.cell(row=idx, column=1, value=act.name)
            ws_ref.cell(row=idx, column=2, value=act.category or "")
            ws_ref.cell(row=idx, column=3, value=unit_name)

        # Set Data Validation using the Reference Data sheet
        dv = DataValidation(type="list", formula1=f"='Reference Data'!$A$2:$A${len(activities)+1}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("E2:E1000")

    file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(file.name)

    return FileResponse(file.name, filename="BOQ_Template.xlsx")


@router.post("/groups/{group_id}/import/excel", response_model=BOQImportResponse)
async def import_boq_excel(
    group_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session)
):
    parent = await db.scalar(
        select(BOQ).where(BOQ.boq_group_id == group_id, BOQ.is_latest == True, BOQ.status != "Deleted")
    )
    if not parent:
        raise NotFoundError("BOQ Group not found")

    content = await file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    # Target the BOQ Entry sheet explicitly, or fallback to active
    ws = wb["BOQ Entry"] if "BOQ Entry" in wb.sheetnames else wb.active

    from app.models.master_data import ActivityType
    activities = (await db.execute(select(ActivityType).where(ActivityType.is_active == True))).scalars().all()
    
    activity_map = {}
    valid_activity_types = []
    for a in activities:
        activity_map[a.name.strip().lower()] = a.id
        valid_activity_types.append(a.name)

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return BOQImportResponse(total_rows=0, valid_rows=0, invalid_rows=0, errors=[], items=[], valid_activity_types=valid_activity_types)

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    
    col_map = {
        "item_name": -1,
        "description": -1,
        "quantity": -1,
        "unit_cost": -1,
        "activity_type": -1
    }

    for idx, h in enumerate(headers):
        if "item name" in h: col_map["item_name"] = idx
        elif "description" in h: col_map["description"] = idx
        elif "quantity" in h: col_map["quantity"] = idx
        elif "unit cost" in h: col_map["unit_cost"] = idx
        elif "activity type" in h: col_map["activity_type"] = idx

    errors = []
    valid_data = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        item_name = str(row[col_map["item_name"]]).strip() if col_map["item_name"] != -1 and row[col_map["item_name"]] is not None else None
        desc = str(row[col_map["description"]]).strip() if col_map["description"] != -1 and row[col_map["description"]] is not None else None
        qty_str = row[col_map["quantity"]] if col_map["quantity"] != -1 and row[col_map["quantity"]] is not None else None
        cost_str = row[col_map["unit_cost"]] if col_map["unit_cost"] != -1 and row[col_map["unit_cost"]] is not None else None
        act_str = str(row[col_map["activity_type"]]).strip() if col_map["activity_type"] != -1 and row[col_map["activity_type"]] is not None else None

        row_errors = []

        act_id = None
        if not act_str:
            row_errors.append("Activity Type is required")
        else:
            clean_act_str = act_str.strip().lower()
            act_id = activity_map.get(clean_act_str)
            if not act_id:
                row_errors.append(f"Activity Type '{act_str}' not found. Please select a valid type from the dropdown.")

        try:
            qty = Decimal(str(qty_str))
            if qty <= 0:
                row_errors.append("Quantity must be greater than 0")
        except:
            row_errors.append("Quantity must be a valid number")
            qty = Decimal("0")

        try:
            cost = Decimal(str(cost_str))
            if cost <= 0:
                row_errors.append("Unit cost must be greater than 0")
        except:
            row_errors.append("Unit cost must be a valid number")
            cost = Decimal("0")

        if not item_name:
            row_errors.append("Item Name is required")

        if row_errors:
            for e in row_errors:
                errors.append(BOQImportError(row=row_idx, message=e))
        else:
            try:
                boq_item = BOQCreate(
                    project_id=parent.project_id,
                    item_name=item_name,
                    description=desc,
                    quantity=qty,
                    unit_cost=cost,
                    activity_type_id=act_id,
                    status="Active"
                )
                valid_data.append(boq_item)
            except ValidationError as ve:
                for err in ve.errors():
                    errors.append(BOQImportError(row=row_idx, message=f"{err['loc'][0]}: {err['msg']}"))

    return BOQImportResponse(
        total_rows=len(valid_data) + len(set([e.row for e in errors])),
        valid_rows=len(valid_data),
        invalid_rows=len(set([e.row for e in errors])),
        errors=errors,
        valid_activity_types=valid_activity_types,
        items=valid_data
    )


@router.get("/{boq_id}", response_model=BOQOut)
async def get_boq(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if obj is None:
        logger.warning(f"BOQ not found id={boq_id}")
        raise NotFoundError("BOQ item not found")

    return BOQOut.model_validate(obj)


# ------------------ UPDATE ------------------


@router.put("/{boq_id}", response_model=BOQOut)
async def update_boq(
    boq_id: int,
    payload: BOQUpdate,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Updating BOQ id={boq_id}")

    obj = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if obj is None:
        logger.warning(f"BOQ not found for update id={boq_id}")
        raise NotFoundError("BOQ item not found")

    # prevent modifying historical versions
    if not obj.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if obj.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    try:
        data = payload.model_dump(exclude_unset=True)

        if payload.activity_type_id is not None:

            activity = await db.get(
                ActivityType,
                payload.activity_type_id
            )

            if not activity:
                raise NotFoundError(
                    "Invalid activity type"
                )

            obj.category = activity.category

            unit_name = "unit"

            if activity.default_unit_id:
                from app.models.master_data import Unit

                unit_obj = await db.get(
                    Unit,
                    activity.default_unit_id
                )

                if unit_obj:
                    unit_name = unit_obj.name

            obj.unit = unit_name

        for k, v in data.items():
            setattr(obj, k, v)

        quantity = Decimal(str(obj.quantity))
        unit_cost = Decimal(str(obj.unit_cost))

        total_cost, variance = calculate_cost(
            quantity, unit_cost, obj.actual_cost or Decimal(0)
        )

        obj.total_cost = total_cost
        obj.variance_cost = variance

        await db.flush()
        await bump_cache_version(redis, VERSION_KEY)

        logger.info(f"BOQ updated id={boq_id}")

        return BOQOut.model_validate(obj)

    except Exception:
        logger.exception(f"BOQ update failed id={boq_id}")
        raise


# ------------------ DELETE ------------------


@router.delete("/{boq_id}")
async def delete_boq(
    boq_id: int,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Deleting BOQ id={boq_id}")

    obj = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if obj is None:
        logger.warning(f"BOQ not found for delete id={boq_id}")
        raise NotFoundError("BOQ item not found")

    if not obj.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if obj.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    obj.status = "Deleted"

    await db.flush()
    await bump_cache_version(redis, VERSION_KEY)

    logger.info(f"BOQ soft-deleted id={boq_id}")

    return {"message": "BOQ deleted successfully", "boq_id": boq_id}


# ------------------ ACTUALS ------------------


@router.post("/{boq_id}/actuals", response_model=BOQOut)
async def update_actuals(
    boq_id: int,
    payload: BOQActualsUpdate,
    redis=Depends(get_request_redis),
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Updating BOQ actuals id={boq_id}")

    obj = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if not obj:
        raise NotFoundError("BOQ not found")

    # prevent modifying historical versions
    if not obj.is_latest:
        raise InvalidStateError("Cannot modify old BOQ version.")

    if obj.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    obj.actual_quantity = Decimal(str(payload.actual_quantity))
    obj.actual_cost = Decimal(str(payload.actual_cost))

    _, variance = calculate_cost(obj.quantity, obj.unit_cost, payload.actual_cost)

    obj.variance_cost = variance

    await db.flush()
    await bump_cache_version(redis, VERSION_KEY)
    logger.info(f"BOQ actuals updated id={boq_id}")

    return BOQOut.model_validate(obj)


# ------------------ SUMMARY ------------------


@router.get("/summary/{project_id}")
async def boq_summary(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(func.count(), func.sum(BOQ.total_cost), func.sum(BOQ.actual_cost)).where(
            BOQ.project_id == project_id, BOQ.is_latest == True, BOQ.status != "Deleted"
        )
    )

    total_items, estimated, actual = result.one()

    return {
        "total_items": total_items or 0,
        "estimated": float(estimated or 0),
        "actual": float(actual or 0),
        "difference": float((estimated or 0) - (actual or 0)),
    }


# ------------------ COMPARISON ------------------


@router.get("/comparison/{project_id}")
async def boq_comparison(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    rows = (
        (
            await db.execute(
                select(BOQ).where(
                    BOQ.project_id == project_id,
                    BOQ.is_latest == True,
                    BOQ.status != "Deleted",
                )
            )
        )
        .scalars()
        .all()
    )

    return [
        {
            "item_name": r.item_name,
            "estimated": float(r.total_cost),
            "actual": float(r.actual_cost),
            "variance": float(r.variance_cost),
        }
        for r in rows
    ]


# ------------------ REPORT ------------------


@router.get("/{boq_id}/report")
async def boq_report(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if not base:
        raise NotFoundError("BOQ not found")

    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.boq_group_id == base.boq_group_id,
                    BOQ.version_no == base.version_no,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    total_estimated = sum(float(r.total_cost) for r in rows)
    total_actual = sum(float(r.actual_cost) for r in rows)

    return {
        "total_items": len(rows),
        "estimated": total_estimated,
        "actual": total_actual,
        "difference": total_estimated - total_actual,
    }


# ------------------ ALERTS ------------------


@router.get("/{boq_id}/alerts")
async def boq_alerts(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(
        select(BOQ).where(
            BOQ.id == boq_id,
            BOQ.status != "Deleted",
        )
    )

    if not base:
        raise NotFoundError("BOQ not found")

    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.boq_group_id == base.boq_group_id,
                    BOQ.version_no == base.version_no,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    alerts = []

    for r in rows:
        if r.actual_cost > r.total_cost:
            alerts.append(
                {
                    "item": r.item_name,
                    "message": "Cost exceeded estimate",
                }
            )

    return {"alerts": alerts}


@router.get("/{boq_id}/versions")
async def get_versions(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    # Get base BOQ to find project
    base = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if not base:
        raise NotFoundError("BOQ not found")

    result = await db.execute(
        select(BOQ.version_no)
        .where(
            BOQ.boq_group_id == base.boq_group_id,
            BOQ.status != "Deleted",
        )
        .distinct()
        .order_by(BOQ.version_no.desc())
    )

    return {"versions": [v[0] for v in result.fetchall()]}


@router.get("/project/{project_id}", response_model=list[BOQOut])
async def get_boq_by_project(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.project_id == project_id,
                    BOQ.is_latest == True,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    return [BOQOut.model_validate(r) for r in rows]


# ------------------ ITEMS ------------------


@router.post("/groups/{group_id}/items", response_model=BOQOut)
async def add_item(
    group_id: int,
    payload: BOQCreate,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    parent = await db.scalar(
        select(BOQ)
        .where(
            BOQ.boq_group_id == group_id,
            BOQ.is_latest == True,
            BOQ.status != "Deleted",
        )
        .order_by(BOQ.id.asc())
    )

    if not parent:
        raise NotFoundError("BOQ not found")

    # prevent modifying old versions
    if not parent.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if parent.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    activity = await db.get(
        ActivityType,
        payload.activity_type_id
    )

    if not activity:
        raise NotFoundError(
            "Invalid activity type"
        )
        
    unit_name = "unit"

    if activity.default_unit_id:
        from app.models.master_data import Unit

        unit_obj = await db.get(
            Unit,
            activity.default_unit_id
        )

        if unit_obj:
            unit_name = unit_obj.name

    quantity = Decimal(str(payload.quantity))
    unit_cost = Decimal(str(payload.unit_cost))

    total_cost, variance = calculate_cost(quantity, unit_cost)

    obj = BOQ(
        project_id=payload.project_id,
        boq_group_id=parent.boq_group_id,
        version_no=parent.version_no,
        is_latest=True,
        item_name=payload.item_name,
        category=activity.category,
        description=payload.description,
        quantity=quantity,
        unit=unit_name,
        unit_cost=unit_cost,
        total_cost=total_cost,
        variance_cost=variance,
        status=payload.status,
        approval_status="Draft",
        activity_type_id=payload.activity_type_id,
    )

    db.add(obj)
    await db.flush()

    await bump_cache_version(redis, VERSION_KEY)

    return BOQOut.model_validate(obj)


@router.get("/groups/{group_id}/items", response_model=list[BOQOut])
async def get_items(
    group_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(
        select(BOQ).where(
            BOQ.boq_group_id == group_id,
            BOQ.is_latest == True,
            BOQ.status != "Deleted",
        )
    )

    if not base:
        raise NotFoundError("BOQ not found")

    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.boq_group_id == base.boq_group_id,
                    BOQ.version_no == base.version_no,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    return [BOQOut.model_validate(r) for r in rows]


@router.put("/items/{item_id}", response_model=BOQOut)
async def update_item(
    item_id: int,
    payload: BOQUpdate,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    obj = await db.scalar(select(BOQ).where(BOQ.id == item_id, BOQ.status != "Deleted"))

    if not obj:
        raise NotFoundError("Item not found")

    # prevent modifying historical versions
    if not obj.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if obj.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    data = payload.model_dump(exclude_unset=True)

    if payload.activity_type_id is not None:

        activity = await db.get(
            ActivityType,
            payload.activity_type_id
        )

        if not activity:
            raise NotFoundError(
                "Invalid activity type"
            )

        obj.category = activity.category

        unit_name = "unit"

        if activity.default_unit_id:
            from app.models.master_data import Unit

            unit_obj = await db.get(
                Unit,
                activity.default_unit_id
            )

            if unit_obj:
                unit_name = unit_obj.name

        obj.unit = unit_name

    for k, v in data.items():
        setattr(obj, k, v)

    quantity = Decimal(str(obj.quantity))
    unit_cost = Decimal(str(obj.unit_cost))

    total, variance = calculate_cost(quantity, unit_cost, obj.actual_cost or Decimal(0))

    obj.total_cost = total
    obj.variance_cost = variance

    await db.flush()

    await bump_cache_version(redis, VERSION_KEY)

    return BOQOut.model_validate(obj)



@router.post("/groups/{group_id}/items/bulk")
async def bulk_add_items(
    group_id: int,
    payload: BOQBulkCreate,
    BOQImportResponse,
    BOQImportError,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    parent = await db.scalar(
        select(BOQ)
        .where(
            BOQ.boq_group_id == group_id,
            BOQ.is_latest == True,
            BOQ.status != "Deleted",
        )
        .order_by(BOQ.id.asc())
    )

    if not parent:
        raise NotFoundError("BOQ not found")

    if not parent.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if parent.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )

    created_items = []

    try:
        for item in payload.items:

            activity = await db.get(
                ActivityType,
                item.activity_type_id
            )

            if not activity:
                raise NotFoundError("Invalid activity type")

            unit_name = "unit"

            if activity.default_unit_id:
                from app.models.master_data import Unit

                unit_obj = await db.get(
                    Unit,
                    activity.default_unit_id
                )

                if unit_obj:
                    unit_name = unit_obj.name

            quantity = Decimal(str(item.quantity))
            unit_cost = Decimal(str(item.unit_cost))

            total_cost, variance = calculate_cost(
                quantity,
                unit_cost
            )

            obj = BOQ(
                project_id=item.project_id,
                boq_group_id=parent.boq_group_id,
                version_no=parent.version_no,
                is_latest=True,
                item_name=item.item_name,
                category=activity.category,
                description=item.description,
                quantity=quantity,
                unit=unit_name,
                unit_cost=unit_cost,
                total_cost=total_cost,
                actual_quantity=Decimal("0"),
                actual_cost=Decimal("0"),
                variance_cost=variance,
                is_completed=False,
                status=item.status,
                approval_status="Draft",
                activity_type_id=item.activity_type_id,
            )

            db.add(obj)
            created_items.append(obj)

        await db.flush()
        await db.commit()

    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(redis, VERSION_KEY)

    return {
        "message": f"{len(created_items)} items created",
        "items": [
            BOQOut.model_validate(item)
            for item in created_items
        ],
    }


@router.delete("/items/{item_id}")
async def delete_item(
    item_id: int,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    obj = await db.scalar(select(BOQ).where(BOQ.id == item_id, BOQ.status != "Deleted"))

    if not obj:
        raise NotFoundError("Item not found")

    # prevent deleting historical versions
    if not obj.is_latest:
        raise InvalidStateError(
            "Cannot modify old BOQ version. Create a new version first."
        )

    if obj.approval_status == "Approved":
        raise InvalidStateError(
            "Approved BOQ cannot be modified. Create a new version first."
        )
    obj.status = "Deleted"

    await db.flush()

    await bump_cache_version(redis, VERSION_KEY)

    return {"message": "BOQ deleted successfully", "item_id": item_id}


# ------------------ CREATE VERSION ------------------


@router.post("/groups/{group_id}/versions")
async def create_version(
    group_id: int,
    current_user: User = Depends(require_roles(WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    async with db.begin():

        base = await db.scalar(
            select(BOQ).where(
                BOQ.boq_group_id == group_id,
                BOQ.is_latest == True,
                BOQ.status != "Deleted",
            )
        )

        if not base:
            raise NotFoundError("BOQ not found")

        if base.approval_status != "Approved":
            raise InvalidStateError(
                "Only approved BOQ versions can create a new version."
            )

        group = await db.get(BOQGroup, base.boq_group_id)

        if not group:
            raise NotFoundError("BOQ group not found")

        new_version = group.current_version + 1

        group.current_version = new_version
        group.name = base.item_name

        await db.execute(
            update(BOQ)
            .where(
                BOQ.boq_group_id == base.boq_group_id,
                BOQ.version_no == base.version_no,
                BOQ.is_latest == True,
            )
            .values(is_latest=False)
        )

        rows = (
            (
                await db.execute(
                    select(BOQ)
                    .where(
                        BOQ.boq_group_id == base.boq_group_id,
                        BOQ.version_no == base.version_no,
                        BOQ.status != "Deleted",
                    )
                    .order_by(BOQ.id.asc())
                )
            )
            .scalars()
            .all()
        )

        for r in rows:
            db.add(
                BOQ(
                    project_id=r.project_id,
                    boq_group_id=base.boq_group_id,
                    version_no=new_version,
                    is_latest=True,
                    item_name=r.item_name,
                    category=r.category,
                    description=r.description,
                    quantity=r.quantity,
                    unit=r.unit,
                    unit_cost=r.unit_cost,
                    total_cost=r.total_cost,
                    actual_quantity=Decimal(0),
                    actual_cost=Decimal(0),
                    variance_cost=Decimal(0),
                    status="Active",
                    approval_status="Draft",
                    activity_type_id=r.activity_type_id,
                )
            )
        await db.flush()

    await bump_cache_version(redis, VERSION_KEY)

    return {
        "message": "Version created successfully",
        "version": new_version,
        "boq_group_id": base.boq_group_id,
    }


# ------------------ EXPORT ------------------


@router.get("/{boq_id}/export/json")
async def export_boq_json(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(
        select(BOQ).where(
            BOQ.id == boq_id,
            BOQ.status != "Deleted",
        )
    )

    if not base:
        raise NotFoundError("BOQ not found")

    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.boq_group_id == base.boq_group_id,
                    BOQ.version_no == base.version_no,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    return [BOQOut.model_validate(r).model_dump() for r in rows]


@router.get("/{boq_id}/export/excel")
async def export_boq_excel(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))
    if not base:
        raise NotFoundError("BOQ not found")

    project = await db.scalar(select(Project).options(selectinload(Project.owner)).where(Project.id == base.project_id))

    rows = (await db.execute(select(BOQ).where(BOQ.boq_group_id == base.boq_group_id, BOQ.version_no == base.version_no, BOQ.status != "Deleted").order_by(BOQ.id.asc()))).scalars().all()
    if not rows:
        raise NotFoundError("No BOQ data found")

    company_settings = await db.scalar(select(CompanySettings))
    company_name = company_settings.company_name if company_settings and company_settings.company_name else "Construction Company"

    file_path = await _generate_boq_excel(base, list(rows), project, current_user, company_name)
    return FileResponse(file_path, filename=f"BOQ_{project.business_id}_v{base.version_no}.xlsx")


@router.get("/{boq_id}/export/pdf")
async def export_boq_pdf(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))
    if not base:
        raise NotFoundError("BOQ not found")

    project = await db.scalar(select(Project).options(selectinload(Project.owner)).where(Project.id == base.project_id))

    rows = (await db.execute(select(BOQ).where(BOQ.boq_group_id == base.boq_group_id, BOQ.version_no == base.version_no, BOQ.status != "Deleted").order_by(BOQ.id.asc()))).scalars().all()
    if not rows:
        raise NotFoundError("No BOQ data found")

    company_settings = await db.scalar(select(CompanySettings))
    company_name = company_settings.company_name if company_settings and company_settings.company_name else "Construction Company"

    file_path = await _generate_boq_pdf(base, list(rows), project, current_user, company_name)
    return FileResponse(file_path, filename=f"BOQ_{project.business_id}_v{base.version_no}.pdf")


# ------------------ OPTIMIZE ------------------


@router.get("/{boq_id}/optimize")
async def boq_optimize(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    base = await db.scalar(select(BOQ).where(BOQ.id == boq_id, BOQ.status != "Deleted"))

    if not base:
        raise NotFoundError("BOQ not found")

    rows = (
        (
            await db.execute(
                select(BOQ)
                .where(
                    BOQ.boq_group_id == base.boq_group_id,
                    BOQ.version_no == base.version_no,
                    BOQ.status != "Deleted",
                )
                .order_by(BOQ.id.asc())
            )
        )
        .scalars()
        .all()
    )

    suggestions = []

    for r in rows:
        if r.actual_cost > r.total_cost:
            suggestions.append(
                {
                    "item": r.item_name,
                    "suggestion": "Reduce cost or renegotiate vendor",
                    "over_budget_by": float(r.actual_cost - r.total_cost),
                }
            )

    return {"suggestions": suggestions}


# ------------------ AUDIT LOGS ------------------


@router.get("/{boq_id}/logs")
async def boq_logs(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    rows = (
        (
            await db.execute(
                select(BOQAudit)
                .where(BOQAudit.boq_id == boq_id)
                .order_by(BOQAudit.id.desc())
            )
        )
        .scalars()
        .all()
    )

    return [
        {
            "action": r.action,
            "message": r.message,
            "user_id": r.user_id,
            "timestamp": r.created_at,
            "changes": r.changes,
        }
        for r in rows
    ]


@router.get("/{boq_id}/logs/export/csv")
async def export_boq_logs_csv(
    boq_id: int,
    current_user: User = Depends(require_roles(READ_ONLY_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    rows = (
        (
            await db.execute(
                select(BOQAudit)
                .where(BOQAudit.boq_id == boq_id)
                .order_by(BOQAudit.id.desc())
            )
        )
        .scalars()
        .all()
    )

    if not rows:
        raise NotFoundError("No audit logs found for this BOQ")

    file = tempfile.NamedTemporaryFile(
        delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8"
    )
    file_path = file.name

    writer = csv.writer(file)
    writer.writerow(["ID", "Action", "Message", "User ID", "Timestamp", "Changes"])

    for r in rows:
        writer.writerow(
            [
                r.id,
                r.action,
                r.message,
                r.user_id,
                r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
                str(r.changes) if r.changes else "",
            ]
        )

    file.close()
    return FileResponse(
        file_path, filename=f"boq_audit_logs_{boq_id}.csv", media_type="text/csv"
    )


@router.post("/{boq_id}/generate-tasks")
async def generate_tasks_from_boq(
    boq_id: int,
    milestone_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_GENERATION_ROLES)),
):
    boq = await db.scalar(
        select(BOQ).where(
            BOQ.id == boq_id,
            BOQ.status != "Deleted",
        )
    )

    if not boq:
        raise NotFoundError("BOQ not found")

    # prevent task generation from historical versions
    if not boq.is_latest:
        raise InvalidStateError("Cannot generate task from old BOQ version.")

    if boq.approval_status != "Approved":
        raise InvalidStateError("BOQ must be approved before generating tasks.")

    # prevent duplicate task generation
    existing_task = await db.scalar(select(Task).where(Task.boq_id == boq.id))

    if milestone_id:
        from app.models.project import Milestone

        milestone = await db.scalar(
            select(Milestone).where(Milestone.id == milestone_id)
        )
        if not milestone:
            raise NotFoundError("Milestone not found")
        if milestone.project_id != boq.project_id:
            raise ValidationError(
                "Milestone does not belong to the same project as the BOQ"
            )

    if existing_task:
        return {
            "message": "Task already exists for this BOQ",
            "task_id": existing_task.id,
        }

    task = Task(
        project_id=boq.project_id,
        boq_id=boq.id,
        milestone_id=milestone_id,
        activity_type_id=boq.activity_type_id,
        title=boq.item_name,
        description=boq.description,
        priority=1,
        status="PLANNED",
        created_by_user_id=current_user.id,
        completion_percentage=0,
    )

    db.add(task)

    await db.flush()
    await db.refresh(task)

    return {
        "message": "Task created from BOQ",
        "task_id": task.id,
        "milestone_id": milestone_id,
    }


# ==============================================================================
# ENTERPRISE BOQ EXCEL EXPORT
# ==============================================================================
async def _generate_boq_excel(base: BOQ, rows: list[BOQ], project: Project, user: User, company_name: str) -> str:
    wb = Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: PROJECT SUMMARY
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Project Summary"
    
    # Styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    bold_font = Font(bold=True)
    
    ws_summary.append(["ENTERPRISE BOQ EXPORT"])
    ws_summary["A1"].font = Font(bold=True, size=18)
    ws_summary.append([])
    
    info_data = [
        ["Company Name", company_name],
        ["Project Name", project.project_name],
        ["Project Code", project.business_id],
        ["Owner Name", project.owner.owner_name if project.owner else "N/A"],
        ["Site Address", project.site_address or "N/A"],
        ["BOQ Version", f"v{base.version_no}"],
        ["Approval Status", base.approval_status],
        ["Generated Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M")],
        ["Generated By", user.full_name]
    ]
    
    for r in info_data:
        ws_summary.append(r)
        ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
        
    ws_summary.append([])
    ws_summary.append(["SUMMARY METRICS"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = header_font
    ws_summary.cell(row=ws_summary.max_row, column=1).fill = header_fill
    
    total_est = sum(float(r.total_cost) for r in rows)
    total_act = sum(float(r.actual_cost) for r in rows)
    total_var = sum(float(r.variance_cost) for r in rows)
    comp_pct = (total_act / total_est * 100) if total_est > 0 else 0
    
    metrics = [
        ["Total BOQ Items", len(rows)],
        ["Total Estimated Cost", total_est],
        ["Total Actual Cost", total_act],
        ["Total Variance", total_var],
        ["Completion %", f"{comp_pct:.2f}%"]
    ]
    
    for m in metrics:
        ws_summary.append(m)
        ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
        
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 40

    # ---------------------------------------------------------
    # SHEET 2: BOQ DETAILS
    # ---------------------------------------------------------
    ws_details = wb.create_sheet(title="BOQ Details")
    
    headers = [
        "Sr No", "Item Name", "Category", "Description", "Quantity", "Unit",
        "Unit Rate", "Estimated Cost", "Actual Quantity", "Actual Cost",
        "Variance", "Completion %", "Status", "Approval Status"
    ]
    ws_details.append(headers)
    
    # Format Headers
    for col in range(1, len(headers) + 1):
        cell = ws_details.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    ws_details.auto_filter.ref = f"A1:N1"
    ws_details.freeze_panes = "A2"
    
    # Group by Category
    categories = {}
    for r in rows:
        categories.setdefault(r.category, []).append(r)
        
    sr_no = 1
    current_row = 2
    
    for cat, items in categories.items():
        cat_est = 0
        cat_act = 0
        cat_var = 0
        
        for r in items:
            est_cost = float(r.total_cost)
            act_cost = float(r.actual_cost)
            var_cost = float(r.variance_cost)
            comp = (act_cost / est_cost) if est_cost > 0 else 0
            
            cat_est += est_cost
            cat_act += act_cost
            cat_var += var_cost
            
            row_data = [
                sr_no,
                r.item_name,
                r.category,
                r.description,
                float(r.quantity),
                r.unit,
                float(r.unit_cost),
                est_cost,
                float(r.actual_quantity),
                act_cost,
                var_cost,
                comp,
                r.status,
                r.approval_status
            ]
            ws_details.append(row_data)
            
            # Format numbers
            ws_details.cell(row=current_row, column=7).number_format = '#,##0.00'
            ws_details.cell(row=current_row, column=8).number_format = '#,##0.00'
            ws_details.cell(row=current_row, column=10).number_format = '#,##0.00'
            ws_details.cell(row=current_row, column=11).number_format = '#,##0.00'
            ws_details.cell(row=current_row, column=12).number_format = '0.00%'
            
            sr_no += 1
            current_row += 1
            
        # Category Subtotal
        ws_details.append([
            "", f"SUBTOTAL: {cat}", "", "", "", "", "",
            cat_est, "", cat_act, cat_var, "", "", ""
        ])
        subtotal_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws_details.cell(row=current_row, column=col)
            cell.fill = subtotal_fill
            cell.font = bold_font
            if col in [8, 10, 11]:
                cell.number_format = '#,##0.00'
        current_row += 1
        
    # Grand Total
    ws_details.append([
        "", "GRAND TOTAL", "", "", "", "", "",
        total_est, "", total_act, total_var, "", "", ""
    ])
    gt_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws_details.cell(row=current_row, column=col)
        cell.fill = gt_fill
        cell.font = Font(bold=True, color="FFFFFF")
        if col in [8, 10, 11]:
            cell.number_format = '#,##0.00'
            
    # Auto width
    for i, col in enumerate(ws_details.columns, 1):
        ws_details.column_dimensions[get_column_letter(i)].width = 15
    ws_details.column_dimensions["B"].width = 30
    ws_details.column_dimensions["D"].width = 40

    # Conditional Formatting for Variance
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C6500")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    # Column 11 is Variance. We need to format based on % variance which is Variance / Estimated
    # Since we can't easily do relative % rule here, we format based on raw numbers:
    # Actually, simpler to format Completion % (Col 12).
    # But user asked: Green: Actual <= Estimated, Yellow: Var 0-10%, Red: Var > 10%
    # We will apply directly to the Variance column if we can, or just loop through.
    # It's easier to apply via a loop for precise rules.
    for row in range(2, current_row):
        est_cell = ws_details.cell(row=row, column=8)
        var_cell = ws_details.cell(row=row, column=11)
        if type(est_cell.value) in (int, float) and type(var_cell.value) in (int, float):
            est = float(est_cell.value)
            var = float(var_cell.value)
            if est > 0:
                var_pct = var / est
                if var_pct <= 0:
                    var_cell.fill = green_fill
                    var_cell.font = green_font
                elif 0 < var_pct <= 0.10:
                    var_cell.fill = yellow_fill
                    var_cell.font = yellow_font
                else:
                    var_cell.fill = red_fill
                    var_cell.font = red_font
                    
    # ---------------------------------------------------------
    # SHEET 3: CATEGORY COST ANALYSIS
    # ---------------------------------------------------------
    ws_cost = wb.create_sheet(title="Category Cost Analysis")
    ws_cost.append(["Category", "Item Count", "Estimated Cost", "Actual Cost", "Variance", "Variance %"])
    for col in range(1, 7):
        ws_cost.cell(row=1, column=col).font = header_font
        ws_cost.cell(row=1, column=col).fill = header_fill
        
    cat_stats = []
    for cat, items in categories.items():
        count = len(items)
        est = sum(float(i.total_cost) for i in items)
        act = sum(float(i.actual_cost) for i in items)
        var = sum(float(i.variance_cost) for i in items)
        var_pct = (var / est) if est > 0 else 0
        cat_stats.append([cat, count, est, act, var, var_pct])
        
    for r in cat_stats:
        ws_cost.append(r)
        row = ws_cost.max_row
        ws_cost.cell(row=row, column=3).number_format = '#,##0.00'
        ws_cost.cell(row=row, column=4).number_format = '#,##0.00'
        ws_cost.cell(row=row, column=5).number_format = '#,##0.00'
        ws_cost.cell(row=row, column=6).number_format = '0.00%'
        
    for i in range(1, 7):
        ws_cost.column_dimensions[get_column_letter(i)].width = 20

    # ---------------------------------------------------------
    # SHEET 4: PROGRESS ANALYSIS
    # ---------------------------------------------------------
    ws_prog = wb.create_sheet(title="Progress Analysis")
    ws_prog.append(["Item Name", "Planned Quantity", "Actual Quantity", "Completion %"])
    for col in range(1, 5):
        ws_prog.cell(row=1, column=col).font = header_font
        ws_prog.cell(row=1, column=col).fill = header_fill
        
    for r in rows:
        est_qty = float(r.quantity)
        act_qty = float(r.actual_quantity)
        comp = (act_qty / est_qty) if est_qty > 0 else 0
        ws_prog.append([r.item_name, est_qty, act_qty, comp])
        row = ws_prog.max_row
        ws_prog.cell(row=row, column=4).number_format = '0.00%'
        
        # Highlight low progress
        if comp < 0.20:
            ws_prog.cell(row=row, column=4).font = red_font
            
    ws_prog.column_dimensions["A"].width = 40
    ws_prog.column_dimensions["B"].width = 20
    ws_prog.column_dimensions["C"].width = 20
    ws_prog.column_dimensions["D"].width = 20

    # ---------------------------------------------------------
    # SHEET 5: MANAGEMENT SUMMARY
    # ---------------------------------------------------------
    ws_mgmt = wb.create_sheet(title="Management Summary")
    
    ws_mgmt.append(["TOP 10 HIGHEST COST ITEMS"])
    ws_mgmt.cell(row=ws_mgmt.max_row, column=1).font = header_font
    ws_mgmt.cell(row=ws_mgmt.max_row, column=1).fill = header_fill
    ws_mgmt.append(["Item Name", "Category", "Estimated Cost", "Actual Cost"])
    
    top_cost = sorted(rows, key=lambda x: float(x.total_cost), reverse=True)[:10]
    for r in top_cost:
        ws_mgmt.append([r.item_name, r.category, float(r.total_cost), float(r.actual_cost)])
        
    ws_mgmt.append([])
    ws_mgmt.append([])
    
    ws_mgmt.append(["TOP 10 HIGHEST VARIANCE ITEMS (BUDGET RISKS)"])
    ws_mgmt.cell(row=ws_mgmt.max_row, column=1).font = header_font
    ws_mgmt.cell(row=ws_mgmt.max_row, column=1).fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    ws_mgmt.append(["Item Name", "Category", "Variance Amount", "Variance %"])
    
    top_var = sorted(rows, key=lambda x: float(x.variance_cost), reverse=True)[:10]
    for r in top_var:
        est = float(r.total_cost)
        var = float(r.variance_cost)
        var_pct = (var / est) if est > 0 else 0
        ws_mgmt.append([r.item_name, r.category, var, var_pct])
        ws_mgmt.cell(row=ws_mgmt.max_row, column=4).number_format = '0.00%'
        
    ws_mgmt.column_dimensions["A"].width = 40
    ws_mgmt.column_dimensions["B"].width = 20
    ws_mgmt.column_dimensions["C"].width = 20
    ws_mgmt.column_dimensions["D"].width = 20
    
    file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(file.name)
    return file.name

# ==============================================================================
# ENTERPRISE BOQ PDF EXPORT
# ==============================================================================
async def _generate_boq_pdf(base: BOQ, rows: list[BOQ], project: Project, user: User, company_name: str) -> str:
    file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    doc = SimpleDocTemplate(file.name, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    h2_style = styles["Heading2"]
    
    # ---------------------------------------------------------
    # PAGE 1: EXECUTIVE SUMMARY
    # ---------------------------------------------------------
    elements.append(Paragraph("Enterprise BOQ Report", title_style))
    elements.append(Spacer(1, 20))
    
    info_data = [
        ["Company Name", company_name],
        ["Project Name", project.project_name],
        ["Project Code", project.business_id],
        ["Owner Name", project.owner.owner_name if project.owner else "N/A"],
        ["BOQ Version", f"v{base.version_no}"],
        ["Generated Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M")]
    ]
    
    info_table = Table(info_data, colWidths=[150, 300])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), FONT_BOLD),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 40))
    
    total_est = sum(float(r.total_cost) for r in rows)
    total_act = sum(float(r.actual_cost) for r in rows)
    total_var = sum(float(r.variance_cost) for r in rows)
    comp_pct = (total_act / total_est * 100) if total_est > 0 else 0
    
    currency_sym = "₹" if FONT_NAME == "Arial" else "Rs."
    
    metrics_data = [
        ["SUMMARY METRICS", ""],
        ["Total Estimated Cost", f"{currency_sym} {total_est:,.2f}"],
        ["Total Actual Cost", f"{currency_sym} {total_act:,.2f}"],
        ["Total Variance", f"{currency_sym} {total_var:,.2f}"],
        ["Completion %", f"{comp_pct:.2f}%"]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[200, 250])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#16a085')),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    
    elements.append(metrics_table)
    elements.append(PageBreak())
    
    # ---------------------------------------------------------
    # PAGE 2+: DETAILED BOQ TABLE
    # ---------------------------------------------------------
    elements.append(Paragraph("Detailed BOQ Items", h2_style))
    elements.append(Spacer(1, 10))
    
    headers = [
        "Sr No", "Item Name", "Category", "Qty", "Unit",
        f"Est Rate ({currency_sym})", f"Est Cost ({currency_sym})", "Act Qty", f"Act Cost ({currency_sym})",
        f"Variance ({currency_sym})", "Comp %"
    ]
    
    table_data = [headers]
    
    categories = {}
    for r in rows:
        categories.setdefault(r.category, []).append(r)
        
    sr_no = 1
    for cat, items in categories.items():
        cat_est = 0
        cat_act = 0
        cat_var = 0
        
        for r in items:
            est = float(r.total_cost)
            act = float(r.actual_cost)
            var = float(r.variance_cost)
            comp = (act / est * 100) if est > 0 else 0
            
            cat_est += est
            cat_act += act
            cat_var += var
            
            table_data.append([
                str(sr_no),
                r.item_name[:25], # truncate for PDF fit
                r.category[:15],
                f"{float(r.quantity):.1f}",
                r.unit,
                f"{float(r.unit_cost):,.2f}",
                f"{est:,.2f}",
                f"{float(r.actual_quantity):.1f}",
                f"{act:,.2f}",
                f"{var:,.2f}",
                f"{comp:.1f}%"
            ])
            sr_no += 1
            
        # Category subtotal
        table_data.append([
            "", f"SUBTOTAL: {cat}", "", "", "", "",
            f"{currency_sym} {cat_est:,.2f}", "", f"{currency_sym} {cat_act:,.2f}", f"{currency_sym} {cat_var:,.2f}", ""
        ])
        
    # Grand Total
    table_data.append([
        "", "GRAND TOTAL", "", "", "", "",
        f"{currency_sym} {total_est:,.2f}", "", f"{currency_sym} {total_act:,.2f}", f"{currency_sym} {total_var:,.2f}", ""
    ])
    
    t = Table(table_data, repeatRows=1, colWidths=[30, 140, 80, 50, 40, 60, 80, 50, 80, 80, 50])
    
    # Complex styling to handle subtotals and totals dynamically
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'), # Item Name
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]
    
    # Apply subtotal / grand total styles by row index
    for idx, row in enumerate(table_data):
        if row[1].startswith("SUBTOTAL"):
            t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#ecf0f1')))
            t_style.append(('FONTNAME', (0, idx), (-1, idx), FONT_BOLD))
        elif row[1] == "GRAND TOTAL":
            t_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#34495e')))
            t_style.append(('TEXTCOLOR', (0, idx), (-1, idx), colors.white))
            t_style.append(('FONTNAME', (0, idx), (-1, idx), FONT_BOLD))
            
    t.setStyle(TableStyle(t_style))
    elements.append(t)
    elements.append(PageBreak())
    
    # ---------------------------------------------------------
    # FINAL PAGE: MANAGEMENT ANALYSIS
    # ---------------------------------------------------------
    elements.append(Paragraph("Management Analysis", title_style))
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("Top 5 Highest Cost Items", h2_style))
    top_cost = sorted(rows, key=lambda x: float(x.total_cost), reverse=True)[:5]
    tc_data = [["Item Name", "Category", f"Estimated Cost ({currency_sym})", f"Actual Cost ({currency_sym})"]]
    for r in top_cost:
        tc_data.append([r.item_name, r.category, f"{float(r.total_cost):,.2f}", f"{float(r.actual_cost):,.2f}"])
    
    tc_table = Table(tc_data, colWidths=[200, 150, 100, 100])
    tc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2980b9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(tc_table)
    elements.append(Spacer(1, 30))
    
    elements.append(Paragraph("Top 5 Highest Variance Items (Risk)", h2_style))
    top_var = sorted(rows, key=lambda x: float(x.variance_cost), reverse=True)[:5]
    tv_data = [["Item Name", "Category", f"Variance ({currency_sym})", "Variance %"]]
    for r in top_var:
        est = float(r.total_cost)
        var = float(r.variance_cost)
        pct = (var/est*100) if est > 0 else 0
        tv_data.append([r.item_name, r.category, f"{var:,.2f}", f"{pct:.1f}%"])
        
    tv_table = Table(tv_data, colWidths=[200, 150, 100, 100])
    tv_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c0392b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(tv_table)
    
    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        canvas.drawString(30, 10, f"{company_name} - Project {project.business_id}")
        canvas.drawRightString(A4[1] - 30, 10, f"Page {doc.page}")
        canvas.restoreState()
        
    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    return file.name

