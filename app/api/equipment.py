from typing import Optional, List
from datetime import date, datetime, timedelta
from decimal import Decimal
import io
import json
from zipfile import Path
from app.models.boq import BOQ

# FastAPI
from fastapi import (
    APIRouter,
    Depends,
    Query,
    HTTPException,
    BackgroundTasks,
    status,
    Request,
)
from fastapi.responses import StreamingResponse
from fastapi.encoders import jsonable_encoder

# SQLAlchemy
from sqlalchemy import exists, select, and_, or_, func, text
from sqlalchemy.ext.asyncio import AsyncSession

# Report / Excel
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Internal - Cache
from app.cache.redis import (
    bump_cache_version,
    cache_get_json,
    cache_set_json,
    get_cache_version,
)

# Internal - Dependencies
from app.core import db
from app.core.dependencies import (
    get_current_active_user,
    get_request_redis,
    require_roles,
)

# Internal - DB / Models
from app.db.session import get_db_session
from app.models.equipment import (
    Equipment,
    EquipmentPurchase,
    EquipmentUsage,
    EquipmentMaintenance,
    EquipmentRental,
    EquipmentAuditLog,
)
from app.models.project import Project
from app.models.user import User, UserRole

# Internal - Enums
from app.core.enums import EquipmentCondition, EquipmentStatus, PurchaseType
from openpyxl.cell.cell import MergedCell

# Internal - Schemas
from app.schemas.base import PaginatedResponse, PaginationMeta
from app.schemas.equipment import (
    DeleteRentalResponse,
    DeleteUsageResponse,
    EquipmentAllocateRequest,
    EquipmentAllocateResponse,
    EquipmentCreate,
    EquipmentDeallocateRequest,
    EquipmentDeallocateResponse,
    EquipmentKPIOut,
    EquipmentMaintenanceUpdate,
    EquipmentPurchaseCreate,
    EquipmentPurchaseOut,
    EquipmentPurchaseUpdate,
    EquipmentPurchaseReportItem,
    EquipmentRentalUpdate,
    EquipmentTransferRequest,
    EquipmentUpdate,
    EquipmentOut,
    EquipmentUsageCreate,
    EquipmentUsageOut,
    EquipmentMaintenanceCreate,
    EquipmentMaintenanceOut,
    EquipmentRentalCreate,
    EquipmentRentalOut,
    EquipmentAuditLogOut,
    AllocationOut,
    EquipmentUsageUpdate,
    UsageReportItem,
    CostReportItem,
    AvailabilityReportItem,
    UtilizationReportItem,
    MaintenanceAlertItem,
)

# Internal - Middleware
from app.middlewares.rate_limiter import default_rate_limiter_dependency

# Utils
from app.utils.helpers import NotFoundError
from app.utils.qr import generate_qr

EQUIPMENT_READ_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.ACCOUNTANT,
        UserRole.CLIENT,
    ]
]

EQUIPMENT_WRITE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
    ]
]

router = APIRouter(
    prefix="/equipment",
    tags=["equipment"],
    dependencies=[default_rate_limiter_dependency()],
)

VERSION_KEY = "cache_version:equipment"


# === UTILITY FUNCTIONS ===
async def get_active_equipment_or_404(db: AsyncSession, equipment_id: int):
    """Get active (not deleted) equipment or 404"""
    stmt = select(Equipment).where(
        and_(Equipment.id == equipment_id, Equipment.is_deleted == False)
    )
    result = await db.execute(stmt)
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail="Equipment not found")
    return obj


async def create_audit_log(
    db: AsyncSession,
    equipment_id: int,
    action: str,
    old_values: Optional[dict] = None,
    new_values: Optional[dict] = None,
    user_id: Optional[int] = None,
    request: Optional[Request] = None,
):
    log = EquipmentAuditLog(
        equipment_id=equipment_id,
        action=action,
        old_values=old_values,
        new_values=new_values,
        user_id=user_id,
        ip_address=request.client.host if request else None,
    )
    db.add(log)


def serialize(data: dict):
    return {
        k: (
            v.isoformat()
            if isinstance(v, (date, datetime))
            else float(v) if isinstance(v, Decimal) else v
        )
        for k, v in data.items()
    }


def safe_parse(value):
    if value is None:
        return None

    if isinstance(value, dict):
        return value

    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return {"raw": value}

    return {"raw": str(value)}


def status_from_row(row):
    today = date.today()

    if row.is_completed:
        return "COMPLETED"
    elif row.next_maintenance_date is None:
        return "NO_SCHEDULE"
    elif row.next_maintenance_date < today:
        return "OVERDUE"
    elif row.next_maintenance_date == today:
        return "TODAY"
    return "UPCOMING"


def convert_decimal(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, dict):
        return {k: convert_decimal(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [convert_decimal(i) for i in obj]
    return obj


def to_decimal(value):
    """Safely convert an incoming numeric (float/int/str/Decimal/None) to Decimal.
    FIX: prevents 'Decimal - Decimal + float' TypeErrors when a Pydantic schema
    field is typed as float but is later combined with Decimal DB columns."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


from datetime import date
from sqlalchemy import exists, or_, select


async def calculate_equipment_status(
    db: AsyncSession,
    equipment: Equipment,
) -> EquipmentStatus:
    """
    Calculate live equipment status without modifying ORM object.
    Safe to use in GET APIs.
    """
    today = date.today()

    if equipment.condition == EquipmentCondition.DAMAGED:
        return EquipmentStatus.DAMAGED

    pending_maintenance = await db.scalar(
        select(
            exists().where(
                EquipmentMaintenance.equipment_id == equipment.id,
                EquipmentMaintenance.is_completed.is_(False),
                EquipmentMaintenance.maintenance_date <= today,
            )
        )
    )

    if pending_maintenance:
        return EquipmentStatus.MAINTENANCE

    active_rental = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment.id,
                EquipmentRental.start_date <= today,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= today,
                ),
            )
        )
    )

    if active_rental:
        return EquipmentStatus.RENTED

    if equipment.project_id:
        return EquipmentStatus.IN_PROJECT

    future_rental = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment.id,
                EquipmentRental.start_date > today,
            )
        )
    )

    if future_rental:
        return EquipmentStatus.IDLE

    return EquipmentStatus.AVAILABLE


async def recalculate_equipment_status(
    db: AsyncSession,
    equipment: Equipment,
):
    """
    NOTE: previously this function computed the DAMAGED and pending-maintenance
    checks TWICE (duplicate blocks). Removed the duplication - behavior is
    unchanged, just cleaned up.
    """
    today = date.today()

    # 1. Damaged check
    if equipment.condition == EquipmentCondition.DAMAGED:
        equipment.status = EquipmentStatus.DAMAGED
        return

    # 2. Pending maintenance check
    pending_maintenance = await db.scalar(
        select(
            exists().where(
                EquipmentMaintenance.equipment_id == equipment.id,
                EquipmentMaintenance.is_completed.is_(False),
                EquipmentMaintenance.maintenance_date <= today,
            )
        )
    )

    if pending_maintenance:
        equipment.status = EquipmentStatus.MAINTENANCE
        return

    # 3. Active Rental
    active_rental = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment.id,
                EquipmentRental.start_date <= today,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= today,
                ),
            )
        )
    )

    if active_rental:
        equipment.status = EquipmentStatus.RENTED
        return

    # 4. Project
    if equipment.project_id:
        equipment.status = EquipmentStatus.IN_PROJECT
        return

    # 5. Future Rental
    future_rental = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment.id,
                EquipmentRental.start_date > today,
            )
        )
    )

    if future_rental:
        equipment.status = EquipmentStatus.IDLE
        return

    equipment.status = await calculate_equipment_status(
        db,
        equipment,
    )


# ============================== EQUIPMENT KPI ========================

MAX_MONTHLY_HOURS = 240


# ============================== EQUIPMENT PURCHASE HISTORY ========================
# Added optional filters + pagination.


@router.get(
    "/purchase/history",
    response_model=List[EquipmentPurchaseOut],
)
async def get_equipment_purchase_history(
    equipment_id: Optional[int] = Query(None),
    purchase_type: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    stmt = select(EquipmentPurchase)

    # Equipment filter (optional)
    if equipment_id is not None:

        await get_active_equipment_or_404(
            db,
            equipment_id,
        )

        stmt = stmt.where(
            EquipmentPurchase.asset_id == equipment_id,
        )

    # Purchase Type filter
    if purchase_type:

        stmt = stmt.where(
            EquipmentPurchase.purchase_type == PurchaseType(purchase_type),
        )

    # Date filters
    if date_from:

        stmt = stmt.where(
            EquipmentPurchase.purchase_date >= date_from,
        )

    if date_to:

        stmt = stmt.where(
            EquipmentPurchase.purchase_date <= date_to,
        )

    stmt = (
        stmt.order_by(
            EquipmentPurchase.purchase_date.desc(),
            EquipmentPurchase.created_at.desc(),
        )
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)

    purchases = result.scalars().all()

    response = []

    for purchase in purchases:

        asset_name = purchase.equipment.equipment_name if purchase.equipment else None

        response.append(
            EquipmentPurchaseOut(
                id=purchase.id,
                project_id=purchase.project_id,
                boq_item_id=purchase.boq_item_id,
                purchase_type=purchase.purchase_type,
                asset_id=purchase.asset_id,
                asset_name=asset_name,
                purchase_date=purchase.purchase_date,
                vendor_name=purchase.vendor_name,
                invoice_number=purchase.invoice_number,
                quantity=purchase.quantity,
                unit_price=float(purchase.unit_price),
                total_amount=float(purchase.total_amount),
                warranty_end_date=purchase.warranty_end_date,
                notes=purchase.notes,
                created_at=purchase.created_at,
            )
        )

    return response


@router.get("/kpi", response_model=EquipmentKPIOut)
async def equipment_kpi(
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    total_equipment = await db.scalar(
        select(func.count()).select_from(Equipment).where(Equipment.is_deleted == False)
    )

    available = await db.scalar(
        select(func.count())
        .select_from(Equipment)
        .where(
            Equipment.status == EquipmentStatus.AVAILABLE,
            Equipment.is_deleted == False,
        )
    )

    allocated = await db.scalar(
        select(func.count())
        .select_from(Equipment)
        .where(
            Equipment.status == EquipmentStatus.IN_PROJECT,
            Equipment.is_deleted == False,
        )
    )

    rented = await db.scalar(
        select(func.count())
        .select_from(Equipment)
        .where(
            Equipment.status == EquipmentStatus.RENTED,
            Equipment.is_deleted == False,
        )
    )

    maintenance = await db.scalar(
        select(func.count())
        .select_from(Equipment)
        .where(
            Equipment.status == EquipmentStatus.MAINTENANCE,
            Equipment.is_deleted == False,
        )
    )

    damaged = await db.scalar(
        select(func.count())
        .select_from(Equipment)
        .where(
            Equipment.condition == EquipmentCondition.DAMAGED,
            Equipment.is_deleted == False,
        )
    )

    total_hours = (
        await db.scalar(
            select(func.sum(EquipmentUsage.working_hours))
            .join(
                Equipment,
                Equipment.id == EquipmentUsage.equipment_id,
            )
            .where(
                Equipment.is_deleted == False,
            )
        )
        or 0
    )

    max_possible_hours = (total_equipment or 0) * MAX_MONTHLY_HOURS

    utilization_rate = (
        (float(total_hours) / max_possible_hours) * 100 if max_possible_hours else 0
    )

    rental_revenue = (
        await db.scalar(
            select(func.sum(EquipmentRental.rental_cost))
            .join(
                Equipment,
                Equipment.id == EquipmentRental.equipment_id,
            )
            .where(
                Equipment.is_deleted == False,
            )
        )
        or 0
    )

    maintenance_cost = (
        await db.scalar(
            select(func.sum(EquipmentMaintenance.cost))
            .join(
                Equipment,
                Equipment.id == EquipmentMaintenance.equipment_id,
            )
            .where(
                Equipment.is_deleted == False,
            )
        )
        or 0
    )

    return EquipmentKPIOut(
        total_equipment=total_equipment or 0,
        available=available or 0,
        allocated=allocated or 0,
        rented=rented or 0,
        maintenance=maintenance or 0,
        damaged=damaged or 0,
        utilization_rate=round(utilization_rate, 2),
        total_rental_revenue=float(rental_revenue),
        total_maintenance_cost=float(maintenance_cost),
    )


# ====================USAGE REPORT====================
# Filters added below are ALL optional (Query default=None) - list behaves
# exactly as before if the caller passes nothing.


@router.get("/usage/report", response_model=List[UsageReportItem])
async def usage_report(
    equipment_id: Optional[int] = Query(
        None, description="Optional: filter by equipment"
    ),
    date_from: Optional[date] = Query(
        None, description="Optional: usage_date >= date_from"
    ),
    date_to: Optional[date] = Query(
        None, description="Optional: usage_date <= date_to"
    ),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    stmt = (
        select(
            EquipmentUsage.equipment_id,
            Equipment.equipment_code,
            func.sum(EquipmentUsage.working_hours).label("total_hours"),
            func.sum(EquipmentUsage.fuel_used).label("total_fuel"),
            func.avg(EquipmentUsage.working_hours).label("avg_hours"),
            func.count().label("usage_count"),
        )
        .join(Equipment)
        .where(Equipment.is_deleted == False)
    )

    if equipment_id:
        stmt = stmt.where(EquipmentUsage.equipment_id == equipment_id)

    if date_from:
        stmt = stmt.where(EquipmentUsage.usage_date >= date_from)

    if date_to:
        stmt = stmt.where(EquipmentUsage.usage_date <= date_to)

    stmt = (
        stmt.group_by(EquipmentUsage.equipment_id, Equipment.equipment_code)
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)

    return [
        UsageReportItem(
            equipment_id=row.equipment_id,
            equipment_code=row.equipment_code,
            total_hours=float(row.total_hours or 0),
            total_fuel=float(row.total_fuel or 0),
            avg_hours=float(row.avg_hours or 0),
            usage_count=int(row.usage_count or 0),
        )
        for row in result.all()
    ]


# ========================== COST REPORT ===========================


@router.get("/cost/report", response_model=List[CostReportItem])
async def cost_report(
    equipment_id: Optional[int] = Query(None),
    date_from: Optional[date] = Query(
        None, description="Optional: rental start_date >= date_from"
    ),
    date_to: Optional[date] = Query(
        None, description="Optional: rental start_date <= date_to"
    ),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    stmt = (
        select(
            EquipmentRental.equipment_id,
            Equipment.equipment_code,
            func.sum(EquipmentRental.rental_cost).label("total_cost"),
            func.count(EquipmentRental.id).label("rental_count"),
            func.sum(
                (
                    func.coalesce(
                        EquipmentRental.end_date,
                        EquipmentRental.start_date,
                    )
                    - EquipmentRental.start_date
                    + 1
                )
            ).label("total_days"),
        )
        .join(
            Equipment,
            Equipment.id == EquipmentRental.equipment_id,
        )
        .where(
            Equipment.is_deleted == False,
        )
    )

    if equipment_id:
        stmt = stmt.where(EquipmentRental.equipment_id == equipment_id)

    if date_from:
        stmt = stmt.where(EquipmentRental.start_date >= date_from)

    if date_to:
        stmt = stmt.where(EquipmentRental.start_date <= date_to)

    stmt = (
        stmt.group_by(
            EquipmentRental.equipment_id,
            Equipment.equipment_code,
        )
        .order_by(
            func.sum(EquipmentRental.rental_cost).desc(),
        )
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)

    response = []

    for row in result.all():

        total_cost = float(row.total_cost or 0)
        rental_count = int(row.rental_count or 0)
        total_days = int(row.total_days or 0)

        avg_cost = total_cost / rental_count if rental_count else 0

        revenue_per_day = total_cost / total_days if total_days else 0

        response.append(
            CostReportItem(
                equipment_id=row.equipment_id,
                equipment_code=row.equipment_code,
                total_cost=round(total_cost, 2),
                rental_count=rental_count,
                avg_cost=round(avg_cost, 2),
                total_days=total_days,
                revenue_per_day=round(revenue_per_day, 2),
            )
        )

    return response


# ============================== PURCHASE REPORT ========================


@router.get(
    "/purchase/report",
    response_model=List[EquipmentPurchaseReportItem],
)
async def purchase_report(
    purchase_type: Optional[str] = Query(None),
    project_id: Optional[int] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    stmt = select(
        EquipmentPurchase.purchase_type,
        EquipmentPurchase.asset_id,
        Equipment.equipment_name.label("asset_name"),
        func.count(EquipmentPurchase.id).label("purchase_count"),
        func.sum(EquipmentPurchase.quantity).label("total_quantity"),
        func.sum(EquipmentPurchase.total_amount).label("total_purchase_amount"),
    ).outerjoin(
        Equipment,
        Equipment.id == EquipmentPurchase.asset_id,
    )

    if purchase_type:
        stmt = stmt.where(
            EquipmentPurchase.purchase_type == PurchaseType(purchase_type)
        )

    if project_id:
        stmt = stmt.where(EquipmentPurchase.project_id == project_id)

    stmt = (
        stmt.group_by(
            EquipmentPurchase.purchase_type,
            EquipmentPurchase.asset_id,
            Equipment.equipment_name,
        )
        .order_by(func.sum(EquipmentPurchase.total_amount).desc())
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)

    rows = result.all()

    return [
        EquipmentPurchaseReportItem(
            purchase_type=row.purchase_type,
            asset_id=row.asset_id,
            asset_name=row.asset_name or "N/A",
            purchase_count=row.purchase_count,
            total_quantity=row.total_quantity or 0,
            total_purchase_amount=float(row.total_purchase_amount or 0),
        )
        for row in rows
    ]


# ===================maintenance_alert=======================


@router.get("/alerts/maintenance", response_model=List[MaintenanceAlertItem])
async def maintenance_alerts(
    equipment_id: Optional[int] = Query(
        None, description="Optional: filter by equipment"
    ),
    days_ahead: int = Query(30, ge=1, le=365, description="Look-ahead window in days"),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):

    today = date.today()
    upcoming_date = today + timedelta(days=days_ahead)

    # Nearest pending maintenance per equipment
    subq_stmt = select(
        EquipmentMaintenance.equipment_id,
        func.min(EquipmentMaintenance.next_maintenance_date).label("next_date"),
    ).where(
        EquipmentMaintenance.next_maintenance_date.isnot(None),
        EquipmentMaintenance.is_completed == False,
    )

    if equipment_id:
        subq_stmt = subq_stmt.where(EquipmentMaintenance.equipment_id == equipment_id)

    subq = subq_stmt.group_by(EquipmentMaintenance.equipment_id).subquery()

    stmt = (
        select(EquipmentMaintenance, Equipment)
        .join(
            subq,
            and_(
                EquipmentMaintenance.equipment_id == subq.c.equipment_id,
                EquipmentMaintenance.next_maintenance_date == subq.c.next_date,
            ),
        )
        .join(
            Equipment,
            Equipment.id == EquipmentMaintenance.equipment_id,
        )
        .where(
            and_(
                EquipmentMaintenance.next_maintenance_date.isnot(None),
                EquipmentMaintenance.is_completed == False,
                Equipment.is_deleted == False,
                # Show all overdue and upcoming maintenance within the window
                EquipmentMaintenance.next_maintenance_date <= upcoming_date,
            )
        )
        .order_by(EquipmentMaintenance.next_maintenance_date.asc())
    )

    result = await db.execute(stmt)
    rows = result.all()

    alerts = []

    for maintenance, equipment in rows:

        days_until = (maintenance.next_maintenance_date - today).days

        if days_until < 0:
            m_status = "OVERDUE"
        elif days_until == 0:
            m_status = "TODAY"
        elif days_until <= 3:
            m_status = "URGENT"
        else:
            m_status = "UPCOMING"

        alerts.append(
            MaintenanceAlertItem(
                equipment_id=equipment.id,
                equipment_code=equipment.equipment_code,
                maintenance_date=maintenance.next_maintenance_date,
                days_until=days_until,
                status=m_status,
            )
        )

    return alerts


# ================== "Availability" =======================


@router.get("/eq/availability", response_model=List[AvailabilityReportItem])
async def availability_report(
    project_id: Optional[int] = Query(None, description="Optional: filter by project"),
    is_available: Optional[bool] = Query(
        None, description="Optional: filter available/unavailable"
    ),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    today = date.today()

    eq_stmt = select(Equipment).where(Equipment.is_deleted == False)

    if project_id:
        eq_stmt = eq_stmt.where(Equipment.project_id == project_id)

    eq_stmt = eq_stmt.limit(limit).offset(offset)

    # Get all active equipments (optionally filtered)
    equipments = (await db.execute(eq_stmt)).scalars().all()

    # Active rentals
    rented_ids = set(
        (
            await db.execute(
                select(EquipmentRental.equipment_id).where(
                    EquipmentRental.start_date <= today,
                    or_(
                        EquipmentRental.end_date.is_(None),
                        EquipmentRental.end_date >= today,
                    ),
                )
            )
        )
        .scalars()
        .all()
    )

    # FIX: compute pending-maintenance set live instead of trusting stored status
    pending_maintenance_ids = set(
        (
            await db.execute(
                select(EquipmentMaintenance.equipment_id).where(
                    EquipmentMaintenance.is_completed.is_(False),
                    EquipmentMaintenance.maintenance_date <= today,
                )
            )
        )
        .scalars()
        .all()
    )

    response = []

    for eq in equipments:

        if eq.condition == EquipmentCondition.DAMAGED:
            eq_status = "DAMAGED"

        elif eq.id in pending_maintenance_ids:
            eq_status = "MAINTENANCE"

        elif eq.id in rented_ids:
            eq_status = "RENTED"

        elif eq.project_id is not None:
            eq_status = "ALLOCATED"

        else:
            eq_status = "AVAILABLE"

        available_flag = eq_status == "AVAILABLE"

        if is_available is not None and available_flag != is_available:
            continue

        response.append(
            AvailabilityReportItem(
                equipment_id=eq.id,
                equipment_code=eq.equipment_code,
                equipment_name=eq.equipment_name,
                is_available=available_flag,
                project_id=eq.project_id,
            )
        )

    return response


# ========== ALLOCATION ===========
@router.post(
    "/allocate",
    response_model=EquipmentAllocateResponse,
)
async def allocate_equipment(
    payload: EquipmentAllocateRequest,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    today = date.today()

    project = await db.get(
        Project,
        payload.project_id,
    )

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found",
        )

    # Prevent allocation to completed project
    if project.end_date and project.end_date < today:
        raise HTTPException(
            status_code=400,
            detail="Cannot allocate equipment to completed project",
        )

    allocated_ids = []
    failed = []

    for equipment_id in payload.equipment_ids:

        obj = await db.scalar(
            select(Equipment).where(
                Equipment.id == equipment_id,
                Equipment.is_deleted == False,
            )
        )

        if not obj:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Equipment not found",
                }
            )
            continue

        # ================= DAMAGED CHECK =================

        if obj.condition == EquipmentCondition.DAMAGED:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Damaged equipment",
                }
            )
            continue

        # ================= RENTAL CHECK =================

        rental_exists = await db.scalar(
            select(
                exists().where(
                    EquipmentRental.equipment_id == equipment_id,
                    or_(
                        # Active rental
                        and_(
                            EquipmentRental.start_date <= today,
                            or_(
                                EquipmentRental.end_date.is_(None),
                                EquipmentRental.end_date >= today,
                            ),
                        ),
                        # Future rental
                        EquipmentRental.start_date > today,
                    ),
                )
            )
        )

        if rental_exists:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Equipment rented or reserved",
                }
            )
            continue

        # ================= MAINTENANCE CHECK =================
        # FIX: was missing `is_completed == False`. Without it, a maintenance
        # record dated in the future but already marked completed would still
        # block allocation, which is wrong - only *pending* maintenance should
        # block allocation.

        maintenance_exists = await db.scalar(
            select(
                exists().where(
                    EquipmentMaintenance.equipment_id == equipment_id,
                    EquipmentMaintenance.maintenance_date >= today,
                    EquipmentMaintenance.is_completed == False,
                )
            )
        )

        if maintenance_exists:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Maintenance scheduled",
                }
            )
            continue

        # ================= SAME PROJECT =================

        if obj.project_id == payload.project_id:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Already allocated to same project",
                }
            )
            continue

        # ================= EXISTING PROJECT =================

        if obj.project_id is not None:

            old_project = await db.get(
                Project,
                obj.project_id,
            )

            if old_project and old_project.end_date and old_project.end_date < today:
                old_project_id = obj.project_id

                obj.project_id = None
                obj.status = EquipmentStatus.AVAILABLE

                await create_audit_log(
                    db=db,
                    equipment_id=obj.id,
                    action="AUTO_DEALLOCATE",
                    old_values={
                        "project_id": old_project_id,
                        "status": EquipmentStatus.IN_PROJECT.value,
                    },
                    new_values={
                        "project_id": None,
                        "status": EquipmentStatus.AVAILABLE.value,
                    },
                    user_id=current_user.id,
                    request=request,
                )

                await db.flush()

            else:
                failed.append(
                    {
                        "equipment_id": equipment_id,
                        "reason": "Already allocated",
                    }
                )
                continue

        # ================= ALLOCATE =================

        old_values = {
            "project_id": obj.project_id,
            "status": obj.status.value if obj.status else None,
        }

        obj.project_id = payload.project_id
        obj.status = EquipmentStatus.IN_PROJECT

        await create_audit_log(
            db=db,
            equipment_id=obj.id,
            action="ALLOCATE",
            old_values=old_values,
            new_values={
                "project_id": payload.project_id,
                "status": EquipmentStatus.IN_PROJECT.value,
            },
            user_id=current_user.id,
            request=request,
        )

        allocated_ids.append(obj.id)

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return EquipmentAllocateResponse(
        equipment_ids=payload.equipment_ids,
        project_id=payload.project_id,
        success_count=len(allocated_ids),
        failed_count=len(failed),
        allocated_ids=allocated_ids,
        failed=failed,
    )


# ================== DEALLOCATE ==================


@router.put(
    "/deallocate",
    response_model=EquipmentDeallocateResponse,
)
async def deallocate_equipment(
    payload: EquipmentDeallocateRequest,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    deallocated_ids = []
    failed = []

    today = date.today()

    for equipment_id in payload.equipment_ids:

        obj = await db.scalar(
            select(Equipment).where(
                Equipment.id == equipment_id,
                Equipment.is_deleted == False,
            )
        )

        if not obj:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Equipment not found",
                }
            )
            continue

        if obj.project_id is None:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Equipment not allocated",
                }
            )
            continue

        if obj.project_id != payload.project_id:
            failed.append(
                {
                    "equipment_id": equipment_id,
                    "reason": "Equipment not allocated to given project",
                }
            )
            continue

        old_values = {
            "project_id": obj.project_id,
            "status": obj.status.value,
        }

        obj.project_id = None

        future_rental = await db.scalar(
            select(
                exists().where(
                    EquipmentRental.equipment_id == equipment_id,
                    EquipmentRental.start_date > today,
                )
            )
        )

        if future_rental:
            obj.status = EquipmentStatus.IDLE
        else:
            obj.status = EquipmentStatus.AVAILABLE

        await create_audit_log(
            db=db,
            equipment_id=obj.id,
            action="DEALLOCATE",
            old_values=old_values,
            new_values={
                "project_id": None,
                "status": obj.status.value,
            },
            user_id=current_user.id,
            request=request,
        )

        deallocated_ids.append(obj.id)

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return EquipmentDeallocateResponse(
        project_id=payload.project_id,
        success_count=len(deallocated_ids),
        failed_count=len(failed),
        deallocated_ids=deallocated_ids,
        failed=failed,
    )


# ================== GET ALLOCATION STATUS ==================


@router.get("/{equipment_id}/allocation", response_model=AllocationOut)
async def get_allocation(
    equipment_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    obj = await get_active_equipment_or_404(db, equipment_id)
    return AllocationOut(
        equipment_id=obj.id,
        project_id=obj.project_id,
        allocated=obj.project_id is not None,
    )


# =========== CREATE EQUIPMENT ====================


@router.post(
    "",
    response_model=EquipmentOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_equipment(
    payload: EquipmentCreate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    # Check duplicate code
    existing = await db.scalar(
        select(Equipment).where(
            and_(
                Equipment.equipment_code == payload.equipment_code,
                Equipment.is_deleted == False,
            )
        )
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Equipment code already exists",
        )

    # Validate project if provided
    if payload.project_id:
        project = await db.get(
            Project,
            payload.project_id,
        )

        if not project:
            raise HTTPException(
                status_code=404,
                detail="Project not found",
            )

    obj = Equipment(**payload.model_dump())

    db.add(obj)

    await db.flush()

    # Auto set status based on project/rental/condition
    await recalculate_equipment_status(
        db,
        obj,
    )

    await create_audit_log(
        db=db,
        equipment_id=obj.id,
        action="CREATE",
        new_values=jsonable_encoder(payload.model_dump()),
        user_id=current_user.id,
    )

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    await db.refresh(obj)

    return EquipmentOut.model_validate(obj)


# ================== LIST EQUIPMENT ==================
# All filters below are optional - none are mandatory.


@router.get("", response_model=PaginatedResponse[EquipmentOut])
async def list_equipment(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
    project_id: Optional[int] = None,
    condition: Optional[str] = None,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    version = await get_cache_version(redis, VERSION_KEY)

    cache_key = (
        f"equipment_list:{version}:{limit}:{offset}:"
        f"{search or ''}:{project_id}:{condition or ''}"
    )

    cached = await cache_get_json(redis, cache_key)

    if cached:
        return PaginatedResponse[EquipmentOut](**cached)

    query = select(Equipment).where(Equipment.is_deleted.is_(False))

    count_query = select(func.count(Equipment.id)).where(
        Equipment.is_deleted.is_(False)
    )

    # ================= SEARCH (optional) =================

    if search:
        query = query.where(Equipment.equipment_name.ilike(f"%{search}%"))

        count_query = count_query.where(Equipment.equipment_name.ilike(f"%{search}%"))

    # ================= PROJECT FILTER (optional) =================

    if project_id:
        query = query.where(Equipment.project_id == project_id)

        count_query = count_query.where(Equipment.project_id == project_id)

    # ================= CONDITION FILTER (optional) =================

    if condition:

        try:
            enum_condition = EquipmentCondition(condition.upper())

        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Invalid condition '{condition}'. "
                    f"Allowed values: "
                    f"{', '.join([c.value for c in EquipmentCondition])}"
                ),
            )

        query = query.where(Equipment.condition == enum_condition)

        count_query = count_query.where(Equipment.condition == enum_condition)

    # ================= PAGINATION =================

    query = query.order_by(Equipment.created_at.desc()).limit(limit).offset(offset)

    result = await db.execute(query)

    items = [EquipmentOut.model_validate(row[0]) for row in result.all()]

    total = await db.scalar(count_query)

    response = PaginatedResponse[EquipmentOut](
        items=[item.model_dump() for item in items],
        meta=PaginationMeta(
            total=total or 0,
            limit=limit,
            offset=offset,
        ),
    ).model_dump()

    await cache_set_json(
        redis,
        cache_key,
        response,
    )

    return PaginatedResponse[EquipmentOut].model_validate(response)


# ================== SOFT DELETE ==================


@router.delete("/{equipment_id}", status_code=204)
async def soft_delete_equipment(
    equipment_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    obj = await get_active_equipment_or_404(
        db,
        equipment_id,
    )

    today = date.today()

    # ================= PROJECT VALIDATION =================

    if obj.project_id is not None:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete allocated equipment",
        )

    # ================= MAINTENANCE VALIDATION =================

    if obj.status == EquipmentStatus.MAINTENANCE:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete equipment under maintenance",
        )

    # ================= RENTAL VALIDATION =================

    rental_exists = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment_id,
                or_(
                    # Future rental
                    EquipmentRental.start_date > today,
                    # Active rental
                    and_(
                        EquipmentRental.start_date <= today,
                        or_(
                            EquipmentRental.end_date.is_(None),
                            EquipmentRental.end_date >= today,
                        ),
                    ),
                ),
            )
        )
    )

    if rental_exists:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete rented or reserved equipment",
        )

    # ================= AUDIT OLD VALUES =================

    old_values = serialize(
        {
            "is_deleted": obj.is_deleted,
            "deleted_at": obj.deleted_at,
            "deleted_by": obj.deleted_by,
        }
    )

    # ================= SOFT DELETE =================

    obj.is_deleted = True
    obj.deleted_at = date.today()
    obj.deleted_by = current_user.id

    # ================= AUDIT NEW VALUES =================

    new_values = serialize(
        {
            "is_deleted": obj.is_deleted,
            "deleted_at": obj.deleted_at,
            "deleted_by": obj.deleted_by,
        }
    )

    await create_audit_log(
        db=db,
        equipment_id=obj.id,
        action="SOFT_DELETE",
        old_values=old_values,
        new_values=new_values,
        user_id=current_user.id,
    )

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )


# ================== RESTORE (NEW) ==================
# NEW API: soft_delete_equipment never had an inverse operation - once
# deleted, equipment could never come back without a manual DB edit.


@router.put("/{equipment_id}/restore", response_model=EquipmentOut)
async def restore_equipment(
    equipment_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    obj = await db.get(Equipment, equipment_id)

    if not obj:
        raise HTTPException(status_code=404, detail="Equipment not found")

    if not obj.is_deleted:
        raise HTTPException(status_code=400, detail="Equipment is not deleted")

    # Prevent restoring into a duplicate active equipment_code
    duplicate = await db.scalar(
        select(Equipment).where(
            Equipment.equipment_code == obj.equipment_code,
            Equipment.is_deleted == False,
            Equipment.id != obj.id,
        )
    )

    if duplicate:
        raise HTTPException(
            status_code=400,
            detail="Cannot restore: another active equipment already uses this code",
        )

    old_values = serialize(
        {
            "is_deleted": obj.is_deleted,
            "deleted_at": obj.deleted_at,
            "deleted_by": obj.deleted_by,
        }
    )

    obj.is_deleted = False
    obj.deleted_at = None
    obj.deleted_by = None

    await recalculate_equipment_status(db, obj)

    new_values = serialize(
        {
            "is_deleted": obj.is_deleted,
            "deleted_at": obj.deleted_at,
            "deleted_by": obj.deleted_by,
            "status": obj.status.value if obj.status else None,
        }
    )

    await create_audit_log(
        db=db,
        equipment_id=obj.id,
        action="RESTORE",
        old_values=old_values,
        new_values=new_values,
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await bump_cache_version(redis, VERSION_KEY)

    await db.refresh(obj)

    return EquipmentOut.model_validate(obj)

    # ========================= CREATE USAGE===========================

    equipment = await get_active_equipment_or_404(
        db,
        equipment_id,
    )

    today = date.today()

    # ================= BASIC VALIDATIONS =================

    if equipment.project_id is None:
        raise HTTPException(
            status_code=400,
            detail="Equipment is not allocated to any project",
        )

    if payload.working_hours <= 0 and payload.fuel_used <= 0:
        raise HTTPException(
            status_code=400,
            detail="Usage cannot be zero",
        )

    if payload.usage_date > today:
        raise HTTPException(
            status_code=400,
            detail="Usage date cannot be in future",
        )

    if equipment.condition == EquipmentCondition.DAMAGED:
        raise HTTPException(
            status_code=400,
            detail="Equipment is damaged and cannot be used",
        )

    # ================= BOQ VALIDATION =================

    boq_item = None

    if payload.boq_item_id:

        boq_item = await db.get(
            BOQ,
            payload.boq_item_id,
        )

        if not boq_item:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        if boq_item.project_id != equipment.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ item does not belong to equipment project",
            )

    # ================= DUPLICATE USAGE =================

    usage_exists = await db.scalar(
        select(
            exists().where(
                EquipmentUsage.equipment_id == equipment_id,
                EquipmentUsage.usage_date == payload.usage_date,
            )
        )
    )

    if usage_exists:
        raise HTTPException(
            status_code=400,
            detail="Usage already exists for this date",
        )

    # ================= RENTAL VALIDATION =================

    rental_active = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment_id,
                EquipmentRental.start_date <= payload.usage_date,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= payload.usage_date,
                ),
            )
        )
    )

    if rental_active:
        raise HTTPException(
            status_code=400,
            detail="Equipment is rented. Cannot log usage",
        )

    # ================= MAINTENANCE VALIDATION =================

    maintenance_active = await db.scalar(
        select(
            exists().where(
                EquipmentMaintenance.equipment_id == equipment_id,
                EquipmentMaintenance.maintenance_date == payload.usage_date,
                EquipmentMaintenance.is_completed.is_(False),
            )
        )
    )

    if maintenance_active:
        raise HTTPException(
            status_code=400,
            detail="Equipment is under maintenance",
        )

    # ================= CREATE USAGE =================

    obj = EquipmentUsage(
        equipment_id=equipment_id,
        **payload.model_dump(),
    )

    db.add(obj)

    old_hours = equipment.working_hours or Decimal("0")
    old_fuel = equipment.fuel_used or Decimal("0")

    equipment.working_hours = old_hours + payload.working_hours

    equipment.fuel_used = old_fuel + payload.fuel_used

    await recalculate_equipment_status(
        db,
        equipment,
    )

    # ================= BOQ ACTUAL COST UPDATE =================

    usage_cost = Decimal("0")

    if boq_item:

        usage_cost = payload.working_hours * (equipment.rental_cost or Decimal("0"))

        boq_item.actual_cost = (boq_item.actual_cost or Decimal("0")) + usage_cost

    await db.flush()

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="USAGE_CREATE",
        old_values={
            "working_hours": float(old_hours),
            "fuel_used": float(old_fuel),
        },
        new_values={
            "working_hours": float(equipment.working_hours),
            "fuel_used": float(equipment.fuel_used),
            "usage_hours_added": float(payload.working_hours),
            "fuel_added": float(payload.fuel_used),
            "usage_date": str(payload.usage_date),
            "boq_item_id": payload.boq_item_id,
            "usage_cost": float(usage_cost),
        },
        user_id=current_user.id,
    )

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    await db.refresh(obj)

    return EquipmentUsageOut.model_validate(obj)


# ========================= GET USAGE===========================


@router.get(
    "/usage/{usage_id}",
    response_model=EquipmentUsageOut,
)
async def get_usage(
    usage_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    usage = await db.get(
        EquipmentUsage,
        usage_id,
    )

    if not usage:
        raise HTTPException(
            status_code=404,
            detail="Usage record not found",
        )

    await get_active_equipment_or_404(
        db,
        usage.equipment_id,
    )

    return EquipmentUsageOut(
        id=usage.id,
        boq_item_id=usage.boq_item_id,
        equipment_id=usage.equipment_id,
        working_hours=float(usage.working_hours or 0),
        fuel_used=float(usage.fuel_used or 0),
        usage_date=usage.usage_date,
        notes=usage.notes,
        created_at=usage.created_at,
    )


# ========================= LIST USAGE===========================
# Added optional filters + pagination (previously returned everything,
# unbounded, with no way to narrow the result set).


@router.get(
    "/usage",
    response_model=List[EquipmentUsageOut],
)
async def list_usage(
    equipment_id: Optional[int] = Query(
        None, description="Optional: Equipment ID to filter usage history"
    ),
    date_from: Optional[date] = Query(
        None, description="Optional: usage_date >= date_from"
    ),
    date_to: Optional[date] = Query(
        None, description="Optional: usage_date <= date_to"
    ),
    boq_item_id: Optional[int] = Query(
        None, description="Optional: filter by BOQ item"
    ),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    stmt = select(EquipmentUsage)

    if equipment_id is not None:
        await get_active_equipment_or_404(db, equipment_id)
        stmt = stmt.where(
            EquipmentUsage.equipment_id == equipment_id,
        )

    if date_from:
        stmt = stmt.where(EquipmentUsage.usage_date >= date_from)

    if date_to:
        stmt = stmt.where(EquipmentUsage.usage_date <= date_to)

    if boq_item_id:
        stmt = stmt.where(EquipmentUsage.boq_item_id == boq_item_id)

    stmt = stmt.order_by(EquipmentUsage.usage_date.desc()).limit(limit).offset(offset)

    result = await db.execute(stmt)
    usages = result.scalars().all()

    return [
        EquipmentUsageOut(
            id=row.id,
            boq_item_id=row.boq_item_id,
            equipment_id=row.equipment_id,
            working_hours=float(row.working_hours or 0),
            fuel_used=float(row.fuel_used or 0),
            usage_date=row.usage_date,
            notes=row.notes,
            created_at=row.created_at,
        )
        for row in usages
    ]


# ======================== UPDATE USAGE===========================


@router.put(
    "/usage/{usage_id}",
    response_model=EquipmentUsageOut,
)
async def update_usage(
    usage_id: int,
    payload: EquipmentUsageUpdate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    usage = await db.get(
        EquipmentUsage,
        usage_id,
    )

    if not usage:
        raise HTTPException(
            status_code=404,
            detail="Usage record not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        usage.equipment_id,
    )

    # ================= OLD VALUES =================

    old_usage_hours = usage.working_hours or Decimal("0")
    old_usage_fuel = usage.fuel_used or Decimal("0")

    old_total_hours = equipment.working_hours or Decimal("0")
    old_total_fuel = equipment.fuel_used or Decimal("0")

    old_boq_item_id = usage.boq_item_id

    update_data = payload.model_dump(exclude_unset=True)

    # FIX: previously `working_hours` / `fuel_used` coming from the payload
    # could be plain floats. Combining `Decimal - Decimal + float` raises
    # TypeError. Coerce these fields to Decimal up-front so downstream math
    # is always Decimal-safe, matching the Equipment/EquipmentUsage columns.
    for numeric_field in ("working_hours", "fuel_used"):
        if numeric_field in update_data and update_data[numeric_field] is not None:
            update_data[numeric_field] = to_decimal(update_data[numeric_field])

    # ================= DUPLICATE DATE CHECK =================

    if "usage_date" in update_data and update_data["usage_date"] != usage.usage_date:
        duplicate = await db.scalar(
            select(
                exists().where(
                    EquipmentUsage.equipment_id == usage.equipment_id,
                    EquipmentUsage.usage_date == update_data["usage_date"],
                    EquipmentUsage.id != usage_id,
                )
            )
        )

        if duplicate:
            raise HTTPException(
                status_code=400,
                detail="Usage already exists for this date",
            )

    # ================= APPLY UPDATE =================

    for field, value in update_data.items():
        setattr(usage, field, value)

    # ================= EQUIPMENT TOTALS UPDATE =================

    equipment.working_hours = max(
        Decimal("0"),
        old_total_hours - old_usage_hours + (usage.working_hours or Decimal("0")),
    )

    equipment.fuel_used = max(
        Decimal("0"),
        old_total_fuel - old_usage_fuel + (usage.fuel_used or Decimal("0")),
    )

    # ================= BOQ COST RECALCULATION =================

    rental_rate = equipment.rental_cost or Decimal("0")

    old_usage_cost = old_usage_hours * rental_rate
    new_usage_cost = (usage.working_hours or Decimal("0")) * rental_rate

    new_boq_item_id = usage.boq_item_id

    # ================= BOQ CHANGED =================

    if old_boq_item_id != new_boq_item_id:

        if old_boq_item_id:

            old_boq = await db.get(
                BOQ,
                old_boq_item_id,
            )

            if old_boq:
                old_boq.actual_cost = max(
                    Decimal("0"),
                    (old_boq.actual_cost or Decimal("0")) - old_usage_cost,
                )

        if new_boq_item_id:

            new_boq = await db.get(
                BOQ,
                new_boq_item_id,
            )

            if not new_boq:
                raise HTTPException(
                    status_code=404,
                    detail="BOQ item not found",
                )

            if equipment.project_id != new_boq.project_id:
                raise HTTPException(
                    status_code=400,
                    detail="BOQ item does not belong to equipment project",
                )

            new_boq.actual_cost = (new_boq.actual_cost or Decimal("0")) + new_usage_cost

    # ================= SAME BOQ =================

    elif new_boq_item_id:

        boq_item = await db.get(
            BOQ,
            new_boq_item_id,
        )

        if boq_item:
            boq_item.actual_cost = (
                (boq_item.actual_cost or Decimal("0")) - old_usage_cost + new_usage_cost
            )

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="USAGE_UPDATE",
        old_values={
            "working_hours": float(old_usage_hours),
            "fuel_used": float(old_usage_fuel),
            "boq_item_id": old_boq_item_id,
            "usage_cost": float(old_usage_cost),
        },
        new_values={
            "working_hours": float(usage.working_hours or 0),
            "fuel_used": float(usage.fuel_used or 0),
            "usage_date": str(usage.usage_date),
            "boq_item_id": usage.boq_item_id,
            "usage_cost": float(new_usage_cost),
        },
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await db.refresh(usage)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return EquipmentUsageOut.model_validate(usage)


# ========================= DELETE USAGE===========================


@router.delete(
    "/usage/{usage_id}",
    response_model=DeleteUsageResponse,
    status_code=status.HTTP_200_OK,
)
async def delete_usage(
    usage_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    usage = await db.get(
        EquipmentUsage,
        usage_id,
    )

    if not usage:
        raise HTTPException(
            status_code=404,
            detail="Usage record not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        usage.equipment_id,
    )

    # ================= BOQ COST ROLLBACK =================

    usage_cost = Decimal("0")

    if usage.boq_item_id:

        boq_item = await db.get(
            BOQ,
            usage.boq_item_id,
        )

        if boq_item:

            usage_cost = usage.working_hours * (equipment.rental_cost or Decimal("0"))

            boq_item.actual_cost = max(
                Decimal("0"),
                (boq_item.actual_cost or Decimal("0")) - usage_cost,
            )

    # ================= EQUIPMENT TOTALS UPDATE =================

    equipment.working_hours = max(
        Decimal("0"),
        (equipment.working_hours or Decimal("0"))
        - (usage.working_hours or Decimal("0")),
    )

    equipment.fuel_used = max(
        Decimal("0"),
        (equipment.fuel_used or Decimal("0")) - (usage.fuel_used or Decimal("0")),
    )

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="USAGE_DELETE",
        old_values={
            "usage_id": usage.id,
            "boq_item_id": usage.boq_item_id,
            "working_hours": float(usage.working_hours or 0),
            "fuel_used": float(usage.fuel_used or 0),
            "usage_date": str(usage.usage_date),
            "usage_cost": float(usage_cost),
        },
        new_values={
            "boq_cost_rolled_back": float(usage_cost),
        },
        user_id=current_user.id,
        request=request,
    )

    # ================= DELETE USAGE =================

    await db.delete(usage)

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return DeleteUsageResponse(
        message="Usage deleted successfully",
        usage_id=usage_id,
        equipment_id=equipment.id,
    )


# ==============CREATE MAINTENANCE =============


@router.post(
    "/{equipment_id}/maintenance",
    response_model=EquipmentMaintenanceOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_maintenance(
    equipment_id: int,
    payload: EquipmentMaintenanceCreate,
    request: Request,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    equipment = await get_active_equipment_or_404(
        db,
        equipment_id,
    )

    today = date.today()

    # ================= DATE VALIDATION =================

    if (
        payload.next_maintenance_date
        and payload.next_maintenance_date <= payload.maintenance_date
    ):
        raise HTTPException(
            status_code=400,
            detail="Next maintenance date must be after maintenance date",
        )

    # ================= PROJECT CHECK =================

    if equipment.project_id is not None:
        raise HTTPException(
            status_code=400,
            detail="Equipment is currently allocated to a project",
        )

    # ================= RENTAL VALIDATION =================

    rental_exists = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment_id,
                EquipmentRental.start_date <= payload.maintenance_date,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= payload.maintenance_date,
                ),
            )
        )
    )

    if rental_exists:
        raise HTTPException(
            status_code=400,
            detail="Equipment is currently rented during maintenance date",
        )

    # ================= DUPLICATE MAINTENANCE =================

    maintenance_exists = await db.scalar(
        select(
            exists().where(
                EquipmentMaintenance.equipment_id == equipment_id,
                EquipmentMaintenance.maintenance_date == payload.maintenance_date,
            )
        )
    )

    if maintenance_exists:
        raise HTTPException(
            status_code=400,
            detail="Maintenance already exists for this date",
        )

    # ================= BOQ VALIDATION =================

    boq_item = None

    if payload.boq_item_id:

        boq_item = await db.get(
            BOQ,
            payload.boq_item_id,
        )

        if not boq_item:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        if boq_item.project_id != payload.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ item does not belong to selected project",
            )

    # ================= CREATE MAINTENANCE =================

    old_status = equipment.status

    obj = EquipmentMaintenance(
        **payload.model_dump(),
        equipment_id=equipment_id,
    )

    db.add(obj)

    await db.flush()

    # ================= BOQ ACTUAL COST UPDATE =================

    maintenance_cost = Decimal(str(payload.cost or 0))

    if boq_item and maintenance_cost > 0:

        boq_item.actual_cost = (boq_item.actual_cost or Decimal("0")) + maintenance_cost

    # ================= STATUS RECALCULATION =================

    await recalculate_equipment_status(
        db,
        equipment,
    )

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment_id,
        action="MAINTENANCE_CREATE",
        old_values={
            "status": old_status.value if old_status else None,
            "project_id": equipment.project_id,
        },
        new_values={
            "maintenance_id": obj.id,
            "description": payload.description,
            "cost": float(payload.cost or 0),
            "maintenance_date": str(payload.maintenance_date),
            "next_maintenance_date": (
                str(payload.next_maintenance_date)
                if payload.next_maintenance_date
                else None
            ),
            "boq_item_id": payload.boq_item_id,
            "status": equipment.status.value,
        },
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    await db.refresh(obj)
    await db.refresh(equipment)

    # ================= RESPONSE STATUS =================

    if obj.is_completed:
        m_status = "COMPLETED"

    elif obj.next_maintenance_date:

        if obj.next_maintenance_date < today:
            m_status = "OVERDUE"

        elif obj.next_maintenance_date == today:
            m_status = "TODAY"

        else:
            m_status = "UPCOMING"

    else:
        m_status = "NO_SCHEDULE"

    return EquipmentMaintenanceOut(
        id=obj.id,
        project_id=obj.project_id,
        boq_item_id=obj.boq_item_id,
        equipment_id=obj.equipment_id,
        description=obj.description,
        maintenance_date=obj.maintenance_date,
        cost=float(obj.cost or 0),
        next_maintenance_date=obj.next_maintenance_date,
        is_completed=obj.is_completed,
        completed_at=obj.completed_at,
        created_at=obj.created_at,
        status=m_status,
    )


# ======================== UPDATE MAINTENANCE =========================


@router.put(
    "/maintenance/{maintenance_id}",
    response_model=EquipmentMaintenanceOut,
)
async def update_maintenance(
    maintenance_id: int,
    payload: EquipmentMaintenanceUpdate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    maintenance = await db.get(
        EquipmentMaintenance,
        maintenance_id,
    )

    if not maintenance:
        raise HTTPException(
            status_code=404,
            detail="Maintenance record not found",
        )

    # ================= SAVE OLD VALUES =================

    old_cost = maintenance.cost or Decimal("0")
    old_boq_id = maintenance.boq_item_id

    update_data = payload.model_dump(exclude_unset=True)

    # ================= VALIDATE BEFORE MUTATING =================

    prospective_maintenance_date = update_data.get(
        "maintenance_date",
        maintenance.maintenance_date,
    )

    prospective_next_date = update_data.get(
        "next_maintenance_date",
        maintenance.next_maintenance_date,
    )

    if (
        prospective_next_date
        and prospective_maintenance_date
        and prospective_next_date <= prospective_maintenance_date
    ):
        raise HTTPException(
            status_code=400,
            detail="Next maintenance date must be after maintenance date",
        )

    # ================= APPLY UPDATE =================

    old_values = {}

    for field, value in update_data.items():
        old_values[field] = getattr(maintenance, field)
        setattr(maintenance, field, value)

    # ================= BOQ RECALCULATION =================

    new_cost = maintenance.cost or Decimal("0")
    new_boq_id = maintenance.boq_item_id

    if old_boq_id != new_boq_id:

        if old_boq_id:

            old_boq = await db.get(
                BOQ,
                old_boq_id,
            )

            if old_boq:
                old_boq.actual_cost = max(
                    Decimal("0"),
                    (old_boq.actual_cost or Decimal("0")) - old_cost,
                )

        if new_boq_id:

            new_boq = await db.get(
                BOQ,
                new_boq_id,
            )

            if not new_boq:
                raise HTTPException(
                    status_code=404,
                    detail="BOQ item not found",
                )

            new_boq.actual_cost = (new_boq.actual_cost or Decimal("0")) + new_cost

    elif new_boq_id:

        boq_item = await db.get(
            BOQ,
            new_boq_id,
        )

        if boq_item:
            boq_item.actual_cost = (
                (boq_item.actual_cost or Decimal("0")) - old_cost + new_cost
            )

    # ================= STATUS RECALCULATION =================

    equipment = await get_active_equipment_or_404(
        db,
        maintenance.equipment_id,
    )

    await recalculate_equipment_status(
        db,
        equipment,
    )

    # ================= AUDIT =================

    await create_audit_log(
        db=db,
        equipment_id=maintenance.equipment_id,
        action="MAINTENANCE_UPDATE",
        old_values=jsonable_encoder(old_values),
        new_values=jsonable_encoder(update_data),
        user_id=current_user.id,
        request=request,
    )

    # ================= COMMIT =================

    await db.commit()

    await db.refresh(maintenance)
    await db.refresh(equipment)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    # ================= RESPONSE STATUS =================

    today = date.today()

    if maintenance.is_completed:
        m_status = "COMPLETED"

    elif maintenance.next_maintenance_date:

        if maintenance.next_maintenance_date < today:
            m_status = "OVERDUE"

        elif maintenance.next_maintenance_date == today:
            m_status = "TODAY"

        else:
            m_status = "UPCOMING"

    else:
        m_status = "NO_SCHEDULE"

    return EquipmentMaintenanceOut(
        id=maintenance.id,
        project_id=maintenance.project_id,
        boq_item_id=maintenance.boq_item_id,
        equipment_id=maintenance.equipment_id,
        description=maintenance.description,
        maintenance_date=maintenance.maintenance_date,
        cost=float(maintenance.cost or 0),
        next_maintenance_date=maintenance.next_maintenance_date,
        created_at=maintenance.created_at,
        status=m_status,
        is_completed=maintenance.is_completed,
        completed_at=maintenance.completed_at,
    )


# ===================== COMPLETE MAINTENANCE =====================


@router.put(
    "/maintenance/{maintenance_id}/complete",
    response_model=EquipmentMaintenanceOut,
)
async def complete_maintenance(
    maintenance_id: int,
    request: Request,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    maintenance = await db.get(
        EquipmentMaintenance,
        maintenance_id,
    )

    if not maintenance:
        raise HTTPException(
            status_code=404,
            detail="Maintenance record not found",
        )

    if maintenance.is_completed:
        raise HTTPException(
            status_code=400,
            detail="Maintenance already completed",
        )

    equipment = await get_active_equipment_or_404(
        db,
        maintenance.equipment_id,
    )

    old_status = equipment.status

    # Complete maintenance
    maintenance.is_completed = True
    maintenance.completed_at = datetime.utcnow()

    # Flush changes before checking status
    await db.flush()

    # Recalculate equipment status
    await recalculate_equipment_status(
        db,
        equipment,
    )

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="MAINTENANCE_COMPLETE",
        old_values={
            "status": old_status.value if old_status else None,
            "is_completed": False,
        },
        new_values={
            "status": equipment.status.value,
            "is_completed": True,
            "completed_at": str(maintenance.completed_at),
        },
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await db.refresh(maintenance)
    await db.refresh(equipment)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return EquipmentMaintenanceOut(
        id=maintenance.id,
        project_id=maintenance.project_id,
        boq_item_id=maintenance.boq_item_id,
        equipment_id=maintenance.equipment_id,
        description=maintenance.description,
        maintenance_date=maintenance.maintenance_date,
        cost=float(maintenance.cost or 0),
        next_maintenance_date=maintenance.next_maintenance_date,
        is_completed=maintenance.is_completed,
        completed_at=maintenance.completed_at,
        created_at=maintenance.created_at,
        status=equipment.status.value,
    )


# ======================== DELETE MAINTENANCE =====================


@router.delete(
    "/maintenance/{maintenance_id}",
    status_code=status.HTTP_200_OK,
)
async def delete_maintenance(
    maintenance_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    maintenance = await db.get(
        EquipmentMaintenance,
        maintenance_id,
    )

    if not maintenance:
        raise HTTPException(
            status_code=404,
            detail="Maintenance record not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        maintenance.equipment_id,
    )

    # ================= BOQ COST ROLLBACK =================

    if maintenance.boq_item_id and maintenance.cost:

        boq_item = await db.get(
            BOQ,
            maintenance.boq_item_id,
        )

        if boq_item:

            boq_item.actual_cost = max(
                Decimal("0"),
                (boq_item.actual_cost or Decimal("0"))
                - (maintenance.cost or Decimal("0")),
            )

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="MAINTENANCE_DELETE",
        old_values={
            "maintenance_id": maintenance.id,
            "boq_item_id": maintenance.boq_item_id,
            "description": maintenance.description,
            "maintenance_date": str(maintenance.maintenance_date),
            "cost": float(maintenance.cost or 0),
        },
        user_id=current_user.id,
        request=request,
    )

    # ================= DELETE =================

    await db.delete(maintenance)

    # ================= STATUS RECALCULATE =================

    await recalculate_equipment_status(
        db,
        equipment,
    )

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return {
        "message": "Maintenance deleted successfully",
        "maintenance_id": maintenance_id,
        "equipment_id": equipment.id,
        "boq_cost_rolled_back": float(maintenance.cost or 0),
    }


# ======================== GET MAINTENANCE =====================


@router.get(
    "/maintenance/{maintenance_id}",
    response_model=EquipmentMaintenanceOut,
)
async def get_maintenance(
    maintenance_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    maintenance = await db.get(
        EquipmentMaintenance,
        maintenance_id,
    )

    if not maintenance:
        raise HTTPException(
            status_code=404,
            detail="Maintenance record not found",
        )

    await get_active_equipment_or_404(
        db,
        maintenance.equipment_id,
    )

    return EquipmentMaintenanceOut(
        id=maintenance.id,
        project_id=maintenance.project_id,
        boq_item_id=maintenance.boq_item_id,
        equipment_id=maintenance.equipment_id,
        description=maintenance.description,
        maintenance_date=maintenance.maintenance_date,
        cost=float(maintenance.cost or 0),
        next_maintenance_date=maintenance.next_maintenance_date,
        is_completed=maintenance.is_completed,
        completed_at=maintenance.completed_at,
        created_at=maintenance.created_at,
        status=status_from_row(maintenance),
    )


# ===================== LIST MAINTENANCE HISTORY =====================
# Added optional filters + pagination.


@router.get(
    "/maintenance",
    response_model=List[EquipmentMaintenanceOut],
)
async def list_maintenance(
    equipment_id: Optional[int] = Query(
        None, description="Optional: filter by equipment"
    ),
    is_completed: Optional[bool] = Query(
        None, description="Optional: filter by completion status"
    ),
    date_from: Optional[date] = Query(
        None, description="Optional: maintenance_date >= date_from"
    ),
    date_to: Optional[date] = Query(
        None, description="Optional: maintenance_date <= date_to"
    ),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    stmt = select(EquipmentMaintenance)

    # Optional equipment filter
    if equipment_id is not None:
        await get_active_equipment_or_404(db, equipment_id)
        stmt = stmt.where(EquipmentMaintenance.equipment_id == equipment_id)

    if is_completed is not None:
        stmt = stmt.where(EquipmentMaintenance.is_completed == is_completed)

    if date_from:
        stmt = stmt.where(EquipmentMaintenance.maintenance_date >= date_from)

    if date_to:
        stmt = stmt.where(EquipmentMaintenance.maintenance_date <= date_to)

    stmt = (
        stmt.order_by(EquipmentMaintenance.maintenance_date.desc())
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)
    maintenances = result.scalars().all()

    return [
        EquipmentMaintenanceOut(
            id=row.id,
            project_id=row.project_id,
            boq_item_id=row.boq_item_id,
            equipment_id=row.equipment_id,
            description=row.description,
            maintenance_date=row.maintenance_date,
            cost=float(row.cost or 0),
            next_maintenance_date=row.next_maintenance_date,
            is_completed=row.is_completed,
            completed_at=row.completed_at,
            created_at=row.created_at,
            status=status_from_row(row),
        )
        for row in maintenances
    ]


# ======================= RENTAL ====================


@router.post(
    "/{equipment_id}/rental",
    response_model=EquipmentRentalOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_rental(
    equipment_id: int,
    payload: EquipmentRentalCreate,
    request: Request,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    equipment = await get_active_equipment_or_404(
        db,
        equipment_id,
    )

    today = date.today()

    start_date = payload.start_date
    end_date = payload.end_date or payload.start_date

    # ================= DATE VALIDATION =================

    if end_date < start_date:
        raise HTTPException(
            status_code=400,
            detail="End date cannot be before start date",
        )

    # ================= RENTAL COST =================

    if payload.rental_cost <= 0:
        raise HTTPException(
            status_code=400,
            detail="Rental cost must be greater than 0",
        )

    # ================= DAMAGED CHECK =================

    if equipment.condition == EquipmentCondition.DAMAGED:
        raise HTTPException(
            status_code=400,
            detail="Damaged equipment cannot be rented",
        )

    # ================= MAINTENANCE STATUS CHECK =================

    if equipment.status == EquipmentStatus.MAINTENANCE:
        raise HTTPException(
            status_code=400,
            detail="Equipment is under maintenance",
        )

    # ================= MAINTENANCE OVERLAP =================

    maintenance_exists = await db.scalar(
        select(
            exists().where(
                EquipmentMaintenance.equipment_id == equipment_id,
                EquipmentMaintenance.maintenance_date >= start_date,
                EquipmentMaintenance.maintenance_date <= end_date,
                EquipmentMaintenance.is_completed == False,
            )
        )
    )

    if maintenance_exists:
        raise HTTPException(
            status_code=400,
            detail="Equipment maintenance scheduled during rental period",
        )

    # ================= PROJECT ALLOCATION CHECK =================

    if equipment.project_id is not None:

        project = await db.get(
            Project,
            equipment.project_id,
        )

        if project:

            if project.end_date and project.end_date < today:

                old_project_id = equipment.project_id

                equipment.project_id = None
                equipment.status = EquipmentStatus.AVAILABLE

                await create_audit_log(
                    db=db,
                    equipment_id=equipment.id,
                    action="AUTO_DEALLOCATE",
                    old_values={
                        "project_id": old_project_id,
                        "status": EquipmentStatus.IN_PROJECT.value,
                    },
                    new_values={
                        "project_id": None,
                        "status": EquipmentStatus.AVAILABLE.value,
                    },
                    user_id=current_user.id,
                    request=request,
                )

                await db.flush()

            else:
                raise HTTPException(
                    status_code=400,
                    detail="Equipment is currently allocated to an active project",
                )

        else:
            raise HTTPException(
                status_code=400,
                detail="Equipment is allocated to a project",
            )

    # ================= RENTAL OVERLAP CHECK =================

    overlap_exists = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment_id,
                EquipmentRental.start_date <= end_date,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= start_date,
                ),
            )
        )
    )

    if overlap_exists:
        raise HTTPException(
            status_code=400,
            detail="Equipment already rented during this period",
        )

    # ================= CREATE RENTAL =================

    rental = EquipmentRental(
        project_id=payload.project_id,
        boq_item_id=payload.boq_item_id,
        equipment_id=equipment_id,
        start_date=start_date,
        end_date=end_date,
        rental_cost=payload.rental_cost,
        client_name=payload.client_name,
        notes=payload.notes,
    )

    db.add(rental)

    # ================= BOQ VALIDATION & COST UPDATE =================

    boq_item = None

    if payload.boq_item_id:

        boq_item = await db.get(
            BOQ,
            payload.boq_item_id,
        )

        if not boq_item:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        if payload.project_id and boq_item.project_id != payload.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ item does not belong to project",
            )

        boq_item.actual_cost = (
            boq_item.actual_cost or Decimal("0")
        ) + payload.rental_cost

    old_status = equipment.status

    # ================= STATUS UPDATE =================

    if start_date <= today <= end_date:
        equipment.status = EquipmentStatus.RENTED

    elif start_date > today:
        equipment.status = EquipmentStatus.IDLE

    elif equipment.project_id:
        equipment.status = EquipmentStatus.IN_PROJECT

    else:
        equipment.status = EquipmentStatus.AVAILABLE

    await db.flush()

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="RENTAL_CREATE",
        old_values={
            "status": old_status.value if old_status else None,
        },
        new_values={
            "start_date": str(start_date),
            "end_date": str(end_date),
            "rental_cost": float(payload.rental_cost),
            "client_name": payload.client_name,
            "status": equipment.status.value,
        },
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    # ================= CACHE VERSION =================

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    await db.refresh(rental)

    # ================= RESPONSE STATUS =================

    if start_date > today:
        rental_status = "UPCOMING"

    elif end_date < today:
        rental_status = "COMPLETED"

    else:
        rental_status = "ACTIVE"

    duration = (end_date - start_date).days + 1

    per_day_cost = float(rental.rental_cost) / duration if duration > 0 else 0

    return EquipmentRentalOut(
        id=rental.id,
        project_id=rental.project_id,
        boq_item_id=rental.boq_item_id,
        equipment_id=rental.equipment_id,
        start_date=rental.start_date,
        end_date=rental.end_date,
        rental_cost=float(rental.rental_cost),
        client_name=rental.client_name,
        notes=rental.notes,
        created_at=rental.created_at,
        status=rental_status,
        duration=duration,
        per_day_cost=round(per_day_cost, 2),
    )


# ========================== RENTAL LIST ===========================
# Added optional status/date filters + pagination.


@router.get(
    "/rental",
    response_model=List[EquipmentRentalOut],
)
async def list_rental(
    equipment_id: Optional[int] = Query(
        None, description="Optional: filter by equipment"
    ),
    rental_status: Optional[str] = Query(
        None,
        description="Optional: ACTIVE | UPCOMING | COMPLETED",
    ),
    date_from: Optional[date] = Query(
        None, description="Optional: start_date >= date_from"
    ),
    date_to: Optional[date] = Query(
        None, description="Optional: start_date <= date_to"
    ),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    today = date.today()

    stmt = select(EquipmentRental)

    # Optional equipment filter
    if equipment_id is not None:
        await get_active_equipment_or_404(db, equipment_id)
        stmt = stmt.where(EquipmentRental.equipment_id == equipment_id)

    if date_from:
        stmt = stmt.where(EquipmentRental.start_date >= date_from)

    if date_to:
        stmt = stmt.where(EquipmentRental.start_date <= date_to)

    if rental_status:
        normalized = rental_status.upper()

        if normalized not in {"ACTIVE", "UPCOMING", "COMPLETED"}:
            raise HTTPException(
                status_code=400,
                detail="Invalid rental_status. Allowed: ACTIVE, UPCOMING, COMPLETED",
            )

        if normalized == "UPCOMING":
            stmt = stmt.where(EquipmentRental.start_date > today)

        elif normalized == "COMPLETED":
            stmt = stmt.where(
                EquipmentRental.end_date.isnot(None),
                EquipmentRental.end_date < today,
            )

        else:  # ACTIVE
            stmt = stmt.where(
                EquipmentRental.start_date <= today,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= today,
                ),
            )

    stmt = (
        stmt.order_by(
            EquipmentRental.start_date.desc(),
            EquipmentRental.created_at.desc(),
        )
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)
    rentals = result.scalars().all()

    return [
        EquipmentRentalOut(
            id=rental.id,
            project_id=rental.project_id,
            boq_item_id=rental.boq_item_id,
            equipment_id=rental.equipment_id,
            start_date=rental.start_date,
            end_date=rental.end_date,
            rental_cost=float(rental.rental_cost or 0),
            client_name=rental.client_name,
            notes=rental.notes,
            created_at=rental.created_at,
            status=(
                "UPCOMING"
                if rental.start_date > today
                else (
                    "COMPLETED"
                    if (rental.end_date or rental.start_date) < today
                    else "ACTIVE"
                )
            ),
            duration=((rental.end_date or rental.start_date) - rental.start_date).days
            + 1,
            per_day_cost=round(
                float(rental.rental_cost or 0)
                / (
                    ((rental.end_date or rental.start_date) - rental.start_date).days
                    + 1
                ),
                2,
            ),
        )
        for rental in rentals
    ]


# =============================== RENTAL GET ========================


@router.get(
    "/rental/{rental_id}",
    response_model=EquipmentRentalOut,
)
async def get_rental(
    rental_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    rental = await db.get(
        EquipmentRental,
        rental_id,
    )

    if not rental:
        raise HTTPException(
            status_code=404,
            detail="Rental not found",
        )

    today = date.today()

    end_date = rental.end_date or rental.start_date

    duration = (end_date - rental.start_date).days + 1

    # Rental Status
    if rental.start_date > today:
        rental_status = "UPCOMING"

    elif end_date < today:
        rental_status = "COMPLETED"

    else:
        rental_status = "ACTIVE"

    per_day_cost = float(rental.rental_cost) / duration if duration > 0 else 0

    return EquipmentRentalOut(
        id=rental.id,
        project_id=rental.project_id,
        boq_item_id=rental.boq_item_id,
        equipment_id=rental.equipment_id,
        start_date=rental.start_date,
        end_date=rental.end_date,
        rental_cost=float(rental.rental_cost),
        client_name=rental.client_name,
        notes=rental.notes,
        created_at=rental.created_at,
        status=rental_status,
        duration=duration,
        per_day_cost=round(per_day_cost, 2),
    )


# ============================== RENTAL UPDATE ========================
# FIX: create_rental validates that a BOQ item belongs to the target project,
# but update_rental didn't repeat that check when boq_item_id changed. A
# caller could attach a rental to a BOQ item from an unrelated project. Added
# the same project-membership validation here.


@router.put(
    "/rental/{rental_id}",
    response_model=EquipmentRentalOut,
)
async def update_rental(
    rental_id: int,
    payload: EquipmentRentalUpdate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    rental = await db.get(
        EquipmentRental,
        rental_id,
    )

    if not rental:
        raise HTTPException(
            status_code=404,
            detail="Rental not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        rental.equipment_id,
    )

    # ================= OLD VALUES =================

    old_boq_item_id = rental.boq_item_id
    old_rental_cost = rental.rental_cost

    update_data = payload.model_dump(exclude_unset=True)

    # ================= VALIDATE BEFORE MUTATING =================

    if rental.end_date and rental.end_date < date.today():
        raise HTTPException(
            status_code=400,
            detail="Completed rental cannot be updated",
        )

    prospective_start = update_data.get(
        "start_date",
        rental.start_date,
    )

    prospective_end = (
        update_data.get(
            "end_date",
            rental.end_date,
        )
        or prospective_start
    )

    if prospective_end < prospective_start:
        raise HTTPException(
            status_code=400,
            detail="End date cannot be before start date",
        )

    overlap_exists = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == rental.equipment_id,
                EquipmentRental.id != rental.id,
                EquipmentRental.start_date <= prospective_end,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= prospective_start,
                ),
            )
        )
    )

    if overlap_exists:
        raise HTTPException(
            status_code=400,
            detail="Rental overlap found",
        )

    # ================= APPLY UPDATE =================

    old_values = {}

    for field, value in update_data.items():
        old_values[field] = getattr(rental, field)
        setattr(rental, field, value)

    start_date = rental.start_date
    end_date = rental.end_date or rental.start_date

    # ================= BOQ COST UPDATE =================

    if old_boq_item_id:

        old_boq = await db.get(
            BOQ,
            old_boq_item_id,
        )

        if old_boq:
            old_boq.actual_cost = max(
                Decimal("0"),
                (old_boq.actual_cost or Decimal("0")) - old_rental_cost,
            )

    if rental.boq_item_id:

        new_boq = await db.get(
            BOQ,
            rental.boq_item_id,
        )

        if not new_boq:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        # FIX: added missing project-membership validation (present in
        # create_rental but absent here before this fix).
        if rental.project_id and new_boq.project_id != rental.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ item does not belong to project",
            )

        new_boq.actual_cost = (new_boq.actual_cost or Decimal("0")) + rental.rental_cost

    # ================= STATUS =================

    await recalculate_equipment_status(
        db,
        equipment,
    )

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="RENTAL_UPDATE",
        old_values=jsonable_encoder(old_values),
        new_values=jsonable_encoder(update_data),
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await db.refresh(rental)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    today = date.today()

    if rental.start_date > today:
        rental_status = "UPCOMING"

    elif rental.end_date and rental.end_date < today:
        rental_status = "COMPLETED"

    else:
        rental_status = "ACTIVE"

    duration = ((rental.end_date or rental.start_date) - rental.start_date).days + 1

    return EquipmentRentalOut(
        id=rental.id,
        project_id=rental.project_id,
        boq_item_id=rental.boq_item_id,
        equipment_id=rental.equipment_id,
        start_date=rental.start_date,
        end_date=rental.end_date,
        rental_cost=float(rental.rental_cost),
        client_name=rental.client_name,
        notes=rental.notes,
        created_at=rental.created_at,
        status=rental_status,
        duration=duration,
        per_day_cost=round(
            float(rental.rental_cost) / duration,
            2,
        ),
    )


# =============================== RENTAL DELETE ========================


@router.delete(
    "/rental/{rental_id}",
    response_model=DeleteRentalResponse,
    status_code=status.HTTP_200_OK,
)
async def delete_rental(
    rental_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    rental = await db.get(
        EquipmentRental,
        rental_id,
    )

    if not rental:
        raise HTTPException(
            status_code=404,
            detail="Rental not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        rental.equipment_id,
    )

    old_status = equipment.status

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="RENTAL_DELETE",
        old_values={
            "rental_id": rental.id,
            "start_date": str(rental.start_date),
            "end_date": (str(rental.end_date) if rental.end_date else None),
            "rental_cost": float(rental.rental_cost or 0),
            "client_name": rental.client_name,
            "boq_item_id": rental.boq_item_id,
            "status": (old_status.value if old_status else None),
        },
        user_id=current_user.id,
        request=request,
    )

    # ================= BOQ COST ROLLBACK =================

    if rental.boq_item_id:

        boq_item = await db.get(
            BOQ,
            rental.boq_item_id,
        )

        if boq_item:

            boq_item.actual_cost = max(
                Decimal("0"),
                (boq_item.actual_cost or Decimal("0"))
                - (rental.rental_cost or Decimal("0")),
            )

    # ================= DELETE RENTAL =================

    await db.delete(rental)

    # Flush delete before status recalculation
    await db.flush()

    # ================= STATUS RECALCULATE =================

    await recalculate_equipment_status(
        db,
        equipment,
    )

    # ================= COMMIT =================

    await db.commit()

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return DeleteRentalResponse(
        message="Rental deleted successfully",
        rental_id=rental_id,
        equipment_id=equipment.id,
        equipment_status=equipment.status.value,
    )


# =================complete_rental=============================


@router.put(
    "/rental/{rental_id}/complete",
    response_model=EquipmentRentalOut,
)
async def complete_rental(
    rental_id: int,
    request: Request,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    rental = await db.get(
        EquipmentRental,
        rental_id,
    )

    if not rental:
        raise HTTPException(
            status_code=404,
            detail="Rental not found",
        )

    equipment = await get_active_equipment_or_404(
        db,
        rental.equipment_id,
    )

    today = date.today()

    # ================= VALIDATIONS =================

    if rental.start_date > today:
        raise HTTPException(
            status_code=400,
            detail="Upcoming rental cannot be completed",
        )

    if rental.end_date and rental.end_date <= today:
        raise HTTPException(
            status_code=400,
            detail="Rental already completed",
        )

    old_status = equipment.status
    old_end_date = rental.end_date

    # ================= COMPLETE RENTAL =================

    rental.end_date = today

    await db.flush()

    # ================= STATUS RECALCULATION =================

    await recalculate_equipment_status(
        db,
        equipment,
    )

    # ================= AUDIT =================

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="RENTAL_COMPLETE",
        old_values={
            "status": old_status.value if old_status else None,
            "end_date": (str(old_end_date) if old_end_date else None),
        },
        new_values={
            "status": equipment.status.value,
            "end_date": str(rental.end_date),
        },
        user_id=current_user.id,
        request=request,
    )

    # ================= COMMIT =================

    await db.commit()

    await db.refresh(rental)
    await db.refresh(equipment)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    # ================= RESPONSE =================

    end_date = rental.end_date or rental.start_date

    duration = (end_date - rental.start_date).days + 1

    per_day_cost = (
        round(
            float(rental.rental_cost) / duration,
            2,
        )
        if duration > 0
        else 0
    )

    return EquipmentRentalOut(
        id=rental.id,
        project_id=rental.project_id,
        boq_item_id=rental.boq_item_id,
        equipment_id=rental.equipment_id,
        start_date=rental.start_date,
        end_date=rental.end_date,
        rental_cost=float(rental.rental_cost),
        client_name=rental.client_name,
        notes=rental.notes,
        created_at=rental.created_at,
        status="COMPLETED",
        duration=duration,
        per_day_cost=per_day_cost,
    )


# =================== ADVANCED APIs ============


@router.get("/report/utilization", response_model=List[UtilizationReportItem])
async def utilization_report(
    equipment_id: Optional[int] = Query(
        None, description="Optional: filter by equipment"
    ),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    MAX_HOURS = 26 * 8  # configurable later

    stmt = (
        select(
            Equipment.id.label("equipment_id"),
            Equipment.equipment_code,
            func.coalesce(func.sum(EquipmentUsage.working_hours), 0).label(
                "total_hours"
            ),
        )
        .outerjoin(EquipmentUsage, Equipment.id == EquipmentUsage.equipment_id)
        .where(Equipment.is_deleted == False)
    )

    if equipment_id:
        stmt = stmt.where(Equipment.id == equipment_id)

    stmt = (
        stmt.group_by(Equipment.id, Equipment.equipment_code)
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)
    rows = result.all()

    response = []

    for row in rows:
        total_hours = float(row.total_hours or 0)

        utilization_rate = (total_hours / MAX_HOURS) * 100 if MAX_HOURS else 0

        response.append(
            UtilizationReportItem(
                equipment_id=row.equipment_id,
                equipment_code=row.equipment_code,
                total_hours=round(total_hours, 2),
                utilization_rate=round(utilization_rate, 2),
            )
        )

    return response


# ===================equipment_alerts======================================

WORKING_HOURS_LIMIT = 1000


@router.get("/alerts/equipment", response_model=list[dict])
async def equipment_alerts(
    severity: Optional[str] = Query(None, description="Optional: CRITICAL | HIGH"),
    project_id: Optional[int] = Query(None, description="Optional: filter by project"),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):

    stmt = select(Equipment).where(
        and_(
            Equipment.is_deleted == False,
            or_(
                Equipment.condition == EquipmentCondition.DAMAGED,
                Equipment.working_hours > WORKING_HOURS_LIMIT,
            ),
        )
    )

    if project_id:
        stmt = stmt.where(Equipment.project_id == project_id)

    stmt = stmt.limit(limit).offset(offset)

    result = await db.execute(stmt)
    rows = result.scalars().all()

    alerts = []

    for row in rows:
        issues = []
        recommendation = None

        if row.condition == EquipmentCondition.DAMAGED:
            issues.append({"type": "DAMAGED", "severity": "CRITICAL"})
            recommendation = "Stop usage and repair immediately"

        # OVERUSED
        if row.working_hours and row.working_hours > WORKING_HOURS_LIMIT:
            if row.working_hours > WORKING_HOURS_LIMIT * 1.5:
                item_severity = "CRITICAL"
            else:
                item_severity = "HIGH"

            issues.append(
                {
                    "type": "OVERUSED",
                    "severity": item_severity,
                    "current_hours": float(row.working_hours),
                    "limit": WORKING_HOURS_LIMIT,
                }
            )

            if not recommendation:
                recommendation = "Schedule maintenance soon"

        if severity and not any(i["severity"] == severity.upper() for i in issues):
            continue

        alerts.append(
            {
                "equipment_id": row.id,
                "equipment_code": row.equipment_code,
                "equipment_name": row.equipment_name,
                "project_id": row.project_id,
                "issues": issues,
                "recommendation": recommendation,
            }
        )

    return alerts


# ================== AUDIT LOGS ==================


@router.get("/{equipment_id}/logs")
async def get_audit_logs(
    equipment_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    action: Optional[str] = None,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await get_active_equipment_or_404(db, equipment_id)

    base_query = select(EquipmentAuditLog).where(
        EquipmentAuditLog.equipment_id == equipment_id
    )

    if action:
        base_query = base_query.where(EquipmentAuditLog.action == action)

    count_stmt = select(func.count()).select_from(base_query.subquery())
    total = await db.scalar(count_stmt)

    stmt = (
        base_query.order_by(EquipmentAuditLog.created_at.desc())
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)
    rows = result.scalars().all()

    items = [
        EquipmentAuditLogOut(
            id=row.id,
            equipment_id=row.equipment_id,
            action=row.action,
            old_values=safe_parse(row.old_values) if row.old_values else None,
            new_values=safe_parse(row.new_values) if row.new_values else None,
            user_id=row.user_id,
            ip_address=row.ip_address,
            created_at=row.created_at,
        )
        for row in rows
    ]

    return {
        "items": items,
        "meta": {
            "total": total or 0,
            "limit": limit,
            "offset": offset,
        },
    }


# ========================CREATE PURCHASE APIs ========================


@router.post(
    "/purchase",
    response_model=EquipmentPurchaseOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_purchase(
    payload: EquipmentPurchaseCreate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    # ================= PROJECT VALIDATION =================

    project = await db.get(
        Project,
        payload.project_id,
    )

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found",
        )

    # ================= EQUIPMENT VALIDATION =================

    equipment = None

    if payload.asset_id is not None:

        equipment = await db.get(
            Equipment,
            payload.asset_id,
        )

        if not equipment or equipment.is_deleted:
            raise HTTPException(
                status_code=404,
                detail="Equipment not found",
            )

        if equipment.project_id and equipment.project_id != payload.project_id:
            raise HTTPException(
                status_code=400,
                detail="Equipment belongs to another project",
            )
    # ================= INVOICE VALIDATION =================

    duplicate = await db.scalar(
        select(EquipmentPurchase).where(
            EquipmentPurchase.invoice_number == payload.invoice_number
        )
    )

    if duplicate:
        raise HTTPException(
            status_code=400,
            detail="Invoice number already exists",
        )

    # ================= WARRANTY VALIDATION =================

    if payload.warranty_end_date and payload.warranty_end_date <= payload.purchase_date:
        raise HTTPException(
            status_code=400,
            detail="Warranty end date must be after purchase date",
        )

    # ================= BOQ VALIDATION =================

    boq_item = None

    if payload.boq_item_id:

        boq_item = await db.get(
            BOQ,
            payload.boq_item_id,
        )

        if not boq_item:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        if boq_item.project_id != payload.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ item does not belong to selected project",
            )

    # ================= TOTAL AMOUNT =================

    total_amount = Decimal(payload.quantity) * Decimal(payload.unit_price)

    try:

        # ================= CREATE PURCHASE =================

        purchase = EquipmentPurchase(
            **payload.model_dump(),
            total_amount=total_amount,
        )

        db.add(purchase)

        await db.flush()

        # ================= BOQ UPDATE =================

        if boq_item:

            boq_item.actual_cost = (boq_item.actual_cost or Decimal("0")) + total_amount

            boq_item.variance_cost = (
                boq_item.total_cost or Decimal("0")
            ) - boq_item.actual_cost

        # ================= AUDIT LOG =================

        await create_audit_log(
            db=db,
            equipment_id=payload.asset_id,
            action="PURCHASE_CREATE",
            new_values={
                "purchase_id": purchase.id,
                "project_id": payload.project_id,
                "boq_item_id": payload.boq_item_id,
                "asset_id": payload.asset_id,
                "purchase_type": payload.purchase_type.value,
                "vendor_name": payload.vendor_name,
                "invoice_number": payload.invoice_number,
                "quantity": payload.quantity,
                "unit_price": float(payload.unit_price),
                "total_amount": float(total_amount),
            },
            user_id=current_user.id,
            request=request,
        )

        await db.commit()

    except Exception:
        await db.rollback()
        raise

    await db.refresh(purchase)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return EquipmentPurchaseOut(
        id=purchase.id,
        project_id=purchase.project_id,
        boq_item_id=purchase.boq_item_id,
        purchase_type=purchase.purchase_type,
        asset_id=purchase.asset_id,
        asset_name=equipment.equipment_name if equipment else None,
        purchase_date=purchase.purchase_date,
        vendor_name=purchase.vendor_name,
        invoice_number=purchase.invoice_number,
        quantity=purchase.quantity,
        unit_price=float(purchase.unit_price or 0),
        total_amount=float(purchase.total_amount or 0),
        warranty_end_date=purchase.warranty_end_date,
        notes=purchase.notes,
        created_at=purchase.created_at,
    )


# =============================== PURCHASE LIST ========================
# Already had proper optional filters + pagination - left as is.


@router.get(
    "/purchase",
    response_model=PaginatedResponse[EquipmentPurchaseOut],
)
async def list_purchase(
    purchase_type: Optional[str] = None,
    asset_id: Optional[int] = None,
    project_id: Optional[int] = None,
    boq_item_id: Optional[int] = None,
    vendor_name: Optional[str] = None,
    purchase_date_from: Optional[date] = None,
    purchase_date_to: Optional[date] = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    stmt = (
        select(
            EquipmentPurchase,
            Equipment.equipment_name,
        )
        .outerjoin(
            Equipment,
            Equipment.id == EquipmentPurchase.asset_id,
        )
        .where(
            or_(
                Equipment.id.is_(None),
                Equipment.is_deleted.is_(False),
            )
        )
    )

    count_stmt = (
        select(func.count())
        .select_from(EquipmentPurchase)
        .outerjoin(
            Equipment,
            Equipment.id == EquipmentPurchase.asset_id,
        )
        .where(
            or_(
                Equipment.id.is_(None),
                Equipment.is_deleted.is_(False),
            )
        )
    )

    if purchase_type:
        stmt = stmt.where(EquipmentPurchase.purchase_type == purchase_type)
        count_stmt = count_stmt.where(EquipmentPurchase.purchase_type == purchase_type)

    if asset_id is not None:
        stmt = stmt.where(EquipmentPurchase.asset_id == asset_id)
        count_stmt = count_stmt.where(EquipmentPurchase.asset_id == asset_id)

    if project_id:
        stmt = stmt.where(EquipmentPurchase.project_id == project_id)
        count_stmt = count_stmt.where(EquipmentPurchase.project_id == project_id)

    if boq_item_id:
        stmt = stmt.where(EquipmentPurchase.boq_item_id == boq_item_id)
        count_stmt = count_stmt.where(EquipmentPurchase.boq_item_id == boq_item_id)

    if vendor_name:
        stmt = stmt.where(EquipmentPurchase.vendor_name.ilike(f"%{vendor_name}%"))
        count_stmt = count_stmt.where(
            EquipmentPurchase.vendor_name.ilike(f"%{vendor_name}%")
        )

    if purchase_date_from:
        stmt = stmt.where(EquipmentPurchase.purchase_date >= purchase_date_from)
        count_stmt = count_stmt.where(
            EquipmentPurchase.purchase_date >= purchase_date_from
        )

    if purchase_date_to:
        stmt = stmt.where(EquipmentPurchase.purchase_date <= purchase_date_to)
        count_stmt = count_stmt.where(
            EquipmentPurchase.purchase_date <= purchase_date_to
        )

    stmt = (
        stmt.order_by(
            EquipmentPurchase.purchase_date.desc(),
            EquipmentPurchase.created_at.desc(),
        )
        .limit(limit)
        .offset(offset)
    )

    result = await db.execute(stmt)
    rows = result.all()

    total = await db.scalar(count_stmt)

    items = [
        EquipmentPurchaseOut(
            id=purchase.id,
            project_id=purchase.project_id,
            boq_item_id=purchase.boq_item_id,
            purchase_type=purchase.purchase_type.value,
            asset_id=purchase.asset_id,
            asset_name=equipment_name,
            purchase_date=purchase.purchase_date,
            vendor_name=purchase.vendor_name,
            invoice_number=purchase.invoice_number,
            quantity=purchase.quantity,
            unit_price=float(purchase.unit_price),
            total_amount=float(purchase.total_amount),
            warranty_end_date=purchase.warranty_end_date,
            notes=purchase.notes,
            created_at=purchase.created_at,
        )
        for purchase, equipment_name in rows
    ]

    return PaginatedResponse[EquipmentPurchaseOut](
        items=items,
        meta=PaginationMeta(
            total=total or 0,
            limit=limit,
            offset=offset,
        ),
    )


# =============================== PURCHASE GET ========================


@router.get(
    "/purchase/{purchase_id}",
    response_model=EquipmentPurchaseOut,
)
async def get_purchase(
    purchase_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(
            EquipmentPurchase,
            Equipment.equipment_name,
        )
        .outerjoin(
            Equipment,
            Equipment.id == EquipmentPurchase.asset_id,
        )
        .where(
            EquipmentPurchase.id == purchase_id,
        )
    )

    row = result.first()

    if not row:
        raise HTTPException(
            status_code=404,
            detail="Purchase not found",
        )

    purchase, equipment_name = row

    return EquipmentPurchaseOut(
        id=purchase.id,
        project_id=purchase.project_id,
        boq_item_id=purchase.boq_item_id,
        purchase_type=purchase.purchase_type,
        asset_id=purchase.asset_id,
        asset_name=equipment_name,
        purchase_date=purchase.purchase_date,
        vendor_name=purchase.vendor_name,
        invoice_number=purchase.invoice_number,
        quantity=purchase.quantity,
        unit_price=float(purchase.unit_price or 0),
        total_amount=float(purchase.total_amount or 0),
        warranty_end_date=purchase.warranty_end_date,
        notes=purchase.notes,
        created_at=purchase.created_at,
    )


# =============================== PURCHASE UPDATE ========================


@router.put(
    "/purchase/{purchase_id}",
    response_model=EquipmentPurchaseOut,
)
async def update_purchase(
    purchase_id: int,
    payload: EquipmentPurchaseUpdate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    purchase = await db.get(
        EquipmentPurchase,
        purchase_id,
    )

    if not purchase:
        raise HTTPException(
            status_code=404,
            detail="Purchase not found",
        )

    # ================= OLD VALUES =================

    old_total_amount = purchase.total_amount or Decimal("0")
    old_boq_item_id = purchase.boq_item_id

    # ================= INVOICE VALIDATION =================

    if payload.invoice_number and payload.invoice_number != purchase.invoice_number:
        duplicate = await db.scalar(
            select(EquipmentPurchase).where(
                EquipmentPurchase.invoice_number == payload.invoice_number,
                EquipmentPurchase.id != purchase_id,
            )
        )

        if duplicate:
            raise HTTPException(
                status_code=400,
                detail="Invoice number already exists",
            )

    # ================= UPDATE DATA =================

    update_data = payload.model_dump(exclude_unset=True)

    old_values = {}

    for field, value in update_data.items():
        old_values[field] = getattr(purchase, field)
        setattr(purchase, field, value)

    # ================= WARRANTY VALIDATION =================

    if (
        purchase.warranty_end_date
        and purchase.purchase_date
        and purchase.warranty_end_date <= purchase.purchase_date
    ):
        raise HTTPException(
            status_code=400,
            detail="Warranty end date must be after purchase date",
        )

    # ================= RECALCULATE TOTAL =================

    if purchase.quantity is not None and purchase.unit_price is not None:
        purchase.total_amount = purchase.quantity * purchase.unit_price

    new_total_amount = purchase.total_amount or Decimal("0")
    new_boq_item_id = purchase.boq_item_id

    # ================= BOQ SYNC =================
    # FIX: previously only actual_cost was kept in sync; variance_cost
    # (total_cost - actual_cost) was never recalculated here, unlike
    # create_purchase. Now both stay consistent on update too.

    if old_boq_item_id != new_boq_item_id:

        # Rollback old BOQ
        if old_boq_item_id:

            old_boq = await db.get(
                BOQ,
                old_boq_item_id,
            )

            if old_boq:

                old_boq.actual_cost = max(
                    Decimal("0"),
                    (old_boq.actual_cost or Decimal("0")) - old_total_amount,
                )
                old_boq.variance_cost = (
                    old_boq.total_cost or Decimal("0")
                ) - old_boq.actual_cost

        # Add amount to new BOQ
        if new_boq_item_id:

            new_boq = await db.get(
                BOQ,
                new_boq_item_id,
            )

            if not new_boq:
                raise HTTPException(
                    status_code=404,
                    detail="BOQ item not found",
                )

            if purchase.project_id and new_boq.project_id != purchase.project_id:
                raise HTTPException(
                    status_code=400,
                    detail="BOQ item does not belong to project",
                )

            new_boq.actual_cost = (
                new_boq.actual_cost or Decimal("0")
            ) + new_total_amount
            new_boq.variance_cost = (
                new_boq.total_cost or Decimal("0")
            ) - new_boq.actual_cost

    # Same BOQ -> adjust amount difference only
    elif new_boq_item_id:

        boq_item = await db.get(
            BOQ,
            new_boq_item_id,
        )

        if boq_item:

            boq_item.actual_cost = max(
                Decimal("0"),
                (
                    (boq_item.actual_cost or Decimal("0"))
                    - old_total_amount
                    + new_total_amount
                ),
            )
            boq_item.variance_cost = (
                boq_item.total_cost or Decimal("0")
            ) - boq_item.actual_cost

    # ================= EQUIPMENT =================

    equipment = None

    if purchase.asset_id is not None:
        equipment = await db.get(
            Equipment,
            purchase.asset_id,
        )

    # ================= AUDIT LOG =================

    await create_audit_log(
        db=db,
        equipment_id=purchase.asset_id,
        action="PURCHASE_UPDATE",
        old_values=jsonable_encoder(convert_decimal(old_values)),
        new_values={
            **jsonable_encoder(convert_decimal(update_data)),
            "old_total_amount": float(old_total_amount),
            "new_total_amount": float(new_total_amount),
            "old_boq_item_id": old_boq_item_id,
            "new_boq_item_id": new_boq_item_id,
        },
        user_id=current_user.id,
        request=request,
    )

    # ================= COMMIT =================

    await db.commit()

    await db.refresh(purchase)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    # ================= RESPONSE =================

    return EquipmentPurchaseOut(
        id=purchase.id,
        project_id=purchase.project_id,
        boq_item_id=purchase.boq_item_id,
        purchase_type=purchase.purchase_type.value,
        asset_id=purchase.asset_id,
        asset_name=(equipment.equipment_name if equipment else None),
        purchase_date=purchase.purchase_date,
        vendor_name=purchase.vendor_name,
        invoice_number=purchase.invoice_number,
        quantity=purchase.quantity,
        unit_price=float(purchase.unit_price or 0),
        total_amount=float(purchase.total_amount or 0),
        warranty_end_date=purchase.warranty_end_date,
        notes=purchase.notes,
        created_at=purchase.created_at,
    )


# =============================== PURCHASE DELETE ========================


@router.delete("/purchase/{purchase_id}", status_code=status.HTTP_200_OK)
async def delete_purchase(
    purchase_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    purchase = await db.get(EquipmentPurchase, purchase_id)

    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")

    asset_id = purchase.asset_id
    invoice_number = purchase.invoice_number
    boq_item_id = purchase.boq_item_id

    try:
        # ================= BOQ ROLLBACK (single execution) =================
        if purchase.boq_item_id:
            boq_item = await db.get(BOQ, purchase.boq_item_id)
            if boq_item:
                boq_item.actual_cost = max(
                    Decimal("0"),
                    (boq_item.actual_cost or Decimal("0"))
                    - (purchase.total_amount or Decimal("0")),
                )
                boq_item.variance_cost = (
                    boq_item.total_cost or Decimal("0")
                ) - boq_item.actual_cost

        # ================= AUDIT LOG (single execution) =================
        await create_audit_log(
            db=db,
            equipment_id=asset_id,
            action="PURCHASE_DELETE",
            old_values={
                "purchase_id": purchase.id,
                "asset_id": asset_id,
                "boq_item_id": boq_item_id,
                "invoice_number": invoice_number,
                "vendor_name": purchase.vendor_name,
                "quantity": purchase.quantity,
                "unit_price": float(purchase.unit_price or 0),
                "total_amount": float(purchase.total_amount or 0),
            },
            user_id=current_user.id,
            request=request,
        )

        await db.delete(purchase)
        await db.commit()

    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(redis, VERSION_KEY)

    return {"message": "Purchase deleted successfully"}


# ============================== EQUIPMENT TRANSFER ========================


@router.post("/transfer")
async def transfer_equipment(
    payload: EquipmentTransferRequest,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    equipment = await get_active_equipment_or_404(
        db,
        payload.equipment_id,
    )

    if equipment.condition == EquipmentCondition.DAMAGED:
        raise HTTPException(
            status_code=400,
            detail="Damaged equipment cannot be transferred",
        )

    if equipment.status == EquipmentStatus.MAINTENANCE:
        raise HTTPException(
            status_code=400,
            detail="Equipment under maintenance",
        )

    if equipment.project_id is None:
        raise HTTPException(
            status_code=400,
            detail="Equipment not allocated",
        )

    if equipment.project_id == payload.to_project_id:
        raise HTTPException(
            status_code=400,
            detail="Already allocated to same project",
        )

    project = await db.get(
        Project,
        payload.to_project_id,
    )

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Target project not found",
        )

    today = date.today()

    if project.end_date and project.end_date < today:
        raise HTTPException(
            status_code=400,
            detail="Cannot transfer to completed project",
        )

    rental_exists = await db.scalar(
        select(
            exists().where(
                EquipmentRental.equipment_id == equipment.id,
                EquipmentRental.start_date <= today,
                or_(
                    EquipmentRental.end_date.is_(None),
                    EquipmentRental.end_date >= today,
                ),
            )
        )
    )

    if rental_exists:
        raise HTTPException(
            status_code=400,
            detail="Equipment currently rented",
        )

    old_project = equipment.project_id

    # Transfer equipment
    equipment.project_id = payload.to_project_id

    # Recalculate equipment status
    await recalculate_equipment_status(
        db=db,
        equipment=equipment,
    )

    await create_audit_log(
        db=db,
        equipment_id=equipment.id,
        action="TRANSFER",
        old_values={
            "project_id": old_project,
        },
        new_values={
            "project_id": payload.to_project_id,
            "status": equipment.status.value,
        },
        user_id=current_user.id,
        request=request,
    )

    await db.commit()

    await db.refresh(equipment)

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return {
        "message": "Equipment transferred successfully",
        "equipment_id": equipment.id,
        "from_project": old_project,
        "to_project": payload.to_project_id,
        "status": equipment.status,
    }


# ============================== TRANSFER HISTORY (NEW) ========================


@router.get("/{equipment_id}/transfer-history")
async def get_transfer_history(
    equipment_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await get_active_equipment_or_404(db, equipment_id)

    base_query = select(EquipmentAuditLog).where(
        EquipmentAuditLog.equipment_id == equipment_id,
        EquipmentAuditLog.action == "TRANSFER",
    )

    total = (
        await db.scalar(select(func.count()).select_from(base_query.subquery())) or 0
    )

    result = await db.execute(
        base_query.order_by(EquipmentAuditLog.created_at.desc())
        .offset(offset)
        .limit(limit)
    )

    logs = result.scalars().all()

    project_ids = set()

    for log in logs:
        old_values = safe_parse(log.old_values) if log.old_values else {}
        new_values = safe_parse(log.new_values) if log.new_values else {}

        if old_values.get("project_id"):
            project_ids.add(old_values["project_id"])

        if new_values.get("project_id"):
            project_ids.add(new_values["project_id"])

    project_name_map = {}

    if project_ids:
        projects = await db.execute(
            select(Project.id, Project.project_name).where(Project.id.in_(project_ids))
        )

        project_name_map = dict(projects.all())

    items = []

    for log in logs:
        old_values = safe_parse(log.old_values) if log.old_values else {}
        new_values = safe_parse(log.new_values) if log.new_values else {}

        from_project_id = old_values.get("project_id")
        to_project_id = new_values.get("project_id")

        items.append(
            {
                "id": log.id,
                "equipment_id": log.equipment_id,
                "from_project_id": from_project_id,
                "from_project_name": project_name_map.get(from_project_id),
                "to_project_id": to_project_id,
                "to_project_name": project_name_map.get(to_project_id),
                "transferred_by": log.user_id,
                "transferred_at": log.created_at,
                "ip_address": log.ip_address,
            }
        )

    return {
        "items": items,
        "meta": {
            "total": total,
            "limit": limit,
            "offset": offset,
        },
    }


# ======================= list_transfer_history =================================


@router.get("/transfer-history")
async def list_transfer_history(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    equipment_id: int | None = Query(None),
    project_id: int | None = Query(None),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(EquipmentAuditLog).where(EquipmentAuditLog.action == "TRANSFER")

    if equipment_id:
        query = query.where(EquipmentAuditLog.equipment_id == equipment_id)

    count_stmt = select(func.count()).select_from(query.subquery())
    total = await db.scalar(count_stmt)

    stmt = (
        query.order_by(EquipmentAuditLog.created_at.desc()).limit(limit).offset(offset)
    )

    result = await db.execute(stmt)
    logs = result.scalars().all()

    equipment_ids = set()
    project_ids = set()
    user_ids = set()

    parsed_logs = []

    for log in logs:
        old_values = safe_parse(log.old_values) if log.old_values else {}
        new_values = safe_parse(log.new_values) if log.new_values else {}

        from_project = old_values.get("project_id")
        to_project = new_values.get("project_id")

        if project_id:
            if from_project != project_id and to_project != project_id:
                continue

        equipment_ids.add(log.equipment_id)

        if from_project:
            project_ids.add(from_project)

        if to_project:
            project_ids.add(to_project)

        if log.user_id:
            user_ids.add(log.user_id)

        parsed_logs.append(
            {
                "id": log.id,
                "equipment_id": log.equipment_id,
                "from_project_id": from_project,
                "to_project_id": to_project,
                "transferred_by": log.user_id,
                "transferred_at": log.created_at,
                "ip_address": log.ip_address,
            }
        )

    equipment_map = {}
    if equipment_ids:
        result = await db.execute(
            select(
                Equipment.id,
                Equipment.equipment_name,
            ).where(Equipment.id.in_(equipment_ids))
        )

        equipment_map = {row.id: row.equipment_name for row in result}

    project_map = {}
    if project_ids:
        result = await db.execute(
            select(
                Project.id,
                Project.project_name,
            ).where(Project.id.in_(project_ids))
        )
        project_map = {row.id: row.project_name for row in result}

    user_map = {}
    if user_ids:
        result = await db.execute(
            select(
                User.id,
                User.full_name,
            ).where(User.id.in_(user_ids))
        )
        user_map = {row.id: row.full_name for row in result}

    items = []

    for row in parsed_logs:
        items.append(
            {
                **row,
                "equipment_name": equipment_map.get(row["equipment_id"]),
                "from_project_name": project_map.get(row["from_project_id"]),
                "to_project_name": project_map.get(row["to_project_id"]),
                "transferred_by_name": user_map.get(row["transferred_by"]),
            }
        )

    return {
        "items": items,
        "meta": {
            "total": total or 0,
            "limit": limit,
            "offset": offset,
        },
    }


# ===================== QR CODE ====================================

@router.get(
    "/{equipment_id}/qr",
    summary="Generate Equipment QR Code",
    description="Stateless endpoint that generates an in-memory QR code PNG for the given equipment. No database writes occur.",
    response_class=StreamingResponse,
)
async def generate_equipment_qr(
    equipment_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    # Verify equipment exists using standard helper
    equipment = await get_active_equipment_or_404(db, equipment_id)
    
    # Generate QR in memory (payload: EQP:ID)
    qr_buf = generate_qr(entity_type="EQP", entity_id=equipment.id)
    
    headers = {
        "Cache-Control": "no-store",
        "Content-Disposition": f'inline; filename="equipment_{equipment.id}.png"'
    }
    
    return StreamingResponse(
        qr_buf, 
        media_type="image/png",
        headers=headers
    )


# =====================get_equipment==============================


@router.get("/{equipment_id}", response_model=EquipmentOut)
async def get_equipment(
    equipment_id: int,
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await get_active_equipment_or_404(
        db,
        equipment_id,
    )

    response = EquipmentOut.model_validate(obj)
    response.status = await calculate_equipment_status(
        db,
        obj,
    )

    return response


# =============================update_equipment=======================


@router.put("/{equipment_id}", response_model=EquipmentOut)
async def update_equipment(
    equipment_id: int,
    payload: EquipmentUpdate,
    current_user: User = Depends(require_roles(EQUIPMENT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    request: Request = None,
):
    obj = await get_active_equipment_or_404(db, equipment_id)

    if payload.equipment_code and payload.equipment_code != obj.equipment_code:
        existing = await db.scalar(
            select(Equipment).where(
                and_(
                    Equipment.equipment_code == payload.equipment_code,
                    Equipment.is_deleted == False,
                    Equipment.id != equipment_id,
                )
            )
        )
        if existing:
            raise HTTPException(
                status_code=400,
                detail="Equipment code already exists",
            )

    update_data = payload.model_dump(exclude_unset=True)

    old_data = {c.name: getattr(obj, c.name) for c in obj.__table__.columns}

    # Apply updates
    for field, value in update_data.items():
        setattr(obj, field, value)

    await recalculate_equipment_status(
        db,
        obj,
    )

    if hasattr(obj, "updated_at"):
        obj.updated_at = datetime.utcnow()

    await db.flush()

    changed_fields = {
        k: {"old": old_data.get(k), "new": getattr(obj, k)}
        for k in update_data
        if old_data.get(k) != getattr(obj, k)
    }

    if not changed_fields:
        await db.commit()
        await db.refresh(obj)
        return EquipmentOut.model_validate(obj)

    old_values = {k: v["old"] for k, v in changed_fields.items()}

    new_values = {k: v["new"] for k, v in changed_fields.items()}

    await create_audit_log(
        db,
        obj.id,
        "UPDATE",
        old_values=jsonable_encoder(old_values),
        new_values=jsonable_encoder(new_values),
        user_id=current_user.id,
        request=request,
    )

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    await db.commit()
    await db.refresh(obj)

    return EquipmentOut.model_validate(obj)


# ======================== REPORTS PDF ========================


@router.get("/reports/pdf")
async def equipment_full_pdf_report(
    project_id: Optional[int] = Query(None),
    equipment_id: Optional[int] = Query(None),
    condition: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    vendor_name: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    """
    NOTE: Make sure these are imported at the top of this router file
    (alongside Equipment, EquipmentMaintenance, EquipmentRental):
        from app.models.equipment import EquipmentUsage, EquipmentPurchase
    """
    try:
        import io
        import os
        from datetime import datetime

        from fastapi import HTTPException
        from fastapi.responses import StreamingResponse

        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            Image,
        )
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch, cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        # ================= COLOR PALETTE (matches INFRA PILOT sample) =====
        NAVY_BLUE = colors.HexColor("#0B2B5C")
        LIGHT_GRAY = colors.HexColor("#F8F9FA")
        BORDER_GRAY = colors.HexColor("#E2E8F0")
        GREEN = colors.HexColor("#27AE60")
        RED = colors.HexColor("#E74C3C")

        # ================= PDF SETUP =================
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=cm,
            leftMargin=cm,
            topMargin=cm,
            bottomMargin=cm,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "EqTitle",
            parent=styles["Heading1"],
            fontSize=20,
            textColor=NAVY_BLUE,
            alignment=0,
            spaceAfter=15,
            fontName="Helvetica-Bold",
        )
        heading2_style = ParagraphStyle(
            "EqH2",
            parent=styles["Heading2"],
            fontSize=12,
            textColor=NAVY_BLUE,
            spaceBefore=6,
            spaceAfter=10,
            fontName="Helvetica-Bold",
        )
        normal_style = ParagraphStyle(
            "EqNormal",
            fontSize=9,
            textColor=colors.black,
            fontName="Helvetica",
            leading=11,
        )
        bold_style = ParagraphStyle(
            "EqBold",
            fontSize=9,
            textColor=colors.black,
            fontName="Helvetica-Bold",
            leading=11,
        )
        small_style = ParagraphStyle(
            "EqSmall",
            fontSize=7.5,
            textColor=colors.black,
            fontName="Helvetica",
            leading=9,
        )

        def safe_val(obj, attr):
            v = getattr(obj, attr, None)
            if v is None:
                return "-"
            return str(getattr(v, "value", v)).upper()

        # ================= FETCH DATA WITH FILTERS =================

        equipment_stmt = select(Equipment).where(Equipment.is_deleted == False)

        if project_id:
            equipment_stmt = equipment_stmt.where(Equipment.project_id == project_id)

        if equipment_id:
            equipment_stmt = equipment_stmt.where(Equipment.id == equipment_id)

        if condition:
            equipment_stmt = equipment_stmt.where(Equipment.condition == condition)

        if status:
            equipment_stmt = equipment_stmt.where(Equipment.status == status)

        equipments = (await db.execute(equipment_stmt)).scalars().all() or []

        usage_stmt = select(EquipmentUsage)

        if equipment_id:
            usage_stmt = usage_stmt.where(EquipmentUsage.equipment_id == equipment_id)

        if start_date:
            usage_stmt = usage_stmt.where(EquipmentUsage.usage_date >= start_date)

        if end_date:
            usage_stmt = usage_stmt.where(EquipmentUsage.usage_date <= end_date)

        usages = (await db.execute(usage_stmt)).scalars().all() or []

        maint_stmt = select(EquipmentMaintenance)

        if equipment_id:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.equipment_id == equipment_id
            )

        if start_date:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.maintenance_date >= start_date
            )

        if end_date:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.maintenance_date <= end_date
            )

        maint = (await db.execute(maint_stmt)).scalars().all() or []

        rental_stmt = select(EquipmentRental)

        if equipment_id:
            rental_stmt = rental_stmt.where(
                EquipmentRental.equipment_id == equipment_id
            )

        if start_date:
            rental_stmt = rental_stmt.where(EquipmentRental.start_date >= start_date)

        if end_date:
            rental_stmt = rental_stmt.where(EquipmentRental.end_date <= end_date)

        rentals = (await db.execute(rental_stmt)).scalars().all() or []

        purchase_stmt = select(EquipmentPurchase)

        if equipment_id:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.asset_id == equipment_id
            )

        if vendor_name:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.vendor_name.ilike(f"%{vendor_name}%")
            )

        if start_date:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.purchase_date >= start_date
            )

        if end_date:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.purchase_date <= end_date
            )

        purchases = (await db.execute(purchase_stmt)).scalars().all() or []

        total_maint_cost = sum(float(m.cost or 0) for m in maint)
        total_rental_cost = sum(float(r.rental_cost or 0) for r in rentals)
        total_purchase_cost = sum(float(p.total_amount or 0) for p in purchases)
        grand_total = total_maint_cost + total_rental_cost + total_purchase_cost

        good_count = sum(1 for e in equipments if safe_val(e, "condition") == "GOOD")
        damaged_count = sum(
            1 for e in equipments if safe_val(e, "condition") == "DAMAGED"
        )

        elements = []

        # ── 1. HEADER (logo left, title right — same as sample) ──────────
        logo_path = "static/logo.png"
        if os.path.exists(logo_path):
            logo_img = Image(logo_path, width=2 * inch, height=0.75 * inch)
        else:
            logo_img = Paragraph("<b>INFRA PILOT</b>", title_style)

        header_data = [
            [logo_img, Paragraph("<b>EQUIPMENT MANAGEMENT REPORT</b>", title_style)]
        ]
        header_table = Table(header_data, colWidths=[2.5 * inch, 5 * inch])
        header_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, 0), "LEFT"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        elements.append(header_table)
        elements.append(Spacer(1, 0.15 * inch))

        # ── 2. REPORT INFORMATION ─────────────────────────────────────────
        ri_data = [
            [
                Paragraph("<b>Report Date</b>", bold_style),
                datetime.now().strftime("%Y-%m-%d"),
                Paragraph("<b>Report Type</b>", bold_style),
                "Equipment Management Report",
            ],
            [
                Paragraph("<b>Total Equipment</b>", bold_style),
                str(len(equipments)),
                Paragraph("<b>Total Records</b>", bold_style),
                str(len(usages) + len(maint) + len(rentals) + len(purchases)),
            ],
        ]
        ri_table = Table(
            ri_data, colWidths=[1.5 * inch, 2.25 * inch, 1.5 * inch, 2.25 * inch]
        )
        ri_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                    ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY),
                    ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(Paragraph("1. REPORT INFORMATION", heading2_style))
        elements.append(ri_table)
        elements.append(Spacer(1, 0.2 * inch))

        # ── 3. FLEET SUMMARY (boxed banner) ────────────────────────────────
        elements.append(Paragraph("2. FLEET SUMMARY", heading2_style))
        summary_box_data = [
            [
                Paragraph(
                    f"<b>Total Equipment:</b> {len(equipments)} | "
                    f"<b>Good Condition:</b> {good_count} | "
                    f"<b>Damaged:</b> {damaged_count}",
                    bold_style,
                )
            ],
            [
                Paragraph(
                    f"<b>Cost Summary:</b> Maintenance Rs. {total_maint_cost:,.2f} | "
                    f"Rental Rs. {total_rental_cost:,.2f} | "
                    f"Purchase Rs. {total_purchase_cost:,.2f}",
                    normal_style,
                )
            ],
            [
                Paragraph(
                    f"<b>Grand Total:</b> Rs. {grand_total:,.2f}",
                    normal_style,
                )
            ],
        ]
        summary_box = Table(summary_box_data, colWidths=[7.25 * inch])
        summary_box.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
                    ("BOX", (0, 0), (-1, -1), 1, NAVY_BLUE),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.append(summary_box)
        elements.append(Spacer(1, 0.2 * inch))

        # ================= SHARED TABLE STYLE (navy header) =================
        def make_table(data, col_widths):
            t = Table(data, colWidths=col_widths, repeatRows=1)
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), NAVY_BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, LIGHT_GRAY],
                        ),
                        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            return t

        def render_section(
            section_no, title_text, headers, rows, col_widths, empty_text
        ):
            elements.append(Paragraph(f"{section_no}. {title_text}", heading2_style))
            if rows:
                header_row = [Paragraph(f"<b>{h}</b>", small_style) for h in headers]
                data_rows = [
                    [Paragraph(str(v), small_style) for v in row] for row in rows
                ]
                elements.append(make_table([header_row] + data_rows, col_widths))
            else:
                elements.append(Paragraph(empty_text, normal_style))
            elements.append(Spacer(1, 0.2 * inch))

        # ---------------- 3. Equipment ----------------
        eq_headers = [
            "#",
            "Code",
            "Name",
            "Operator",
            "Condition",
            "Status",
            "Hours",
            "Fuel",
            "Rental Cost",
        ]
        eq_rows = []
        for i, e in enumerate(equipments, 1):
            eq_rows.append(
                [
                    i,
                    e.equipment_code or "-",
                    e.equipment_name or "-",
                    e.operator_name or "-",
                    safe_val(e, "condition"),
                    safe_val(e, "status"),
                    f"{float(e.working_hours or 0):,.1f}",
                    f"{float(e.fuel_used or 0):,.1f}",
                    f"Rs. {float(e.rental_cost or 0):,.2f}",
                ]
            )
        render_section(
            "3",
            "EQUIPMENT DETAILS",
            eq_headers,
            eq_rows,
            [
                0.25 * inch,
                0.7 * inch,
                1.15 * inch,
                0.9 * inch,
                0.7 * inch,
                0.75 * inch,
                0.55 * inch,
                0.55 * inch,
                0.7 * inch,
            ],
            "No equipment found.",
        )

        # ---------------- 4. Usage ----------------
        u_headers = ["Equip ID", "Hours", "Fuel", "Date", "Notes"]
        u_rows = [
            [
                u.equipment_id or "-",
                f"{float(u.working_hours or 0):,.1f}",
                f"{float(u.fuel_used or 0):,.1f}",
                str(u.usage_date or "-"),
                u.notes or "-",
            ]
            for u in usages
        ]
        render_section(
            "4",
            "USAGE RECORDS",
            u_headers,
            u_rows,
            [0.7 * inch, 0.7 * inch, 0.7 * inch, 1.0 * inch, 3.4 * inch],
            "No usage records found.",
        )

        # ---------------- 5. Maintenance ----------------
        m_headers = ["Equip ID", "Description", "Date", "Cost", "Next Due"]
        m_rows = [
            [
                m.equipment_id or "-",
                m.description or "-",
                str(m.maintenance_date or "-"),
                f"Rs. {float(m.cost or 0):,.2f}",
                str(m.next_maintenance_date or "-"),
            ]
            for m in maint
        ]
        render_section(
            "5",
            "MAINTENANCE RECORDS",
            m_headers,
            m_rows,
            [0.7 * inch, 2.55 * inch, 1.0 * inch, 0.95 * inch, 1.0 * inch],
            "No maintenance records found.",
        )

        # ---------------- 6. Rentals ----------------
        r_headers = ["Equip ID", "Client", "Start", "End", "Cost"]
        r_rows = [
            [
                r.equipment_id or "-",
                r.client_name or "-",
                str(r.start_date or "-"),
                str(r.end_date or "-"),
                f"Rs. {float(r.rental_cost or 0):,.2f}",
            ]
            for r in rentals
        ]
        render_section(
            "6",
            "RENTAL RECORDS",
            r_headers,
            r_rows,
            [0.7 * inch, 2.35 * inch, 1.0 * inch, 1.0 * inch, 1.15 * inch],
            "No rental records found.",
        )

        # ---------------- 7. Purchases ----------------
        p_headers = [
            "Asset ID",
            "Type",
            "Vendor",
            "Invoice",
            "Qty",
            "Unit Price",
            "Total",
            "Date",
        ]
        p_rows = [
            [
                p.asset_id or "-",
                safe_val(p, "purchase_type"),
                p.vendor_name or "-",
                p.invoice_number or "-",
                p.quantity or 0,
                f"Rs. {float(p.unit_price or 0):,.2f}",
                f"Rs. {float(p.total_amount or 0):,.2f}",
                str(p.purchase_date or "-"),
            ]
            for p in purchases
        ]
        render_section(
            "7",
            "PURCHASE RECORDS",
            p_headers,
            p_rows,
            [
                0.65 * inch,
                0.7 * inch,
                1.2 * inch,
                0.9 * inch,
                0.45 * inch,
                0.85 * inch,
                0.85 * inch,
                0.9 * inch,
            ],
            "No purchase records found.",
        )

        # ── 8. COST SUMMARY ────────────────────────────────────────────────
        elements.append(Paragraph("8. COST SUMMARY", heading2_style))
        cost_data = [
            [
                Paragraph("<b>Total Maintenance Cost</b>", bold_style),
                f"Rs. {total_maint_cost:,.2f}",
            ],
            [
                Paragraph("<b>Total Rental Cost</b>", bold_style),
                f"Rs. {total_rental_cost:,.2f}",
            ],
            [
                Paragraph("<b>Total Purchase Cost</b>", bold_style),
                f"Rs. {total_purchase_cost:,.2f}",
            ],
            [Paragraph("<b>Grand Total</b>", bold_style), f"Rs. {grand_total:,.2f}"],
        ]
        cost_table = Table(cost_data, colWidths=[3.6 * inch, 3.6 * inch])
        cost_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                    ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(cost_table)

        # ── FOOTER (signature lines — identical to sample) ───────────────
        def add_footer(canvas, doc_):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setStrokeColor(NAVY_BLUE)
            canvas.setLineWidth(1)
            canvas.line(cm, 1.5 * cm, A4[0] - cm, 1.5 * cm)

            canvas.drawString(cm, 1.2 * cm, "Prepared By: ______________")
            canvas.drawString(
                A4[0] / 2 - 1.5 * cm, 1.2 * cm, "Reviewed By: ______________"
            )
            canvas.drawString(A4[0] - 5 * cm, 1.2 * cm, "Approved By: ______________")

            canvas.drawString(
                cm, 0.8 * cm, "Generated by InfraPilot Construction Management System"
            )
            canvas.drawRightString(A4[0] - cm, 0.8 * cm, f"Page {doc_.page}")
            canvas.restoreState()

        doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)

        buffer.seek(0)
        filename = f"equipment_report_{datetime.now().strftime('%Y%m%d')}.pdf"

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


# ================================ EXCEL REPORT ================================


@router.get("/reports/excel")
async def equipment_excel_report(
    project_id: Optional[int] = Query(None),
    equipment_id: Optional[int] = Query(None),
    condition: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    vendor_name: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(EQUIPMENT_READ_ROLES)),
):
    try:
        import io
        from datetime import datetime

        from fastapi import HTTPException
        from fastapi.responses import StreamingResponse

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ================= COLOR PALETTE (matches INFRA PILOT sample) =====
        NAVY_HEX = "0B2B5C"
        LIGHT_GRAY_HEX = "F8F9FA"
        BORDER_GRAY_HEX = "E2E8F0"

        # ================= FETCH DATA WITH FILTERS =================

        equipment_stmt = select(Equipment).where(Equipment.is_deleted == False)

        if project_id:
            equipment_stmt = equipment_stmt.where(Equipment.project_id == project_id)

        if equipment_id:
            equipment_stmt = equipment_stmt.where(Equipment.id == equipment_id)

        if condition:
            equipment_stmt = equipment_stmt.where(Equipment.condition == condition)

        if status:
            equipment_stmt = equipment_stmt.where(Equipment.status == status)

        equipments = (await db.execute(equipment_stmt)).scalars().all() or []

        usage_stmt = select(EquipmentUsage)

        if equipment_id:
            usage_stmt = usage_stmt.where(EquipmentUsage.equipment_id == equipment_id)

        if start_date:
            usage_stmt = usage_stmt.where(EquipmentUsage.usage_date >= start_date)

        if end_date:
            usage_stmt = usage_stmt.where(EquipmentUsage.usage_date <= end_date)

        usages = (await db.execute(usage_stmt)).scalars().all() or []

        maint_stmt = select(EquipmentMaintenance)

        if equipment_id:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.equipment_id == equipment_id
            )

        if start_date:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.maintenance_date >= start_date
            )

        if end_date:
            maint_stmt = maint_stmt.where(
                EquipmentMaintenance.maintenance_date <= end_date
            )

        maint = (await db.execute(maint_stmt)).scalars().all() or []

        rental_stmt = select(EquipmentRental)

        if equipment_id:
            rental_stmt = rental_stmt.where(
                EquipmentRental.equipment_id == equipment_id
            )

        if start_date:
            rental_stmt = rental_stmt.where(EquipmentRental.start_date >= start_date)

        if end_date:
            rental_stmt = rental_stmt.where(EquipmentRental.end_date <= end_date)

        rentals = (await db.execute(rental_stmt)).scalars().all() or []

        purchase_stmt = select(EquipmentPurchase)

        if equipment_id:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.asset_id == equipment_id
            )

        if vendor_name:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.vendor_name.ilike(f"%{vendor_name}%")
            )

        if start_date:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.purchase_date >= start_date
            )

        if end_date:
            purchase_stmt = purchase_stmt.where(
                EquipmentPurchase.purchase_date <= end_date
            )

        purchases = (await db.execute(purchase_stmt)).scalars().all() or []

        total_maint_cost = sum(float(m.cost or 0) for m in maint)
        total_rental_cost = sum(float(r.rental_cost or 0) for r in rentals)
        total_purchase_cost = sum(float(p.total_amount or 0) for p in purchases)
        grand_total = total_maint_cost + total_rental_cost + total_purchase_cost

        now_str = datetime.now().strftime("%d %b %Y %I:%M %p")
        CURRENCY_FMT = '"Rs." #,##0.00'

        # ================= NAVY-THEMED STYLES (matches PDF) ================
        HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        HEADER_FILL = PatternFill("solid", fgColor=NAVY_HEX)
        TITLE_FONT = Font(name="Arial", bold=True, size=14, color=NAVY_HEX)
        SUBTITLE_FONT = Font(name="Arial", italic=True, size=9, color="6B7280")
        LABEL_FONT = Font(name="Arial", bold=True, size=10, color=NAVY_HEX)
        CELL_FONT = Font(name="Arial", size=10)
        THIN = Side(style="thin", color=BORDER_GRAY_HEX)
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CENTER = Alignment(horizontal="center", vertical="center")
        ACCENT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)

        def safe_val(obj, attr):
            v = getattr(obj, attr, None)
            if v is None:
                return "-"
            return getattr(v, "value", v)

        def write_table(ws, headers, rows, currency_cols=None, title=None):
            """Writes an optional title row, then header row + zebra data rows,
            navy-themed to match the PDF / sample design."""
            currency_cols = currency_cols or []
            start_row = 1

            if title:
                ws.merge_cells(
                    start_row=1, start_column=1, end_row=1, end_column=len(headers)
                )
                ws.cell(row=1, column=1, value=title).font = TITLE_FONT
                ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
                start_row = 3

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
                cell.border = BORDER

            for r, row_values in enumerate(rows, start_row + 1):
                fill = ACCENT_FILL if (r - start_row) % 2 == 0 else None
                for col, val in enumerate(row_values, 1):
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = CELL_FONT
                    cell.alignment = CENTER
                    cell.border = BORDER
                    if fill:
                        cell.fill = fill
                    if col in currency_cols:
                        cell.number_format = CURRENCY_FMT

            for col, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row_values in rows:
                    val = row_values[col - 1]
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col)].width = min(
                    max(max_len + 4, 12), 40
                )

            ws.freeze_panes = f"A{start_row + 1}"

        # ================= WORKBOOK =================
        wb = Workbook()
        wb.remove(wb.active)

        # ---------------- Summary (first sheet) ----------------
        good_count = sum(1 for e in equipments if safe_val(e, "condition") == "GOOD")
        damaged_count = sum(
            1 for e in equipments if safe_val(e, "condition") == "DAMAGED"
        )

        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.merge_cells("A1:B1")
        ws_summary["A1"] = "EQUIPMENT MANAGEMENT REPORT"
        ws_summary["A1"].font = TITLE_FONT
        ws_summary.merge_cells("A2:B2")
        ws_summary["A2"] = f"Generated: {now_str}"
        ws_summary["A2"].font = SUBTITLE_FONT

        summary_rows = [
            ("Total Equipment", len(equipments)),
            ("Good Condition", good_count),
            ("Damaged", damaged_count),
            ("Usage Records", len(usages)),
            ("Maintenance Records", len(maint)),
            ("Rental Records", len(rentals)),
            ("Purchase Records", len(purchases)),
            ("Total Maintenance Cost", total_maint_cost),
            ("Total Rental Cost", total_rental_cost),
            ("Total Purchase Cost", total_purchase_cost),
            ("Grand Total (Maint + Rental + Purchase)", grand_total),
        ]

        row_num = 4
        for label, value in summary_rows:
            label_cell = ws_summary.cell(row=row_num, column=1, value=label)
            label_cell.font = LABEL_FONT
            label_cell.fill = ACCENT_FILL
            value_cell = ws_summary.cell(row=row_num, column=2, value=value)
            value_cell.font = CELL_FONT
            if "Cost" in label or "Total" in label:
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    value_cell.number_format = CURRENCY_FMT
            row_num += 1

        ws_summary.column_dimensions["A"].width = 38
        ws_summary.column_dimensions["B"].width = 20

        # Footer signature block on Summary sheet
        row_num += 2
        ws_summary.cell(
            row=row_num, column=1, value="Prepared By: ______________"
        ).font = CELL_FONT
        row_num += 1
        ws_summary.cell(
            row=row_num, column=1, value="Reviewed By: ______________"
        ).font = CELL_FONT
        row_num += 1
        ws_summary.cell(
            row=row_num, column=1, value="Approved By: ______________"
        ).font = CELL_FONT

        # ---------------- Equipment ----------------
        ws = wb.create_sheet("Equipment")
        headers = [
            "ID",
            "Project ID",
            "Code",
            "Name",
            "Operator",
            "Condition",
            "Status",
            "Working Hours",
            "Fuel Used",
            "Rental Cost",
            "Maintenance Date",
        ]
        rows = [
            [
                e.id,
                e.project_id or "-",
                e.equipment_code or "-",
                e.equipment_name or "-",
                e.operator_name or "-",
                safe_val(e, "condition"),
                safe_val(e, "status"),
                float(e.working_hours or 0),
                float(e.fuel_used or 0),
                float(e.rental_cost or 0),
                str(e.maintenance_date) if e.maintenance_date else "-",
            ]
            for e in equipments
        ]
        write_table(ws, headers, rows, currency_cols=[10], title="Equipment")

        # ---------------- Usage ----------------
        ws = wb.create_sheet("Usage")
        headers = ["Equip ID", "Working Hours", "Fuel Used", "Usage Date", "Notes"]
        rows = [
            [
                u.equipment_id,
                float(u.working_hours or 0),
                float(u.fuel_used or 0),
                str(u.usage_date or "-"),
                u.notes or "-",
            ]
            for u in usages
        ]
        write_table(ws, headers, rows, title="Usage Records")

        # ---------------- Maintenance ----------------
        ws = wb.create_sheet("Maintenance")
        headers = ["Equip ID", "Description", "Date", "Cost", "Next Due"]
        rows = [
            [
                m.equipment_id,
                m.description or "-",
                str(m.maintenance_date or "-"),
                float(m.cost or 0),
                str(m.next_maintenance_date or "-"),
            ]
            for m in maint
        ]
        write_table(ws, headers, rows, currency_cols=[4], title="Maintenance Records")

        # ---------------- Rentals ----------------
        ws = wb.create_sheet("Rentals")
        headers = ["Equip ID", "Client", "Start", "End", "Cost", "Notes"]
        rows = [
            [
                r.equipment_id,
                r.client_name or "-",
                str(r.start_date or "-"),
                str(r.end_date or "-"),
                float(r.rental_cost or 0),
                r.notes or "-",
            ]
            for r in rentals
        ]
        write_table(ws, headers, rows, currency_cols=[5], title="Rental Records")

        # ---------------- Purchases ----------------
        ws = wb.create_sheet("Purchases")
        headers = [
            "Asset ID",
            "Type",
            "Vendor",
            "Invoice",
            "Qty",
            "Unit Price",
            "Total",
            "Purchase Date",
            "Warranty End",
        ]
        rows = [
            [
                p.asset_id,
                safe_val(p, "purchase_type"),
                p.vendor_name or "-",
                p.invoice_number or "-",
                p.quantity or 0,
                float(p.unit_price or 0),
                float(p.total_amount or 0),
                str(p.purchase_date or "-"),
                str(p.warranty_end_date or "-"),
            ]
            for p in purchases
        ]
        write_table(ws, headers, rows, currency_cols=[6, 7], title="Purchase Records")

        # ================= SAVE FILE =================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"equipment_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Excel generation failed: {str(e)}"
        )
