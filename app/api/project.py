from __future__ import annotations
from datetime import date, datetime
import json
import mimetypes
import pathlib, re, io, os, uuid
from ezdxf import colors
from openpyxl import Workbook
from typing import Annotated, List, Optional, Union
from fastapi import APIRouter, Depends, Query, Request, Form
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.enums import (
    PRIORITY_MAP,
    REVERSE_PRIORITY_MAP,
    DocumentStatus,
    LabourStatus,
    MilestoneStatus,
    SkillType,
    TaskPriority,
    WorkActivityStatus,
)
from app.core.validators import validate_drawing_file
from app.db.session import get_db_session
from sqlalchemy.orm import selectinload
import traceback
from app.models.approval import Approval
from app.models.labour import Labour
from app.models.user import UserAttendance
from app.middlewares.rate_limiter import default_rate_limiter_dependency
from app.cache.redis import (
    bump_cache_version,
    cache_get_json,
    cache_set_json,
    get_cache_version,
)
from sqlalchemy.orm import aliased
from PIL import Image
from app.core.dependencies import (
    get_current_active_user,
    get_request_redis,
    require_roles,
)
import shutil
import uuid
import os
from app.services.notification_service import create_notification
from app.models.contractor import Contractor
from sqlalchemy import delete, select, func, or_, update
from app.models import project as m
from app.models.messages import MessageStatus
from app.models.user import User, UserRole
from app.models.owner import Owner
from app.models.expense import Expense
from app.models.invoice import Invoice
from app.models.master_data import ActivityType, LabourType
from app.schemas.base import PaginatedResponse, PaginationMeta
from app.schemas import project as s
from app.core.logger import logger
from fastapi.responses import FileResponse, StreamingResponse
from reportlab.platypus import (
    PageBreak,
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.exc import IntegrityError
from fastapi import UploadFile, File
from app.utils.helpers import (
    AppError,
    BadRequestError,
    DataIntegrityError,
    NotFoundError,
    ConflictError,
    PermissionDeniedError,
    ValidationError,
)
from app.utils.pagination import PaginationParams
from app.utils.common import (
    assert_project_access,
    generate_business_id,
    assert_task_project,
)


def compute_project_status(project):
    today = date.today()

    if project.status == s.ProjectStatus.COMPLETED:
        return "Completed"

    if project.status == s.ProjectStatus.ON_HOLD:
        return "On Hold"

    if project.status == s.ProjectStatus.PLANNED:
        return "Planned"

    if (
        project.status == s.ProjectStatus.ONGOING
        and project.end_date
        and today > project.end_date
    ):
        return "Delayed"

    return "Ongoing"


def compute_milestone_status(milestone):
    today = date.today()

    if milestone.status == MilestoneStatus.COMPLETED:
        return "Completed"

    if milestone.start_date and today < milestone.start_date:
        return "Planned"

    if milestone.end_date and today > milestone.end_date:
        return "Delayed"

    return "In Progress"


def get_pagination(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    search: Optional[str] = Query(None),
) -> PaginationParams:
    return PaginationParams(limit=limit, offset=offset, search=search).normalized()


router = APIRouter(
    prefix="/projects",
    tags=["project_management"],
    dependencies=[default_rate_limiter_dependency()],
)


VERSION_KEY = "cache_version:projects"

PROJECT_WRITE_ROLES = [r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER]]
PROJECT_DELETE_ROLES = [UserRole.ADMIN.value]

TASK_WRITE_ROLES = [
    r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER, UserRole.SITE_ENGINEER]
]
TASK_DELETE_ROLES = [r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER]]

DSR_WRITE_ROLES = [
    r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER, UserRole.SITE_ENGINEER]
]
DSR_READ_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.CLIENT,
    ]
]
DSR_DELETE_ROLES = [UserRole.ADMIN.value]
DSR_APPROVE_ROLES = [
    r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER, UserRole.CLIENT]
]

ISSUE_CREATE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.CLIENT,
    ]
]
ISSUE_UPDATE_ROLES = [r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER]]
ISSUE_DELETE_ROLES = [UserRole.ADMIN.value]

FINANCIAL_ROLES = [
    r.value for r in [UserRole.ADMIN, UserRole.PROJECT_MANAGER, UserRole.ACCOUNTANT]
]

READ_ROLES = [r.value for r in UserRole]

DRAWING_WRITE_ROLES = TASK_WRITE_ROLES
DRAWING_READ_ROLES = READ_ROLES

DRAWING_DELETE_ROLES = [
    UserRole.ADMIN.value,
    UserRole.PROJECT_MANAGER.value,
]


class ProjectsRepository:
    async def create_project(self, db: AsyncSession, data: dict) -> m.Project:
        obj = m.Project(**data)
        db.add(obj)
        await db.flush()
        return obj

    async def get_project(
        self, db: AsyncSession, project_id: int
    ) -> Optional[m.Project]:
        return await db.scalar(
            select(m.Project)
            .options(
                selectinload(m.Project.milestones).selectinload(m.Milestone.tasks),
                selectinload(m.Project.tasks),
            )
            .where(m.Project.id == project_id)
        )

    async def list_projects(
        self,
        db: AsyncSession,
        *,
        limit: int,
        offset: int,
        search: Optional[str] = None,
        status: Optional[s.ProjectStatus] = None,
    ) -> tuple[list[m.Project], int]:
        query = select(m.Project).options(
            selectinload(m.Project.milestones).selectinload(m.Milestone.tasks),
            selectinload(m.Project.tasks),
        )
        count_query = select(func.count()).select_from(m.Project)

        if search:
            like = f"%{search}%"
            query = query.where(m.Project.project_name.ilike(like))
            count_query = count_query.where(m.Project.project_name.ilike(like))

        if status:
            query = query.where(m.Project.status == status)
            count_query = count_query.where(m.Project.status == status)

        query = query.order_by(m.Project.id.desc()).limit(limit).offset(offset)

        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()
        return rows, int(total or 0)

    async def update_project(
        self, db: AsyncSession, obj: m.Project, data: dict
    ) -> m.Project:
        for k, v in data.items():
            setattr(obj, k, v)
        await db.flush()
        return obj

    async def delete_project(self, db: AsyncSession, obj: m.Project) -> None:
        await db.delete(obj)
        await db.flush()


class ProjectMembersRepository:
    async def get_member(
        self, db: AsyncSession, *, project_id: int, user_id: int
    ) -> Optional[m.ProjectMember]:
        return await db.scalar(
            select(m.ProjectMember).where(
                m.ProjectMember.project_id == project_id,
                m.ProjectMember.user_id == user_id,
            )
        )

    async def assign_member(
        self, db: AsyncSession, *, project_id: int, user_id: int
    ) -> m.ProjectMember:
        obj = m.ProjectMember(project_id=project_id, user_id=user_id)
        db.add(obj)
        await db.flush()
        return obj

    async def remove_member(
        self, db: AsyncSession, *, project_id: int, user_id: int
    ) -> None:
        obj = await self.get_member(db, project_id=project_id, user_id=user_id)
        if obj is None:
            return
        await db.delete(obj)
        await db.flush()

    async def is_member(
        self, db: AsyncSession, *, project_id: int, user_id: int
    ) -> bool:
        exists_query = (
            select(func.count())
            .select_from(m.ProjectMember)
            .where(
                m.ProjectMember.project_id == project_id,
                m.ProjectMember.user_id == user_id,
            )
        )
        count = await db.scalar(exists_query)
        return bool(count and int(count) > 0)

    async def list_members(
        self,
        db: AsyncSession,
        *,
        project_id: int,
        limit: int,
        offset: int,
    ) -> tuple[list[User], int]:
        from app.models.user import User  # local import to avoid circulars

        count_query = (
            select(func.count())
            .select_from(m.ProjectMember)
            .where(m.ProjectMember.project_id == project_id)
        )

        query = (
            select(User)
            .join(m.ProjectMember, m.ProjectMember.user_id == User.id)
            .where(m.ProjectMember.project_id == project_id)
            .order_by(User.id.desc())
            .limit(limit)
            .offset(offset)
        )

        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()
        return rows, int(total or 0)


class MilestonesRepository:
    async def create_milestone(
        self, db: AsyncSession, *, project_id: int, data: dict
    ) -> m.Milestone:
        obj = m.Milestone(project_id=project_id, **data)
        db.add(obj)
        await db.flush()
        return obj

    async def get_milestone(
        self, db: AsyncSession, *, project_id: int, milestone_id: int
    ) -> Optional[m.Milestone]:
        from sqlalchemy.orm import selectinload

        return await db.scalar(
            select(m.Milestone)
            .where(m.Milestone.project_id == project_id, m.Milestone.id == milestone_id)
            .options(selectinload(m.Milestone.tasks))
        )

    async def list_milestones(
        self,
        db: AsyncSession,
        *,
        project_id: int,
        limit: int,
        offset: int,
    ) -> tuple[list[m.Milestone], int]:

        count_query = select(func.count()).where(m.Milestone.project_id == project_id)

        from sqlalchemy.orm import selectinload

        query = (
            select(m.Milestone)
            .where(m.Milestone.project_id == project_id)
            .order_by(m.Milestone.id.desc())
            .limit(limit)
            .offset(offset)
            .options(selectinload(m.Milestone.tasks))
        )

        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()

        return rows, int(total or 0)

    async def update_milestone(
        self, db: AsyncSession, *, obj: m.Milestone, data: dict
    ) -> m.Milestone:
        for k, v in data.items():
            setattr(obj, k, v)
        await db.flush()
        return obj

    async def delete_milestone(self, db: AsyncSession, *, obj: m.Milestone) -> None:
        await db.delete(obj)
        await db.flush()


class TasksRepository:
    async def create_task(
        self, db: AsyncSession, *, project_id: int, data: dict
    ) -> m.Task:
        obj = m.Task(project_id=project_id, **data)
        db.add(obj)
        await db.flush()
        return obj

    async def get_task(
        self, db: AsyncSession, *, project_id: int, task_id: int
    ) -> Optional[m.Task]:
        from sqlalchemy.orm import selectinload, joinedload

        return await db.scalar(
            select(m.Task)
            .options(selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user))
            .where(m.Task.project_id == project_id, m.Task.id == task_id)
        )

    async def list_tasks(
        self,
        db: AsyncSession,
        *,
        project_id: int,
        status: Optional[s.TaskStatus],
        assigned_user_id: Optional[int],
        limit: int,
        offset: int,
    ) -> tuple[list[m.Task], int]:
        from sqlalchemy.orm import selectinload, joinedload

        query = (
            select(m.Task)
            .options(selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user))
            .where(m.Task.project_id == project_id)
        )
        count_query = (
            select(func.count(m.Task.id.distinct()))
            .select_from(m.Task)
            .where(m.Task.project_id == project_id)
        )

        if status is not None:
            query = query.where(m.Task.status == status)
            count_query = count_query.where(m.Task.status == status)

        if assigned_user_id is not None:
            query = query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == assigned_user_id)
            )
            count_query = count_query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == assigned_user_id)
            )

        query = query.order_by(m.Task.id.desc()).limit(limit).offset(offset)

        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()
        return rows, int(total or 0)

    async def update_task(self, db: AsyncSession, *, obj: m.Task, data: dict) -> m.Task:
        for k, v in data.items():
            setattr(obj, k, v)
        await db.flush()
        return obj

    async def delete_task(self, db: AsyncSession, *, obj: m.Task) -> None:
        await db.delete(obj)
        await db.flush()

    async def list_task_completion_by_project_ids(
        self, db: AsyncSession, project_ids: list[int]
    ) -> list[tuple[int, int]]:
        if not project_ids:
            return []

        query = select(m.Task.project_id, m.Task.completion_percentage).where(
            m.Task.project_id.in_(project_ids)
        )
        rows = (await db.execute(query)).all()
        return [(int(pid), int(pct)) for pid, pct in rows]


class TaskProgressRepository:
    async def create_progress(
        self,
        db: AsyncSession,
        *,
        task_id: int,
        percentage: int,
        remarks: Optional[str],
        created_by_user_id: Optional[int],
    ) -> m.TaskProgress:
        obj = m.TaskProgress(
            task_id=task_id,
            percentage=percentage,
            remarks=remarks,
            created_by_user_id=created_by_user_id,
        )
        db.add(obj)
        await db.flush()
        return obj

    async def list_progress_history(
        self,
        db: AsyncSession,
        *,
        task_id: int,
        limit: int,
        offset: int,
    ) -> tuple[list[m.TaskProgress], int]:
        count_query = (
            select(func.count())
            .select_from(m.TaskProgress)
            .where(m.TaskProgress.task_id == task_id)
        )
        query = (
            select(m.TaskProgress)
            .where(m.TaskProgress.task_id == task_id)
            .order_by(m.TaskProgress.created_at.desc())
            .limit(limit)
            .offset(offset)
        )
        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()
        return rows, int(total or 0)


class CommentsRepository:
    async def create_comment(
        self,
        db: AsyncSession,
        *,
        task_id: int,
        author_user_id: int,
        content: str,
    ) -> m.Comment:
        obj = m.Comment(task_id=task_id, author_user_id=author_user_id, content=content)
        db.add(obj)
        await db.flush()
        return obj

    async def list_comments(
        self,
        db: AsyncSession,
        *,
        task_id: int,
        limit: int,
        offset: int,
    ) -> tuple[list[m.Comment], int]:
        count_query = (
            select(func.count())
            .select_from(m.Comment)
            .where(m.Comment.task_id == task_id)
        )
        query = (
            select(m.Comment)
            .where(m.Comment.task_id == task_id)
            .order_by(m.Comment.id.desc())
            .limit(limit)
            .offset(offset)
        )
        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()
        return rows, int(total or 0)


class ProjectsService:
    def __init__(
        self,
        projects_repo: ProjectsRepository,
        tasks_repo: TasksRepository,
    ) -> None:
        self.projects_repo = projects_repo
        self.tasks_repo = tasks_repo

    def _assert_project_mutation_role(self, current_user: User) -> None:
        if current_user.role not in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            raise PermissionDeniedError("Insufficient permissions")

    async def _compute_completion_percentage_by_project_ids(
        self, db: AsyncSession, project_ids: list[int]
    ) -> dict[int, float]:
        completion_rows = await self.tasks_repo.list_task_completion_by_project_ids(
            db, project_ids
        )
        completion_map: dict[int, list[int]] = {pid: [] for pid in project_ids}
        for pid, pct in completion_rows:
            completion_map[pid].append(pct)

        out: dict[int, float] = {}
        for pid in project_ids:
            values = completion_map.get(pid) or []
            out[pid] = (float(sum(values)) / len(values)) if values else 0.0
        return out

    async def create_project(
        self, db: AsyncSession, current_user: User, payload: s.ProjectCreate
    ) -> s.ProjectOut:
        self._assert_project_mutation_role(current_user)
        data = payload.model_dump(exclude_unset=True)
        if "status" not in data:
            data["status"] = s.ProjectStatus.PLANNED
        owner = await db.scalar(select(Owner).where(Owner.id == payload.owner_id))
        if not owner:
            raise NotFoundError("Owner not found")

        if payload.start_date and payload.end_date:
            if payload.end_date < payload.start_date:
                raise ValidationError("end_date cannot be before start_date")

        for _ in range(3):
            try:
                data["business_id"] = await generate_business_id(
                    db, m.Project, "business_id", "PRJ"
                )

                obj = await self.projects_repo.create_project(db, data)
                break

            except IntegrityError as e:
                await db.rollback()

                # Optional: if name conflict (keep your logic)
                if "project_name" in str(e.orig):
                    raise ConflictError("Project with this name already exists")

                continue
        else:
            raise Exception("Failed to generate unique project ID")

        completion_map = await self._compute_completion_percentage_by_project_ids(
            db, [obj.id]
        )
        completion = completion_map.get(obj.id, 0.0)
        return s.ProjectOut(
            id=obj.id,
            business_id=obj.business_id,
            project_name=obj.project_name,
            owner_id=obj.owner_id,
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            status=compute_project_status(obj),
            completion_percentage=completion,
            type=obj.type,
            location_type=obj.location_type,
            site_address=obj.site_address,
            city=obj.city,
            state=obj.state,
            country=obj.country,
            pincode=obj.pincode,
            latitude=obj.latitude,
            longitude=obj.longitude,
            shift_start_time=obj.shift_start_time,
            shift_end_time=obj.shift_end_time,
            grace_period_minutes=obj.grace_period_minutes,
            quotation_id=obj.quotation_id,
            budget_amount=float(obj.budget_amount),
        )

    async def list_projects(
        self,
        db: AsyncSession,
        *,
        current_user: User,
        limit: int,
        offset: int,
        search: Optional[str] = None,
        status: Optional[s.ProjectStatus] = None,
    ) -> PaginatedResponse[s.ProjectOut]:

        if current_user.role in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            base_query = select(m.Project)
        else:
            base_query = (
                select(m.Project)
                .join(m.ProjectMember, m.ProjectMember.project_id == m.Project.id)
                .where(m.ProjectMember.user_id == current_user.id)
            )

        if search:
            base_query = base_query.where(
                m.Project.project_name.ilike(f"%{search.strip()}%")
            )

        today = date.today()

        if status:

            if status == s.ProjectStatus.DELAYED:

                base_query = base_query.where(
                    m.Project.status == s.ProjectStatus.ONGOING,
                    m.Project.end_date.is_not(None),
                    m.Project.end_date < today,
                )

            else:
                base_query = base_query.where(m.Project.status == status)

        count_query = select(func.count()).select_from(
            base_query.order_by(None).subquery()
        )
        total = await db.scalar(count_query)

        from sqlalchemy.orm import selectinload

        query = (
            base_query.order_by(m.Project.id.desc())
            .limit(limit)
            .offset(offset)
            .options(
                selectinload(m.Project.milestones).selectinload(m.Milestone.tasks),
                selectinload(m.Project.tasks),
            )
        )

        rows = (await db.execute(query)).scalars().all()

        project_ids = [p.id for p in rows]
        completion_map = await self._compute_completion_percentage_by_project_ids(
            db, project_ids
        )

        items = [
            s.ProjectOut(
                id=p.id,
                business_id=p.business_id,
                project_name=p.project_name,
                owner_id=p.owner_id,
                description=p.description,
                start_date=p.start_date,
                end_date=p.end_date,
                status=compute_project_status(p),
                completion_percentage=completion_map.get(p.id, 0.0),
                execution_completion_percentage=p.execution_completion_percentage,
                total_milestones=p.total_milestones,
                total_tasks=p.total_tasks,
                completed_tasks=p.completed_tasks,
                delayed_tasks=p.delayed_tasks,
                type=p.type,
                location_type=p.location_type,
                site_address=p.site_address,
                city=p.city,
                state=p.state,
                country=p.country,
                pincode=p.pincode,
                latitude=p.latitude,
                longitude=p.longitude,
                shift_start_time=p.shift_start_time,
                shift_end_time=p.shift_end_time,
                grace_period_minutes=p.grace_period_minutes,
                quotation_id=p.quotation_id,
                budget_amount=float(p.budget_amount),
            )
            for p in rows
        ]

        meta = PaginationMeta(total=int(total or 0), limit=limit, offset=offset)

        return PaginatedResponse[s.ProjectOut](items=items, meta=meta)

    async def get_project(
        self,
        db: AsyncSession,
        project_id: int,
        current_user: User,
    ) -> s.ProjectOut:
        obj = await self.projects_repo.get_project(db, project_id=project_id)
        if obj is None:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=obj.id,
            current_user=current_user,
        )

        completion_map = await self._compute_completion_percentage_by_project_ids(
            db, [obj.id]
        )
        completion = completion_map.get(obj.id, 0.0)

        return s.ProjectOut(
            id=obj.id,
            business_id=obj.business_id,
            project_name=obj.project_name,
            owner_id=obj.owner_id,
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            status=compute_project_status(obj),
            completion_percentage=completion,
            execution_completion_percentage=obj.execution_completion_percentage,
            total_milestones=obj.total_milestones,
            total_tasks=obj.total_tasks,
            completed_tasks=obj.completed_tasks,
            delayed_tasks=obj.delayed_tasks,
            type=obj.type,
            location_type=obj.location_type,
            site_address=obj.site_address,
            city=obj.city,
            state=obj.state,
            country=obj.country,
            pincode=obj.pincode,
            latitude=obj.latitude,
            longitude=obj.longitude,
            shift_start_time=obj.shift_start_time,
            shift_end_time=obj.shift_end_time,
            grace_period_minutes=obj.grace_period_minutes,
            quotation_id=obj.quotation_id,
            budget_amount=float(obj.budget_amount),
        )

    async def update_project(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        payload: s.ProjectUpdate,
    ) -> s.ProjectOut:
        self._assert_project_mutation_role(current_user)
        obj = await self.projects_repo.get_project(db, project_id=project_id)
        if obj is None:
            raise NotFoundError("Project not found")
        data = payload.model_dump(exclude_unset=True)
        if "project_name" in data and data["project_name"] is None:
            raise ValidationError("project_name cannot be null")
        if "status" in data:
            allowed_statuses = [
                s.ProjectStatus.PLANNED,
                s.ProjectStatus.ONGOING,
                s.ProjectStatus.ON_HOLD,
                s.ProjectStatus.COMPLETED,
            ]

            if data["status"] not in allowed_statuses:
                raise ValidationError("Invalid project status")
        try:
            await self.projects_repo.update_project(db, obj, data)
            await db.refresh(obj)
        except Exception:
            await db.rollback()
            logger.exception(f"Project update failed id={project_id}")
            raise
        completion_map = await self._compute_completion_percentage_by_project_ids(
            db, [obj.id]
        )
        completion = completion_map.get(obj.id, 0.0)
        return s.ProjectOut(
            id=obj.id,
            business_id=obj.business_id,
            project_name=obj.project_name,
            owner_id=obj.owner_id,
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            status=compute_project_status(obj),
            completion_percentage=completion,
            type=obj.type,
            location_type=obj.location_type,
            site_address=obj.site_address,
            city=obj.city,
            state=obj.state,
            country=obj.country,
            pincode=obj.pincode,
            latitude=obj.latitude,
            longitude=obj.longitude,
            shift_start_time=obj.shift_start_time,
            shift_end_time=obj.shift_end_time,
            grace_period_minutes=obj.grace_period_minutes,
        )

    async def delete_project(
        self, db: AsyncSession, current_user: User, *, project_id: int
    ) -> None:
        self._assert_project_mutation_role(current_user)
        obj = await self.projects_repo.get_project(db, project_id=project_id)
        if obj is None:
            raise NotFoundError("Project not found")
        try:
            await self.projects_repo.delete_project(db, obj)
        except Exception:
            await db.rollback()
            logger.exception(f"Project delete failed id={project_id}")
            raise


class ProjectMembersService:
    def __init__(
        self,
        projects_repo: ProjectsRepository,
        members_repo: ProjectMembersRepository,
    ) -> None:
        self.projects_repo = projects_repo
        self.members_repo = members_repo

    def _assert_member_mutation_role(self, current_user: User) -> None:
        if current_user.role not in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            raise PermissionDeniedError("Insufficient permissions")

    async def assign_member(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        user_id: int,
    ) -> s.ProjectMemberOut:

        self._assert_member_mutation_role(current_user)

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        user = await db.scalar(select(User).where(User.id == user_id))
        if user is None:
            raise NotFoundError("User not found")

        existing = await self.members_repo.get_member(
            db, project_id=project_id, user_id=user_id
        )
        if existing is not None:
            raise ConflictError("User is already assigned to this project")

        try:
            await self.members_repo.assign_member(
                db, project_id=project_id, user_id=user_id
            )
        except IntegrityError:
            await db.rollback()
            raise ConflictError("User is already assigned to this project")

        role = user.role

        return s.ProjectMemberOut(
            user_id=user.id,
            full_name=user.full_name,
            email=user.email,
            role=role,
        )

    async def list_members(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        limit: int,
        offset: int,
    ) -> PaginatedResponse[s.ProjectMemberOut]:
        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        users, total = await self.members_repo.list_members(
            db, project_id=project_id, limit=limit, offset=offset
        )
        items: list[s.ProjectMemberOut] = []
        for user in users:
            role = user.role
            items.append(
                s.ProjectMemberOut(
                    user_id=user.id,
                    full_name=user.full_name,
                    email=user.email,
                    role=role,
                )
            )
        meta = PaginationMeta(total=int(total), limit=limit, offset=offset)
        return PaginatedResponse[s.ProjectMemberOut](items=items, meta=meta)

    async def remove_member(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        user_id: int,
    ) -> None:
        self._assert_member_mutation_role(current_user)

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        existing = await self.members_repo.get_member(
            db, project_id=project_id, user_id=user_id
        )
        if existing is None:
            raise NotFoundError("Member not found")

        await self.members_repo.remove_member(
            db, project_id=project_id, user_id=user_id
        )


class MilestonesService:
    def __init__(
        self, projects_repo: ProjectsRepository, milestones_repo: MilestonesRepository
    ) -> None:
        self.projects_repo = projects_repo
        self.milestones_repo = milestones_repo

    def _assert_milestone_mutation_role(self, current_user: User) -> None:
        if current_user.role not in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            raise PermissionDeniedError("Insufficient permissions")

    async def create_milestone(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        payload: s.MilestoneCreate,
    ) -> s.MilestoneOut:
        self._assert_milestone_mutation_role(current_user)
        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        data = payload.model_dump(exclude_unset=True)

        if "status" not in data:
            data["status"] = MilestoneStatus.PLANNED

        try:
            obj = await self.milestones_repo.create_milestone(
                db, project_id=project_id, data=data
            )
        except IntegrityError:
            await db.rollback()
            raise ConflictError(
                "Milestone with this title already exists in this project"
            )
        except Exception:
            await db.rollback()
            logger.exception(f"Milestone create failed")
            raise

        return s.MilestoneOut(
            id=obj.id,
            project_id=obj.project_id,
            title=obj.title,
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            actual_start_date=obj.actual_start_date,
            actual_end_date=obj.actual_end_date,
            status=compute_milestone_status(obj),  # IMPORTANT
            total_tasks=0,
            completed_tasks=0,
            pending_tasks=0,
            delayed_tasks=0,
            is_delayed=False,
            completion_percentage=0.0,
            execution_completion_percentage=0.0,
        )

    async def list_milestones(
        self,
        db: AsyncSession,
        *,
        project_id: int,
        pagination: PaginationParams,
    ) -> PaginatedResponse[s.MilestoneOut]:

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        rows, total = await self.milestones_repo.list_milestones(
            db,
            project_id=project_id,
            limit=pagination.limit,
            offset=pagination.offset,
        )

        items = [
            s.MilestoneOut(
                id=m.id,
                project_id=m.project_id,
                title=m.title,
                description=m.description,
                start_date=m.start_date,
                end_date=m.end_date,
                status=compute_milestone_status(m),
            )
            for m in rows
        ]

        return PaginatedResponse(
            items=items,
            meta=PaginationMeta(
                total=total,
                limit=pagination.limit,
                offset=pagination.offset,
            ),
        )

    async def get_milestone(
        self, db: AsyncSession, *, project_id: int, milestone_id: int
    ) -> s.MilestoneOut:
        obj = await self.milestones_repo.get_milestone(
            db, project_id=project_id, milestone_id=milestone_id
        )
        if obj is None:
            raise NotFoundError("Milestone not found")
        return s.MilestoneOut(
            id=obj.id,
            project_id=obj.project_id,
            title=obj.title,
            status=compute_milestone_status(obj),
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            actual_start_date=obj.actual_start_date,
            actual_end_date=obj.actual_end_date,
            total_tasks=obj.total_tasks,
            completed_tasks=obj.completed_tasks,
            pending_tasks=obj.pending_tasks,
            delayed_tasks=obj.delayed_tasks,
            is_delayed=obj.is_delayed,
            completion_percentage=obj.completion_percentage,
            execution_completion_percentage=obj.execution_completion_percentage,
        )

    async def update_milestone(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        milestone_id: int,
        payload: s.MilestoneUpdate,
    ) -> s.MilestoneOut:
        self._assert_milestone_mutation_role(current_user)
        obj = await self.milestones_repo.get_milestone(
            db, project_id=project_id, milestone_id=milestone_id
        )
        if obj is None:
            raise NotFoundError("Milestone not found")

        data = payload.model_dump(exclude_unset=True)
        if "title" in data and data["title"] is None:
            raise ValidationError("title cannot be null")

        try:
            await self.milestones_repo.update_milestone(db, obj=obj, data=data)
            await db.refresh(obj)
        except IntegrityError:
            await db.rollback()
            raise ConflictError(
                "Milestone with this title already exists in this project"
            )
        except Exception:
            await db.rollback()
            logger.exception(f"Milestone update failed id={milestone_id}")
            raise

        return s.MilestoneOut(
            id=obj.id,
            project_id=obj.project_id,
            title=obj.title,
            status=compute_milestone_status(obj),
            description=obj.description,
            start_date=obj.start_date,
            end_date=obj.end_date,
            actual_start_date=obj.actual_start_date,
            actual_end_date=obj.actual_end_date,
            total_tasks=obj.total_tasks,
            completed_tasks=obj.completed_tasks,
            pending_tasks=obj.pending_tasks,
            delayed_tasks=obj.delayed_tasks,
            is_delayed=obj.is_delayed,
            completion_percentage=obj.completion_percentage,
            execution_completion_percentage=obj.execution_completion_percentage,
        )

    async def delete_milestone(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        milestone_id: int,
    ) -> None:
        self._assert_milestone_mutation_role(current_user)
        obj = await self.milestones_repo.get_milestone(
            db, project_id=project_id, milestone_id=milestone_id
        )
        if obj is None:
            raise NotFoundError("Milestone not found")
        try:
            await self.milestones_repo.delete_milestone(db, obj=obj)
        except Exception:
            await db.rollback()
            logger.exception(f"Milestone delete failed id={milestone_id}")
            raise


class TasksService:
    def __init__(
        self,
        projects_repo: ProjectsRepository,
        members_repo: ProjectMembersRepository,
        tasks_repo: TasksRepository,
        progress_repo: TaskProgressRepository,
        comments_repo: CommentsRepository,
    ) -> None:
        self.projects_repo = projects_repo
        self.members_repo = members_repo
        self.tasks_repo = tasks_repo
        self.progress_repo = progress_repo
        self.comments_repo = comments_repo

    def _assert_task_mutation_role(self, current_user: User) -> None:
        if current_user.role not in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
            UserRole.SITE_ENGINEER.value,
        ):
            raise PermissionDeniedError("Insufficient permissions")

    def _is_delayed(self, *, task: m.Task, current_date: date) -> bool:
        if task.end_date is None:
            return False
        return (current_date > task.end_date) and (
            task.status != s.TaskStatus.COMPLETED
        )

    async def _assert_progress_or_comment_auth(
        self,
        db: AsyncSession,
        *,
        current_user: User,
        project_id: int,
        task: m.Task,
    ) -> None:
        if current_user.role in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            return

        if any(a.user_id == current_user.id for a in task.assignments):
            return

        allowed = await self.members_repo.is_member(
            db,
            project_id=project_id,
            user_id=current_user.id,
        )

        if not allowed:
            raise PermissionDeniedError("Insufficient permissions")

    def _task_to_out(self, *, task: m.Task, is_delayed: bool) -> s.TaskOut:
        assigned_users_list = (
            [
                s.AssignedUserOut(
                    id=a.user.id,
                    name=a.user.full_name or str(a.user.id),
                    role=(
                        a.user.role.value
                        if hasattr(a.user.role, "value")
                        else str(a.user.role)
                    ),
                )
                for a in task.assignments
            ]
            if hasattr(task, "assignments")
            else []
        )

        return s.TaskOut(
            id=task.id,
            project_id=task.project_id,
            milestone_id=task.milestone_id,
            boq_id=task.boq_id,
            title=task.title,
            description=task.description,
            priority=PRIORITY_MAP.get(task.priority, TaskPriority.MEDIUM),
            status=task.status,
            start_date=task.start_date,
            end_date=task.end_date,
            created_by_user_id=task.created_by_user_id,
            assigned_users=assigned_users_list,
            completion_percentage=task.completion_percentage,
            is_delayed=is_delayed,
            audio_instruction_url=task.audio_instruction_url,
            instruction_image_url=task.instruction_image_url,
            task_icon=task.task_icon,
        )

    async def create_task(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        payload: s.TaskCreate,
        audio_instruction_url: Optional[str] = None,
        instruction_image_url: Optional[str] = None,
    ) -> s.TaskOut:

        self._assert_task_mutation_role(current_user)

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        # =========================
        # MASTER DATA VALIDATION
        # =========================
        if payload.activity_type_id is not None:
            activity = await db.get(ActivityType, payload.activity_type_id)
            if not activity:
                raise NotFoundError("Invalid activity type")

        data = payload.model_dump(exclude_unset=True)

        # =========================
        # MEDIA FILES
        # =========================

        data["audio_instruction_url"] = audio_instruction_url

        data["instruction_image_url"] = instruction_image_url

        if "priority" in data:
            if isinstance(data["priority"], TaskPriority):
                data["priority"] = REVERSE_PRIORITY_MAP[data["priority"]]
            elif isinstance(data["priority"], str):
                try:
                    enum_val = TaskPriority(data["priority"])
                    data["priority"] = REVERSE_PRIORITY_MAP[enum_val]
                except ValueError:
                    raise ValidationError("Invalid priority value")

        # =========================
        # MULTI-ASSIGN LOGIC (FIXED)
        # =========================
        assigned_ids = payload.assigned_user_ids

        data.pop("assigned_user_ids", None)

        if assigned_ids == []:
            raise ValidationError("assigned_user_ids cannot be empty")

        # Resolve all unique user IDs from either input method
        final_user_ids = set()

        if assigned_ids is not None:
            for uid in assigned_ids:
                if uid is not None:
                    final_user_ids.add(uid)

        final_user_ids_list = list(final_user_ids)

        # Validate all users
        for uid in final_user_ids_list:
            assigned_user = await db.scalar(select(User).where(User.id == uid))
            if assigned_user is None:
                raise NotFoundError(f"User {uid} not found")

            is_member = await db.scalar(
                select(m.ProjectMember).where(
                    m.ProjectMember.project_id == project_id,
                    m.ProjectMember.user_id == uid,
                )
            )
            if not is_member:
                raise ValidationError(f"User {uid} not part of project")

        data["created_by_user_id"] = current_user.id

        # We no longer populate assigned_user_id

        # 1. Create ONE Task
        try:
            obj = await self.tasks_repo.create_task(
                db,
                project_id=project_id,
                data=data,
            )
        except IntegrityError:
            await db.rollback()
            raise ConflictError("Task with this title already exists in this project")

        # 2. Create TaskAssignments
        for uid in final_user_ids_list:
            assignment = m.TaskAssignment(
                task_id=obj.id, user_id=uid, assigned_by_user_id=current_user.id
            )
            db.add(assignment)

            await create_notification(
                db,
                user_id=uid,
                title="New Task Assigned",
                message=f"You have been assigned a new task: {obj.title}",
                type="info",
            )

        await db.flush()

        from sqlalchemy.orm import selectinload, joinedload

        # Refetch with relations
        obj = await db.scalar(
            select(m.Task)
            .options(selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user))
            .where(m.Task.id == obj.id)
        )

        return self._task_to_out(
            task=obj,
            is_delayed=self._is_delayed(task=obj, current_date=date.today()),
        )

    async def list_tasks(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        status: Optional[s.TaskStatus],
        assigned_user_id: Optional[int],
        limit: int,
        offset: int,
        search: Optional[str] = None,
        view: Optional[str] = None,
    ) -> PaginatedResponse[s.TaskOut]:

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        from sqlalchemy.orm import selectinload, joinedload

        #  base query
        query = select(m.Task).options(
            selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user)
        )
        count_query = select(func.count(m.Task.id.distinct())).select_from(m.Task)

        #  mandatory filter
        query = query.where(m.Task.project_id == project_id)
        count_query = count_query.where(m.Task.project_id == project_id)

        #  optional filters

        if status:
            query = query.where(m.Task.status == status)
            count_query = count_query.where(m.Task.status == status)

        if assigned_user_id is not None:
            query = query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == assigned_user_id)
            )
            count_query = count_query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == assigned_user_id)
            )

        if search:
            query = query.where(m.Task.title.ilike(f"%{search}%"))
            count_query = count_query.where(m.Task.title.ilike(f"%{search}%"))

        if view == "created":
            query = query.where(m.Task.created_by_user_id == current_user.id)
            count_query = count_query.where(
                m.Task.created_by_user_id == current_user.id
            )

        elif view == "received":
            query = query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == current_user.id)
            )
            count_query = count_query.where(
                m.Task.assignments.any(m.TaskAssignment.user_id == current_user.id)
            )

        #  ordering + pagination
        query = query.order_by(m.Task.id.desc()).limit(limit).offset(offset)

        #  execute
        rows = (await db.execute(query)).scalars().unique().all()
        total = await db.scalar(count_query)

        current_date = date.today()

        items = [
            self._task_to_out(
                task=t,
                is_delayed=self._is_delayed(task=t, current_date=current_date),
            )
            for t in rows
        ]

        meta = PaginationMeta(
            total=int(total or 0),
            limit=limit,
            offset=offset,
        )

        return PaginatedResponse[s.TaskOut](items=items, meta=meta)

    async def get_task(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
    ) -> s.TaskOut:

        project = await self.projects_repo.get_project(db, project_id=project_id)
        if project is None:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        obj = await self.tasks_repo.get_task(
            db,
            project_id=project_id,
            task_id=task_id,
        )
        if obj is None:
            raise NotFoundError("Task not found")

        is_delayed = self._is_delayed(task=obj, current_date=date.today())

        return self._task_to_out(task=obj, is_delayed=is_delayed)

    async def update_task(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        payload: s.TaskUpdate,
        audio_instruction_url: Optional[str] = None,
        instruction_image_url: Optional[str] = None,
        remove_audio: bool = False,
        remove_image: bool = False,
    ) -> s.TaskOut:
        self._assert_task_mutation_role(current_user)

        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        data = payload.model_dump(exclude_unset=True)

        # =====================================
        # MEDIA FILES
        # =====================================

        if audio_instruction_url:

            data["audio_instruction_url"] = audio_instruction_url

        if instruction_image_url:

            data["instruction_image_url"] = instruction_image_url

        if remove_audio:

            data["audio_instruction_url"] = None

        if remove_image:

            data["instruction_image_url"] = None

        if "priority" in data:
            if isinstance(data["priority"], TaskPriority):
                data["priority"] = REVERSE_PRIORITY_MAP[data["priority"]]
            elif isinstance(data["priority"], str):
                try:
                    enum_val = TaskPriority(data["priority"])
                    data["priority"] = REVERSE_PRIORITY_MAP[enum_val]
                except ValueError:
                    raise ValidationError("Invalid priority value")

        if "title" in data and data["title"] is None:
            raise ValidationError("title cannot be null")

        if "priority" in data and data["priority"] is None:
            raise ValidationError("priority cannot be null")

        if "status" in data and data["status"] is None:
            raise ValidationError("status cannot be null")

        # Resolve user IDs
        assigned_ids = data.pop("assigned_user_ids", None)

        has_assignment_update = assigned_ids is not None

        if has_assignment_update:
            final_user_ids = set()
            if assigned_ids is not None:
                for uid in assigned_ids:
                    if uid is not None:
                        final_user_ids.add(uid)

            final_user_ids_list = list(final_user_ids)

            if not final_user_ids_list:
                raise ValidationError("assigned_user_ids cannot be empty")

            for uid in final_user_ids_list:
                assigned_user = await db.scalar(select(User).where(User.id == uid))
                if assigned_user is None:
                    raise NotFoundError(f"User {uid} not found")

                is_member = await db.scalar(
                    select(m.ProjectMember).where(
                        m.ProjectMember.project_id == project_id,
                        m.ProjectMember.user_id == uid,
                    )
                )
                if not is_member:
                    raise ValidationError(f"User {uid} not part of project")

        try:
            await self.tasks_repo.update_task(db, obj=obj, data=data)
            await db.refresh(obj)

            # Sync Assignments
            if has_assignment_update:
                from sqlalchemy.orm import selectinload

                # Load current assignments
                obj_with_assignments = await db.scalar(
                    select(m.Task)
                    .options(selectinload(m.Task.assignments))
                    .where(m.Task.id == obj.id)
                )

                current_uids = {a.user_id for a in obj_with_assignments.assignments}
                target_uids = set(final_user_ids_list)

                to_remove = current_uids - target_uids
                to_add = target_uids - current_uids

                if to_remove:
                    await db.execute(
                        m.TaskAssignment.__table__.delete().where(
                            m.TaskAssignment.task_id == obj.id,
                            m.TaskAssignment.user_id.in_(to_remove),
                        )
                    )

                if to_add:
                    for uid in to_add:
                        assignment = m.TaskAssignment(
                            task_id=obj.id,
                            user_id=uid,
                            assigned_by_user_id=current_user.id,
                        )
                        db.add(assignment)

                        await create_notification(
                            db,
                            user_id=uid,
                            title="New Task Assigned",
                            message=f"You have been assigned a new task: {obj.title}",
                            type="info",
                        )
                await db.flush()

        except IntegrityError:
            await db.rollback()
            raise ConflictError("Task with this title already exists in this project")
        except Exception:
            await db.rollback()
            logger.exception(f"Task update failed id={task_id}")
            raise

        from sqlalchemy.orm import selectinload, joinedload

        obj = await db.scalar(
            select(m.Task)
            .options(selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user))
            .where(m.Task.id == obj.id)
        )

        is_delayed = self._is_delayed(task=obj, current_date=date.today())

        return self._task_to_out(task=obj, is_delayed=is_delayed)

    async def pass_task(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        new_user_id: int,
    ):
        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if not obj:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        new_user = await db.scalar(select(User).where(User.id == new_user_id))
        if not new_user:
            raise NotFoundError("User not found")

        #  FIX: ensure user belongs to project
        is_member = await db.scalar(
            select(m.ProjectMember).where(
                m.ProjectMember.project_id == project_id,
                m.ProjectMember.user_id == new_user_id,
            )
        )
        if not is_member:
            raise ValidationError("User not part of project")

        # Overwrite assignments with this single new user
        await db.execute(
            m.TaskAssignment.__table__.delete().where(
                m.TaskAssignment.task_id == obj.id,
            )
        )
        assignment = m.TaskAssignment(
            task_id=obj.id, user_id=new_user_id, assigned_by_user_id=current_user.id
        )
        db.add(assignment)
        await db.flush()

        from sqlalchemy.orm import selectinload, joinedload

        obj = await db.scalar(
            select(m.Task)
            .options(selectinload(m.Task.assignments).joinedload(m.TaskAssignment.user))
            .where(m.Task.id == obj.id)
        )

        return self._task_to_out(
            task=obj,
            is_delayed=self._is_delayed(task=obj, current_date=date.today()),
        )

    async def update_task_status(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        status: s.TaskStatus,
    ):
        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if not obj:
            raise NotFoundError("Task not found")

        #  FIX: access check
        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        await self.tasks_repo.update_task(
            db,
            obj=obj,
            data={"status": status},
        )

        await db.refresh(obj)

        return self._task_to_out(
            task=obj,
            is_delayed=self._is_delayed(task=obj, current_date=date.today()),
        )

    async def delete_task(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
    ) -> None:
        self._assert_task_mutation_role(current_user)
        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )
        try:
            await self.tasks_repo.delete_task(db, obj=obj)
        except Exception:
            await db.rollback()
            logger.exception(f"Task delete failed id={task_id}")
            raise

    async def update_task_progress(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        payload: s.TaskProgressUpdate,
    ) -> s.TaskProgressOut:
        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        await self._assert_progress_or_comment_auth(
            db, current_user=current_user, project_id=project_id, task=obj
        )

        if payload.percentage < obj.completion_percentage:
            raise ValidationError("Progress cannot decrease")

        progress_obj = await self.progress_repo.create_progress(
            db,
            task_id=obj.id,
            percentage=int(payload.percentage),
            remarks=payload.remarks,
            created_by_user_id=current_user.id,
        )

        await db.refresh(progress_obj)

        await self.tasks_repo.update_task(
            db,
            obj=obj,
            data={"completion_percentage": int(payload.percentage)},
        )

        return s.TaskProgressOut(
            id=progress_obj.id,
            task_id=progress_obj.task_id,
            percentage=progress_obj.percentage,
            remarks=progress_obj.remarks,
            created_at=progress_obj.created_at,
        )

    async def list_task_progress_history(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        limit: int,
        offset: int,
    ) -> PaginatedResponse[s.TaskProgressOut]:

        obj = await self.tasks_repo.get_task(
            db,
            project_id=project_id,
            task_id=task_id,
        )
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        rows, total = await self.progress_repo.list_progress_history(
            db,
            task_id=obj.id,
            limit=limit,
            offset=offset,
        )

        items = [
            s.TaskProgressOut(
                id=p.id,
                task_id=p.task_id,
                percentage=p.percentage,
                remarks=p.remarks,
                created_at=p.created_at,
            )
            for p in rows
        ]

        meta = PaginationMeta(total=int(total), limit=limit, offset=offset)

        return PaginatedResponse[s.TaskProgressOut](items=items, meta=meta)

    async def create_comment(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        payload: s.CommentCreate,
    ) -> s.CommentOut:
        obj = await self.tasks_repo.get_task(db, project_id=project_id, task_id=task_id)
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        await self._assert_progress_or_comment_auth(
            db,
            current_user=current_user,
            project_id=project_id,
            task=obj,
        )

        comment_obj = await self.comments_repo.create_comment(
            db,
            task_id=obj.id,
            author_user_id=current_user.id,
            content=payload.content,
        )

        return s.CommentOut(
            id=comment_obj.id,
            task_id=comment_obj.task_id,
            author_user_id=comment_obj.author_user_id,
            content=comment_obj.content,
        )

    async def list_comments(
        self,
        db: AsyncSession,
        current_user: User,
        *,
        project_id: int,
        task_id: int,
        limit: int,
        offset: int,
    ) -> PaginatedResponse[s.CommentOut]:

        obj = await self.tasks_repo.get_task(
            db,
            project_id=project_id,
            task_id=task_id,
        )
        if obj is None:
            raise NotFoundError("Task not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        rows, total = await self.comments_repo.list_comments(
            db,
            task_id=obj.id,
            limit=limit,
            offset=offset,
        )

        items = [
            s.CommentOut(
                id=c.id,
                task_id=c.task_id,
                author_user_id=c.author_user_id,
                content=c.content,
            )
            for c in rows
        ]

        meta = PaginationMeta(total=int(total), limit=limit, offset=offset)

        return PaginatedResponse[s.CommentOut](items=items, meta=meta)


class SchedulingService:
    def __init__(self, projects_repo: ProjectsRepository):
        self.projects_repo = projects_repo

    async def set_schedule(
        self,
        db: AsyncSession,
        *,
        project_id: int,
        start_date: date,
        end_date: date,
    ):
        project = await self.projects_repo.get_project(db, project_id)
        if not project:
            raise NotFoundError("Project not found")

        if end_date < start_date:
            raise ValidationError("End date cannot be before start date")

        try:
            await self.projects_repo.update_project(
                db,
                project,
                {"start_date": start_date, "end_date": end_date},
            )
        except Exception:
            await db.rollback()
            logger.exception(f"Schedule update failed project_id={project_id}")
            raise

        return {
            "project_id": project_id,
            "start_date": start_date,
            "end_date": end_date,
        }

    async def get_schedule(self, db: AsyncSession, *, project_id: int):
        project = await self.projects_repo.get_project(db, project_id)
        if not project:
            raise NotFoundError("Project not found")

        return {
            "project_id": project_id,
            "start_date": project.start_date,
            "end_date": project.end_date,
        }


class AlertsService:
    def __init__(self, projects_repo: ProjectsRepository, tasks_repo: TasksRepository):
        self.projects_repo = projects_repo
        self.tasks_repo = tasks_repo

    async def get_project_alerts(
        self,
        db: AsyncSession,
        current_user: User,
        pagination: PaginationParams,
    ):
        today = date.today()

        if current_user.role in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            base_query = select(m.Project).where(
                m.Project.end_date < today,
                m.Project.status != s.ProjectStatus.COMPLETED,
            )
        else:
            base_query = (
                select(m.Project)
                .join(m.ProjectMember, m.ProjectMember.project_id == m.Project.id)
                .where(
                    m.ProjectMember.user_id == current_user.id,
                    m.Project.end_date < today,
                    m.Project.status != s.ProjectStatus.COMPLETED,
                )
            )

        base_query = base_query.distinct()

        count_query = select(func.count()).select_from(
            base_query.order_by(None).subquery()
        )
        total = await db.scalar(count_query)

        query = (
            base_query.order_by(m.Project.end_date.asc())
            .limit(pagination.limit)
            .offset(pagination.offset)
        )

        rows = (await db.execute(query)).scalars().all()

        items = [
            {
                "project_id": p.id,
                "project_name": p.project_name,
                "end_date": p.end_date,
                "status": "Delayed",
            }
            for p in rows
        ]

        return PaginatedResponse(
            items=items,
            meta=PaginationMeta(
                total=int(total or 0),
                limit=pagination.limit,
                offset=pagination.offset,
            ),
        )

    async def get_task_alerts(
        self,
        db: AsyncSession,
        current_user: User,
        pagination: PaginationParams,
    ):
        today = date.today()

        if current_user.role in (
            UserRole.ADMIN.value,
            UserRole.PROJECT_MANAGER.value,
        ):
            base_query = select(m.Task).where(
                m.Task.end_date < today,
                m.Task.status != s.TaskStatus.COMPLETED,
            )
        else:
            base_query = (
                select(m.Task)
                .join(m.ProjectMember, m.ProjectMember.project_id == m.Task.project_id)
                .where(
                    m.ProjectMember.user_id == current_user.id,
                    m.Task.end_date < today,
                    m.Task.status != s.TaskStatus.COMPLETED,
                )
            )

        count_query = select(func.count()).select_from(
            base_query.order_by(None).subquery()
        )
        total = await db.scalar(count_query)

        query = (
            base_query.order_by(m.Task.end_date.asc())
            .limit(pagination.limit)
            .offset(pagination.offset)
        )

        rows = (await db.execute(query)).scalars().all()

        items = [
            {
                "task_id": t.id,
                "project_id": t.project_id,
                "title": t.title,
                "end_date": t.end_date,
                "status": "Delayed",
            }
            for t in rows
        ]

        return PaginatedResponse(
            items=items,
            meta=PaginationMeta(
                total=int(total or 0),
                limit=pagination.limit,
                offset=pagination.offset,
            ),
        )


class ReportsService:
    def __init__(self, projects_repo: ProjectsRepository):
        self.projects_repo = projects_repo

    async def get_project_data(
        self,
        db: AsyncSession,
        project_id: int,
        current_user: User,
    ):
        project = await self.projects_repo.get_project(db, project_id)
        if not project:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        return {
            "id": project.id,
            "name": project.project_name,
            "status": project.status,
            "start_date": project.start_date,
            "end_date": project.end_date,
        }

    async def export_excel(
        self,
        db: AsyncSession,
        project_id: int,
        current_user: User,
    ):
        data = await self.get_project_data(db, project_id, current_user)

        wb = Workbook()
        ws = wb.active
        ws.title = "Project Report"

        headers = ["ID", "Name", "Status", "Start Date", "End Date"]
        ws.append(headers)

        ws.append(
            [
                data["id"],
                data["name"],
                str(data["status"]),
                str(data["start_date"]),
                str(data["end_date"]),
            ]
        )

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=project.xlsx"},
        )

    async def export_pdf(
        self,
        db: AsyncSession,
        project_id: int,
        current_user: User,
    ):
        from app.utils.project_report_pdf import generate_project_report_pdf
        from app.models.boq import BOQ
        from app.models.expense import Expense
        from app.models.invoice import Invoice
        from app.models.owner import Owner
        from sqlalchemy import select, func
        from app.models import project as m
        from app.models.user import User as UserModel, UserRole

        project = await self.projects_repo.get_project(db, project_id)
        if not project:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db, project_id=project_id, current_user=current_user
        )

        owner = None
        if getattr(project, "owner_id", None):
            owner = await db.scalar(select(Owner).where(Owner.id == project.owner_id))

        # Financials
        total_boq = await db.scalar(
            select(func.sum(BOQ.total_cost)).where(BOQ.project_id == project_id)
        )
        total_invoiced = await db.scalar(
            select(func.sum(Invoice.total_amount)).where(
                Invoice.project_id == project_id
            )
        )
        total_expenses = await db.scalar(
            select(func.sum(Expense.amount)).where(Expense.project_id == project_id)
        )

        boq_val = float(total_boq or 0)
        invoiced_val = float(total_invoiced or 0)
        expense_val = float(total_expenses or 0)
        profit = invoiced_val - expense_val
        outstanding = boq_val - invoiced_val

        # Tasks
        from sqlalchemy.orm import selectinload

        tasks = (
            (
                await db.execute(
                    select(m.Task)
                    .options(
                        selectinload(m.Task.assignments).joinedload(
                            m.TaskAssignment.user
                        )
                    )
                    .where(m.Task.project_id == project_id)
                )
            )
            .scalars()
            .unique()
            .all()
        )
        total_tasks = len(tasks)
        completed_tasks = sum(
            1
            for t in tasks
            if str(t.status) == "Completed"
            or (hasattr(t.status, "value") and t.status.value == "Completed")
        )
        pending_tasks = sum(
            1
            for t in tasks
            if str(t.status) in ["Pending", "In Progress"]
            or (
                hasattr(t.status, "value")
                and t.status.value in ["Pending", "In Progress"]
            )
        )
        delayed_tasks = sum(
            1
            for t in tasks
            if str(t.status) == "Delayed"
            or (hasattr(t.status, "value") and t.status.value == "Delayed")
        )
        avg_progress = (
            sum(getattr(t, "completion_percentage", 0) for t in tasks) / total_tasks
            if total_tasks
            else 0
        )

        # Milestones
        milestones = (
            (
                await db.execute(
                    select(m.Milestone).where(m.Milestone.project_id == project_id)
                )
            )
            .scalars()
            .all()
        )
        total_milestones = len(milestones)
        completed_milestones = sum(
            1
            for m_obj in milestones
            if str(m_obj.status) == "Completed"
            or (hasattr(m_obj.status, "value") and m_obj.status.value == "Completed")
        )

        # Members
        members_query = (
            select(UserModel)
            .join(m.ProjectMember, m.ProjectMember.user_id == UserModel.id)
            .where(m.ProjectMember.project_id == project_id)
        )
        members_result = await db.execute(members_query)
        members_list = []
        manager = "N/A"
        supervisor = "N/A"
        for user in members_result.scalars().all():
            role_str = (
                user.role.value if hasattr(user.role, "value") else str(user.role)
            )
            members_list.append(
                {
                    "name": user.full_name,
                    "role": role_str,
                    "phone": getattr(user, "phone", "N/A"),
                    "email": user.email,
                }
            )
            if role_str == UserRole.PROJECT_MANAGER.value:
                manager = user.full_name
            elif role_str == UserRole.SITE_ENGINEER.value:
                supervisor = user.full_name

        data = {
            "project": {
                "name": project.project_name,
                "code": project.business_id,
                "client": owner.owner_name if owner else "N/A",
                "type": getattr(project, "type", "Residential"),
                "location": getattr(
                    project,
                    "location",
                    getattr(project, "address", "Ranchi, Jharkhand"),
                ),
                "start_date": project.start_date,
                "end_date": project.end_date,
                "status": "In Progress" if avg_progress < 100 else "Completed",
                "manager": manager,
                "supervisor": supervisor,
            },
            "summary": {
                "progress": round(avg_progress),
                "total_tasks": total_tasks,
                "completed_tasks": completed_tasks,
                "pending_tasks": pending_tasks,
                "delayed_tasks": delayed_tasks,
                "milestones_total": total_milestones,
                "milestones_completed": completed_milestones,
                "team_members": len(members_list),
                "boq_value": boq_val,
                "invoiced": invoiced_val,
                "expenses": expense_val,
                "net_profit": profit,
                "outstanding": outstanding,
            },
            "tasks": [
                {
                    "name": t.title,
                    "assignee": (
                        ", ".join(
                            [
                                a.user.full_name
                                for a in t.assignments
                                if a.user and a.user.full_name
                            ]
                        )
                        if t.assignments
                        else "Unassigned"
                    ),
                    "start_date": t.start_date,
                    "end_date": t.end_date,
                    "status": (
                        t.status.value if hasattr(t.status, "value") else str(t.status)
                    ),
                    "progress": getattr(t, "completion_percentage", 0),
                }
                for t in tasks
            ],
            "milestones": [
                {
                    "name": ms.title,
                    "end_date": ms.end_date,
                    "status": (
                        ms.status.value
                        if hasattr(ms.status, "value")
                        else str(ms.status)
                    ),
                    "completion": (
                        ms.completion_percentage
                        if hasattr(ms, "completion_percentage")
                        else (
                            100
                            if (
                                hasattr(ms.status, "value")
                                and ms.status.value == "Completed"
                            )
                            or str(ms.status) == "Completed"
                            else 0
                        )
                    ),
                }
                for ms in milestones
            ],
            "members": members_list,
        }

        buffer = generate_project_report_pdf(data)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=Project_Report_{project.business_id}.pdf"
            },
        )


def get_tasks_service():
    return TasksService(
        ProjectsRepository(),
        ProjectMembersRepository(),
        TasksRepository(),
        TaskProgressRepository(),
        CommentsRepository(),
    )


def get_projects_service():
    return ProjectsService(ProjectsRepository(), TasksRepository())


def get_milestones_service():
    return MilestonesService(ProjectsRepository(), MilestonesRepository())


def get_scheduling_service():
    return SchedulingService(ProjectsRepository())


def get_alerts_service():
    return AlertsService(ProjectsRepository(), TasksRepository())


def get_project_members_service():
    return ProjectMembersService(ProjectsRepository(), ProjectMembersRepository())


def get_reports_service():
    return ReportsService(ProjectsRepository())


@router.get("/module-summary", response_model=s.ProjectsModuleResponse)
async def projects_module_summary(
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    today = date.today()

    # 1. Summary
    total = await db.scalar(select(func.count(m.Project.id)))
    ongoing = await db.scalar(
        select(func.count(m.Project.id)).where(m.Project.status == "ONGOING")
    )
    completed = await db.scalar(
        select(func.count(m.Project.id)).where(m.Project.status == "COMPLETED")
    )
    delayed = await db.scalar(
        select(func.count(m.Project.id)).where(
            m.Project.status == "ONGOING", m.Project.end_date < today
        )
    )

    summary = s.ProjectsModuleSummary(
        total_projects=total or 0,
        ongoing_sites=ongoing or 0,
        completed_projects=completed or 0,
        delayed_projects=delayed or 0,
    )

    # 2. Activities (Aggregated Feed)
    activities = []

    # a. Task Progress
    task_p = await db.execute(
        select(m.TaskProgress, m.Task.title, m.Project.project_name, User.full_name)
        .join(m.Task, m.TaskProgress.task_id == m.Task.id)
        .join(m.Project, m.Task.project_id == m.Project.id)
        .join(User, m.TaskProgress.created_by_user_id == User.id)
        .order_by(m.TaskProgress.created_at.desc())
        .limit(5)
    )
    for row in task_p.all():
        activities.append(
            s.ProjectActivityItem(
                type="task_completion",
                user_name=row[3],
                description=f"updated progress on {row[1]} to {row[0].percentage}%",
                project_name=row[2],
                timestamp=row[0].created_at,
            )
        )

    # b. Invoices
    invoices = await db.execute(
        select(Invoice, m.Project.project_name)
        .join(m.Project, Invoice.project_id == m.Project.id)
        .order_by(Invoice.created_at.desc())
        .limit(5)
    )
    for row in invoices.all():
        activities.append(
            s.ProjectActivityItem(
                type="invoice",
                user_name="Financial Team",
                description=f"submitted Invoice #{row[0].id} for {row[0].total_amount}",
                project_name=row[1],
                timestamp=row[0].created_at,
            )
        )

    # c. Site Photos
    photos = await db.execute(
        select(m.SitePhoto, m.Project.project_name)
        .join(m.Project, m.SitePhoto.project_id == m.Project.id)
        .order_by(m.SitePhoto.created_at.desc())
        .limit(5)
    )
    for row in photos.all():
        activities.append(
            s.ProjectActivityItem(
                type="photo",
                user_name="Site Bot",
                description="uploaded a new site photo",
                project_name=row[1],
                timestamp=row[0].created_at,
            )
        )

    # d. Issues
    issues = await db.execute(
        select(m.Issue, m.Project.project_name)
        .join(m.Project, m.Issue.project_id == m.Project.id)
        .order_by(m.Issue.created_at.desc())
        .limit(5)
    )
    for row in issues.all():
        activities.append(
            s.ProjectActivityItem(
                type="issue",
                user_name="Site Manager",
                description=f"reported {row[0].priority} issue: {row[0].title}",
                project_name=row[1],
                timestamp=row[0].created_at,
            )
        )

    # Sort and return
    activities.sort(key=lambda x: x.timestamp, reverse=True)

    return s.ProjectsModuleResponse(summary=summary, activities=activities[:15])


@router.post("", response_model=s.ProjectOut)
async def create_project(
    payload: s.ProjectCreate,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectsService = Depends(get_projects_service),
):
    logger.info(f"Creating project name={payload.project_name}")

    try:
        out = await service.create_project(db, current_user, payload=payload)
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception("Project creation failed")
        raise

    logger.info(f"Project created id={out.id}")

    return out


@router.get("", response_model=PaginatedResponse[s.ProjectOut])
async def list_projects(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    search: Optional[str] = None,
    status: Optional[s.ProjectStatus] = None,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectsService = Depends(get_projects_service),
):
    version = await get_cache_version(redis, VERSION_KEY)
    cache_key = f"cache:projects:list:{version}:{current_user.id}:{current_user.role}:{limit}:{offset}:{search}:{status}"
    cached = await cache_get_json(redis, cache_key)
    if cached is not None:
        items = cached.get("items") if isinstance(cached, dict) else None
        if items and isinstance(items, list) and "completion_percentage" in items[0]:
            return PaginatedResponse[s.ProjectOut].model_validate(cached)

    result = await service.list_projects(
        db,
        current_user=current_user,
        limit=limit,
        offset=offset,
        search=search,
        status=status,
    )
    await cache_set_json(redis, cache_key, result.model_dump())
    return result


@router.get("/{project_id}", response_model=s.ProjectOut)
async def get_project(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectsService = Depends(get_projects_service),
):
    version = await get_cache_version(redis, VERSION_KEY)
    cache_key = f"cache:projects:get:{version}:{current_user.id}:{current_user.role}:{project_id}"
    cached_json = await cache_get_json(redis, cache_key)
    if (
        cached_json is not None
        and isinstance(cached_json, dict)
        and "completion_percentage" in cached_json
    ):
        return s.ProjectOut.model_validate(cached_json)

    out = await service.get_project(
        db,
        project_id=project_id,
        current_user=current_user,
    )
    await cache_set_json(redis, cache_key, out.model_dump())
    return out


@router.put("/{project_id}", response_model=s.ProjectOut)
async def update_project(
    project_id: int,
    payload: s.ProjectUpdate,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectsService = Depends(get_projects_service),
):
    logger.info(f"Updating project id={project_id}")

    try:
        out = await service.update_project(
            db, current_user, project_id=project_id, payload=payload
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Project update failed id={project_id}")
        raise

    logger.info(f"Project updated id={project_id}")

    return out


@router.post("/{project_id}/schedule")
async def set_project_schedule(
    project_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: SchedulingService = Depends(get_scheduling_service),
):

    result = await service.set_schedule(
        db, project_id=project_id, start_date=start_date, end_date=end_date
    )

    await bump_cache_version(redis, VERSION_KEY)

    return result


@router.get("/{project_id}/schedule")
async def get_project_schedule(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: SchedulingService = Depends(get_scheduling_service),
):
    return await service.get_schedule(db, project_id=project_id)


@router.get("/{project_id}/progress")
async def get_project_progress(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: ProjectsService = Depends(get_projects_service),
):
    project = await service.get_project(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    return {
        "project_id": project_id,
        "completion_percentage": project.completion_percentage,
        "status": project.status,
    }


@router.get("/alerts/projects")
async def get_project_alerts(
    pagination: PaginationParams = Depends(get_pagination),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
    service: AlertsService = Depends(get_alerts_service),
):
    return await service.get_project_alerts(db, current_user, pagination)


@router.get("/alerts/tasks")
async def get_task_alerts(
    pagination: PaginationParams = Depends(get_pagination),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
    service: AlertsService = Depends(get_alerts_service),
):
    return await service.get_task_alerts(db, current_user, pagination)


@router.delete("/{project_id}", status_code=200)
async def delete_project(
    project_id: int,
    current_user: User = Depends(require_roles(PROJECT_DELETE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectsService = Depends(get_projects_service),
):
    logger.info(f"Deleting project id={project_id}")

    try:
        await service.delete_project(db, current_user, project_id=project_id)
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Project delete failed id={project_id}")
        raise

    logger.info(f"Project deleted id={project_id}")

    return {"success": True, "message": f"Project_id {project_id} deleted successfully"}


@router.post(
    "/{project_id}/members/{user_id}",
    response_model=s.ProjectMemberOut,
    status_code=201,
)
async def assign_project_member(
    project_id: int,
    user_id: int,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectMembersService = Depends(get_project_members_service),
):
    logger.info(f"Assigning member user_id={user_id} project_id={project_id}")

    try:
        out = await service.assign_member(
            db, current_user, project_id=project_id, user_id=user_id
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(
            f"Assign member failed user_id={user_id} project_id={project_id}"
        )
        raise

    logger.info(f"Member assigned user_id={user_id} project_id={project_id}")

    return out


@router.get(
    "/{project_id}/members", response_model=PaginatedResponse[s.ProjectMemberOut]
)
async def list_project_members(
    project_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: ProjectMembersService = Depends(get_project_members_service),
):
    return await service.list_members(
        db, current_user, project_id=project_id, limit=limit, offset=offset
    )


@router.delete("/{project_id}/members/{user_id}", status_code=200)
async def remove_project_member(
    project_id: int,
    user_id: int,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: ProjectMembersService = Depends(get_project_members_service),
):
    logger.info(f"Removing member user_id={user_id} project_id={project_id}")

    try:
        await service.remove_member(
            db, current_user, project_id=project_id, user_id=user_id
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(
            f"Remove member failed user_id={user_id} project_id={project_id}"
        )
        raise

    logger.info(f"Member removed user_id={user_id} project_id={project_id}")

    return {"success": True, "message": "Member Remove successfully"}


@router.get("/{project_id}/report/excel")
async def export_project_excel(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: ReportsService = Depends(get_reports_service),
):
    return await service.export_excel(db, project_id, current_user)


@router.get("/{project_id}/report/pdf")
async def export_project_pdf(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: ReportsService = Depends(get_reports_service),
):
    return await service.export_pdf(db, project_id, current_user)


@router.get("/{project_id}/logs")
async def get_project_logs(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    return {
        "project_id": project_id,
        "message": "Logs available in logging system (file/ELK)",
    }


@router.get("/projects/{project_id}/photos")
async def get_project_photos(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    #  Access check
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    result = await db.execute(
        select(m.SitePhoto)
        .where(m.SitePhoto.project_id == project_id)
        .order_by(m.SitePhoto.date.desc())
    )

    photos = result.scalars().all()

    return [
        {
            "id": p.id,
            "photo_url": p.photo_url,
            "date": p.date,
            "activity": p.activity_tag,
            "description": p.description,
        }
        for p in photos
    ]


@router.put(
    "/{project_id}/ot-policy",
    response_model=s.ProjectOTPolicyOut,
)
async def create_or_update_ot_policy(
    project_id: int,
    payload: s.ProjectOTPolicyCreate,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    # CHECK PROJECT
    project = await db.scalar(select(m.Project).where(m.Project.id == project_id))

    if not project:
        raise NotFoundError("Project not found")

    # ACCESS CHECK
    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    # EXISTING POLICY
    policy = await db.scalar(
        select(m.ProjectOTPolicy).where(m.ProjectOTPolicy.project_id == project_id)
    )

    data = payload.model_dump(exclude_unset=True)

    # UPDATE
    if policy:

        for k, v in data.items():
            setattr(policy, k, v)

    # CREATE
    else:

        policy = m.ProjectOTPolicy(project_id=project_id, **data)

        db.add(policy)

    await db.flush()
    await db.refresh(policy)

    return s.ProjectOTPolicyOut.model_validate(policy, from_attributes=True)


milestones_router = APIRouter(
    prefix="",
    tags=["project_management"],
    dependencies=[default_rate_limiter_dependency()],
)
tasks_router = APIRouter(
    prefix="",
    tags=["project_management"],
    dependencies=[default_rate_limiter_dependency()],
)


@milestones_router.post("/{project_id}/milestones", response_model=s.MilestoneOut)
async def create_milestone(
    project_id: int,
    payload: s.MilestoneCreate,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: MilestonesService = Depends(get_milestones_service),
):
    logger.info(f"Creating milestone project_id={project_id}")

    try:
        out = await service.create_milestone(
            db, current_user, project_id=project_id, payload=payload
        )
        await bump_cache_version(redis, VERSION_KEY)

    except Exception as e:
        # This will print the complete error and stack trace in server logs
        logger.exception(
            f"Milestone creation failed project_id={project_id}. Error: {repr(e)}"
        )

        # This will return the actual error message in the API response
        raise HTTPException(status_code=500, detail=str(e))

    logger.info(f"Milestone created id={out.id}")

    return out


@milestones_router.get(
    "/{project_id}/milestones",
    response_model=PaginatedResponse[s.MilestoneOut],
)
async def list_milestones(
    project_id: int,
    pagination: PaginationParams = Depends(get_pagination),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: MilestonesService = Depends(get_milestones_service),
):
    return await service.list_milestones(
        db,
        project_id=project_id,
        pagination=pagination,
    )


@milestones_router.get(
    "/{project_id}/milestones/{milestone_id}", response_model=s.MilestoneOut
)
async def get_milestone(
    project_id: int,
    milestone_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: MilestonesService = Depends(get_milestones_service),
):
    return await service.get_milestone(
        db, project_id=project_id, milestone_id=milestone_id
    )


@milestones_router.put(
    "/{project_id}/milestones/{milestone_id}", response_model=s.MilestoneOut
)
async def update_milestone(
    project_id: int,
    milestone_id: int,
    payload: s.MilestoneUpdate,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: MilestonesService = Depends(get_milestones_service),
):
    logger.info(f"Updating milestone id={milestone_id}")

    try:
        out = await service.update_milestone(
            db,
            current_user,
            project_id=project_id,
            milestone_id=milestone_id,
            payload=payload,
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Milestone update failed id={milestone_id}")
        raise

    logger.info(f"Milestone updated id={milestone_id}")

    return out


@milestones_router.delete("/{project_id}/milestones/{milestone_id}")
async def delete_milestone(
    project_id: int,
    milestone_id: int,
    current_user: User = Depends(require_roles(PROJECT_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: MilestonesService = Depends(get_milestones_service),
):
    logger.info(f"Deleting milestone id={milestone_id}")

    try:
        await service.delete_milestone(
            db, current_user, project_id=project_id, milestone_id=milestone_id
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Milestone delete failed id={milestone_id}")
        raise

    logger.info(f"Milestone deleted id={milestone_id}")

    return {
        "success": True,
        "message": f"Milestone_id {milestone_id}  deleted successfully",
    }


IMAGE_DIR = "uploads/task_images"

os.makedirs(IMAGE_DIR, exist_ok=True)

AUDIO_DIR = "uploads/task_audio"

os.makedirs(AUDIO_DIR, exist_ok=True)


@tasks_router.post("/{project_id}/tasks", response_model=s.TaskOut)
async def create_task(
    project_id: int,
    payload: s.TaskCreateForm = Depends(),
    audio_file: Optional[UploadFile] = File(None),
    instruction_image: Optional[UploadFile] = File(None),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: TasksService = Depends(get_tasks_service),
):

    logger.info(f"Creating task project_id={project_id}")

    task_payload = payload.to_schema()

    # =========================================
    # SAVE AUDIO FILE
    # =========================================

    audio_instruction_url = None

    if audio_file:

        allowed_audio = ["mp3", "wav", "m4a", "webm", "aac", "ogg"]

        if "." not in audio_file.filename:

            raise ValidationError("Audio file must have extension")

        ext = audio_file.filename.rsplit(".", 1)[-1].lower()

        if ext not in allowed_audio:

            raise ValidationError("Invalid audio format")

        filename = f"{uuid.uuid4().hex}.{ext}"

        filepath = os.path.join(AUDIO_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(audio_file.file, buffer)

        audio_instruction_url = filepath.replace("\\", "/")

    # =========================================
    # SAVE IMAGE FILE
    # =========================================

    instruction_image_url = None

    if instruction_image:

        allowed_images = ["jpg", "jpeg", "png", "webp"]

        if "." not in instruction_image.filename:

            raise ValidationError("Image file must have extension")

        ext = instruction_image.filename.rsplit(".", 1)[-1].lower()

        if ext not in allowed_images:

            raise ValidationError("Invalid image format")

        filename = f"{uuid.uuid4().hex}.{ext}"

        filepath = os.path.join(IMAGE_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(instruction_image.file, buffer)

        instruction_image_url = filepath.replace("\\", "/")

    try:

        out = await service.create_task(
            db,
            current_user,
            project_id=project_id,
            payload=task_payload,
            audio_instruction_url=audio_instruction_url,
            instruction_image_url=instruction_image_url,
        )
        await bump_cache_version(redis, VERSION_KEY)

    except Exception:

        logger.exception(f"Task creation failed project_id={project_id}")

        raise

    if isinstance(out, list):
        logger.info(f"Tasks created count={len(out)}")
    else:
        logger.info(f"Task created id={out.id}")

    return out


@tasks_router.get("/{project_id}/tasks", response_model=PaginatedResponse[s.TaskOut])
async def list_tasks(
    project_id: int,
    status: Optional[s.TaskStatus] = Query(None),
    assigned_user_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    view: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):
    return await service.list_tasks(
        db,
        current_user,
        project_id=project_id,
        status=status,
        assigned_user_id=assigned_user_id,
        limit=limit,
        offset=offset,
        search=search,
        view=view,
    )


@tasks_router.get("/{project_id}/tasks/{task_id}", response_model=s.TaskOut)
async def get_task(
    project_id: int,
    task_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):

    return await service.get_task(
        db, current_user, project_id=project_id, task_id=task_id
    )


@tasks_router.put("/{project_id}/tasks/{task_id}", response_model=s.TaskOut)
async def update_task(
    project_id: int,
    task_id: int,
    payload: s.TaskUpdateForm = Depends(),
    audio_file: Optional[UploadFile] = File(None),
    instruction_image: Optional[UploadFile] = File(None),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: TasksService = Depends(get_tasks_service),
):

    logger.info(f"Updating task id={task_id}")

    task_payload = payload.to_schema()

    # =========================================
    # SAVE AUDIO FILE
    # =========================================

    audio_instruction_url = None

    if audio_file:

        allowed_audio = ["mp3", "wav", "m4a", "webm"]

        if "." not in audio_file.filename:

            raise BadRequestError("Audio file must have extension")

        ext = audio_file.filename.rsplit(".", 1)[-1].lower()

        if ext not in allowed_audio:

            raise BadRequestError("Invalid audio format")

        filename = f"{uuid.uuid4().hex}.{ext}"

        filepath = os.path.join(AUDIO_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(audio_file.file, buffer)

        audio_instruction_url = filepath.replace("\\", "/")

    # =========================================
    # SAVE IMAGE FILE
    # =========================================

    instruction_image_url = None
    if instruction_image:

        allowed_images = ["jpg", "jpeg", "png", "webp"]

        if "." not in instruction_image.filename:

            raise BadRequestError("Image file must have extension")

        ext = instruction_image.filename.rsplit(".", 1)[-1].lower()

        if ext not in allowed_images:

            raise BadRequestError("Invalid image format")

        filename = f"{uuid.uuid4().hex}.{ext}"

        filepath = os.path.join(IMAGE_DIR, filename)

        with open(filepath, "wb") as buffer:

            shutil.copyfileobj(instruction_image.file, buffer)

        instruction_image_url = filepath.replace("\\", "/")

    try:

        out = await service.update_task(
            db,
            current_user,
            project_id=project_id,
            task_id=task_id,
            payload=task_payload,
            audio_instruction_url=audio_instruction_url,
            instruction_image_url=instruction_image_url,
            remove_audio=payload.remove_audio,
            remove_image=payload.remove_image,
        )

        await bump_cache_version(redis, VERSION_KEY)

    except Exception:

        logger.exception(f"Task update failed id={task_id}")

        raise

    logger.info(f"Task updated id={task_id}")

    return out


@tasks_router.patch("/{project_id}/tasks/{task_id}/status")
async def update_status(
    project_id: int,
    task_id: int,
    payload: s.TaskStatusUpdate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):
    return await service.update_task_status(
        db,
        current_user,
        project_id=project_id,
        task_id=task_id,
        status=payload.status,
    )


@tasks_router.post("/{project_id}/tasks/{task_id}/pass")
async def pass_task(
    project_id: int,
    task_id: int,
    payload: s.TaskPass,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):
    return await service.pass_task(
        db,
        current_user,
        project_id=project_id,
        task_id=task_id,
        new_user_id=payload.new_user_id,
    )


@tasks_router.delete("/{project_id}/tasks/{task_id}")
async def delete_task(
    project_id: int,
    task_id: int,
    current_user: User = Depends(require_roles(TASK_DELETE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: TasksService = Depends(get_tasks_service),
):
    logger.info(f"Deleting task id={task_id}")

    try:
        await service.delete_task(
            db, current_user, project_id=project_id, task_id=task_id
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Task delete failed id={task_id}")
        raise

    logger.info(f"Task deleted id={task_id}")

    return {"success": True, "message": f"Task_id {task_id} deleted successfully"}


@tasks_router.post(
    "/{project_id}/tasks/{task_id}/progress", response_model=s.TaskProgressOut
)
async def update_task_progress(
    project_id: int,
    task_id: int,
    payload: s.TaskProgressUpdate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: TasksService = Depends(get_tasks_service),
):
    logger.info(f"Updating task progress task_id={task_id}")

    try:
        out = await service.update_task_progress(
            db, current_user, project_id=project_id, task_id=task_id, payload=payload
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Task progress update failed task_id={task_id}")
        raise

    logger.info(f"Task progress updated task_id={task_id}")

    return out


@tasks_router.get(
    "/{project_id}/tasks/{task_id}/progress",
    response_model=PaginatedResponse[s.TaskProgressOut],
)
async def list_task_progress_history(
    project_id: int,
    task_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):

    return await service.list_task_progress_history(
        db,
        current_user,
        project_id=project_id,
        task_id=task_id,
        limit=limit,
        offset=offset,
    )


@tasks_router.post(
    "/{project_id}/tasks/{task_id}/comments", response_model=s.CommentOut
)
async def create_comment(
    project_id: int,
    task_id: int,
    payload: s.CommentCreate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    service: TasksService = Depends(get_tasks_service),
):
    logger.info(f"Creating comment task_id={task_id}")

    try:
        out = await service.create_comment(
            db,
            current_user,
            project_id=project_id,
            task_id=task_id,
            payload=payload,
        )
        await bump_cache_version(redis, VERSION_KEY)
    except Exception:
        logger.exception(f"Comment creation failed task_id={task_id}")
        raise

    logger.info(f"Comment created id={out.id}")

    return out


@tasks_router.get(
    "/{project_id}/tasks/{task_id}/comments",
    response_model=PaginatedResponse[s.CommentOut],
)
async def list_comments(
    project_id: int,
    task_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    service: TasksService = Depends(get_tasks_service),
):
    return await service.list_comments(
        db,
        current_user,
        project_id=project_id,
        task_id=task_id,
        limit=limit,
        offset=offset,
    )


@router.get("/{project_id}/profit-loss")
async def project_profit_loss(
    project_id: int,
    current_user: User = Depends(require_roles(FINANCIAL_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    project = await db.get(m.Project, project_id)
    if not project:
        raise NotFoundError("Project not found")

    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    total_expense = await db.scalar(
        select(func.sum(Expense.amount)).where(Expense.project_id == project_id)
    )

    total_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(Invoice.project_id == project_id)
    )

    total_expense = float(total_expense or 0)
    total_invoice = float(total_invoice or 0)

    profit = total_invoice - total_expense

    return {
        "project_id": project_id,
        "total_invoice": total_invoice,
        "total_expense": total_expense,
        "profit": profit,
        "status": "profit" if profit >= 0 else "loss",
    }


dsr_router = APIRouter(
    prefix="/dsr",
    tags=["DSR"],
    dependencies=[default_rate_limiter_dependency()],
)


# =========================
# CREATE DSR
# =========================
@dsr_router.post("", response_model=s.DSROut)
async def create_dsr(
    request: Request,
    payload: s.DSRCreate = Depends(),
    photos: Optional[UploadFile] = File(None),
    current_user: User = Depends(require_roles(DSR_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(
        f"Creating DSR project_id={payload.project_id} date={payload.report_date}"
    )

    project = await db.get(m.Project, payload.project_id)
    if not project:
        raise NotFoundError("Project not found")

    await assert_project_access(
        db,
        project_id=payload.project_id,
        current_user=current_user,
    )

    existing = await db.scalar(
        select(m.DailySiteReport).where(
            m.DailySiteReport.project_id == payload.project_id,
            m.DailySiteReport.report_date == payload.report_date,
        )
    )

    if existing:
        raise BadRequestError("DSR already exists for this date")

    # Contractor validation
    if payload.contractor_id:
        contractor = await db.get(Contractor, payload.contractor_id)
        if not contractor:
            raise ValidationError("Invalid contractor_id")

    labour_result = await db.execute(
        select(
            LabourType.skill_category,
            func.count(func.distinct(Labour.id)),
        )
        .join(
            LabourType,
            Labour.labour_type_id == LabourType.id,
        )
        .join(
            UserAttendance,
            Labour.user_id == UserAttendance.user_id,
        )
        .where(
            UserAttendance.project_id == payload.project_id,
            Labour.status == LabourStatus.ACTIVE,
        )
        .group_by(LabourType.skill_category)
    )

    skilled = 0
    unskilled = 0

    for skill_category, count in labour_result.all():
        if skill_category == SkillType.SKILLED:
            skilled += count
        else:
            unskilled += count

        total_labour = skilled + unskilled

    data = payload.model_dump()

    data["created_by_id"] = current_user.id

    data["total_labour"] = total_labour
    data["skilled_labour"] = skilled
    data["unskilled_labour"] = unskilled

    for _ in range(3):
        try:
            data["business_id"] = await generate_business_id(
                db, m.DailySiteReport, "business_id", "DSR"
            )

            obj = m.DailySiteReport(**data)

            db.add(obj)
            await db.flush()
            break

        except IntegrityError:
            await db.rollback()
            continue
    else:
        raise Exception("Failed to generate unique DSR ID")

    # Handle Photos
    if photos:
        upload_dir = "uploads/dsr"
        os.makedirs(upload_dir, exist_ok=True)

        file = photos
        if file.content_type and file.content_type.startswith("image/"):
            content = await file.read()
            if len(content) <= 5 * 1024 * 1024:
                try:
                    img = Image.open(io.BytesIO(content))
                    img.verify()

                    safe_name = pathlib.Path(file.filename or "file").name
                    safe_name = re.sub(r"[^a-zA-Z0-9_.-]", "_", safe_name)

                    ext = pathlib.Path(safe_name).suffix.lower().replace(".", "")
                    if ext in {"jpg", "jpeg", "png"}:
                        filename = f"{uuid.uuid4()}_{safe_name}"
                        path = os.path.join(upload_dir, filename).replace("\\", "/")

                        with open(path, "wb") as f:
                            f.write(content)

                        photo = m.DSRPhoto(dsr_id=obj.id, file_url=path)
                        db.add(photo)
                except Exception:
                    pass

    try:
        await db.flush()
        await db.refresh(obj)

        # Reload with relationships to avoid async lazy-loading issue
        result = await db.execute(
            select(m.DailySiteReport)
            .options(
                selectinload(m.DailySiteReport.contractor),
                selectinload(m.DailySiteReport.created_by),
            )
            .where(m.DailySiteReport.id == obj.id)
        )

        obj = result.scalar_one()

        await bump_cache_version(redis, "cache_version:dsr")

    except Exception:
        await db.rollback()
        logger.exception("DSR creation failed")
        raise

    dsr_out = s.DSROut.model_validate(obj)

    if obj.contractor:
        dsr_out.contractor_name = obj.contractor.name

    if obj.created_by:
        dsr_out.created_by_name = obj.created_by.full_name

    # Add photo URLs to output
    base_url = str(request.base_url).rstrip("/")
    result_photos = await db.execute(
        select(m.DSRPhoto).where(m.DSRPhoto.dsr_id == obj.id)
    )
    dsr_out.photos = [f"{base_url}/{p.file_url}" for p in result_photos.scalars().all()]

    return dsr_out


# =========================
# GET PROJECT DSR
# =========================
@dsr_router.get("/project/{project_id}", response_model=PaginatedResponse[s.DSROut])
async def get_project_dsr(
    project_id: int,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Fetching DSR for project_id={project_id}")

    try:
        project = await db.get(m.Project, project_id)
        if not project:
            raise NotFoundError("Project not found")

        await assert_project_access(
            db,
            project_id=project_id,
            current_user=current_user,
        )

        version = await get_cache_version(redis, "cache_version:dsr")
        cache_key = f"cache:dsr:list:{version}:{project_id}:{limit}:{offset}"

        cached = await cache_get_json(redis, cache_key)
        if cached:
            return PaginatedResponse[s.DSROut].model_validate(cached)

        query = (
            select(m.DailySiteReport)
            .options(
                selectinload(m.DailySiteReport.contractor),
                selectinload(m.DailySiteReport.created_by),
            )
            .where(m.DailySiteReport.project_id == project_id)
            .order_by(m.DailySiteReport.report_date.desc())
            .limit(limit)
            .offset(offset)
        )

        count_query = (
            select(func.count())
            .select_from(m.DailySiteReport)
            .where(m.DailySiteReport.project_id == project_id)
        )

        total = await db.scalar(count_query)
        rows = (await db.execute(query)).scalars().all()

        items = []
        for row in rows:
            dsr = s.DSROut.model_validate(row, from_attributes=True)

            if row.contractor:
                dsr.contractor_name = row.contractor.name

            if row.created_by:
                dsr.created_by_name = row.created_by.full_name

            items.append(dsr.model_dump())

        meta = PaginationMeta(
            total=int(total or 0),
            limit=limit,
            offset=offset,
        )

        result = {
            "items": items,
            "meta": meta.model_dump(),
        }

        await cache_set_json(redis, cache_key, result)

        return PaginatedResponse[s.DSROut].model_validate(result)

    except Exception as e:
        traceback.print_exc()
        raise DataIntegrityError("Data integrity issue")


# =========================
# GET DSR BY ID
# =========================
@dsr_router.get("/{id}", response_model=s.DSROut)
async def get_dsr(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    redis=Depends(get_request_redis),
):
    logger.info(f"Fetching DSR id={id}")

    version = await get_cache_version(redis, "cache_version:dsr")
    cache_key = f"cache:dsr:get:{version}:{id}"

    cached = await cache_get_json(redis, cache_key)
    if cached:
        return s.DSROut.model_validate(cached)

    result = await db.execute(
        select(m.DailySiteReport)
        .options(
            selectinload(m.DailySiteReport.contractor),
            selectinload(m.DailySiteReport.created_by),
        )
        .where(m.DailySiteReport.id == id)
    )

    obj = result.scalar_one_or_none()

    if not obj:
        raise NotFoundError("DSR not found")

    await assert_project_access(
        db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    dsr_out = s.DSROut.model_validate(obj, from_attributes=True)

    if obj.contractor:
        dsr_out.contractor_name = obj.contractor.name

    if obj.created_by:
        dsr_out.created_by_name = obj.created_by.full_name

    await cache_set_json(redis, cache_key, dsr_out.model_dump())

    return dsr_out


# =========================
# UPDATE DSR
# =========================
@dsr_router.put("/{id}", response_model=s.DSROut)
async def update_dsr(
    id: int,
    payload: s.DSRUpdate,
    current_user: User = Depends(require_roles(DSR_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Updating DSR id={id}")

    result = await db.execute(
        select(m.DailySiteReport)
        .options(
            selectinload(m.DailySiteReport.contractor),
            selectinload(m.DailySiteReport.created_by),
        )
        .where(m.DailySiteReport.id == id)
    )

    obj = result.scalar_one_or_none()

    if not obj:
        raise NotFoundError("DSR not found")

    if obj.status == "Approved":
        raise ValidationError("Cannot update approved DSR")

    await assert_project_access(
        db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if payload.report_date:
        existing = await db.scalar(
            select(m.DailySiteReport).where(
                m.DailySiteReport.project_id == obj.project_id,
                m.DailySiteReport.report_date == payload.report_date,
                m.DailySiteReport.id != id,
            )
        )
        if existing:
            raise BadRequestError("DSR already exists for this date")

    update_data = payload.model_dump(exclude_unset=True)

    for k, v in update_data.items():
        if k not in ["project_id", "created_by_id"]:
            setattr(obj, k, v)

    try:
        await db.flush()
    except Exception:
        await db.rollback()
        logger.exception(f"DSR update failed id={id}")
        raise

    await db.refresh(obj)
    await bump_cache_version(redis, "cache_version:dsr")

    dsr_out = s.DSROut.model_validate(obj, from_attributes=True)

    if obj.contractor:
        dsr_out.contractor_name = obj.contractor.name

    if obj.created_by:
        dsr_out.created_by_name = obj.created_by.full_name

    return dsr_out


@dsr_router.get("/project/{project_id}/map")
async def get_dsr_map_points(
    project_id: int,
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(
            m.DailySiteReport.latitude,
            m.DailySiteReport.longitude,
            m.DailySiteReport.report_date,
        ).where(
            m.DailySiteReport.project_id == project_id,
            m.DailySiteReport.latitude.isnot(None),
            m.DailySiteReport.longitude.isnot(None),
        )
    )

    rows = result.all()

    return [
        {
            "lat": r[0],
            "lng": r[1],
            "date": r[2],
        }
        for r in rows
    ]


@dsr_router.get("/project/{project_id}/analytics/labour")
async def labour_trend(
    project_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    if start_date and end_date and end_date < start_date:
        raise BadRequestError("end_date cannot be before start_date")

    query = select(
        m.DailySiteReport.report_date,
        func.sum(m.DailySiteReport.total_labour),
    ).where(m.DailySiteReport.project_id == project_id)

    if start_date:
        query = query.where(m.DailySiteReport.report_date >= start_date)

    if end_date:
        query = query.where(m.DailySiteReport.report_date <= end_date)

    query = query.group_by(m.DailySiteReport.report_date).order_by(
        m.DailySiteReport.report_date
    )

    result = await db.execute(query)
    rows = result.all()

    return [
        {
            "date": r[0],
            "labour": int(r[1] or 0),
        }
        for r in rows
    ]


@dsr_router.get("/project/{project_id}/analytics/contractor")
async def contractor_analytics(
    project_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    if start_date and end_date and end_date < start_date:
        raise BadRequestError("end_date cannot be before start_date")

    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    query = (
        select(
            Contractor.name,
            func.count(m.DailySiteReport.id),
        )
        .select_from(m.DailySiteReport)
        .join(
            Contractor, Contractor.id == m.DailySiteReport.contractor_id, isouter=True
        )
        .where(m.DailySiteReport.project_id == project_id)
    )

    if start_date:
        query = query.where(m.DailySiteReport.report_date >= start_date)

    if end_date:
        query = query.where(m.DailySiteReport.report_date <= end_date)

    query = query.group_by(Contractor.name)

    result = await db.execute(query)
    rows = result.all()

    return [
        {
            "contractor": r[0] or "Unknown",
            "entries": r[1],
        }
        for r in rows
    ]


@dsr_router.delete("/{id}")
async def delete_dsr(
    id: int,
    current_user: User = Depends(require_roles(DSR_DELETE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    logger.info(f"Deleting DSR id={id}")

    obj = await db.get(m.DailySiteReport, id)

    if not obj:
        logger.warning(f"DSR not found id={id}")
        raise NotFoundError("DSR not found")

    try:
        await db.delete(obj)
        await db.flush()

        await bump_cache_version(redis, "cache_version:dsr")

    except Exception:
        await db.rollback()
        logger.exception(f"DSR delete failed id={id}")
        raise

    logger.info(f"DSR deleted id={id}")

    return {"success": True, "message": "DSR deleted successfully"}


@dsr_router.get("/{dsr_id}/photos")
async def get_dsr_photos(
    dsr_id: int,
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    dsr = await db.get(m.DailySiteReport, dsr_id)
    if not dsr:
        raise NotFoundError("DSR not found")

    result = await db.execute(select(m.DSRPhoto).where(m.DSRPhoto.dsr_id == dsr_id))
    rows = result.scalars().all()

    return [{"id": p.id, "url": p.file_url} for p in rows]


@dsr_router.delete("/photo/{photo_id}")
async def delete_dsr_photo(
    photo_id: int,
    current_user: User = Depends(require_roles(DSR_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.DSRPhoto, photo_id)

    if not obj:
        raise NotFoundError("Photo not found")

    try:
        await db.delete(obj)
        await db.flush()
    except Exception:
        await db.rollback()
        raise

    return {"status": "success"}


@dsr_router.get("/project/{project_id}/export")
async def export_dsr_excel(
    project_id: int,
    start_date: Optional[date] = Query(default=None),
    end_date: Optional[date] = Query(default=None),
    contractor_name: Optional[str] = Query(default=None),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Exporting DSR Excel project_id={project_id}")

    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    query = (
        select(m.DailySiteReport, Contractor.name, User.full_name)
        .join(
            Contractor, Contractor.id == m.DailySiteReport.contractor_id, isouter=True
        )
        .join(User, User.id == m.DailySiteReport.created_by_id, isouter=True)
        .where(m.DailySiteReport.project_id == project_id)
    )

    if start_date:
        query = query.where(m.DailySiteReport.report_date >= start_date)

    if end_date:
        query = query.where(m.DailySiteReport.report_date <= end_date)

    if contractor_name:
        contractor_name = contractor_name.strip()
        query = query.where(Contractor.name.ilike(f"%{contractor_name}%"))

    query = query.order_by(m.DailySiteReport.report_date.desc())

    result = await db.execute(query)
    rows = result.all()  # ❗ NOT scalars()

    if not rows:
        raise NotFoundError("No DSR data found")

    wb = Workbook()
    ws = wb.active
    ws.title = "DSR Report"

    headers = [
        "Date",
        "Project ID",
        "Contractor",
        "Weather",
        "Work Done",
        "Work Planned",
        "Labour Count",
        "Material Used",
        "Issues",
        "Remarks",
        "Created By",
    ]
    ws.append(headers)

    for r, contractor_name, created_by_name in rows:
        ws.append(
            [
                str(r.report_date),
                r.project_id,
                contractor_name,
                r.weather,
                r.work_done,
                r.work_planned,
                r.total_labour,
                r.skilled_labour,
                r.unskilled_labour,
                r.material_used,
                r.issues,
                r.remarks,
                created_by_name,
            ]
        )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=dsr_project_{project_id}.xlsx"
        },
    )


@dsr_router.put("/{id}/submit")
async def submit_dsr(
    id: int,
    current_user: User = Depends(require_roles(DSR_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.DailySiteReport, id)

    if not obj:
        raise NotFoundError("DSR not found")

    await assert_project_access(
        db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status != "Draft":
        raise ValidationError("Only draft DSR can be submitted")

    obj.status = "Submitted"

    await db.flush()

    return {"message": "DSR submitted successfully"}


@dsr_router.put("/{id}/approve")
async def approve_dsr(
    id: int,
    current_user: User = Depends(require_roles(DSR_APPROVE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.DailySiteReport, id)

    if not obj:
        raise NotFoundError("DSR not found")

    await assert_project_access(
        db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status != "Submitted":
        raise ValidationError("DSR must be submitted before approval")

    obj.status = "Approved"

    await db.commit()
    await db.refresh(obj)

    return {"message": "DSR approved successfully"}


@dsr_router.put("/{id}/reject")
async def reject_dsr(
    id: int,
    current_user: User = Depends(require_roles(DSR_APPROVE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.DailySiteReport, id)

    if not obj:
        raise NotFoundError("DSR not found")

    await assert_project_access(
        db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status != "Submitted":
        raise ValidationError("Only submitted DSR can be rejected")

    obj.status = "Draft"

    await db.commit()
    await db.refresh(obj)

    return {"message": "DSR rejected and moved to draft"}


issues_router = APIRouter(
    prefix="/issues",
    tags=["Issues"],
    dependencies=[default_rate_limiter_dependency()],
)


@issues_router.post("", response_model=s.IssueOut)
async def create_issue(
    payload: s.IssueCreate,
    redis=Depends(get_request_redis),
    current_user: User = Depends(require_roles(ISSUE_CREATE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Issue create start project_id={payload.project_id}")

    project = await db.get(m.Project, payload.project_id)
    if not project:
        logger.warning(f"Project not found id={payload.project_id}")
        raise NotFoundError("Project not found")

    await assert_project_access(
        db, project_id=payload.project_id, current_user=current_user
    )

    if not payload.title or not payload.title.strip():
        raise ValidationError("title is required")

    title = payload.title.strip()

    existing = await db.scalar(
        select(m.Issue).where(
            m.Issue.project_id == payload.project_id, m.Issue.title == title
        )
    )
    if existing:
        raise ConflictError("Issue with same title already exists in this project")

    try:
        data = payload.model_dump()
        data["title"] = title

        for _ in range(3):
            try:
                data["business_id"] = await generate_business_id(
                    db, m.Issue, "business_id", "ISS"
                )

                obj = m.Issue(**data)

                db.add(obj)
                await db.flush()

                if getattr(obj.priority, "value", str(obj.priority)) == "HIGH":
                    pm = await db.scalar(
                        select(m.ProjectMember.user_id)
                        .join(User, User.id == m.ProjectMember.user_id)
                        .where(
                            m.ProjectMember.project_id == payload.project_id,
                            User.role == UserRole.PROJECT_MANAGER.value,
                        )
                        .limit(1)
                    )
                    if pm:
                        await create_notification(
                            db,
                            user_id=pm,
                            title="Critical Issue Logged",
                            message=f"CRITICAL ISSUE: {obj.title} logged at {project.project_name}",
                            type="alert",
                        )

                break

            except IntegrityError:
                await db.rollback()
                continue
        else:
            raise Exception("Failed to generate unique ISSUE ID")

        await db.refresh(obj)

    except IntegrityError:
        await db.rollback()
        raise ConflictError("Issue with this title already exists in this project")

    except Exception:
        await db.rollback()
        logger.exception("Issue creation failed")
        raise

    logger.info(f"Issue created id={obj.id}")

    await bump_cache_version(redis, VERSION_KEY)

    return s.IssueOut.model_validate(obj)


@issues_router.get("", response_model=PaginatedResponse[s.IssueOut])
async def list_issues(
    pagination: PaginationParams = Depends(),
    status: Optional[s.IssueStatus] = Query(None),
    priority: Optional[s.IssuePriority] = Query(None),
    assigned_to: Optional[int] = Query(None),
    project_id: Optional[int] = Query(None),  # Added
    category: Optional[s.IssueCategory] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("id"),
    order: Optional[str] = Query("desc"),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
):
    pagination = pagination.normalized()

    if current_user.role in (
        UserRole.ADMIN.value,
        UserRole.PROJECT_MANAGER.value,
    ):
        base_query = select(m.Issue)
    else:
        subquery = select(m.ProjectMember.project_id).where(
            m.ProjectMember.user_id == current_user.id
        )

        base_query = select(m.Issue).where(m.Issue.project_id.in_(subquery))

    # Project Filter
    if project_id is not None:
        base_query = base_query.where(m.Issue.project_id == project_id)

    if status is not None:
        base_query = base_query.where(m.Issue.status == status)

    if priority is not None:
        base_query = base_query.where(m.Issue.priority == priority)

    if assigned_to is not None:
        base_query = base_query.where(m.Issue.assigned_to == assigned_to)

    if category is not None:
        base_query = base_query.where(m.Issue.category == category)

    if search and search.strip():
        search_term = f"%{search.strip()}%"
        base_query = base_query.where(
            or_(
                m.Issue.title.ilike(search_term),
                func.coalesce(m.Issue.description, "").ilike(search_term),
            )
        )

    sort_mapping = {
        "id": m.Issue.id,
        "priority": m.Issue.priority,
        "reported_date": m.Issue.reported_date,
        "status": m.Issue.status,
    }

    sort_column = sort_mapping.get(
        sort_by,
        m.Issue.id,
    )

    if order.lower() == "asc":
        base_query = base_query.order_by(sort_column.asc())
    else:
        base_query = base_query.order_by(sort_column.desc())

    count_query = select(func.count()).select_from(base_query.order_by(None).subquery())

    total = await db.scalar(count_query)

    query = base_query.offset(pagination.offset).limit(pagination.limit)

    rows = (await db.execute(query)).scalars().all()

    items = [s.IssueOut.model_validate(row) for row in rows]

    return PaginatedResponse(
        items=items,
        meta=PaginationMeta(
            total=int(total or 0),
            limit=pagination.limit,
            offset=pagination.offset,
        ),
    )


@issues_router.get(
    "/project/{project_id}", response_model=PaginatedResponse[s.IssueOut]
)
async def get_issues_by_project(
    project_id: int,
    pagination: PaginationParams = Depends(),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(
        db,
        project_id=project_id,
        current_user=current_user,
    )

    pagination = pagination.normalized()

    total = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(m.Issue.project_id == project_id)
    )

    query = (
        select(m.Issue)
        .where(m.Issue.project_id == project_id)
        .order_by(m.Issue.id.desc())
        .offset(pagination.offset)
        .limit(pagination.limit)
    )

    rows = (await db.execute(query)).scalars().all()

    items = [s.IssueOut.model_validate(row) for row in rows]

    return PaginatedResponse(
        items=items,
        meta=PaginationMeta(
            total=int(total or 0),
            limit=pagination.limit,
            offset=pagination.offset,
        ),
    )


@issues_router.get("/{id}", response_model=s.IssueOut)
async def get_issue(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
):
    obj = await db.get(m.Issue, id)

    if not obj:
        raise NotFoundError("Issue not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    return s.IssueOut.model_validate(obj)


@issues_router.put("/{id}", response_model=s.IssueOut)
async def update_issue(
    id: int,
    payload: s.IssueUpdate,
    current_user: User = Depends(require_roles(ISSUE_UPDATE_ROLES)),
    redis=Depends(get_request_redis),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Updating issue id={id}")

    obj = await db.get(m.Issue, id)

    if not obj:
        logger.warning(f"Issue not found id={id}")
        raise NotFoundError("Issue not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    data = payload.model_dump(exclude_unset=True)

    if "title" in data and data["title"]:
        title = data["title"].strip()

        existing = await db.scalar(
            select(m.Issue).where(
                m.Issue.project_id == obj.project_id,
                m.Issue.title == title,
                m.Issue.id != obj.id,
            )
        )
        if existing:
            raise ConflictError("Issue with same title already exists in this project")

        data["title"] = title

    if "assigned_to" in data and data["assigned_to"] is not None:
        user = await db.scalar(select(User).where(User.id == data["assigned_to"]))
        if not user:
            raise NotFoundError("Assigned user not found")

        is_member = await db.scalar(
            select(func.count())
            .select_from(m.ProjectMember)
            .where(
                m.ProjectMember.project_id == obj.project_id,
                m.ProjectMember.user_id == data["assigned_to"],
            )
        )
        if not is_member:
            raise ValidationError("Assigned user is not part of this project")

    if data.get("status") == s.IssueStatus.CLOSED and not data.get("resolution"):
        raise ValidationError("Resolution is required to close the issue")

    for k, v in data.items():
        setattr(obj, k, v)

    try:
        await db.flush()

    except IntegrityError:
        await db.rollback()
        raise ConflictError("Issue with this title already exists in this project")

    except Exception:
        await db.rollback()
        logger.exception(f"Issue update failed id={id}")
        raise

    await db.refresh(obj)

    logger.info(f"Issue updated id={id}")

    await bump_cache_version(redis, VERSION_KEY)

    return s.IssueOut.model_validate(obj)


@issues_router.delete("/{id}")
async def delete_issue(
    id: int,
    current_user: User = Depends(require_roles(ISSUE_DELETE_ROLES)),
    redis=Depends(get_request_redis),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Deleting issue id={id}")

    obj = await db.get(m.Issue, id)

    if not obj:
        logger.warning(f"Issue not found id={id}")
        raise NotFoundError("Issue not found")

    if current_user.role not in (
        UserRole.ADMIN.value,
        UserRole.PROJECT_MANAGER.value,
    ):
        is_member = await db.scalar(
            select(func.count())
            .select_from(m.ProjectMember)
            .where(
                m.ProjectMember.project_id == obj.project_id,
                m.ProjectMember.user_id == current_user.id,
            )
        )
        if not is_member:
            raise PermissionDeniedError("User is not part of this project")

    try:
        await db.delete(obj)
        await db.flush()
    except Exception:
        await db.rollback()
        logger.exception(f"Issue delete failed id={id}")
        raise

    logger.info(f"Issue deleted id={id}")

    await bump_cache_version(redis, VERSION_KEY)

    return {"success": True, "message": "Issue deleted successfully"}


@issues_router.get(
    "/project/{project_id}", response_model=PaginatedResponse[s.IssueOut]
)
async def issues_by_project(
    project_id: int,
    pagination: PaginationParams = Depends(),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    pagination = pagination.normalized()

    total = await db.scalar(
        select(func.count()).where(m.Issue.project_id == project_id)
    )

    query = (
        select(m.Issue)
        .where(m.Issue.project_id == project_id)
        .order_by(m.Issue.id.desc())
        .offset(pagination.offset)
        .limit(pagination.limit)
    )

    rows = (await db.execute(query)).scalars().all()

    items = [s.IssueOut.model_validate(r) for r in rows]

    return PaginatedResponse(
        items=items,
        meta=PaginationMeta(
            total=int(total or 0),
            limit=pagination.limit,
            offset=pagination.offset,
        ),
    )


@dsr_router.get("/project/{project_id}/analytics/issues")
async def issue_analytics(
    project_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(require_roles(DSR_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    if start_date and end_date and end_date < start_date:
        raise BadRequestError("end_date cannot be before start_date")

    base_query = select(m.DailySiteReport).where(
        m.DailySiteReport.project_id == project_id
    )

    if start_date:
        base_query = base_query.where(m.DailySiteReport.report_date >= start_date)

    if end_date:
        base_query = base_query.where(m.DailySiteReport.report_date <= end_date)

    total = await db.scalar(select(func.count()).select_from(base_query.subquery()))

    issues = await db.scalar(
        select(func.count()).select_from(
            base_query.where(m.DailySiteReport.issues.isnot(None)).subquery()
        )
    )

    return {
        "total_reports": int(total or 0),
        "reports_with_issues": int(issues or 0),
    }


# =======================work_progress===================================

work_progress_router = APIRouter(
    prefix="/work-progress",
    tags=["Work Progress"],
    dependencies=[default_rate_limiter_dependency()],
)

import json
from decimal import ROUND_HALF_UP, Decimal
from app.models.project import (
    WorkActivity,
    DailyProgressEntry,
    ActivityHistory,
    Project,
    Task,
    TaskProgress,
    Comment,
)


def json_serializer(obj):

    if isinstance(obj, Decimal):
        return float(obj)

    if isinstance(obj, (date, datetime)):
        return obj.isoformat()

    if hasattr(obj, "value"):
        return obj.value

    return str(obj)


async def create_activity_log(
    db,
    activity_id,
    action,
    changed_by,
    old_value=None,
    new_value=None,
    remarks=None,
):

    log = ActivityHistory(
        activity_id=activity_id,
        action=action,
        old_value=(
            json.loads(json.dumps(old_value, default=json_serializer))
            if old_value
            else None
        ),
        new_value=(
            json.loads(json.dumps(new_value, default=json_serializer))
            if new_value
            else None
        ),
        changed_by=changed_by,
        remarks=remarks,
    )

    db.add(log)


def update_activity_status(activity):

    if (
        activity.end_date
        and activity.end_date < date.today()
        and activity.completion_percentage < Decimal("100")
    ):

        activity.status = WorkActivityStatus.DELAY

    elif activity.completion_percentage >= Decimal("100"):

        activity.status = WorkActivityStatus.COMPLETED

    elif activity.completion_percentage > Decimal("0"):

        activity.status = WorkActivityStatus.ON_TRACK

    else:

        activity.status = WorkActivityStatus.NOT_STARTED


# ==============work progress===========================================
# 1. CREATE ACTIVITY

from decimal import Decimal, ROUND_HALF_UP

from fastapi import Depends, HTTPException
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db_session
from app.models.project import Project, ProjectMember
from app.models.user import User
from app.models.work_order import WorkOrder
from app.models.project import WorkActivity


@work_progress_router.post("/activities")
async def create_activity(
    data: s.WorkActivityCreate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    try:

        # ================= CLEAN ACTIVITY NAME =================

        data.activity_name = data.activity_name.strip()

        # ================= VALIDATE ACTIVITY NAME =================

        if not data.activity_name:

            raise HTTPException(
                status_code=400,
                detail="Activity name cannot be empty",
            )

        # ================= VALIDATE PLANNED QUANTITY =================

        if data.planned_quantity <= 0:

            raise HTTPException(
                status_code=400,
                detail="Planned quantity must be greater than 0",
            )

        # ================= VALIDATE DATES =================

        if data.end_date < data.start_date:

            raise HTTPException(
                status_code=400,
                detail="End date cannot be before start date",
            )

        # ================= VALIDATE PROJECT =================

        project = await db.get(
            Project,
            data.project_id,
        )

        if not project:

            raise HTTPException(
                status_code=404,
                detail="Project not found",
            )

        # ================= VALIDATE WORK ORDER =================

        work_order = await db.get(
            WorkOrder,
            data.work_order_id,
        )

        if not work_order:

            raise HTTPException(
                status_code=404,
                detail="Work order not found",
            )

        # ================= CHECK DUPLICATE ACTIVITY =================

        duplicate_stmt = select(WorkActivity).where(
            WorkActivity.project_id == data.project_id,
            WorkActivity.work_order_id == data.work_order_id,
            WorkActivity.boq_code == data.boq_code,
            func.lower(WorkActivity.activity_name) == data.activity_name.lower(),
        )

        duplicate_result = await db.execute(duplicate_stmt)

        existing_activity = duplicate_result.scalars().first()

        # ================= DUPLICATE FOUND =================

        if existing_activity:

            raise HTTPException(
                status_code=400,
                detail="Activity already exists for this BOQ and Work Order",
            )

        # ================= VALIDATE ENGINEER =================

        if data.engineer_id is not None:

            engineer = await db.get(
                User,
                data.engineer_id,
            )

            if not engineer:

                raise HTTPException(
                    status_code=404,
                    detail="Engineer not found",
                )

            # ================= CHECK ENGINEER ASSIGNED TO PROJECT =================

            member_stmt = select(ProjectMember).where(
                ProjectMember.project_id == data.project_id,
                ProjectMember.user_id == data.engineer_id,
            )

            member_result = await db.execute(member_stmt)

            member = member_result.scalar_one_or_none()

            if not member:

                raise HTTPException(
                    status_code=400,
                    detail="Engineer is not assigned to this project",
                )

        # ================= CREATE ACTIVITY =================

        activity = WorkActivity(
            **data.model_dump(),
            total_completed=Decimal("0.00"),
            remaining_quantity=(data.planned_quantity - Decimal("0.00")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            ),
            completion_percentage=Decimal("0.00"),
        )

        # ================= UPDATE STATUS =================

        update_activity_status(activity)

        # ================= ADD TO DB =================

        db.add(activity)

        # ================= GENERATE ID BEFORE COMMIT =================

        await db.flush()

        # ================= AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="CREATE",
            changed_by=current_user.id,
            new_value={
                "activity_name": activity.activity_name,
                "planned_quantity": str(activity.planned_quantity),
                "status": activity.status.value,
                "engineer_id": activity.engineer_id,
            },
        )

        # ================= SAVE =================

        await db.commit()

        # ================= REFRESH =================

        await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "message": "Activity Created",
            "data": activity,
        }

    # ================= HANDLE CUSTOM VALIDATION ERRORS =================

    except HTTPException:

        await db.rollback()

        raise

    # ================= HANDLE DB CONSTRAINT ERRORS =================

    except IntegrityError as e:

        await db.rollback()

        print(
            "INTEGRITY ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=400,
            detail="Duplicate or invalid database entry",
        )

    # ================= HANDLE UNKNOWN ERRORS =================

    except Exception as e:

        await db.rollback()

        print(
            "ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 2. LIST ACTIVITIES

from sqlalchemy import select, func
from fastapi import Query, Depends, HTTPException

# =========================================================
# LIST ACTIVITIES


@work_progress_router.get("/activities")
async def list_activities(
    project_id: int | None = None,
    status: WorkActivityStatus | None = None,
    engineer_id: int | None = None,
    search: str | None = None,
    limit: int = Query(default=10, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # =====================================================
        # REALTIME STATUS REFRESH FIRST
        # =====================================================

        activity_result = await db.execute(select(WorkActivity))

        all_activities = activity_result.scalars().all()

        status_changed = False

        for activity in all_activities:

            old_status = activity.status

            update_activity_status(activity)

            if old_status != activity.status:

                status_changed = True

        if status_changed:

            await db.commit()

        # =====================================================
        # BASE QUERY
        # =====================================================

        stmt = select(WorkActivity)

        count_stmt = select(func.count()).select_from(WorkActivity)

        # =====================================================
        # PROJECT FILTER
        # =====================================================

        if project_id is not None:

            stmt = stmt.where(WorkActivity.project_id == project_id)

            count_stmt = count_stmt.where(WorkActivity.project_id == project_id)

        # =====================================================
        # STATUS FILTER
        # =====================================================

        if status is not None:

            stmt = stmt.where(WorkActivity.status == status)

            count_stmt = count_stmt.where(WorkActivity.status == status)

        # =====================================================
        # ENGINEER FILTER
        # =====================================================

        if engineer_id is not None:

            stmt = stmt.where(WorkActivity.engineer_id == engineer_id)

            count_stmt = count_stmt.where(WorkActivity.engineer_id == engineer_id)

        # =====================================================
        # SEARCH
        # =====================================================

        if search:

            search = search.strip()

            stmt = stmt.where(WorkActivity.activity_name.ilike(f"%{search}%"))

            count_stmt = count_stmt.where(
                WorkActivity.activity_name.ilike(f"%{search}%")
            )

        # =====================================================
        # ORDERING
        # =====================================================

        stmt = stmt.order_by(WorkActivity.created_at.desc())

        # =====================================================
        # PAGINATION
        # =====================================================

        stmt = stmt.offset(offset).limit(limit)

        # =====================================================
        # EXECUTE QUERY
        # =====================================================

        result = await db.execute(stmt)

        activities = result.scalars().all()

        # =====================================================
        # TOTAL COUNT
        # =====================================================

        total_result = await db.execute(count_stmt)

        total_count = total_result.scalar() or 0

        # =====================================================
        # RESPONSE
        # =====================================================

        return {
            "success": True,
            "limit": limit,
            "offset": offset,
            "page_count": len(activities),
            "total_count": total_count,
            "data": activities,
        }

    except Exception as e:

        await db.rollback()

        print(
            "LIST ACTIVITIES ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 3. GET SINGLE ACTIVITY

from sqlalchemy import select
from fastapi import Depends, HTTPException

# =========================================================
# GET SINGLE ACTIVITY


@work_progress_router.get("/activities/{id}")
async def get_activity(
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ID =================

        if id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid activity ID",
            )

        # ================= FETCH ACTIVITY =================

        result = await db.execute(select(WorkActivity).where(WorkActivity.id == id))

        activity = result.scalars().first()

        # ================= NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= REALTIME STATUS UPDATE =================

        old_status = activity.status

        update_activity_status(activity)

        # ================= SAVE IF STATUS CHANGED =================

        if old_status != activity.status:

            await db.commit()

            await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "success": True,
            "data": activity,
        }

    # ================= HANDLE CUSTOM ERRORS =================

    except HTTPException:

        raise

    # ================= HANDLE UNKNOWN ERRORS =================

    except Exception as e:

        await db.rollback()

        print("GET ACTIVITY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 4. UPDATE ACTIVITY

from sqlalchemy import select, func
from sqlalchemy.exc import IntegrityError
from fastapi import Depends, HTTPException
from decimal import Decimal, ROUND_HALF_UP

# =========================================================
# UPDATE ACTIVITY


@work_progress_router.put("/activities/{id}")
async def update_activity(
    id: int,
    data: s.WorkActivityUpdate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= GET ACTIVITY =================

        result = await db.execute(select(WorkActivity).where(WorkActivity.id == id))

        activity = result.scalars().first()

        # ================= NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= CLEAN ACTIVITY NAME =================

        if data.activity_name is not None:

            data.activity_name = data.activity_name.strip()

            if not data.activity_name:

                raise HTTPException(
                    status_code=400,
                    detail="Activity name cannot be empty",
                )

        # ================= STORE OLD DATA FOR AUDIT =================

        old_data = {
            "activity_name": activity.activity_name,
            "planned_quantity": str(activity.planned_quantity),
            "status": activity.status.value,
            "engineer_id": activity.engineer_id,
        }

        # ================= VALIDATE DATES =================

        new_start_date = data.start_date or activity.start_date

        new_end_date = data.end_date or activity.end_date

        if new_end_date < new_start_date:

            raise HTTPException(
                status_code=400,
                detail="End date cannot be before start date",
            )

        # ================= VALIDATE PLANNED QUANTITY =================

        if data.planned_quantity is not None:

            if data.planned_quantity <= 0:

                raise HTTPException(
                    status_code=400,
                    detail="Planned quantity must be greater than 0",
                )

            if data.planned_quantity < activity.total_completed:

                raise HTTPException(
                    status_code=400,
                    detail="Planned quantity cannot be less than completed quantity",
                )

        # ================= CHECK DUPLICATE ACTIVITY =================

        duplicate_name = (
            data.activity_name
            if data.activity_name is not None
            else activity.activity_name
        )

        duplicate_stmt = select(WorkActivity).where(
            WorkActivity.project_id == activity.project_id,
            WorkActivity.work_order_id == activity.work_order_id,
            WorkActivity.boq_code == activity.boq_code,
            func.lower(WorkActivity.activity_name) == duplicate_name.lower(),
            WorkActivity.id != activity.id,
        )

        duplicate_result = await db.execute(duplicate_stmt)

        existing_activity = duplicate_result.scalars().first()

        if existing_activity:

            raise HTTPException(
                status_code=400,
                detail="Another activity with same name already exists",
            )

        # ================= VALIDATE ENGINEER =================

        if data.engineer_id is not None:

            engineer = await db.get(
                User,
                data.engineer_id,
            )

            if not engineer:

                raise HTTPException(
                    status_code=404,
                    detail="Engineer not found",
                )

            # ================= CHECK ENGINEER ASSIGNED TO PROJECT =================

            member_stmt = select(ProjectMember).where(
                ProjectMember.project_id == activity.project_id,
                ProjectMember.user_id == data.engineer_id,
            )

            member_result = await db.execute(member_stmt)

            member = member_result.scalar_one_or_none()

            if not member:

                raise HTTPException(
                    status_code=400,
                    detail="Engineer is not assigned to this project",
                )

        # ================= UPDATE FIELDS =================

        update_data = data.model_dump(exclude_unset=True)

        for key, value in update_data.items():

            setattr(
                activity,
                key,
                value,
            )

        # ================= RECALCULATE VALUES =================

        if activity.planned_quantity > 0:

            activity.remaining_quantity = (
                activity.planned_quantity - activity.total_completed
            ).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            percentage = (
                (activity.total_completed / activity.planned_quantity) * Decimal("100")
            ).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            activity.completion_percentage = min(
                percentage,
                Decimal("100.00"),
            )

        else:

            activity.remaining_quantity = Decimal("0.00")

            activity.completion_percentage = Decimal("0.00")

        # ================= STATUS UPDATE =================

        update_activity_status(activity)

        # ================= STORE NEW DATA FOR AUDIT =================

        new_data = {
            "activity_name": activity.activity_name,
            "planned_quantity": str(activity.planned_quantity),
            "status": activity.status.value,
            "engineer_id": activity.engineer_id,
        }

        # ================= AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="UPDATE",
            changed_by=current_user.id,
            old_value=old_data,
            new_value=new_data,
        )

        # ================= SAVE =================

        await db.commit()

        await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "message": "Activity Updated",
            "data": activity,
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        await db.rollback()

        raise

    # ================= HANDLE DB ERRORS =================

    except IntegrityError as e:

        await db.rollback()

        print(
            "INTEGRITY ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=400,
            detail="Database integrity error",
        )

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        await db.rollback()

        print(
            "UPDATE ACTIVITY ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 5. DELETE ACTIVITY


@work_progress_router.delete("/activities/{id}")
async def delete_activity(
    id: int,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ID =================

        if id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid activity ID",
            )

        # ================= GET ACTIVITY =================

        result = await db.execute(select(WorkActivity).where(WorkActivity.id == id))

        activity = result.scalars().first()

        # ================= NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= CHECK DAILY PROGRESS EXISTS =================

        progress_result = await db.execute(
            select(DailyProgressEntry).where(
                DailyProgressEntry.activity_id == activity.id
            )
        )

        existing_progress = progress_result.scalars().first()

        # ================= PREVENT DELETE =================

        if existing_progress:

            raise HTTPException(
                status_code=400,
                detail="Cannot delete activity with daily progress entries",
            )

        # ================= CREATE DELETE AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="DELETE",
            changed_by=current_user.id,
            old_value={
                "activity_name": activity.activity_name,
                "planned_quantity": str(activity.planned_quantity),
                "status": activity.status.value,
            },
            remarks="Activity deleted",
        )

        # ================= SAVE LOG BEFORE DELETE =================

        await db.flush()

        # ================= DELETE ACTIVITY =================

        await db.delete(activity)

        # ================= SAVE =================

        await db.commit()

        # ================= RESPONSE =================

        return {
            "message": "Activity Deleted Successfully",
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        await db.rollback()

        raise

    # ================= HANDLE DB ERRORS =================

    except IntegrityError as e:

        await db.rollback()

        print("INTEGRITY ERROR =>", str(e))

        raise HTTPException(
            status_code=400,
            detail="Database integrity error",
        )

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        await db.rollback()

        print("DELETE ACTIVITY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 6. ADD DAILY PROGRESS


@work_progress_router.post(
    "/daily-entry",
    response_model=s.DailyProgressWithActivityResponse,
)
async def add_daily_progress(
    data: s.DailyProgressCreate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ACTIVITY ID =================

        if data.activity_id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid activity ID",
            )

        # ================= VALIDATE TODAY PROGRESS =================

        if data.today_progress <= 0:

            raise HTTPException(
                status_code=400,
                detail="Today progress must be greater than 0",
            )

        # ================= CHECK ACTIVITY EXISTS =================

        result = await db.execute(
            select(WorkActivity)
            .where(WorkActivity.id == data.activity_id)
            .with_for_update()
        )

        activity = result.scalars().first()

        # ================= ACTIVITY NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= VALIDATE ENTRY DATE =================

        if data.entry_date < activity.start_date:

            raise HTTPException(
                status_code=400,
                detail="Progress date cannot be before activity start date",
            )

        if activity.end_date and data.entry_date > activity.end_date:

            raise HTTPException(
                status_code=400,
                detail="Progress date cannot be after activity end date",
            )

        # ================= CHECK COMPLETED ACTIVITY =================

        if activity.completion_percentage >= Decimal("100"):

            raise HTTPException(
                status_code=400,
                detail="Cannot add progress to completed activity",
            )

        # ================= STORE OLD DATA =================

        old_data = {
            "total_completed": str(activity.total_completed),
            "completion_percentage": str(activity.completion_percentage),
            "status": activity.status.value,
        }

        # ================= DECIMAL VALUES =================

        current_completed = Decimal(str(activity.total_completed or 0))

        today_progress = Decimal(str(data.today_progress))

        planned_quantity = Decimal(str(activity.planned_quantity or 0))

        # ================= CALCULATE NEW COMPLETION =================

        new_completed = current_completed + today_progress

        # ================= VALIDATE OVER COMPLETION =================

        if new_completed > planned_quantity:

            raise HTTPException(
                status_code=400,
                detail="Progress cannot exceed planned quantity",
            )

        # ================= CREATE ENTRY =================

        entry = DailyProgressEntry(
            activity_id=data.activity_id,
            entry_date=data.entry_date,
            today_progress=data.today_progress,
            remarks=data.remarks,
            created_by=current_user.id,
        )

        db.add(entry)

        # ================= UPDATE ACTIVITY =================

        activity.total_completed = new_completed.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        activity.remaining_quantity = (planned_quantity - new_completed).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        # ================= COMPLETION PERCENTAGE =================

        if planned_quantity > 0:

            percentage = ((new_completed / planned_quantity) * Decimal("100")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            activity.completion_percentage = min(
                percentage,
                Decimal("100.00"),
            )

        else:

            activity.completion_percentage = Decimal("0.00")

        # ================= UPDATE STATUS =================

        update_activity_status(activity)

        # ================= UPDATE WORK ORDER COMPLETED QUANTITY =================

        if activity.work_order_id:

            from app.models.work_order import WorkOrder
            from sqlalchemy import func

            work_order = await db.get(
                WorkOrder,
                activity.work_order_id,
            )

            if work_order:

                result = await db.execute(
                    select(func.sum(WorkActivity.total_completed)).where(
                        WorkActivity.work_order_id == activity.work_order_id
                    )
                )

                total_completed = result.scalar() or Decimal("0")

                work_order.completed_quantity = Decimal(str(total_completed))

        # ================= STORE NEW DATA =================

        new_data = {
            "total_completed": str(activity.total_completed),
            "completion_percentage": str(activity.completion_percentage),
            "status": activity.status.value,
        }

        # ================= AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="DAILY_PROGRESS_ADD",
            changed_by=current_user.id,
            old_value=old_data,
            new_value=new_data,
        )

        # ================= SAVE =================

        await db.commit()

        # ================= REFRESH =================

        await db.refresh(entry)

        await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "message": "Progress Added Successfully",
            "progress": entry,
            "activity": activity,
        }

    except HTTPException:

        await db.rollback()

        raise

    except IntegrityError as e:

        await db.rollback()

        print("INTEGRITY ERROR =>", str(e))

        raise HTTPException(
            status_code=400,
            detail="Progress entry already exists for this activity on this date",
        )

    except Exception as e:

        await db.rollback()

        print("ADD DAILY PROGRESS ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 7. LIST DAILY ENTRIES


@work_progress_router.get("/daily-entry")
async def list_daily_entries(
    activity_id: int | None = None,
    entry_date: date | None = None,
    limit: int = Query(default=10, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= BASE QUERY =================

        stmt = select(DailyProgressEntry)

        count_stmt = select(func.count()).select_from(DailyProgressEntry)

        # ================= FILTER : ACTIVITY =================

        if activity_id is not None:

            if activity_id <= 0:

                raise HTTPException(
                    status_code=400,
                    detail="Invalid activity ID",
                )

            stmt = stmt.where(DailyProgressEntry.activity_id == activity_id)

            count_stmt = count_stmt.where(DailyProgressEntry.activity_id == activity_id)

        # ================= FILTER : ENTRY DATE =================

        if entry_date is not None:

            stmt = stmt.where(DailyProgressEntry.entry_date == entry_date)

            count_stmt = count_stmt.where(DailyProgressEntry.entry_date == entry_date)

        # ================= ORDERING =================

        stmt = stmt.order_by(
            DailyProgressEntry.entry_date.desc(),
            DailyProgressEntry.id.desc(),
        )

        # ================= PAGINATION =================

        stmt = stmt.offset(offset).limit(limit)

        # ================= EXECUTE QUERY =================

        result = await db.execute(stmt)

        entries = result.scalars().all()

        # ================= TOTAL COUNT =================

        total_result = await db.execute(count_stmt)

        total_count = total_result.scalar() or 0

        # ================= RESPONSE =================

        return {
            "success": True,
            "limit": limit,
            "offset": offset,
            "page_count": len(entries),
            "total_count": total_count,
            "data": entries,
        }

    except HTTPException:

        raise

    except Exception as e:

        print("LIST DAILY ENTRIES ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 8. UPDATE DAILY ENTRY


@work_progress_router.put("/daily-entry/{id}")
async def update_daily_entry(
    id: int,
    data: s.DailyProgressUpdate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ID =================

        if id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid daily entry ID",
            )

        # ================= VALIDATE TODAY PROGRESS =================

        if data.today_progress is not None and data.today_progress <= 0:

            raise HTTPException(
                status_code=400,
                detail="Today progress must be greater than 0",
            )

        # ================= LOCK DAILY ENTRY =================

        result = await db.execute(
            select(DailyProgressEntry)
            .where(DailyProgressEntry.id == id)
            .with_for_update()
        )

        entry = result.scalars().first()

        # ================= DAILY ENTRY NOT FOUND =================

        if not entry:

            raise HTTPException(
                status_code=404,
                detail="Daily Entry Not Found",
            )

        # ================= LOCK ACTIVITY =================

        result = await db.execute(
            select(WorkActivity)
            .where(WorkActivity.id == entry.activity_id)
            .with_for_update()
        )

        activity = result.scalars().first()

        # ================= ACTIVITY NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= OLD AUDIT DATA =================

        old_data = {
            "today_progress": str(entry.today_progress),
            "total_completed": str(activity.total_completed),
            "status": activity.status.value,
        }

        # ================= DECIMAL VALUES =================

        old_progress = Decimal(str(entry.today_progress or 0))

        current_total = Decimal(str(activity.total_completed or 0))

        planned_quantity = Decimal(str(activity.planned_quantity or 0))

        # ================= UPDATE ENTRY =================

        update_data = data.model_dump(exclude_unset=True)

        for key, value in update_data.items():

            setattr(entry, key, value)

        # ================= NEW PROGRESS =================

        new_progress = Decimal(str(entry.today_progress or 0))

        difference = new_progress - old_progress

        updated_total = current_total + difference

        # ================= VALIDATE NEGATIVE TOTAL =================

        if updated_total < 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid progress calculation",
            )

        # ================= VALIDATE OVER PROGRESS =================

        if updated_total > planned_quantity:

            raise HTTPException(
                status_code=400,
                detail="Progress cannot exceed planned quantity",
            )

        # ================= UPDATE ACTIVITY =================

        activity.total_completed = updated_total.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        activity.remaining_quantity = (planned_quantity - updated_total).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        # ================= COMPLETION PERCENTAGE =================

        if planned_quantity > 0:

            percentage = ((updated_total / planned_quantity) * Decimal("100")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            # ================= PREVENT ABOVE 100 =================

            activity.completion_percentage = min(
                percentage,
                Decimal("100.00"),
            )

        else:

            activity.completion_percentage = Decimal("0.00")

        # ================= STATUS UPDATE =================

        update_activity_status(activity)

        # ================= NEW AUDIT DATA =================

        new_data = {
            "today_progress": str(entry.today_progress),
            "total_completed": str(activity.total_completed),
            "status": activity.status.value,
        }

        # ================= CREATE AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="DAILY_PROGRESS_UPDATE",
            changed_by=current_user.id,
            old_value=old_data,
            new_value=new_data,
        )

        # ================= SAVE =================

        await db.commit()

        # ================= REFRESH =================

        await db.refresh(entry)

        await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "message": "Daily Entry Updated",
            "data": entry,
            "activity": activity,
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        await db.rollback()

        raise

    # ================= HANDLE DB ERRORS =================

    except IntegrityError as e:

        await db.rollback()

        print("INTEGRITY ERROR =>", str(e))

        raise HTTPException(
            status_code=400,
            detail="Database integrity error",
        )

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        await db.rollback()

        print("UPDATE DAILY ENTRY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 9. DELETE DAILY ENTRY


@work_progress_router.delete("/daily-entry/{id}")
async def delete_daily_entry(
    id: int,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ID =================

        if id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid daily entry ID",
            )

        # ================= LOCK DAILY ENTRY =================

        result = await db.execute(
            select(DailyProgressEntry)
            .where(DailyProgressEntry.id == id)
            .with_for_update()
        )

        entry = result.scalars().first()

        # ================= DAILY ENTRY NOT FOUND =================

        if not entry:

            raise HTTPException(
                status_code=404,
                detail="Daily Entry Not Found",
            )

        # ================= LOCK RELATED ACTIVITY =================

        activity_result = await db.execute(
            select(WorkActivity)
            .where(WorkActivity.id == entry.activity_id)
            .with_for_update()
        )

        activity = activity_result.scalars().first()

        # ================= ACTIVITY NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Related Activity Not Found",
            )

        # ================= OLD DATA FOR AUDIT =================

        old_data = {
            "deleted_progress": str(entry.today_progress),
            "old_total_completed": str(activity.total_completed),
            "old_completion_percentage": str(activity.completion_percentage),
            "old_status": activity.status.value,
        }

        # ================= DECIMAL VALUES =================

        deleted_progress = Decimal(str(entry.today_progress or 0))

        current_total = Decimal(str(activity.total_completed or 0))

        planned_quantity = Decimal(str(activity.planned_quantity or 0))

        # ================= REVERSE PROGRESS =================

        new_total = max(
            Decimal("0.00"),
            current_total - deleted_progress,
        )

        # ================= UPDATE ACTIVITY =================

        activity.total_completed = new_total.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        activity.remaining_quantity = (planned_quantity - new_total).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        # ================= RECALCULATE PERCENTAGE =================

        if planned_quantity > 0:

            percentage = ((new_total / planned_quantity) * Decimal("100")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

            # ================= PREVENT ABOVE 100 =================

            activity.completion_percentage = min(
                percentage,
                Decimal("100.00"),
            )

        else:

            activity.completion_percentage = Decimal("0.00")

        # ================= UPDATE STATUS =================

        update_activity_status(activity)

        # ================= NEW DATA FOR AUDIT =================

        new_data = {
            "new_total_completed": str(activity.total_completed),
            "new_completion_percentage": str(activity.completion_percentage),
            "new_status": activity.status.value,
        }

        # ================= CREATE AUDIT LOG =================

        await create_activity_log(
            db=db,
            activity_id=activity.id,
            action="DAILY_PROGRESS_DELETE",
            changed_by=current_user.id,
            old_value=old_data,
            new_value=new_data,
            remarks="Daily progress entry deleted",
        )

        # ================= DELETE ENTRY =================

        await db.delete(entry)

        # ================= SAVE CHANGES =================

        await db.commit()

        # ================= REFRESH ACTIVITY =================

        await db.refresh(activity)

        # ================= RESPONSE =================

        return {
            "message": "Daily Entry Deleted Successfully",
            "activity": activity,
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        await db.rollback()

        raise

    # ================= HANDLE DB ERRORS =================

    except IntegrityError as e:

        await db.rollback()

        print("INTEGRITY ERROR =>", str(e))

        raise HTTPException(
            status_code=400,
            detail="Database integrity error",
        )

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        await db.rollback()

        print("DELETE DAILY ENTRY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 10. PROJECT SUMMARY


@work_progress_router.get("/project-summary/{project_id}")
async def project_summary(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    try:

        # ================= VALIDATE PROJECT ID =================

        if project_id <= 0:
            raise HTTPException(
                status_code=400,
                detail="Invalid project ID",
            )

        # ================= CHECK PROJECT EXISTS =================

        project = await db.get(Project, project_id)

        if not project:
            raise HTTPException(
                status_code=404,
                detail="Project not found",
            )

        # ================= LOAD ONLY REQUIRED COLUMNS =================

        activity_result = await db.execute(
            select(
                WorkActivity.id,
                WorkActivity.end_date,
                WorkActivity.completion_percentage,
                WorkActivity.status,
            ).where(WorkActivity.project_id == project_id)
        )

        activities = activity_result.all()

        status_changed = False

        # ================= REALTIME STATUS UPDATE =================

        for activity_row in activities:

            activity_id = activity_row.id

            activity = await db.get(
                WorkActivity,
                activity_id,
            )

            old_status = activity.status

            update_activity_status(activity)

            if old_status != activity.status:
                status_changed = True

        if status_changed:
            await db.commit()

        # ================= TOTAL ACTIVITIES =================

        total_result = await db.execute(
            select(func.count())
            .select_from(WorkActivity)
            .where(WorkActivity.project_id == project_id)
        )

        total_activities = total_result.scalar() or 0

        # ================= COMPLETED =================

        completed_result = await db.execute(
            select(func.count())
            .select_from(WorkActivity)
            .where(
                WorkActivity.project_id == project_id,
                WorkActivity.status == WorkActivityStatus.COMPLETED,
            )
        )

        completed = completed_result.scalar() or 0

        # ================= DELAYED =================

        delayed_result = await db.execute(
            select(func.count())
            .select_from(WorkActivity)
            .where(
                WorkActivity.project_id == project_id,
                WorkActivity.status == WorkActivityStatus.DELAY,
            )
        )

        delayed = delayed_result.scalar() or 0

        # ================= ON TRACK =================

        on_track_result = await db.execute(
            select(func.count())
            .select_from(WorkActivity)
            .where(
                WorkActivity.project_id == project_id,
                WorkActivity.status == WorkActivityStatus.ON_TRACK,
            )
        )

        on_track = on_track_result.scalar() or 0

        # ================= NOT STARTED =================

        not_started_result = await db.execute(
            select(func.count())
            .select_from(WorkActivity)
            .where(
                WorkActivity.project_id == project_id,
                WorkActivity.status == WorkActivityStatus.NOT_STARTED,
            )
        )

        not_started = not_started_result.scalar() or 0

        # ================= COMPLETION % =================

        completion_percentage = Decimal("0.00")

        if total_activities > 0:

            completion_percentage = (
                (Decimal(completed) / Decimal(total_activities)) * Decimal("100")
            ).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        # ================= RESPONSE =================

        return {
            "project_id": project_id,
            "total_activities": total_activities,
            "completed_activities": completed,
            "delayed_activities": delayed,
            "on_track_activities": on_track,
            "not_started_activities": not_started,
            "completion_percentage": completion_percentage,
        }

    except HTTPException:
        raise

    except Exception as e:

        await db.rollback()

        print("PROJECT SUMMARY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 11. DELAY REPORT


@work_progress_router.get("/delay-report")
async def delay_report(
    project_id: int | None = None,
    limit: int = Query(default=10, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE PROJECT =================

        if project_id is not None:

            project = await db.get(Project, project_id)

            if not project:

                raise HTTPException(
                    status_code=404,
                    detail="Project not found",
                )

        # ================= REFRESH NON-COMPLETED ACTIVITY STATUS =================

        activity_stmt = select(WorkActivity).where(
            WorkActivity.status != WorkActivityStatus.COMPLETED
        )

        if project_id is not None:

            activity_stmt = activity_stmt.where(WorkActivity.project_id == project_id)

        activity_result = await db.execute(activity_stmt)

        all_activities = activity_result.scalars().all()

        status_changed = False

        for activity in all_activities:

            old_status = activity.status

            update_activity_status(activity)

            if old_status != activity.status:

                status_changed = True

        # ================= SAVE STATUS CHANGES =================

        if status_changed:

            await db.commit()

        # ================= TOTAL COUNT QUERY =================

        count_stmt = (
            select(func.count())
            .select_from(WorkActivity)
            .where(WorkActivity.status == WorkActivityStatus.DELAY)
        )

        # ================= MAIN QUERY =================

        stmt = select(WorkActivity).where(
            WorkActivity.status == WorkActivityStatus.DELAY
        )

        # ================= PROJECT FILTER =================

        if project_id is not None:

            count_stmt = count_stmt.where(WorkActivity.project_id == project_id)

            stmt = stmt.where(WorkActivity.project_id == project_id)

        # ================= PAGINATION + ORDERING =================

        stmt = stmt.order_by(WorkActivity.created_at.desc()).offset(offset).limit(limit)

        # ================= EXECUTE QUERY =================

        result = await db.execute(stmt)

        activities = result.scalars().all()

        # ================= TOTAL COUNT =================

        total_result = await db.execute(count_stmt)

        total_count = total_result.scalar() or 0

        # ================= RESPONSE =================

        return {
            "success": True,
            "project_id": project_id,
            "limit": limit,
            "offset": offset,
            "page_count": len(activities),
            "total_count": total_count,
            "data": activities,
        }

    except HTTPException:

        raise

    except Exception as e:

        await db.rollback()

        print(
            "DELAY REPORT ERROR =>",
            str(e),
        )

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 12. SITE ENGINEER TODAY PROGRESS


@work_progress_router.get("/site-engineer/today-progress")
async def today_progress(
    engineer_id: int,
    # ================= PAGINATION =================
    limit: int = Query(default=10, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ENGINEER ID =================

        if engineer_id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid engineer ID",
            )

        # ================= TOTAL COUNT QUERY =================

        count_stmt = (
            select(func.count())
            .select_from(DailyProgressEntry)
            .join(
                WorkActivity,
                WorkActivity.id == DailyProgressEntry.activity_id,
            )
            .where(
                WorkActivity.engineer_id == engineer_id,
                DailyProgressEntry.entry_date == date.today(),
            )
        )

        # ================= MAIN QUERY =================

        stmt = (
            select(DailyProgressEntry)
            .join(
                WorkActivity,
                WorkActivity.id == DailyProgressEntry.activity_id,
            )
            .where(
                WorkActivity.engineer_id == engineer_id,
                DailyProgressEntry.entry_date == date.today(),
            )
            .order_by(DailyProgressEntry.created_at.desc())
            .offset(offset)
            .limit(limit)
        )

        # ================= EXECUTE MAIN QUERY =================

        result = await db.execute(stmt)

        entries = result.scalars().all()

        # ================= EXECUTE COUNT QUERY =================

        total_result = await db.execute(count_stmt)

        total_count = total_result.scalar() or 0

        # ================= RESPONSE =================

        return {
            "engineer_id": engineer_id,
            "entry_date": date.today(),
            "limit": limit,
            "offset": offset,
            "page_count": len(entries),
            "total_count": total_count,
            "data": entries,
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        raise

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        print("TODAY PROGRESS ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# =========================================================
# 13. ACTIVITY HISTORY


@work_progress_router.get("/activities/{id}/history")
async def activity_history(
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ================= VALIDATE ID =================

        if id <= 0:

            raise HTTPException(
                status_code=400,
                detail="Invalid activity ID",
            )

        # ================= CHECK ACTIVITY EXISTS =================

        activity_result = await db.execute(
            select(WorkActivity).where(WorkActivity.id == id)
        )

        activity = activity_result.scalars().first()

        # ================= ACTIVITY NOT FOUND =================

        if not activity:

            raise HTTPException(
                status_code=404,
                detail="Activity Not Found",
            )

        # ================= FETCH HISTORY =================

        result = await db.execute(
            select(ActivityHistory)
            .where(ActivityHistory.activity_id == id)
            .order_by(ActivityHistory.created_at.desc())
        )

        logs = result.scalars().all()

        # ================= RESPONSE =================

        return {
            "activity_id": id,
            "total_logs": len(logs),
            "data": logs,
        }

    # ================= HANDLE VALIDATION ERRORS =================

    except HTTPException:

        raise

    # ================= HANDLE OTHER ERRORS =================

    except Exception as e:

        print("ACTIVITY HISTORY ERROR =>", str(e))

        raise HTTPException(
            status_code=500,
            detail="Something went wrong",
        )


# 14.==================== WORK PROGRESS LOGS=================

from datetime import date
from sqlalchemy import select, func
from fastapi import Depends, HTTPException, Query


@work_progress_router.get("/logs")
async def get_work_progress_logs(
    activity_id: int | None = None,
    changed_by: int | None = None,
    action: str | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    limit: int = Query(
        default=20,
        ge=1,
        le=100,
    ),
    offset: int = Query(
        default=0,
        ge=0,
    ),
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # =====================
        # VALIDATIONS
        # =====================

        if activity_id is not None and activity_id <= 0:

            raise HTTPException(status_code=400, detail="Invalid activity id")

        if changed_by is not None and changed_by <= 0:

            raise HTTPException(status_code=400, detail="Invalid user id")

        if from_date and to_date and from_date > to_date:

            raise HTTPException(
                status_code=400, detail="From date cannot be greater than To date"
            )

        # =====================
        # BASE QUERY
        # =====================

        stmt = select(ActivityHistory)

        count_stmt = select(func.count()).select_from(ActivityHistory)

        # =====================
        # FILTER ACTIVITY
        # =====================

        if activity_id:

            stmt = stmt.where(ActivityHistory.activity_id == activity_id)

            count_stmt = count_stmt.where(ActivityHistory.activity_id == activity_id)

        # =====================
        # FILTER USER
        # =====================

        if changed_by:

            stmt = stmt.where(ActivityHistory.changed_by == changed_by)

            count_stmt = count_stmt.where(ActivityHistory.changed_by == changed_by)

        # =====================
        # FILTER ACTION
        # =====================

        if action:

            stmt = stmt.where(ActivityHistory.action == action)

            count_stmt = count_stmt.where(ActivityHistory.action == action)

        # =====================
        # DATE FILTER
        # =====================

        if from_date:

            stmt = stmt.where(func.date(ActivityHistory.created_at) >= from_date)

            count_stmt = count_stmt.where(
                func.date(ActivityHistory.created_at) >= from_date
            )

        if to_date:

            stmt = stmt.where(func.date(ActivityHistory.created_at) <= to_date)

            count_stmt = count_stmt.where(
                func.date(ActivityHistory.created_at) <= to_date
            )

        # =====================
        # ORDERING
        # =====================

        stmt = stmt.order_by(ActivityHistory.created_at.desc())

        # =====================
        # PAGINATION
        # =====================

        stmt = stmt.offset(offset).limit(limit)

        # =====================
        # EXECUTE QUERY
        # =====================

        result = await db.execute(stmt)

        logs = result.scalars().all()

        # =====================
        # TOTAL COUNT
        # =====================

        total_result = await db.execute(count_stmt)

        total_count = total_result.scalar()

        # =====================
        # RESPONSE
        # =====================

        return {
            "success": True,
            "limit": limit,
            "offset": offset,
            "page_count": len(logs),
            "total_count": total_count,
            "data": logs,
        }

    except HTTPException:

        raise

    except Exception as e:

        print("WORK PROGRESS LOG ERROR =>", str(e))

        raise HTTPException(status_code=500, detail="Something went wrong")


# 15.======work progress pdf report ============

import io
from decimal import Decimal
from datetime import datetime

from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse

from sqlalchemy import select

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table as PdfTable,
    TableStyle,
    PageBreak,
)

from reportlab.lib import colors as pdf_colors
from reportlab.lib.styles import getSampleStyleSheet


@work_progress_router.get("/reports/pdf")
async def work_progress_pdf_report(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    try:

        # ==================================
        # VALIDATE PROJECT
        # ==================================

        if project_id <= 0:
            raise HTTPException(status_code=400, detail="Invalid project id")

        project_result = await db.execute(
            select(Project).where(Project.id == project_id)
        )

        project = project_result.scalars().first()

        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        # ==================================
        # GET ACTIVITIES
        # ==================================

        activity_result = await db.execute(
            select(WorkActivity)
            .where(WorkActivity.project_id == project_id)
            .order_by(WorkActivity.created_at.desc())
        )

        activities = activity_result.scalars().all()

        # ==================================
        # PDF BUFFER
        # ==================================

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()

        elements = []

        # ==================================
        # PROJECT NAME SAFE
        # ==================================

        project_name = (
            getattr(project, "project_name", None)
            or getattr(project, "name", None)
            or getattr(project, "title", None)
            or f"Project-{project.id}"
        )

        # ==================================
        # HEADER
        # ==================================

        elements.append(Paragraph("WORK PROGRESS REPORT", styles["Title"]))

        elements.append(Spacer(1, 10))

        elements.append(Paragraph(f"Project : {project_name}", styles["Heading2"]))

        elements.append(Paragraph(f"Project ID : {project.id}", styles["Normal"]))

        elements.append(
            Paragraph(
                f"Generated On : {datetime.now().strftime('%d-%m-%Y %H:%M')}",
                styles["Normal"],
            )
        )

        elements.append(Spacer(1, 20))

        # ==================================
        # TABLE HEADER
        # ==================================

        table_data = [["Activity", "Planned", "Completed", "Remaining", "%", "Status"]]

        total_planned = Decimal("0")
        total_completed = Decimal("0")

        delayed_activities = []

        # ==================================
        # LOOP
        # ==================================

        for activity in activities:

            planned_qty = Decimal(str(activity.planned_quantity or 0))

            completed_qty = Decimal(str(activity.total_completed or 0))

            remaining_qty = Decimal(str(activity.remaining_quantity or 0))

            completion_pct = Decimal(str(activity.completion_percentage or 0))

            status_value = (
                activity.status.value
                if hasattr(activity.status, "value")
                else str(activity.status or "UNKNOWN")
            )

            total_planned += planned_qty
            total_completed += completed_qty

            table_data.append(
                [
                    str(activity.activity_name or ""),
                    str(planned_qty),
                    str(completed_qty),
                    str(remaining_qty),
                    str(completion_pct),
                    status_value,
                ]
            )

            if status_value == "DELAY":
                delayed_activities.append(activity)

        # ==================================
        # MAIN TABLE
        # ==================================

        table = PdfTable(table_data)

        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        pdf_colors.HexColor("#1F4E78"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        pdf_colors.white,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        1,
                        pdf_colors.black,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        9,
                    ),
                ]
            )
        )

        elements.append(table)

        elements.append(Spacer(1, 20))

        # ==================================
        # SUMMARY
        # ==================================

        overall_completion = Decimal("0")

        if total_planned > 0:

            overall_completion = (total_completed / total_planned) * Decimal("100")

        elements.append(Paragraph("PROJECT SUMMARY", styles["Heading2"]))

        elements.append(
            Paragraph(f"Total Activities : {len(activities)}", styles["Normal"])
        )

        elements.append(
            Paragraph(f"Total Planned Quantity : {total_planned}", styles["Normal"])
        )

        elements.append(
            Paragraph(f"Total Completed Quantity : {total_completed}", styles["Normal"])
        )

        elements.append(
            Paragraph(
                f"Overall Completion : {round(float(overall_completion),2)}%",
                styles["Normal"],
            )
        )

        # ==================================
        # DELAY PAGE
        # ==================================

        elements.append(PageBreak())

        elements.append(Paragraph("DELAYED ACTIVITIES", styles["Heading1"]))

        delay_table = [
            [
                "Activity",
                "End Date",
                "Completion %",
            ]
        ]

        for item in delayed_activities:

            delay_table.append(
                [
                    str(item.activity_name),
                    str(item.end_date),
                    str(item.completion_percentage),
                ]
            )

        if len(delay_table) > 1:

            dt = PdfTable(delay_table)

            dt.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            pdf_colors.lightgrey,
                        ),
                        (
                            "GRID",
                            (0, 0),
                            (-1, -1),
                            1,
                            pdf_colors.black,
                        ),
                        (
                            "FONTNAME",
                            (0, 0),
                            (-1, 0),
                            "Helvetica-Bold",
                        ),
                    ]
                )
            )

            elements.append(dt)

        else:

            elements.append(Paragraph("No delayed activities found.", styles["Normal"]))

        # ==================================
        # BUILD PDF
        # ==================================

        doc.build(elements)

        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=work_progress_{project_id}.pdf"
            },
        )

    except HTTPException:
        raise

    except Exception as e:

        import traceback

        traceback.print_exc()

        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)}")


# 16. ===========work progress excel report========

import io

from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse

from sqlalchemy import select

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)

from openpyxl.utils import get_column_letter


@work_progress_router.get("/reports/excel")
async def work_progress_excel_report(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    try:

        # =====================================
        # VALIDATE PROJECT
        # =====================================

        if project_id <= 0:
            raise HTTPException(
                status_code=400,
                detail="Invalid project id",
            )

        # =====================================
        # GET PROJECT
        # =====================================

        project_result = await db.execute(
            select(Project).where(Project.id == project_id)
        )

        project = project_result.scalars().first()

        if not project:
            raise HTTPException(
                status_code=404,
                detail="Project not found",
            )

        # =====================================
        # PROJECT NAME SAFE
        # =====================================

        project_name = (
            getattr(project, "project_name", None)
            or getattr(project, "name", None)
            or getattr(project, "title", None)
            or f"Project-{project.id}"
        )

        # =====================================
        # GET ACTIVITIES
        # =====================================

        result = await db.execute(
            select(WorkActivity)
            .where(WorkActivity.project_id == project_id)
            .order_by(WorkActivity.created_at.desc())
        )

        activities = result.scalars().all()

        # =====================================
        # WORKBOOK
        # =====================================

        wb = Workbook()

        # =====================================
        # SHEET 1 : ACTIVITIES
        # =====================================

        ws = wb.active
        ws.title = "Activities"

        headers = [
            "Activity Name",
            "Planned Qty",
            "Completed Qty",
            "Remaining Qty",
            "Completion %",
            "Status",
            "Start Date",
            "End Date",
        ]

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid",
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
        )

        for col_num, header in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col_num,
            )

            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row = 2

        for activity in activities:

            status_value = (
                activity.status.value
                if hasattr(activity.status, "value")
                else str(activity.status or "")
            )

            ws.cell(
                row=row,
                column=1,
            ).value = (
                activity.activity_name or ""
            )

            ws.cell(
                row=row,
                column=2,
            ).value = float(activity.planned_quantity or 0)

            ws.cell(
                row=row,
                column=3,
            ).value = float(activity.total_completed or 0)

            ws.cell(
                row=row,
                column=4,
            ).value = float(activity.remaining_quantity or 0)

            ws.cell(
                row=row,
                column=5,
            ).value = float(activity.completion_percentage or 0)

            ws.cell(
                row=row,
                column=6,
            ).value = status_value

            ws.cell(
                row=row,
                column=7,
            ).value = (
                activity.start_date.strftime("%d-%m-%Y") if activity.start_date else ""
            )

            ws.cell(
                row=row,
                column=8,
            ).value = (
                activity.end_date.strftime("%d-%m-%Y") if activity.end_date else ""
            )

            row += 1

        # =====================================
        # SHEET 2 : SUMMARY
        # =====================================

        summary_sheet = wb.create_sheet(title="Summary")

        total_activities = len(activities)

        total_planned = sum(float(x.planned_quantity or 0) for x in activities)

        total_completed = sum(float(x.total_completed or 0) for x in activities)

        total_remaining = sum(float(x.remaining_quantity or 0) for x in activities)

        total_delayed = len(
            [
                x
                for x in activities
                if (hasattr(x.status, "value") and x.status.value == "DELAY")
            ]
        )

        completion_percentage = 0

        if total_planned > 0:
            completion_percentage = round(
                (total_completed / total_planned) * 100,
                2,
            )

        summary_sheet.append(["Project Name", project_name])

        summary_sheet.append(["Project ID", project.id])

        summary_sheet.append(["Total Activities", total_activities])

        summary_sheet.append(["Total Planned Qty", total_planned])

        summary_sheet.append(["Total Completed Qty", total_completed])

        summary_sheet.append(["Total Remaining Qty", total_remaining])

        summary_sheet.append(["Completion %", completion_percentage])

        summary_sheet.append(["Delayed Activities", total_delayed])

        # =====================================
        # SHEET 3 : DELAY REPORT
        # =====================================

        delay_sheet = wb.create_sheet(title="Delayed Activities")

        delay_sheet.append(
            [
                "Activity",
                "End Date",
                "Completion %",
                "Status",
            ]
        )

        for item in activities:

            status_value = (
                item.status.value
                if hasattr(item.status, "value")
                else str(item.status or "")
            )

            if status_value == "DELAY":

                delay_sheet.append(
                    [
                        item.activity_name or "",
                        (item.end_date.strftime("%d-%m-%Y") if item.end_date else ""),
                        float(item.completion_percentage or 0),
                        status_value,
                    ]
                )

        # =====================================
        # AUTO WIDTH ALL SHEETS
        # =====================================

        for sheet in wb.worksheets:

            for column in sheet.columns:

                max_length = 0

                column_letter = get_column_letter(column[0].column)

                for cell in column:

                    try:

                        if cell.value:

                            max_length = max(
                                max_length,
                                len(str(cell.value)),
                            )

                    except Exception:
                        pass

                sheet.column_dimensions[column_letter].width = max_length + 5

        # =====================================
        # SAVE EXCEL
        # =====================================

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        # =====================================
        # RESPONSE
        # =====================================

        return StreamingResponse(
            output,
            media_type=(
                "application/vnd.openxmlformats-" "officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename=work_progress_{project_id}.xlsx"
            },
        )

    except HTTPException:
        raise

    except Exception as e:

        import traceback

        traceback.print_exc()

        raise HTTPException(
            status_code=500,
            detail=f"{type(e).__name__}: {str(e)}",
        )


# ===================== QC =====================

from fastapi import Form, File, UploadFile, Depends, HTTPException
from uuid import uuid4
import os, shutil

UPLOAD_DIR_QC = "uploads/qc"
os.makedirs(UPLOAD_DIR_QC, exist_ok=True)


def save_qc_file(file: UploadFile) -> str:
    ext = file.filename.split(".")[-1].lower()
    filename = f"{uuid4()}.{ext}"

    path = os.path.join(UPLOAD_DIR_QC, filename)

    with open(path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return f"/uploads/qc/{filename}"


async def validate_and_save_qc_file(file: UploadFile) -> str:
    if not file.content_type.startswith("image/"):
        raise AppError(400, "Only image files allowed")

    allowed_extensions = {"jpg", "jpeg", "png", "webp"}
    ext = file.filename.split(".")[-1].lower()

    if ext not in allowed_extensions:
        raise AppError(400, "Invalid file format")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise AppError(400, "File too large")

    file.file.seek(0)

    return save_qc_file(file)


qc_router = APIRouter(prefix="/qc", tags=["QC"])


@qc_router.post("", response_model=s.QCOut)
async def create_qc(
    payload: s.QCCreate = Depends(),
    report_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    file_url = None

    if report_file:
        file_url = await validate_and_save_qc_file(report_file)

    obj = m.QCRecord(**payload.model_dump(), report_file_url=file_url)
    await assert_task_project(db, payload.task_id, payload.project_id)

    db.add(obj)
    await db.commit()
    await db.refresh(obj)

    return obj


@qc_router.get("/{qc_id}", response_model=s.QCOut)
async def get_qc(
    qc_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
):
    return await db.scalar(select(m.QCRecord).where(m.QCRecord.id == qc_id))


@qc_router.get("", response_model=PaginatedResponse[s.QCOut])
async def list_qc(
    project_id: Optional[int] = None,
    task_id: Optional[int] = None,
    status: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(m.QCRecord)

    if project_id:
        query = query.where(m.QCRecord.project_id == project_id)
    if task_id:
        query = query.where(m.QCRecord.task_id == task_id)
    if status:
        query = query.where(m.QCRecord.status == status)

    count = await db.scalar(select(func.count()).select_from(query.subquery()))
    rows = (await db.execute(query.limit(limit).offset(offset))).scalars().all()

    return PaginatedResponse(
        items=rows, meta=PaginationMeta(total=count, limit=limit, offset=offset)
    )


@qc_router.put("/{qc_id}", response_model=s.QCOut)
async def update_qc(
    qc_id: int,
    data: s.QCCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.QCRecord, qc_id)
    for k, v in data.dict().items():
        setattr(obj, k, v)

    await db.commit()
    await db.refresh(obj)
    return obj


@qc_router.delete("/{qc_id}")
async def delete_qc(
    qc_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.QCRecord, qc_id)
    await db.delete(obj)
    await db.commit()
    return {"message": "QC deleted"}


# ===================== SAFETY =====================

safety_router = APIRouter(prefix="/safety", tags=["Safety"])


@safety_router.post("", response_model=s.SafetyOut)
async def create_incident(
    data: s.SafetyCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    await assert_task_project(db, data.task_id, data.project_id)
    obj = m.SafetyIncident(**data.dict())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


@safety_router.get("/{id}", response_model=s.SafetyOut)
async def get_incident(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
):
    incident = await db.scalar(
        select(m.SafetyIncident).where(m.SafetyIncident.id == id)
    )

    if not incident:
        raise HTTPException(status_code=404, detail="Safety incident not found")

    return incident


@safety_router.get("", response_model=PaginatedResponse[s.SafetyOut])
async def list_incidents(
    project_id: Optional[int] = None,
    violation_type: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(m.SafetyIncident)

    if project_id:
        query = query.where(m.SafetyIncident.project_id == project_id)
    if violation_type:
        query = query.where(m.SafetyIncident.violation_type == violation_type)

    count = await db.scalar(select(func.count()).select_from(query.subquery()))
    rows = (await db.execute(query.limit(limit).offset(offset))).scalars().all()

    return PaginatedResponse(
        items=rows, meta=PaginationMeta(total=count, limit=limit, offset=offset)
    )


@safety_router.put("/{id}", response_model=s.SafetyOut)
async def update_incident(
    id: int,
    data: s.SafetyCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.SafetyIncident, id)
    for k, v in data.dict().items():
        setattr(obj, k, v)

    await db.commit()
    await db.refresh(obj)
    return obj


@safety_router.delete("/{id}")
async def delete_incident(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.SafetyIncident, id)
    await db.delete(obj)
    await db.commit()
    return {"message": "Incident deleted"}


# ===================== CHECKLIST =====================

checklist_router = APIRouter(prefix="/checklists", tags=["Checklist"])


@checklist_router.post("")
async def create_checklist(
    data: s.ChecklistCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = m.Checklist(**data.dict())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


@checklist_router.post("/items")
async def add_item(
    data: s.ChecklistItemCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = m.ChecklistItem(**data.dict())
    db.add(obj)
    await db.commit()
    return obj


@checklist_router.get("")
async def list_checklists(
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(READ_ROLES)),
):
    return (await db.execute(select(m.Checklist))).scalars().all()


@checklist_router.post("/execute")
async def execute_checklist(
    data: s.ChecklistLogCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = m.ChecklistLog(**data.dict())
    db.add(obj)
    await db.commit()
    return obj


@checklist_router.get("/logs", response_model=PaginatedResponse[s.ChecklistLogOut])
async def list_logs(
    project_id: Optional[int] = None,
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(m.ChecklistLog)

    if project_id:
        query = query.where(m.ChecklistLog.project_id == project_id)

    count = await db.scalar(select(func.count()).select_from(query.subquery()))
    rows = (await db.execute(query.limit(limit).offset(offset))).scalars().all()

    items = [s.ChecklistLogOut.model_validate(x) for x in rows]  # ✅ FIX

    return PaginatedResponse(
        items=items, meta=PaginationMeta(total=count, limit=limit, offset=offset)
    )


@checklist_router.delete("/{id}")
async def delete_checklist(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.Checklist, id)
    await db.delete(obj)
    await db.commit()
    return {"message": "Checklist deleted"}


# ===================== Site Photos =====================
UPLOAD_DIR = "uploads/site_photos"
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB


site_photo_router = APIRouter(prefix="/site-photos", tags=["Site Photos"])


#  Upload Photo
@site_photo_router.post("/upload", response_model=s.SitePhotoOut)
async def upload_photo(
    project_id: int = Form(...),
    task_id: Optional[int] = Form(None),
    dsr_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    date: Optional[date] = Form(None),
    activity_tag: Optional[str] = Form(None),
    location_tag: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    #  Validate file type
    await assert_task_project(db, task_id, project_id)
    ext = file.filename.split(".")[-1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, "Invalid file type")

    #  Validate size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "File too large")

    file.file.seek(0)

    #  Unique filename
    filename = f"{uuid4()}.{ext}"
    file_path = os.path.join(UPLOAD_DIR, filename)

    #  Save file
    with open(file_path, "wb") as f:
        f.write(content)

    #  Store URL (NOT raw path)
    file_url = f"/uploads/site_photos/{filename}"

    obj = m.SitePhoto(
        project_id=project_id,
        task_id=task_id,
        dsr_id=dsr_id,
        photo_url=file_url,
        date=date,
        activity_tag=activity_tag,
        location_tag=location_tag,
        description=description,
    )

    db.add(obj)
    await db.commit()
    await db.refresh(obj)

    return obj


#  Filter Photos (IMPORTANT FEATURE)
@site_photo_router.get("", response_model=list[s.SitePhotoOut])
async def list_photos(
    project_id: int,
    activity_tag: Optional[str] = None,
    location_tag: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(m.SitePhoto).where(m.SitePhoto.project_id == project_id)

    if activity_tag:
        query = query.where(m.SitePhoto.activity_tag == activity_tag)

    if location_tag:
        query = query.where(m.SitePhoto.location_tag == location_tag)

    #  Date range filter
    if start_date:
        query = query.where(m.SitePhoto.date >= start_date)

    if end_date:
        query = query.where(m.SitePhoto.date <= end_date)

    result = (await db.execute(query)).scalars().all()
    return result


#  Delete
@site_photo_router.delete("/{photo_id}")
async def delete_photo(
    photo_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
):
    obj = await db.get(m.SitePhoto, photo_id)

    if not obj:
        raise HTTPException(404, "Photo not found")

    #  Delete file from disk
    if obj.photo_url:
        file_path = obj.photo_url.replace("/uploads/", "uploads/")
        if os.path.exists(file_path):
            os.remove(file_path)

    await db.delete(obj)
    await db.commit()

    return {"message": "Photo deleted"}


# ===================== Drawings & Documents =====================

drawing_router = APIRouter(prefix="/drawings", tags=["Drawings & Documents"])


# ===================== Upload =====================


@drawing_router.post("/upload", response_model=s.DrawingOut)
async def upload_drawing(
    project_id: int = Form(...),
    drawing_name: str = Form(...),
    version: str = Form(...),
    date: Optional[date] = Form(None),
    remarks: Optional[str] = Form(None),
    file: UploadFile = File(...),
    current_user: User = Depends(require_roles(DRAWING_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    os.makedirs("uploads/drawings", exist_ok=True)

    validate_drawing_file(file.filename)

    MAX_DRAWING_SIZE = 20 * 1024 * 1024

    content = await file.read()

    if len(content) > MAX_DRAWING_SIZE:
        raise HTTPException(status_code=400, detail="Drawing size cannot exceed 20 MB")

    await file.seek(0)

    ext = os.path.splitext(file.filename)[1].lower()

    unique_name = f"{uuid.uuid4().hex}{ext}"

    file_path = f"uploads/drawings/{unique_name}"

    try:

        # ================= OLD LATEST VERSION FALSE =================

        await db.execute(
            update(m.DrawingDocument)
            .where(
                m.DrawingDocument.project_id == project_id,
                m.DrawingDocument.drawing_name == drawing_name,
                m.DrawingDocument.is_latest_version == True,
            )
            .values(is_latest_version=False)
        )

        # ================= SAVE FILE =================

        with open(file_path, "wb") as f:
            f.write(await file.read())

        # ================= GET NEXT REVISION =================

        latest_revision = await db.scalar(
            select(func.max(m.DrawingDocument.revision_no)).where(
                m.DrawingDocument.project_id == project_id,
                m.DrawingDocument.drawing_name == drawing_name,
            )
        )

        next_revision = (latest_revision or 0) + 1

        # ================= CREATE DRAWING =================

        obj = m.DrawingDocument(
            project_id=project_id,
            drawing_name=drawing_name,
            version=version,
            file_url=file_path,
            date=date,
            remarks=remarks,
            approval_status=DocumentStatus.UNDER_REVIEW,
            revision_no=next_revision,
            is_latest_version=True,
        )

        db.add(obj)

        await db.flush()

        # ================= CREATE APPROVAL =================

        approval = Approval(
            entity_type="drawing",
            entity_id=obj.id,
            requested_by=current_user.id,
            remarks=f"Approval requested for drawing: {drawing_name}",
            status="Pending",
        )

        db.add(approval)

        await db.flush()

        # ================= UPDATE DRAWING APPROVAL REF =================

        obj.approval_id = approval.id

        await db.commit()

        await db.refresh(obj)

        return obj

    except Exception:

        await db.rollback()

        if os.path.exists(file_path):
            os.remove(file_path)

        raise


# ===================== Update =====================


@drawing_router.put("/{id}", response_model=s.DrawingOut)
async def update_drawing(
    id: int,
    payload: s.DrawingUpdate,
    current_user: User = Depends(require_roles(DRAWING_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.DrawingDocument, id)

    if not obj:
        raise NotFoundError("Drawing not found")

    # ================= LOCK APPROVED DRAWINGS =================

    if obj.approval_status == DocumentStatus.APPROVED:
        raise ValidationError("Approved drawing cannot be edited. Create new revision.")

    update_data = payload.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        setattr(obj, field, value)

    await db.commit()

    await db.refresh(obj)

    return obj


# ===================== Approval History =====================


@drawing_router.get("/{id}/approval-history")
async def get_drawing_approval_history(
    id: int,
    current_user: User = Depends(require_roles(DRAWING_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    drawing = await db.get(m.DrawingDocument, id)

    if not drawing:
        raise NotFoundError("Drawing not found")

    result = await db.execute(
        select(Approval)
        .where(
            Approval.entity_type == "drawing",
            Approval.entity_id == id,
        )
        .order_by(Approval.id.desc())
    )

    approvals = result.scalars().all()

    return [
        {
            "id": approval.id,
            "entity_type": approval.entity_type,
            "entity_id": approval.entity_id,
            "requested_by": approval.requested_by,
            "approved_by": approval.approved_by,
            "status": approval.status,
            "remarks": approval.remarks,
            "created_at": approval.created_at,
            "updated_at": approval.updated_at,
        }
        for approval in approvals
    ]


# ===================== Version History =====================


@drawing_router.get("/{project_id}/versions", response_model=list[s.DrawingOut])
async def get_versions(
    project_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DRAWING_READ_ROLES)),
    skip: int = 0,
    limit: int = 50,
):
    result = await db.execute(
        select(m.DrawingDocument)
        .where(m.DrawingDocument.project_id == project_id)
        .order_by(
            m.DrawingDocument.drawing_name.asc(),
            m.DrawingDocument.revision_no.desc(),
            m.DrawingDocument.id.desc(),
        )
        .offset(skip)
        .limit(limit)
    )

    drawings = result.scalars().all()

    return drawings


# ===================== Latest =====================


@drawing_router.get("/{project_id}/latest", response_model=list[s.DrawingOut])
async def get_latest(
    project_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DRAWING_READ_ROLES)),
):
    result = await db.execute(
        select(m.DrawingDocument)
        .where(
            m.DrawingDocument.project_id == project_id,
            m.DrawingDocument.is_latest_version == True,
        )
        .order_by(
            m.DrawingDocument.drawing_name.asc(),
            m.DrawingDocument.revision_no.desc(),
        )
    )

    drawings = result.scalars().all()

    if not drawings:
        raise HTTPException(status_code=404, detail="No drawings found")

    return drawings


# ===================== Delete =====================


@drawing_router.delete("/{id}")
async def delete_drawing(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DRAWING_DELETE_ROLES)),
):
    obj = await db.get(m.DrawingDocument, id)

    if not obj:
        raise NotFoundError("Drawing not found")

    if obj.file_url and os.path.exists(obj.file_url):
        os.remove(obj.file_url)

    await db.execute(
        delete(Approval).where(
            Approval.entity_type == "drawing",
            Approval.entity_id == id,
        )
    )

    await db.delete(obj)

    await db.commit()

    return {"message": "Deleted"}


# ===================== Download =====================


@drawing_router.get("/documents/download/{id}")
async def download_document(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DRAWING_READ_ROLES)),
):
    doc = await db.get(m.DrawingDocument, id)

    if not doc:
        raise NotFoundError("Document not found")

    if not os.path.exists(doc.file_url):
        raise NotFoundError("File not found on server")

    return FileResponse(
        path=doc.file_url,
        filename=os.path.basename(doc.file_url),
        media_type="application/octet-stream",
    )


# ===================== View =====================


@drawing_router.get("/documents/view/{id}")
async def view_document(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(DRAWING_READ_ROLES)),
):
    doc = await db.get(m.DrawingDocument, id)

    if not doc:
        raise NotFoundError("Document not found")

    if not os.path.exists(doc.file_url):
        raise NotFoundError("File not found on server")

    media_type, _ = mimetypes.guess_type(doc.file_url)

    return FileResponse(
        path=doc.file_url,
        filename=os.path.basename(doc.file_url),
        media_type=media_type or "application/octet-stream",
        headers={"Content-Disposition": "inline"},
    )


# ===================== Site Requests =====================


site_request_router = APIRouter(prefix="/site-requests", tags=["Site Requests"])


@site_request_router.post("", response_model=s.SiteRequestOut)
async def create_request(
    payload: s.SiteRequestCreate,
    current_user: User = Depends(require_roles(TASK_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = m.SiteRequest(
        **payload.dict(),
        requested_by=current_user.id,
        status="Pending",
    )

    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


@site_request_router.get("", response_model=list[s.SiteRequestOut])
async def list_requests(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(m.SiteRequest).where(m.SiteRequest.project_id == project_id)
    )
    return result.scalars().all()


@site_request_router.put("/{id}/approve")
async def approve_request(
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.SiteRequest, id)
    obj.status = "Approved"
    obj.approved_by = current_user.id

    await db.commit()
    return {"message": "Approved"}


@site_request_router.put("/{id}/reject")
async def reject_request(
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(m.SiteRequest, id)
    obj.status = "Rejected"
    obj.approved_by = current_user.id

    await db.commit()
    return {"message": "Rejected"}


# ===================== COMMUNICATION =====================
from app.models.messages import Message

communication_router = APIRouter(prefix="/communication", tags=["Communication"])

# ===================== 1. SEND MESSAGE =====================


@communication_router.post("/{project_id}/messages")
async def send_message(
    request: Request,
    project_id: int,
    payload: s.MessageCreate,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    #  FIX: validate parent_id
    parent = None
    if payload.parent_id:
        parent = await db.get(Message, payload.parent_id)
        if not parent:
            raise HTTPException(status_code=400, detail="Parent message not found")

    obj = Message(
        project_id=project_id,
        message=payload.message,
        parent_id=parent.id if parent else None,  #  FIXED
        attachment_url=payload.attachment_url,
        created_by=current_user.id,
        status=MessageStatus.SENT,
    )

    db.add(obj)
    await db.flush()
    await db.refresh(obj)

    redis = request.app.state.redis

    if redis:
        await redis.publish(
            f"project:{project_id}",
            json.dumps(
                {
                    "id": obj.id,
                    "message": obj.message,
                    "parent_id": obj.parent_id,  # added (helpful)
                    "user": current_user.id,
                    "created_at": obj.created_at.isoformat(),
                    "status": obj.status.value,
                    "timestamp": obj.created_at.isoformat(),
                }
            ),
        )

    return obj


# ===================== 2. GET MESSAGES =====================


@communication_router.get("/{project_id}/messages")
async def get_messages(
    project_id: int,
    limit: int = 20,
    offset: int = 0,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    result = await db.execute(
        select(Message)
        .where(Message.project_id == project_id)
        .order_by(Message.created_at.desc())
        .limit(limit)
        .offset(offset)
    )

    return result.scalars().all()


# ===================== 3. GET REPLIES =====================


@communication_router.get("/messages/{message_id}/replies")
async def get_replies(
    message_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(Message, message_id)

    if not obj:
        raise NotFoundError("Message not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    result = await db.execute(
        select(Message)
        .where(Message.parent_id == message_id)
        .order_by(Message.created_at.asc())
    )

    return result.scalars().all()


# ===================== 4. MARK AS READ =====================


@communication_router.put("/messages/{id}/read")
async def mark_read(
    request: Request,
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(Message, id)

    if not obj:
        raise NotFoundError("Message not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    obj.status = MessageStatus.READ
    await db.flush()

    redis = request.app.state.redis

    if redis:
        await redis.publish(
            f"project:{obj.project_id}",
            json.dumps(
                {
                    "type": "read",
                    "message_id": obj.id,
                    "user": current_user.id,
                    "timestamp": datetime.utcnow().isoformat(),
                }
            ),
        )

    return {"message": "read"}


# ===================== 5. MARK AS DELIVERED =====================


@communication_router.put("/messages/{id}/delivered")
async def mark_delivered(
    request: Request,
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(Message, id)

    if not obj:
        raise NotFoundError("Message not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    obj.status = MessageStatus.DELIVERED
    await db.flush()

    redis = request.app.state.redis

    if redis:
        await redis.publish(
            f"project:{obj.project_id}",
            json.dumps(
                {
                    "type": "delivered",
                    "message_id": obj.id,
                    "user": current_user.id,
                    "timestamp": datetime.utcnow().isoformat(),
                }
            ),
        )

    return {"message": "delivered"}


# ===================== 6. UNREAD COUNT =====================


@communication_router.get("/{project_id}/messages/unread-count")
async def unread_count(
    project_id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    count = await db.scalar(
        select(func.count()).where(
            Message.project_id == project_id, Message.status != "read"
        )
    )

    return {"unread": count}


# ===================== 7. DELETE MESSAGE =====================


@communication_router.delete("/messages/{id}")
async def delete_message(
    id: int,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(Message, id)

    if not obj:
        raise NotFoundError("Message not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    if obj.created_by != current_user.id:
        raise ValidationError("Not allowed")

    result = await db.execute(select(Message).where(Message.parent_id == id))
    child = result.scalars().first()

    if child:
        raise ValidationError("Cannot delete message with replies")

    await db.delete(obj)
    await db.flush()

    return {"message": "deleted"}


# ===================== 8. UPDATE MESSAGE =====================


@communication_router.put("/messages/{id}")
async def update_message(
    id: int,
    payload: s.MessageCreate,
    current_user: User = Depends(require_roles(READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(Message, id)

    if not obj:
        raise NotFoundError("Message not found")

    await assert_project_access(
        db, project_id=obj.project_id, current_user=current_user
    )

    if obj.created_by != current_user.id:
        raise ValidationError("Not allowed")

    obj.message = payload.message
    obj.attachment_url = payload.attachment_url

    await db.flush()

    return obj


router.include_router(milestones_router)
router.include_router(tasks_router)
