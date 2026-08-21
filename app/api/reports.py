# ======================================================================
# SECTION 1: REPORT THEME (PdfReportBuilder / ExcelReportBuilder)
# ======================================================================

from app.models.final_measurement import FinalMeasurement
import os
from datetime import datetime
from typing import Optional, Sequence

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────── COLOR PALETTE ──────────────────────────────────
NAVY_BLUE = colors.HexColor("#0B2B5C")
LIGHT_GRAY = colors.HexColor("#F8F9FA")
BORDER_GRAY = colors.HexColor("#E2E8F0")
GREEN = colors.HexColor("#27AE60")
RED = colors.HexColor("#E74C3C")
ORANGE = colors.HexColor("#F39C12")

NAVY_HEX = "0B2B5C"
LIGHT_GRAY_HEX = "F8F9FA"
BORDER_GRAY_HEX = "E2E8F0"
GREEN_HEX = "27AE60"
RED_HEX = "E74C3C"
ORANGE_HEX = "F39C12"

LOGO_PATH = "static/logo.png"


# ═══════════════════════════ PDF HELPERS ═════════════════════════════════════
class PdfReportBuilder:
    """
    Usage:
        b = PdfReportBuilder("EQUIPMENT MANAGEMENT REPORT", landscape_mode=True)
        b.add_info_table([
            ("Report Date", "2026-07-28", "Report Type", "Audit Summary"),
        ])
        b.add_summary_box("2. SUMMARY", [
            "<b>Total Logs:</b> 128 | <b>Period:</b> 2026-07-01 to 2026-07-28",
        ])
        b.add_section_table("3. LOG DETAILS", headers, rows, col_widths)
        pdf_bytes_io = b.build()  # BytesIO, ready for StreamingResponse
    """

    def __init__(
        self, title: str, landscape_mode: bool = False, subtitle: Optional[str] = None
    ):
        import io

        self.stream = io.BytesIO()
        pagesize = landscape(A4) if landscape_mode else A4
        self.doc = SimpleDocTemplate(
            self.stream,
            pagesize=pagesize,
            rightMargin=cm,
            leftMargin=cm,
            topMargin=cm,
            bottomMargin=cm,
        )
        self.page_width = pagesize[0]
        self.usable_width = self.page_width - 2 * cm
        self._section_no = 0
        self.elements = []

        styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            "RptTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=NAVY_BLUE,
            alignment=0,
            spaceAfter=15,
            fontName="Helvetica-Bold",
        )
        self.heading2_style = ParagraphStyle(
            "RptH2",
            parent=styles["Heading2"],
            fontSize=12,
            textColor=NAVY_BLUE,
            spaceBefore=6,
            spaceAfter=10,
            fontName="Helvetica-Bold",
        )
        self.normal_style = ParagraphStyle(
            "RptNormal",
            fontSize=9,
            textColor=colors.black,
            fontName="Helvetica",
            leading=11,
        )
        self.bold_style = ParagraphStyle(
            "RptBold",
            fontSize=9,
            textColor=colors.black,
            fontName="Helvetica-Bold",
            leading=11,
        )
        self.small_style = ParagraphStyle(
            "RptSmall",
            fontSize=7.5,
            textColor=colors.black,
            fontName="Helvetica",
            leading=9,
        )

        self._add_header(title, subtitle)

    # ---------------------------------------------------------------
    def _add_header(self, title: str, subtitle: Optional[str]):
        if os.path.exists(LOGO_PATH):
            logo_img = Image(LOGO_PATH, width=2 * inch, height=0.75 * inch)
        else:
            logo_img = Paragraph("<b>INFRA PILOT</b>", self.title_style)

        title_para = Paragraph(f"<b>{title}</b>", self.title_style)
        header_data = [[logo_img, title_para]]
        header_table = Table(
            header_data, colWidths=[2.5 * inch, self.usable_width - 2.5 * inch]
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, 0), "LEFT"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        self.elements.append(header_table)
        if subtitle:
            self.elements.append(Paragraph(subtitle, self.normal_style))
        self.elements.append(Spacer(1, 0.15 * inch))

    def _next_no(self) -> int:
        self._section_no += 1
        return self._section_no

    # ---------------------------------------------------------------
    def add_info_table(
        self, rows: Sequence[tuple], heading: str = "REPORT INFORMATION"
    ):
        """rows: list of 4-tuples (label, value, label, value)."""
        no = self._next_no()
        data = []
        for label1, value1, label2, value2 in rows:
            data.append(
                [
                    Paragraph(f"<b>{label1}</b>", self.bold_style),
                    str(value1),
                    Paragraph(f"<b>{label2}</b>", self.bold_style),
                    str(value2),
                ]
            )
        col_w = self.usable_width / 6
        table = Table(data, colWidths=[col_w, 2 * col_w, col_w, 2 * col_w])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                    ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY),
                    ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        self.elements.append(Paragraph(f"{no}. {heading}", self.heading2_style))
        self.elements.append(table)
        self.elements.append(Spacer(1, 0.2 * inch))
        return self

    def add_summary_box(self, heading: str, lines: Sequence[str]):
        """lines: list of HTML strings (already containing <b> tags as needed)."""
        no = self._next_no()
        data = [
            [Paragraph(line, self.bold_style if i == 0 else self.normal_style)]
            for i, line in enumerate(lines)
        ]
        box = Table(data, colWidths=[self.usable_width])
        box.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
                    ("BOX", (0, 0), (-1, -1), 1, NAVY_BLUE),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        self.elements.append(Paragraph(f"{no}. {heading}", self.heading2_style))
        self.elements.append(box)
        self.elements.append(Spacer(1, 0.2 * inch))
        return self

    def add_section_table(
        self,
        heading: str,
        headers: Sequence[str],
        rows: Sequence[Sequence],
        col_widths: Optional[Sequence[float]] = None,
        empty_text: str = "No records found.",
        total_row: Optional[Sequence] = None,
    ):
        no = self._next_no()
        self.elements.append(Paragraph(f"{no}. {heading}", self.heading2_style))
        if not rows:
            self.elements.append(Paragraph(empty_text, self.normal_style))
            self.elements.append(Spacer(1, 0.2 * inch))
            return self

        if col_widths is None:
            col_widths = [self.usable_width / len(headers)] * len(headers)

        header_row = [Paragraph(f"<b>{h}</b>", self.small_style) for h in headers]
        data_rows = [[Paragraph(str(v), self.small_style) for v in row] for row in rows]
        table_data = [header_row] + data_rows

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, len(data_rows)),
                [colors.white, LIGHT_GRAY],
            ),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]

        if total_row is not None:
            total_row_p = [
                Paragraph(f"<b>{v}</b>", self.small_style) for v in total_row
            ]
            table_data.append(total_row_p)
            total_idx = len(table_data) - 1
            style_cmds.append(
                ("BACKGROUND", (0, total_idx), (-1, total_idx), LIGHT_GRAY)
            )
            style_cmds.append(
                ("LINEABOVE", (0, total_idx), (-1, total_idx), 1, NAVY_BLUE)
            )

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(style_cmds))
        self.elements.append(table)
        self.elements.append(Spacer(1, 0.2 * inch))
        return self

    def add_paragraph(self, text: str, bold: bool = False):
        self.elements.append(
            Paragraph(text, self.bold_style if bold else self.normal_style)
        )
        return self

    def add_spacer(self, height: float = 0.15):
        self.elements.append(Spacer(1, height * inch))
        return self

    # ---------------------------------------------------------------
    def build(
        self,
        footer_note: str = "Generated by InfraPilot Construction Management System",
    ):
        page_w = self.page_width

        def add_footer(canvas, doc_):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setStrokeColor(NAVY_BLUE)
            canvas.setLineWidth(1)
            canvas.line(cm, 1.5 * cm, page_w - cm, 1.5 * cm)

            canvas.drawString(cm, 1.2 * cm, "Prepared By: ______________")
            canvas.drawString(
                page_w / 2 - 1.5 * cm, 1.2 * cm, "Reviewed By: ______________"
            )
            canvas.drawString(page_w - 5 * cm, 1.2 * cm, "Approved By: ______________")

            canvas.drawString(cm, 0.8 * cm, footer_note)
            canvas.drawRightString(page_w - cm, 0.8 * cm, f"Page {doc_.page}")
            canvas.restoreState()

        self.doc.build(self.elements, onFirstPage=add_footer, onLaterPages=add_footer)
        self.stream.seek(0)
        return self.stream


# ═══════════════════════════ EXCEL HELPERS ═══════════════════════════════════
class ExcelReportBuilder:
    """
    Usage:
        b = ExcelReportBuilder("Audit Summary Report", project_line="All Projects")
        b.add_summary_sheet([("Total Logs", 128), ("Period", "Jul 2026")])
        b.add_data_sheet("Logs", headers, rows, currency_cols=[5])
        xlsx_bytes_io = b.build()
    """

    def __init__(self, title: str, project_line: Optional[str] = None):
        self.title = title
        self.project_line = project_line
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        self.header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        self.header_fill = PatternFill("solid", fgColor=NAVY_HEX)
        self.title_font = Font(name="Arial", bold=True, size=14, color=NAVY_HEX)
        self.subtitle_font = Font(name="Arial", italic=True, size=9, color="6B7280")
        self.label_font = Font(name="Arial", bold=True, size=10, color=NAVY_HEX)
        self.cell_font = Font(name="Arial", size=10)
        thin = Side(style="thin", color=BORDER_GRAY_HEX)
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.center = Alignment(horizontal="center", vertical="center")
        self.accent_fill = PatternFill("solid", fgColor=LIGHT_GRAY_HEX)
        self.currency_fmt = '"Rs." #,##0.00'

        self._summary_rows = []

    def add_summary_row(self, label: str, value, is_currency: bool = False):
        self._summary_rows.append((label, value, is_currency))
        return self

    def build_summary_sheet(self):
        ws = self.wb.create_sheet("Summary", 0)
        ws.merge_cells("A1:B1")
        ws["A1"] = self.title.upper()
        ws["A1"].font = self.title_font
        row = 2
        if self.project_line:
            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"] = self.project_line
            ws[f"A{row}"].font = self.subtitle_font
            row += 1
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
        ws[f"A{row}"].font = self.subtitle_font
        row += 2

        for label, value, is_currency in self._summary_rows:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = self.label_font
            label_cell.fill = self.accent_fill
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = self.cell_font
            if is_currency:
                value_cell.number_format = self.currency_fmt
            row += 1

        row += 2
        ws.cell(row=row, column=1, value="Prepared By: ______________").font = (
            self.cell_font
        )
        row += 1
        ws.cell(row=row, column=1, value="Reviewed By: ______________").font = (
            self.cell_font
        )
        row += 1
        ws.cell(row=row, column=1, value="Approved By: ______________").font = (
            self.cell_font
        )

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 24
        return self

    def add_data_sheet(
        self,
        sheet_name: str,
        headers: Sequence[str],
        rows: Sequence[Sequence],
        currency_cols: Optional[Sequence[int]] = None,
        title: Optional[str] = None,
    ):
        currency_cols = currency_cols or []
        ws = self.wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
        start_row = 1

        if title:
            ws.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1)
            )
            ws.cell(row=1, column=1, value=title).font = self.title_font
            start_row = 3

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center
            cell.border = self.border

        for r, row_values in enumerate(rows, start_row + 1):
            fill = self.accent_fill if (r - start_row) % 2 == 0 else None
            for col, val in enumerate(row_values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = self.cell_font
                cell.alignment = self.center
                cell.border = self.border
                if fill:
                    cell.fill = fill
                if col in currency_cols:
                    cell.number_format = self.currency_fmt

        for col, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_values in rows:
                val = row_values[col - 1]
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(
                max(max_len + 4, 12), 40
            )

        ws.freeze_panes = f"A{start_row + 1}"
        return self

    def build(self):
        import io

        if "Summary" not in self.wb.sheetnames:
            self.build_summary_sheet()
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output


# ======================================================================
# SECTION 2: REPORT API ENDPOINTS
# ======================================================================

from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime, timedelta
from typing import Optional
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, case, extract
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.core.dependencies import require_roles
from app.core.enums import (
    InvoiceStatus,
    IssueStatus,
    IssuePriority,
    LabourStatus,
    ProjectStatus,
    TaskStatus,
)
from app.db.session import get_db_session
from app.models import project as m
from app.models.accountant import FixedAsset
from app.models.expense import Expense
from app.models.invoice import Invoice, Transaction
from app.models.master_data import LabourType, MaterialMaster
from app.models.material import Material
from app.models.user import User, UserRole, UserAttendance, ActivityLog
from app.utils.common import assert_project_access


REPORT_READ_ROLES = [role.value for role in UserRole]

router = APIRouter(prefix="/reports", tags=["Reports"])


# ===================== PROJECT REPORTS =====================


@router.get("/projects/excel")
async def export_projects_excel(
    project_id: Optional[int] = Query(
        None,
        description="Project ID to filter. If none, exports all projects.",
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    from app.api.project import get_reports_service

    service = get_reports_service()

    # ======================================================
    # Single Project Report (Same Logic as PDF)
    # ======================================================
    if project_id:
        return await service.export_excel(
            db=db,
            project_id=project_id,
            current_user=current_user,
        )

    # ======================================================
    # Company Portfolio Report
    # ======================================================

    projects = (
        (await db.execute(select(m.Project).order_by(m.Project.id.desc())))
        .scalars()
        .all()
    )

    total_projects = len(projects)
    ongoing = sum(
        1 for p in projects if str(getattr(p.status, "value", p.status)) == "Ongoing"
    )
    completed = sum(
        1 for p in projects if str(getattr(p.status, "value", p.status)) == "Completed"
    )

    eb = ExcelReportBuilder("Company Portfolio Overview")
    eb.add_summary_row("Total Projects", total_projects)
    eb.add_summary_row("Ongoing", ongoing)
    eb.add_summary_row("Completed", completed)
    eb.build_summary_sheet()

    headers = ["Business ID", "Project Name", "Status", "Start Date", "End Date"]
    rows = [
        [
            p.business_id,
            (
                p.project_name[:30] + "..."
                if p.project_name and len(p.project_name) > 30
                else (p.project_name or "N/A")
            ),
            (p.status.value if hasattr(p.status, "value") else str(p.status)),
            str(p.start_date) if p.start_date else "N/A",
            str(p.end_date) if p.end_date else "N/A",
        ]
        for p in projects
    ]
    eb.add_data_sheet("Portfolio", headers, rows, title="Company Portfolio")

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": ("attachment; filename=all_projects_report.xlsx")
        },
    )


# =====================================================================


@router.get("/projects/pdf")
async def export_projects_pdf(
    project_id: Optional[int] = Query(
        None, description="Project ID to filter. If none, exports all projects."
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    from app.api.project import get_reports_service

    service = get_reports_service()

    if project_id:
        return await service.export_pdf(db, project_id, current_user)

    # Export ALL projects
    projects_query = select(m.Project).order_by(m.Project.id.desc())
    projects = (await db.execute(projects_query)).scalars().all()

    total_projects = len(projects)
    ongoing = sum(
        1 for p in projects if str(getattr(p.status, "value", p.status)) == "Ongoing"
    )
    completed = sum(
        1 for p in projects if str(getattr(p.status, "value", p.status)) == "Completed"
    )

    b = PdfReportBuilder("COMPANY PORTFOLIO OVERVIEW")
    b.add_info_table(
        [
            ("Generated On", str(date.today()), "Report Type", "Company Portfolio"),
        ]
    )
    b.add_summary_box(
        "SUMMARY",
        [
            f"<b>Total Projects:</b> {total_projects} | "
            f"<b>Ongoing:</b> {ongoing} | <b>Completed:</b> {completed}",
        ],
    )

    headers = ["ID", "Name", "Status", "Start Date", "End Date"]
    rows = [
        [
            p.business_id,
            (
                p.project_name[:30] + "..."
                if p.project_name and len(p.project_name) > 30
                else (p.project_name or "N/A")
            ),
            str(getattr(p.status, "value", p.status)),
            str(p.start_date) if p.start_date else "N/A",
            str(p.end_date) if p.end_date else "N/A",
        ]
        for p in projects
    ]
    b.add_section_table("PROJECT LIST", headers, rows)

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=all_projects_portfolio.pdf"
        },
    )


# ===================== AUDIT REPORTS =====================


@router.get("/audit/excel")
async def export_audit_excel(
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    user_id: Optional[int] = Query(None, description="Filter by user ID"),
    module: Optional[str] = Query(None, description="Filter by entity/module"),
    action: Optional[str] = Query(
        None, description="Filter by action (CREATE, UPDATE, DELETE)"
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(ActivityLog, User).outerjoin(
        User, ActivityLog.performed_by == User.id
    )

    if start_date:
        query = query.where(ActivityLog.created_at >= start_date)
    if end_date:
        query = query.where(ActivityLog.created_at <= end_date + timedelta(days=1))
    if user_id:
        query = query.where(ActivityLog.performed_by == user_id)
    if module:
        query = query.where(ActivityLog.entity == module)
    if action:
        query = query.where(ActivityLog.action == action)

    query = query.order_by(ActivityLog.created_at.desc())
    result = await db.execute(query)
    logs = result.all()

    eb = ExcelReportBuilder("System Audit Summary Report")
    eb.add_summary_row("Total Logs", len(logs))
    if start_date or end_date:
        eb.add_summary_row("Period", f"{start_date or 'Start'} to {end_date or 'End'}")
    if module:
        eb.add_summary_row("Module", module)
    if action:
        eb.add_summary_row("Action", action)
    eb.build_summary_sheet()

    headers = ["Timestamp", "User Name", "Module", "Action", "Entity ID", "Details"]
    rows = []
    for log, user in logs:
        details_str = str(log.details) if log.details else ""
        user_name = user.full_name if user else "System/Unknown"
        rows.append(
            [
                str(log.created_at),
                user_name,
                log.entity,
                log.action,
                log.entity_id or "",
                details_str,
            ]
        )
    eb.add_data_sheet("Audit Log", headers, rows, title="Audit Log")

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=audit_summary_report.xlsx"
        },
    )



# =====================================================================
# PROCUREMENT EFFICIENCY REPORT
# =====================================================================
@router.get("/procurement-efficiency")
async def procurement_efficiency_report(
    project_id: int = Query(..., description="Project ID (required)"),
    supplier_id: Optional[int] = Query(None, description="Filter by supplier ID"),
    status: Optional[str] = Query(None, description="VendorBill status filter"),
    payment_status: Optional[str] = Query(None, description="PAID|PARTIAL|UNPAID"),
    date_from: Optional[date] = Query(None, description="Start date filter"),
    date_to: Optional[date] = Query(None, description="End date filter"),
    search: Optional[str] = Query(None, description="Search term for bill number or supplier name"),
    format: str = Query("json", description="Response format: json|pdf|csv"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    """Generate Procurement Efficiency Report.

    The underlying aggregation is performed in ReportService. The endpoint
    supports three response formats, all sharing the same data.
    """
    filters: Dict[str, Any] = {
        "project_id": project_id,
        "supplier_id": supplier_id,
        "status": status,
        "payment_status": payment_status,
        "date_from": date_from.isoformat() if date_from else None,
        "date_to": date_to.isoformat() if date_to else None,
        "search": search,
    }
    filters = {k: v for k, v in filters.items() if v is not None}
    from app.services.report_service import ReportService
    try:
        report_dto = await ReportService.get_procurement_efficiency_report(db, filters)
    except Exception as e:
        if str(e) == "Report data missing":
            raise HTTPException(status_code=404, detail="Project not found")
        raise
    if format.lower() == "json":
        return report_dto
    if format.lower() == "pdf":
        b = PdfReportBuilder("Procurement Efficiency Report")
        b.add_info_table([
            (
                "Project", report_dto.summary.project_name,
                "Budget", str(report_dto.summary.budget_amount)
            ),
            (
                "Total Spend", str(report_dto.summary.total_spend),
                "Budget Utilisation", report_dto.summary.budget_vs_actual
            ),
        ])
        b.add_section_table(
            "Totals",
            ["Metric", "Amount"],
            [
                ["Total Spend", str(report_dto.procurement.total_spend)],
                ["Total Paid", str(report_dto.procurement.total_paid)],
                ["Total Pending", str(report_dto.procurement.total_pending)],
                ["Materials Qty", str(report_dto.procurement.materials_procured.total_quantity)],
                ["Materials Value", str(report_dto.procurement.materials_procured.total_value)],
            ]
        )
        supplier_headers = ["Supplier", "Bills", "Spend", "Paid", "Pending", "Avg Days"]
        supplier_rows = [
            [
                s.supplier_name,
                str(s.bill_count),
                str(s.total_spend),
                str(s.paid_amount),
                str(s.pending_amount),
                str(s.avg_payment_days),
            ]
            for s in report_dto.suppliers
        ]
        b.add_section_table("Supplier Performance", supplier_headers, supplier_rows)
        po = report_dto.purchase_orders
        b.add_section_table(
            "Outstanding Purchase Orders",
            ["Count", "Value"],
            [[str(po.outstanding_count), str(po.outstanding_value)]],
        )
        stream = b.build()
        return StreamingResponse(
            stream,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=procurement_efficiency.pdf"},
        )
    if format.lower() == "csv":
        from app.utils.csv_report_builder import CsvReportBuilder
        csv_builder = CsvReportBuilder(
            filename="procurement_efficiency.csv",
            headers=["Project Summary", "", "", "", "", "", ""]
        )
        csv_builder.add_row(["Project ID", str(report_dto.summary.project_id)])
        csv_builder.add_row(["Project Name", report_dto.summary.project_name])
        csv_builder.add_row(["Budget Amount", str(report_dto.summary.budget_amount)])
        csv_builder.add_row(["Total Spend", str(report_dto.summary.total_spend)])
        csv_builder.add_row(["Budget Vs Actual", report_dto.summary.budget_vs_actual])
        csv_builder.add_row([])
        csv_builder.add_row(["Procurement Totals", "", "", "", "", "", ""])
        csv_builder.add_row(["Total Spend", str(report_dto.procurement.total_spend)])
        csv_builder.add_row(["Total Paid", str(report_dto.procurement.total_paid)])
        csv_builder.add_row(["Total Pending", str(report_dto.procurement.total_pending)])
        csv_builder.add_row(["Materials Qty", str(report_dto.procurement.materials_procured.total_quantity)])
        csv_builder.add_row(["Materials Value", str(report_dto.procurement.materials_procured.total_value)])
        csv_builder.add_row([])
        csv_builder.add_row(["Outstanding Purchase Orders", "", "", "", "", "", ""])
        csv_builder.add_row(["Count", str(report_dto.purchase_orders.outstanding_count)])
        csv_builder.add_row(["Value", str(report_dto.purchase_orders.outstanding_value)])
        csv_builder.add_row([])
        csv_builder.add_row(["Supplier Performance", "", "", "", "", "", ""])
        csv_builder.add_row([
            "Supplier ID",
            "Supplier Name",
            "Bill Count",
            "Total Spend",
            "Paid Amount",
            "Pending Amount",
            "Avg Payment Days",
        ])
        for s in report_dto.suppliers:
            csv_builder.add_row([
                s.supplier_id,
                s.supplier_name,
                s.bill_count,
                str(s.total_spend),
                str(s.paid_amount),
                str(s.pending_amount),
                s.avg_payment_days,
            ])
        csv_builder.add_row([])
        csv_builder.add_row(["Filters Applied", "", "", "", "", "", ""])
        f = report_dto.filters_applied
        csv_builder.add_row(["Project ID", str(f.project_id)])
        if f.supplier_id: csv_builder.add_row(["Supplier ID", str(f.supplier_id)])
        if f.status: csv_builder.add_row(["Status", f.status])
        if f.date_from: csv_builder.add_row(["Date From", f.date_from])
        if f.date_to: csv_builder.add_row(["Date To", f.date_to])
        if f.search: csv_builder.add_row(["Search", f.search])
        if f.payment_status: csv_builder.add_row(["Payment Status", f.payment_status])
        return csv_builder.build()
    raise HTTPException(status_code=400, detail="Invalid format. Supported: json, pdf, csv")

@router.get("/audit/pdf")
async def export_audit_pdf(
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    user_id: Optional[int] = Query(None, description="Filter by user ID"),
    module: Optional[str] = Query(None, description="Filter by entity/module"),
    action: Optional[str] = Query(
        None, description="Filter by action (CREATE, UPDATE, DELETE)"
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(ActivityLog, User).outerjoin(
        User, ActivityLog.performed_by == User.id
    )

    if start_date:
        query = query.where(ActivityLog.created_at >= start_date)
    if end_date:
        query = query.where(ActivityLog.created_at <= end_date + timedelta(days=1))
    if user_id:
        query = query.where(ActivityLog.performed_by == user_id)
    if module:
        query = query.where(ActivityLog.entity == module)
    if action:
        query = query.where(ActivityLog.action == action)

    query = query.order_by(ActivityLog.created_at.desc()).limit(
        1000
    )  # Limit PDF to 1000 rows for performance
    result = await db.execute(query)
    logs = result.all()

    filter_bits = [f"Generated: {date.today()}"]
    if start_date or end_date:
        filter_bits.append(f"Period: {start_date or 'Start'} to {end_date or 'End'}")
    if module:
        filter_bits.append(f"Module: {module}")
    if action:
        filter_bits.append(f"Action: {action}")

    b = PdfReportBuilder("SYSTEM AUDIT SUMMARY", landscape_mode=True)
    b.add_info_table(
        [
            ("Generated", str(date.today()), "Total Logs", f"{len(logs)} (Max 1000)"),
        ]
    )
    b.add_summary_box("FILTERS APPLIED", [" | ".join(filter_bits)])

    headers = ["Date/Time", "User", "Module", "Action", "Details"]
    rows = []
    for log, user in logs:
        details_str = (
            str(log.details)[:50] + "..."
            if log.details and len(str(log.details)) > 50
            else (str(log.details) if log.details else "N/A")
        )
        user_name = user.full_name if user else "System"
        rows.append(
            [
                log.created_at.strftime("%Y-%m-%d %H:%M"),
                user_name,
                log.entity or "N/A",
                log.action or "N/A",
                details_str,
            ]
        )
    b.add_section_table("AUDIT LOG DETAILS", headers, rows, [100, 100, 100, 100, 300])

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=audit_summary_report.pdf"
        },
    )


# ===================== ASSET REPORTS =====================


@router.get("/assets/excel")
async def export_assets_excel(
    project_id: Optional[int] = Query(
        None, description="Filter by allocated project ID"
    ),
    start_date: Optional[date] = Query(None, description="Purchase start date"),
    end_date: Optional[date] = Query(None, description="Purchase end date"),
    min_value: Optional[float] = Query(None, description="Minimum current value"),
    max_value: Optional[float] = Query(None, description="Maximum current value"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(FixedAsset, m.Project).outerjoin(
        m.Project, FixedAsset.project_id == m.Project.id
    )

    if project_id:
        query = query.where(FixedAsset.project_id == project_id)
    if start_date:
        query = query.where(FixedAsset.purchase_date >= start_date)
    if end_date:
        query = query.where(FixedAsset.purchase_date <= end_date)
    if min_value is not None:
        query = query.where(FixedAsset.current_value >= min_value)
    if max_value is not None:
        query = query.where(FixedAsset.current_value <= max_value)

    result = await db.execute(query)
    assets = result.all()

    headers = [
        "Asset ID",
        "Asset Name",
        "Allocated Project",
        "Purchase Date",
        "Purchase Value",
        "Depreciation Rate (%)",
        "Accumulated Depreciation",
        "Current Net Book Value",
    ]

    rows = []
    total_purchase = 0.0
    total_current = 0.0
    for asset, project in assets:
        proj_name = project.project_name if project else "Unallocated"
        purch_val = float(asset.purchase_value or 0)
        curr_val = float(asset.current_value or 0)
        depr_acc = purch_val - curr_val
        total_purchase += purch_val
        total_current += curr_val

        rows.append(
            [
                asset.id,
                asset.name,
                proj_name,
                str(asset.purchase_date) if asset.purchase_date else "N/A",
                purch_val,
                float(asset.depreciation_rate or 0),
                depr_acc,
                curr_val,
            ]
        )

    eb = ExcelReportBuilder("Fixed Asset Register")
    eb.add_summary_row("Total Assets", len(rows))
    eb.add_summary_row(
        "Total Original Value", round(total_purchase, 2), is_currency=True
    )
    eb.add_summary_row(
        "Total Current Net Book Value", round(total_current, 2), is_currency=True
    )
    eb.build_summary_sheet()
    eb.add_data_sheet("Fixed Asset Register", headers, rows, currency_cols=[5, 7, 8])

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=fixed_asset_register.xlsx"
        },
    )


@router.get("/assets/pdf")
async def export_assets_pdf(
    project_id: Optional[int] = Query(
        None, description="Filter by allocated project ID"
    ),
    start_date: Optional[date] = Query(None, description="Purchase start date"),
    end_date: Optional[date] = Query(None, description="Purchase end date"),
    min_value: Optional[float] = Query(None, description="Minimum current value"),
    max_value: Optional[float] = Query(None, description="Maximum current value"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = select(FixedAsset, m.Project).outerjoin(
        m.Project, FixedAsset.project_id == m.Project.id
    )

    if project_id:
        query = query.where(FixedAsset.project_id == project_id)
    if start_date:
        query = query.where(FixedAsset.purchase_date >= start_date)
    if end_date:
        query = query.where(FixedAsset.purchase_date <= end_date)
    if min_value is not None:
        query = query.where(FixedAsset.current_value >= min_value)
    if max_value is not None:
        query = query.where(FixedAsset.current_value <= max_value)

    result = await db.execute(query)
    assets = result.all()

    total_assets = len(assets)
    total_purchase = sum(float(a.FixedAsset.purchase_value or 0) for a in assets)
    total_current = sum(float(a.FixedAsset.current_value or 0) for a in assets)
    total_depr = total_purchase - total_current

    b = PdfReportBuilder("Fixed Asset & Depreciation Report", landscape_mode=True)
    filters_applied = (
        "Yes"
        if (project_id or start_date or end_date or min_value or max_value)
        else "No"
    )
    b.add_info_table(
        [("Generated", str(date.today()), "Filters Applied", filters_applied)]
    )
    b.add_summary_box(
        "SUMMARY",
        [
            f"<b>Total Assets:</b> {total_assets}",
            f"Total Original Value: Rs. {total_purchase:,.2f}",
            f"Total Accumulated Depreciation: Rs. {total_depr:,.2f}",
            f"Total Current Net Book Value: Rs. {total_current:,.2f}",
        ],
    )

    headers = [
        "ID",
        "Name",
        "Project",
        "Purchase Date",
        "Orig Value",
        "Depr",
        "Net Book Value",
    ]
    rows = []
    for asset, project in assets:
        proj_name = (
            project.project_name[:20] + "..."
            if project and len(project.project_name) > 20
            else (project.project_name if project else "Unallocated")
        )
        purch_val = float(asset.purchase_value or 0)
        curr_val = float(asset.current_value or 0)
        depr_acc = purch_val - curr_val

        rows.append(
            [
                str(asset.id),
                asset.name[:25] + "..." if len(asset.name) > 25 else asset.name,
                proj_name,
                str(asset.purchase_date) if asset.purchase_date else "N/A",
                f"Rs. {purch_val:,.2f}",
                f"Rs. {depr_acc:,.2f}",
                f"Rs. {curr_val:,.2f}",
            ]
        )

    b.add_section_table("ASSET DETAILS", headers, rows)

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=fixed_asset_depreciation_report.pdf"
        },
    )


# ===================== ISSUE REPORTS =====================


@router.get("/issues/excel")
async def export_issues_excel(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    status: Optional[str] = Query(
        None, description="Filter by status (e.g., OPEN, RESOLVED)"
    ),
    priority: Optional[str] = Query(
        None, description="Filter by priority (e.g., HIGH, LOW)"
    ),
    start_date: Optional[date] = Query(None, description="Reported start date"),
    end_date: Optional[date] = Query(None, description="Reported end date"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = (
        select(m.Issue, m.Project, User)
        .join(m.Project, m.Issue.project_id == m.Project.id)
        .outerjoin(User, m.Issue.assigned_to == User.id)
    )

    if project_id:
        query = query.where(m.Issue.project_id == project_id)
    if status:
        query = query.where(m.Issue.status == status)
    if priority:
        query = query.where(m.Issue.priority == priority)
    if start_date:
        query = query.where(m.Issue.reported_date >= start_date)
    if end_date:
        query = query.where(m.Issue.reported_date <= end_date)

    query = query.order_by(m.Issue.reported_date.desc())
    result = await db.execute(query)
    issues = result.all()

    headers = [
        "Issue ID",
        "Project Name",
        "Title",
        "Category",
        "Reported Date",
        "Priority",
        "Status",
        "Assigned To",
        "Description",
        "Resolution Notes",
    ]
    rows = []
    for issue, project, user in issues:
        assigned_name = user.full_name if user else "Unassigned"
        rows.append(
            [
                issue.business_id or str(issue.id),
                project.project_name,
                issue.title,
                str(
                    issue.category.value
                    if hasattr(issue.category, "value")
                    else issue.category
                ),
                str(issue.reported_date) if issue.reported_date else "N/A",
                str(
                    issue.priority.value
                    if hasattr(issue.priority, "value")
                    else issue.priority
                ),
                str(
                    issue.status.value
                    if hasattr(issue.status, "value")
                    else issue.status
                ),
                assigned_name,
                issue.description or "",
                issue.resolution or "",
            ]
        )

    eb = ExcelReportBuilder("Site Issue Log")
    eb.add_summary_row("Total Issues", len(rows))
    eb.build_summary_sheet()
    eb.add_data_sheet("Site Issue Log", headers, rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=site_issue_log.xlsx"},
    )


@router.get("/issues/pdf")
async def export_issues_pdf(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    status: Optional[str] = Query(
        None, description="Filter by status (e.g., OPEN, RESOLVED)"
    ),
    priority: Optional[str] = Query(
        None, description="Filter by priority (e.g., HIGH, LOW)"
    ),
    start_date: Optional[date] = Query(None, description="Reported start date"),
    end_date: Optional[date] = Query(None, description="Reported end date"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = (
        select(m.Issue, m.Project, User)
        .join(m.Project, m.Issue.project_id == m.Project.id)
        .outerjoin(User, m.Issue.assigned_to == User.id)
    )

    if project_id:
        query = query.where(m.Issue.project_id == project_id)
    if status:
        query = query.where(m.Issue.status == status)
    if priority:
        query = query.where(m.Issue.priority == priority)
    if start_date:
        query = query.where(m.Issue.reported_date >= start_date)
    if end_date:
        query = query.where(m.Issue.reported_date <= end_date)

    query = query.order_by(m.Issue.reported_date.desc()).limit(1000)
    result = await db.execute(query)
    issues = result.all()

    total_issues = len(issues)
    open_count = sum(
        1
        for i in issues
        if str(getattr(i.Issue.status, "value", i.Issue.status))
        in ["OPEN", "IN_PROGRESS"]
    )
    resolved_count = sum(
        1
        for i in issues
        if str(getattr(i.Issue.status, "value", i.Issue.status))
        in ["RESOLVED", "CLOSED"]
    )
    critical_count = sum(
        1
        for i in issues
        if str(getattr(i.Issue.priority, "value", i.Issue.priority))
        in ["HIGH", "CRITICAL"]
    )

    b = PdfReportBuilder("Executive Site Issue Report", landscape_mode=True)
    filters_applied = (
        "Yes" if (project_id or status or priority or start_date or end_date) else "No"
    )
    b.add_info_table(
        [("Generated", str(date.today()), "Filters Applied", filters_applied)]
    )
    b.add_summary_box(
        "SUMMARY",
        [
            f"<b>Total Issues:</b> {total_issues}",
            f"Open/In-Progress: {open_count}",
            f"Resolved/Closed: {resolved_count}",
            f"High Priority: {critical_count}",
        ],
    )

    headers = ["ID", "Date", "Project", "Title", "Priority", "Status", "Assigned To"]
    rows = []
    for issue, project, user in issues:
        proj_name = (
            project.project_name[:15] + "..."
            if len(project.project_name) > 15
            else project.project_name
        )
        title = issue.title[:25] + "..." if len(issue.title) > 25 else issue.title
        assigned_name = (
            user.full_name[:15] + "..."
            if user and len(user.full_name) > 15
            else (user.full_name if user else "Unassigned")
        )

        rows.append(
            [
                issue.business_id or str(issue.id),
                str(issue.reported_date) if issue.reported_date else "N/A",
                proj_name,
                title,
                str(getattr(issue.priority, "value", issue.priority)),
                str(getattr(issue.status, "value", issue.status)),
                assigned_name,
            ]
        )

    b.add_section_table("ISSUE DETAILS", headers, rows)

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=executive_issue_report.pdf"
        },
    )


# ===================== FINANCIAL REPORTS =====================


@router.get("/finance/excel")
async def export_finance_excel(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    start_date: Optional[date] = Query(
        None, description="Start date for financial period"
    ),
    end_date: Optional[date] = Query(None, description="End date for financial period"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    # 1. Fetch Projects
    proj_query = select(m.Project)
    if project_id:
        proj_query = proj_query.where(m.Project.id == project_id)
    projects = (await db.execute(proj_query)).scalars().all()
    project_map = {p.id: p.project_name for p in projects}

    # 2. Fetch Expenses (grouped by project and category)
    exp_query = select(
        Expense.project_id, Expense.category, func.sum(Expense.amount)
    ).group_by(Expense.project_id, Expense.category)
    if project_id:
        exp_query = exp_query.where(Expense.project_id == project_id)
    if start_date:
        exp_query = exp_query.where(Expense.expense_date >= start_date)
    if end_date:
        exp_query = exp_query.where(Expense.expense_date <= end_date)

    exp_result = await db.execute(exp_query)

    project_expenses = defaultdict(lambda: defaultdict(float))
    all_categories = set()
    for pid, cat, amount in exp_result.all():
        if pid in project_map:
            project_expenses[pid][cat] += float(amount or 0)
            all_categories.add(cat)

    # 3. Fetch Invoices (grouped by project and status)
    inv_query = select(
        Invoice.project_id, Invoice.status, func.sum(Invoice.total_amount)
    ).group_by(Invoice.project_id, Invoice.status)
    if project_id:
        inv_query = inv_query.where(Invoice.project_id == project_id)
    if start_date:
        inv_query = inv_query.where(Invoice.created_at >= start_date)
    if end_date:
        # Cast created_at to Date for accurate comparison, or just add days
        inv_query = inv_query.where(Invoice.created_at <= end_date + timedelta(days=1))

    inv_result = await db.execute(inv_query)

    project_invoices = defaultdict(lambda: defaultdict(float))
    for pid, status, amount in inv_result.all():
        if pid in project_map:
            status_str = status.value if hasattr(status, "value") else str(status)
            project_invoices[pid][status_str] += float(amount or 0)

    # 4. Generate Excel
    sorted_categories = sorted(list(all_categories))

    headers = [
        "Project ID",
        "Project Name",
        "Total Invoiced",
        "Amount Paid",
        "Amount Pending",
        "Total Expenses",
    ]
    for cat in sorted_categories:
        headers.append(f"Exp: {cat}")
    headers.append("Net Profit / Loss")
    headers.append("Profit Margin (%)")

    rows = []
    grand_invoice = 0.0
    grand_expense = 0.0
    for pid, p_name in project_map.items():
        inv_totals = project_invoices[pid]
        total_inv = sum(inv_totals.values())
        paid_inv = inv_totals.get("PAID", 0.0) + inv_totals.get(
            "PARTIAL", 0.0
        )  # simplify
        pending_inv = inv_totals.get("PENDING", 0.0)

        exp_totals = project_expenses[pid]
        total_exp = sum(exp_totals.values())

        net_profit = total_inv - total_exp
        margin = (net_profit / total_inv * 100) if total_inv > 0 else 0.0

        grand_invoice += total_inv
        grand_expense += total_exp

        row = [pid, p_name, total_inv, paid_inv, pending_inv, total_exp]
        for cat in sorted_categories:
            row.append(exp_totals.get(cat, 0.0))
        row.append(net_profit)
        row.append(round(margin, 2))

        rows.append(row)

    eb = ExcelReportBuilder("Financial Ledger")
    eb.add_summary_row("Total Invoiced", round(grand_invoice, 2), is_currency=True)
    eb.add_summary_row("Total Expenses", round(grand_expense, 2), is_currency=True)
    eb.add_summary_row(
        "Net Profit / Loss", round(grand_invoice - grand_expense, 2), is_currency=True
    )
    eb.build_summary_sheet()
    eb.add_data_sheet("Financial Ledger", headers, rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=financial_ledger_report.xlsx"
        },
    )


@router.get("/finance/pdf")
async def export_finance_pdf(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    start_date: Optional[date] = Query(
        None, description="Start date for financial period"
    ),
    end_date: Optional[date] = Query(None, description="End date for financial period"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    proj_query = select(m.Project)
    if project_id:
        proj_query = proj_query.where(m.Project.id == project_id)
    projects = (await db.execute(proj_query)).scalars().all()
    project_map = {p.id: p.project_name for p in projects}

    exp_query = select(Expense.project_id, func.sum(Expense.amount)).group_by(
        Expense.project_id
    )
    if project_id:
        exp_query = exp_query.where(Expense.project_id == project_id)
    if start_date:
        exp_query = exp_query.where(Expense.expense_date >= start_date)
    if end_date:
        exp_query = exp_query.where(Expense.expense_date <= end_date)

    exp_result = await db.execute(exp_query)
    project_expenses = {pid: float(amt or 0) for pid, amt in exp_result.all()}

    inv_query = select(
        Invoice.project_id, Invoice.status, func.sum(Invoice.total_amount)
    ).group_by(Invoice.project_id, Invoice.status)
    if project_id:
        inv_query = inv_query.where(Invoice.project_id == project_id)
    if start_date:
        inv_query = inv_query.where(Invoice.created_at >= start_date)
    if end_date:
        inv_query = inv_query.where(Invoice.created_at <= end_date + timedelta(days=1))

    inv_result = await db.execute(inv_query)
    project_invoices = defaultdict(lambda: defaultdict(float))
    for pid, status, amount in inv_result.all():
        status_str = status.value if hasattr(status, "value") else str(status)
        project_invoices[pid][status_str] += float(amount or 0)

    global_exp = sum(project_expenses.values())
    global_inv = sum(sum(inv.values()) for inv in project_invoices.values())
    global_profit = global_inv - global_exp
    global_margin = (global_profit / global_inv * 100) if global_inv > 0 else 0.0

    b = PdfReportBuilder("Executive Financial Summary", landscape_mode=True)
    filters_applied = "Yes" if (project_id or start_date or end_date) else "No"
    b.add_info_table(
        [("Generated", str(date.today()), "Filters Applied", filters_applied)]
    )
    b.add_summary_box(
        "SUMMARY",
        [
            f"<b>Total Company Expenses:</b> Rs. {global_exp:,.2f}",
            f"Total Company Invoiced: Rs. {global_inv:,.2f}",
            f"Total Net Profit: Rs. {global_profit:,.2f}",
            f"Overall Profit Margin: {global_margin:,.2f}%",
        ],
    )

    headers = [
        "Project",
        "Total Expenses",
        "Total Invoiced",
        "Pending",
        "Net Profit",
        "Margin",
    ]
    rows = []
    for pid, p_name in project_map.items():
        total_exp = project_expenses.get(pid, 0.0)
        inv_totals = project_invoices[pid]
        total_inv = sum(inv_totals.values())
        pending_inv = inv_totals.get("PENDING", 0.0)

        net_profit = total_inv - total_exp
        margin = (net_profit / total_inv * 100) if total_inv > 0 else 0.0

        p_name_short = p_name[:25] + "..." if len(p_name) > 25 else p_name

        rows.append(
            [
                p_name_short,
                f"Rs. {total_exp:,.2f}",
                f"Rs. {total_inv:,.2f}",
                f"Rs. {pending_inv:,.2f}",
                f"Rs. {net_profit:,.2f}",
                f"{margin:,.2f}%",
            ]
        )

    b.add_section_table("PROJECT FINANCIALS", headers, rows)

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=executive_financial_summary.pdf"
        },
    )


# ===================== PROFIT & LOSS REPORTS =====================


@router.get("/profit-loss/excel")
async def export_profit_loss_excel(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    year: Optional[int] = Query(None, description="Financial Year"),
    quarter: Optional[int] = Query(None, description="Quarter (1-4)"),
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    # Base queries
    inv_query = select(Invoice)
    exp_query = select(Expense)

    if project_id:
        inv_query = inv_query.where(Invoice.project_id == project_id)
        exp_query = exp_query.where(Expense.project_id == project_id)

    if year:
        inv_query = inv_query.where(extract("year", Invoice.created_at) == year)
        exp_query = exp_query.where(extract("year", Expense.expense_date) == year)
        if quarter:
            q_months = {1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)}
            sm, em = q_months[quarter]
            inv_query = inv_query.where(
                extract("month", Invoice.created_at).between(sm, em)
            )
            exp_query = exp_query.where(
                extract("month", Expense.expense_date).between(sm, em)
            )

    if start_date:
        inv_query = inv_query.where(Invoice.created_at >= start_date)
        exp_query = exp_query.where(Expense.expense_date >= start_date)
    if end_date:
        inv_query = inv_query.where(Invoice.created_at <= end_date + timedelta(days=1))
        exp_query = exp_query.where(Expense.expense_date <= end_date)

    invoices = (await db.execute(inv_query)).scalars().all()
    expenses = (await db.execute(exp_query)).scalars().all()

    # Group by YYYY-MM
    monthly_data = defaultdict(
        lambda: {
            "revenue": 0.0,
            "cogs_labour": 0.0,
            "cogs_material": 0.0,
            "overhead": defaultdict(float),
        }
    )
    all_months = set()

    for inv in invoices:
        month_key = inv.created_at.strftime("%Y-%m")
        all_months.add(month_key)
        amt = float(inv.total_amount or 0)

        # 'owner' invoices are revenue. others are COGS
        if inv.type == "owner":
            monthly_data[month_key]["revenue"] += amt
        elif inv.type == "labour":
            monthly_data[month_key]["cogs_labour"] += amt
        elif inv.type == "material":
            monthly_data[month_key]["cogs_material"] += amt

    for exp in expenses:
        month_key = exp.expense_date.strftime("%Y-%m")
        all_months.add(month_key)
        amt = float(exp.amount or 0)
        monthly_data[month_key]["overhead"][exp.category] += amt

    sorted_months = sorted(list(all_months))
    all_overhead_cats = set()
    for data in monthly_data.values():
        all_overhead_cats.update(data["overhead"].keys())
    sorted_overhead_cats = sorted(list(all_overhead_cats))

    headers = ["Category", "Total"] + sorted_months
    rows = []

    def append_row(name, data_dict, overhead_cat=None):
        total = 0.0
        month_vals = []
        for mth in sorted_months:
            if overhead_cat:
                val = data_dict[mth]["overhead"].get(overhead_cat, 0.0)
            elif name == "Revenue":
                val = data_dict[mth]["revenue"]
            elif name == "Labour Costs":
                val = data_dict[mth]["cogs_labour"]
            elif name == "Material Costs":
                val = data_dict[mth]["cogs_material"]
            else:
                val = 0.0
            total += val
            month_vals.append(val)
        rows.append([name, total] + month_vals)
        return total, month_vals

    def blank_row():
        rows.append([""] * len(headers))

    def divider_row(label):
        rows.append([label] + [""] * (len(headers) - 1))

    divider_row("--- REVENUE ---")
    total_rev, rev_months = append_row("Revenue", monthly_data)

    blank_row()
    divider_row("--- COST OF GOODS SOLD (COGS) ---")
    t_labour, m_labour = append_row("Labour Costs", monthly_data)
    t_material, m_material = append_row("Material Costs", monthly_data)

    total_cogs = t_labour + t_material
    cogs_months = [l + mt for l, mt in zip(m_labour, m_material)]

    gross_profit = total_rev - total_cogs
    gp_months = [r - c for r, c in zip(rev_months, cogs_months)]
    rows.append(["Gross Profit", gross_profit] + gp_months)

    blank_row()
    divider_row("--- OPERATING EXPENSES (OVERHEAD) ---")
    total_op_ex = 0.0
    op_ex_months = [0.0] * len(sorted_months)
    for cat in sorted_overhead_cats:
        t_cat, m_cat = append_row(cat, monthly_data, overhead_cat=cat)
        total_op_ex += t_cat
        op_ex_months = [o + c for o, c in zip(op_ex_months, m_cat)]

    rows.append(["Total Operating Expenses", total_op_ex] + op_ex_months)

    blank_row()
    divider_row("--- NET INCOME ---")
    net_income = gross_profit - total_op_ex
    ni_months = [g - o for g, o in zip(gp_months, op_ex_months)]
    rows.append(["Net Income", net_income] + ni_months)

    eb = ExcelReportBuilder("Profit and Loss")
    eb.add_summary_row("Total Revenue", round(total_rev, 2), is_currency=True)
    eb.add_summary_row("Total COGS", round(total_cogs, 2), is_currency=True)
    eb.add_summary_row("Gross Profit", round(gross_profit, 2), is_currency=True)
    eb.add_summary_row(
        "Total Operating Expenses", round(total_op_ex, 2), is_currency=True
    )
    eb.add_summary_row("Net Income", round(net_income, 2), is_currency=True)
    eb.build_summary_sheet()
    eb.add_data_sheet("Profit and Loss", headers, rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=profit_and_loss.xlsx"},
    )


@router.get("/profit-loss/pdf")
async def export_profit_loss_pdf(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    year: Optional[int] = Query(None, description="Financial Year"),
    quarter: Optional[int] = Query(None, description="Quarter (1-4)"),
    start_date: Optional[date] = Query(None, description="Start date"),
    end_date: Optional[date] = Query(None, description="End date"),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    inv_query = select(Invoice)
    exp_query = select(Expense)

    if project_id:
        inv_query = inv_query.where(Invoice.project_id == project_id)
        exp_query = exp_query.where(Expense.project_id == project_id)

    if year:
        inv_query = inv_query.where(extract("year", Invoice.created_at) == year)
        exp_query = exp_query.where(extract("year", Expense.expense_date) == year)
        if quarter:
            q_months = {1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)}
            sm, em = q_months[quarter]
            inv_query = inv_query.where(
                extract("month", Invoice.created_at).between(sm, em)
            )
            exp_query = exp_query.where(
                extract("month", Expense.expense_date).between(sm, em)
            )

    if start_date:
        inv_query = inv_query.where(Invoice.created_at >= start_date)
        exp_query = exp_query.where(Expense.expense_date >= start_date)
    if end_date:
        inv_query = inv_query.where(Invoice.created_at <= end_date + timedelta(days=1))
        exp_query = exp_query.where(Expense.expense_date <= end_date)

    invoices = (await db.execute(inv_query)).scalars().all()
    expenses = (await db.execute(exp_query)).scalars().all()

    revenue = 0.0
    cogs_labour = 0.0
    cogs_material = 0.0
    overhead = defaultdict(float)

    for inv in invoices:
        amt = float(inv.total_amount or 0)
        if inv.type == "owner":
            revenue += amt
        elif inv.type == "labour":
            cogs_labour += amt
        elif inv.type == "material":
            cogs_material += amt

    total_overhead = 0.0
    for exp in expenses:
        amt = float(exp.amount or 0)
        overhead[exp.category] += amt
        total_overhead += amt

    cogs = cogs_labour + cogs_material
    gross_profit = revenue - cogs
    net_income = gross_profit - total_overhead
    margin = (net_income / revenue * 100) if revenue > 0 else 0.0

    b = PdfReportBuilder("Profit & Loss Statement")
    info_bits = [
        (
            "Generated",
            str(date.today()),
            "Project",
            str(project_id) if project_id else "All Projects",
        )
    ]
    b.add_info_table(info_bits)

    filter_bits = []
    if year:
        filter_bits.append(f"Year: {year}")
    if quarter:
        filter_bits.append(f"Quarter: {quarter}")
    if filter_bits:
        b.add_summary_box("FILTERS APPLIED", [" | ".join(filter_bits)])

    headers = ["Particulars", "Amount (Rs.)"]
    rows = [
        ["Revenue", f"Rs. {revenue:,.2f}"],
        ["Labour Costs", f"Rs. {cogs_labour:,.2f}"],
        ["Material Costs", f"Rs. {cogs_material:,.2f}"],
        ["Total COGS", f"Rs. {cogs:,.2f}"],
        ["Gross Profit", f"Rs. {gross_profit:,.2f}"],
    ]
    for cat, amt in overhead.items():
        rows.append([cat, f"Rs. {amt:,.2f}"])
    rows.append(["Total Operating Expenses", f"Rs. {total_overhead:,.2f}"])
    rows.append(["Net Income", f"Rs. {net_income:,.2f}"])
    rows.append(["Net Profit Margin", f"{margin:,.2f}%"])

    b.add_section_table("PROFIT & LOSS STATEMENT", headers, rows, col_widths=[300, 150])

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=profit_and_loss_statement.pdf"
        },
    )


# ===================== DAILY REPORT =====================


@router.get("/daily")
async def daily_report(
    project_id: int,
    report_date: date,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    dsr = await db.scalar(
        select(m.DailySiteReport).where(
            m.DailySiteReport.project_id == project_id,
            m.DailySiteReport.report_date == report_date,
        )
    )

    return {"dsr": dsr}


# ===================== DAILY REPORT PDF =====================


@router.get("/daily/export/pdf")
async def export_daily_pdf(
    project_id: int,
    report_date: date,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    dsr = await db.scalar(
        select(m.DailySiteReport).where(
            m.DailySiteReport.project_id == project_id,
            m.DailySiteReport.report_date == report_date,
        )
    )

    b = PdfReportBuilder(f"Daily Report - {report_date}")
    b.add_info_table([("Project ID", str(project_id), "Report Date", str(report_date))])

    if dsr:
        b.add_summary_box(
            "SITE DETAILS",
            [
                f"<b>Work Done:</b> {dsr.work_done}",
                f"Weather: {dsr.weather}",
                f"Remarks: {dsr.remarks}",
            ],
        )
    else:
        b.add_summary_box("SITE DETAILS", ["No data available"])

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=daily_report.pdf"},
    )


# ===================== WEEKLY PROGRESS =====================


@router.get("/weekly")
async def weekly_progress(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    week_ago = datetime.utcnow() - timedelta(days=7)

    result = await db.execute(
        select(m.Task.id, func.max(m.TaskProgress.percentage))
        .join(m.TaskProgress, m.Task.id == m.TaskProgress.task_id)
        .where(
            m.Task.project_id == project_id,
            m.TaskProgress.created_at >= week_ago,
        )
        .group_by(m.Task.id)
    )

    rows = result.all()

    #  safe calculation
    progress = (
        sum(float(r[1]) for r in rows if r[1] is not None) / len(rows) if rows else 0
    )

    return {"weekly_progress_percent": round(progress, 2), "tasks_count": len(rows)}


# ===================== LABOUR REPORT =====================


@router.get("/labour")
async def labour_report(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    from app.models.labour import Labour

    result = await db.execute(
        select(LabourType.skill_category, func.count(func.distinct(Labour.id)))
        .join(Labour, Labour.labour_type_id == LabourType.id)
        .join(UserAttendance, Labour.user_id == UserAttendance.user_id)
        .where(
            UserAttendance.project_id == project_id,
            Labour.status == LabourStatus.ACTIVE,
        )
        .group_by(LabourType.skill_category)
    )

    rows = result.all()

    return {"labour_summary": [{"skill_type": row[0], "count": row[1]} for row in rows]}


# ===================== LABOUR DISTRIBUTION REPORTS =====================


@router.get("/labour-distribution/excel")
async def export_labour_distribution_excel(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    date: Optional[date] = Query(
        None, description="Specific date for attendance filter"
    ),
    skill_category: Optional[str] = Query(
        None, description="Filter by SKILLED, UNSKILLED, etc"
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    from app.models.labour import Labour, LabourProject

    query = (
        select(Labour, m.Project, LabourType)
        .join(LabourProject, Labour.id == LabourProject.labour_id)
        .join(m.Project, LabourProject.project_id == m.Project.id)
        .outerjoin(LabourType, Labour.labour_type_id == LabourType.id)
    )

    if project_id:
        query = query.where(m.Project.id == project_id)
    if skill_category:
        query = query.where(LabourType.skill_category == skill_category)

    query = query.where(Labour.status == LabourStatus.ACTIVE)
    results = (await db.execute(query)).all()

    attendance_map = {}
    if date:
        att_query = select(UserAttendance.user_id, UserAttendance.status).where(
            UserAttendance.date == date
        )
        if project_id:
            att_query = att_query.where(UserAttendance.project_id == project_id)
        att_results = (await db.execute(att_query)).all()
        attendance_map = {user_id: status for user_id, status in att_results}

    agg_headers = ["Project Name", "Skill Category", "Trade", "Total Count"]
    det_headers = [
        "Project Name",
        "Worker Code",
        "Worker Name",
        "Skill Category",
        "Trade",
        "Status",
    ]
    if date:
        det_headers.append(f"Attendance ({date})")

    agg_data = defaultdict(int)
    det_rows = []

    for labour, project, ltype in results:
        skill = (
            str(getattr(ltype.skill_category, "value", ltype.skill_category))
            if ltype and getattr(ltype, "skill_category", None)
            else "Unclassified"
        )
        trade = ltype.name if ltype else "N/A"

        att_status = "Not Logged"
        if date and labour.user_id:
            att_status_raw = attendance_map.get(labour.user_id)
            att_status = (
                str(getattr(att_status_raw, "value", att_status_raw))
                if att_status_raw
                else "Absent"
            )

        row = [
            project.project_name,
            labour.worker_code,
            labour.labour_name,
            skill,
            trade,
            str(getattr(labour.status, "value", labour.status)),
        ]
        if date:
            row.append(att_status)
        det_rows.append(row)

        agg_key = (project.project_name, skill, trade)
        agg_data[agg_key] += 1

    agg_rows = [
        [proj, skill, trade, count]
        for (proj, skill, trade), count in sorted(agg_data.items())
    ]

    eb = ExcelReportBuilder("Labour Distribution Report")
    eb.add_summary_row("Total Active Workers", len(det_rows))
    eb.build_summary_sheet()
    eb.add_data_sheet("Distribution Summary", agg_headers, agg_rows)
    eb.add_data_sheet("Detailed Roster", det_headers, det_rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=labour_distribution_report.xlsx"
        },
    )


@router.get("/labour-distribution/pdf")
async def export_labour_distribution_pdf(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    skill_category: Optional[str] = Query(
        None, description="Filter by SKILLED, UNSKILLED, etc"
    ),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    from app.models.labour import Labour, LabourProject

    query = (
        select(Labour, m.Project, LabourType)
        .join(LabourProject, Labour.id == LabourProject.labour_id)
        .join(m.Project, LabourProject.project_id == m.Project.id)
        .outerjoin(LabourType, Labour.labour_type_id == LabourType.id)
        .where(Labour.status == LabourStatus.ACTIVE)
    )

    if project_id:
        query = query.where(m.Project.id == project_id)
    if skill_category:
        query = query.where(LabourType.skill_category == skill_category)

    results = (await db.execute(query)).all()

    project_stats = defaultdict(
        lambda: {"SKILLED": 0, "UNSKILLED": 0, "SEMI_SKILLED": 0, "OTHER": 0}
    )
    total_workers = 0
    total_skilled = 0
    total_unskilled = 0

    for labour, project, ltype in results:
        skill = (
            str(getattr(ltype.skill_category, "value", ltype.skill_category)).upper()
            if ltype and getattr(ltype, "skill_category", None)
            else "OTHER"
        )
        if skill not in project_stats[project.project_name]:
            skill = "OTHER"

        project_stats[project.project_name][skill] += 1
        total_workers += 1
        if skill == "SKILLED":
            total_skilled += 1
        elif skill == "UNSKILLED":
            total_unskilled += 1

    b = PdfReportBuilder("Executive Labour Distribution Summary")
    filters_applied = "Yes" if project_id else "No"
    b.add_info_table(
        [
            (
                "Generated",
                datetime.now().strftime("%Y-%m-%d"),
                "Filters Applied",
                filters_applied,
            )
        ]
    )

    summary_lines = [f"<b>Total Active Workforce:</b> {total_workers}"]
    if total_workers > 0:
        summary_lines.append(
            f"Skilled: {total_skilled} ({total_skilled/total_workers*100:.1f}%) | "
            f"Unskilled: {total_unskilled} ({total_unskilled/total_workers*100:.1f}%)"
        )
    summary_lines.append(f"Total Active Projects: {len(project_stats)}")
    b.add_summary_box("SUMMARY", summary_lines)

    headers = [
        "Project Name",
        "Skilled",
        "Semi",
        "Unskilled",
        "Other",
        "Total",
        "% of Company",
    ]
    rows = []
    for proj, stats in sorted(project_stats.items()):
        p_total = sum(stats.values())
        pct = (p_total / total_workers * 100) if total_workers > 0 else 0
        rows.append(
            [
                proj[:25] + "..." if len(proj) > 25 else proj,
                stats["SKILLED"],
                stats["SEMI_SKILLED"],
                stats["UNSKILLED"],
                stats["OTHER"],
                p_total,
                f"{pct:.1f}%",
            ]
        )

    b.add_section_table("LABOUR DISTRIBUTION", headers, rows)

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=labour_distribution_summary.pdf"
        },
    )


# ===================== MATERIAL REPORT =====================


@router.get("/material")
async def material_report(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(Material)
        .where(Material.project_id == project_id)
        .order_by(Material.created_at.desc())
    )

    materials = result.scalars().all()

    return {"materials": materials}


# ===================== MATERIAL EXCEL =====================


@router.get("/material/export/excel")
async def export_material_excel(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    stmt = (
        select(Material)
        .options(joinedload(Material.material_master).joinedload(MaterialMaster.unit))
        .where(Material.project_id == project_id)
        .order_by(Material.created_at.desc())
    )

    result = await db.execute(stmt)
    materials = result.scalars().all()

    headers = [
        "Material Name",
        "Category",
        "Unit",
        "Purchased Qty",
        "Used Qty",
        "Remaining Stock",
        "Purchase Rate",
        "Total Amount",
    ]

    rows = []
    total_amount = 0.0
    for mat in materials:
        unit_name = ""
        if mat.material_master and mat.material_master.unit:
            unit_name = mat.material_master.unit.name

        amt = float(mat.total_amount or 0)
        total_amount += amt

        rows.append(
            [
                mat.material_name or "",
                mat.category or "",
                unit_name,
                float(mat.quantity_purchased or 0),
                float(mat.quantity_used or 0),
                float(mat.remaining_stock or 0),
                float(mat.purchase_rate or 0),
                amt,
            ]
        )

    eb = ExcelReportBuilder("Material Report")
    eb.add_summary_row("Total Materials", len(rows))
    eb.add_summary_row("Total Amount", round(total_amount, 2), is_currency=True)
    eb.build_summary_sheet()
    eb.add_data_sheet("Materials", headers, rows, currency_cols=[7, 8])

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=material_report_{project_id}.xlsx"
        },
    )


# ===================== ISSUE REPORT =====================


@router.get("/issues")
async def issue_report(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    open_issues = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(
            m.Issue.project_id == project_id,
            m.Issue.status == IssueStatus.OPEN.value,
        )
    )

    closed_issues = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(
            m.Issue.project_id == project_id,
            m.Issue.status == IssueStatus.CLOSED.value,
        )
    )

    return {"open": open_issues, "closed": closed_issues}


# ===================== ISSUE EXCEL =====================


@router.get("/issues/export/excel")
async def export_issue_excel(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(m.Issue)
        .where(m.Issue.project_id == project_id)
        .order_by(m.Issue.created_at.desc())
    )
    issues = result.scalars().all()

    headers = [
        "Issue ID",
        "Title",
        "Category",
        "Priority",
        "Status",
        "Reported Date",
        "Description",
    ]
    rows = [
        [
            issue.business_id or str(issue.id),
            issue.title,
            str(getattr(issue.category, "value", issue.category)),
            str(getattr(issue.priority, "value", issue.priority)),
            str(getattr(issue.status, "value", issue.status)),
            str(issue.reported_date) if issue.reported_date else "N/A",
            issue.description or "",
        ]
        for issue in issues
    ]

    eb = ExcelReportBuilder("Project Issue Report")
    eb.add_summary_row("Total Issues", len(rows))
    eb.build_summary_sheet()
    eb.add_data_sheet("Issues", headers, rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=project_{project_id}_issues.xlsx"
        },
    )


# ===================== FILTERED REPORT DOWNLOAD =====================


@router.get("/download")
async def client_report_download(
    project_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(m.DailySiteReport)
        .where(
            m.DailySiteReport.project_id == project_id,
            m.DailySiteReport.report_date >= start_date,
            m.DailySiteReport.report_date <= end_date,
        )
        .order_by(m.DailySiteReport.report_date.desc())
    )
    reports = result.scalars().all()

    b = PdfReportBuilder(f"Report ({start_date} to {end_date})")
    b.add_info_table(
        [("Project ID", str(project_id), "Period", f"{start_date} to {end_date}")]
    )

    headers = ["Date", "Work Done"]
    rows = [[str(r.report_date), r.work_done] for r in reports]
    b.add_section_table(
        "DAILY SITE REPORTS", headers, rows, empty_text="No data available"
    )

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=filtered_report.pdf"},
    )


@router.get("/combined")
async def combined_report(
    project_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    #  Work Progress
    progress = await db.scalar(
        select(func.avg(m.Task.completion_percentage)).where(
            m.Task.project_id == project_id
        )
    )

    #  Financials
    total_paid = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PAID
        )
    )

    total_pending = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PENDING
        )
    )

    #  DSR (work summary)
    reports = await db.execute(
        select(m.DailySiteReport).where(
            m.DailySiteReport.project_id == project_id,
            m.DailySiteReport.report_date >= start_date,
            m.DailySiteReport.report_date <= end_date,
        )
    )
    dsr_list = reports.scalars().all()

    b = PdfReportBuilder("Combined Project Report")
    b.add_info_table(
        [("Project ID", str(project_id), "Period", f"{start_date} to {end_date}")]
    )
    b.add_summary_box(
        "PROGRESS & FINANCIAL SUMMARY",
        [
            f"<b>Progress:</b> {round(progress or 0, 2)}%",
            f"Total Paid: Rs. {float(total_paid or 0):,.2f}",
            f"Pending: Rs. {float(total_pending or 0):,.2f}",
        ],
    )

    headers = ["Date", "Work Done"]
    rows = [[str(r.report_date), r.work_done] for r in dsr_list]
    b.add_section_table("WORK SUMMARY", headers, rows, empty_text="No data available")

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=combined_report.pdf"},
    )


@router.get("/contractor-performance")
async def contractor_performance(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    # 1. Total tasks
    total_tasks = await db.scalar(
        select(func.count(m.Task.id)).where(m.Task.project_id == project_id)
    )

    # 2. Avg progress
    avg_progress = await db.scalar(
        select(func.avg(m.Task.completion_percentage)).where(
            m.Task.project_id == project_id
        )
    )

    # 3. Total paid invoices
    total_paid = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id,
            Invoice.status == InvoiceStatus.PAID,
        )
    )

    progress_val = float(avg_progress or 0)

    # 4. Performance logic
    if progress_val >= 75:
        rating = "Excellent"
    elif progress_val >= 50:
        rating = "Good"
    elif progress_val > 0:
        rating = "Average"
    else:
        rating = "Low"

    return {
        "project_id": project_id,
        "total_tasks": int(total_tasks or 0),
        "avg_progress": round(progress_val, 2),
        "total_paid": float(total_paid or 0),
        "performance": rating,
    }


@router.get("/profit-loss")
async def profit_loss(
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
):

    # Income (owner invoices)
    income = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(Invoice.type == "owner")
    )

    # Expense (labour + material)
    expense = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.type.in_(["labour", "material"])
        )
    )

    income_val = float(income or 0)
    expense_val = float(expense or 0)

    return {
        "income": income_val,
        "expense": expense_val,
        "profit": income_val - expense_val,
    }


@router.get("/project/{project_id}")
async def project_financial_summary_by_id(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    #  Revenue (owner invoices only)
    revenue = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.type == "owner"
        )
    )

    #  Expense (labour + material invoices)
    expense = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.type.in_(["labour", "material"])
        )
    )

    revenue_val = float(revenue or 0)
    expense_val = float(expense or 0)

    return {
        "project_id": project_id,
        "revenue": revenue_val,
        "expense": expense_val,
        "profit": revenue_val - expense_val,
    }


@router.get("/cashflow")
async def cashflow(
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
):
    inflow = await db.scalar(
        select(func.sum(Transaction.amount)).where(Transaction.type == "receipt")
    )

    outflow = await db.scalar(
        select(func.sum(Transaction.amount)).where(Transaction.type == "payment")
    )

    return {
        "inflow": float(inflow or 0),
        "outflow": float(outflow or 0),
        "balance": float((inflow or 0) - (outflow or 0)),
    }


@router.get("/assets")
async def asset_report(
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
):
    result = await db.execute(select(FixedAsset))
    assets = result.scalars().all()

    return assets


# ===================== FINANCIAL SUMMARY =====================
@router.get("/financial-summary")
async def financial_summary(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)
    total_expense = await db.scalar(
        select(func.sum(Expense.amount)).where(Expense.project_id == project_id)
    )
    total_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(Invoice.project_id == project_id)
    )
    paid_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PAID
        )
    )
    pending_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PENDING
        )
    )
    expense_val = round(float(total_expense or 0), 2)
    invoice_val = round(float(total_invoice or 0), 2)
    return {
        "project_id": project_id,
        "total_expense": expense_val,
        "total_invoice": invoice_val,
        "paid_invoice": round(float(paid_invoice or 0), 2),
        "pending_invoice": round(float(pending_invoice or 0), 2),
        "profit": round(invoice_val - expense_val, 2),
    }


# ===================== QUARTERLY AUDIT SUMMARY =====================
@router.get("/quarterly-audit-summary")
async def quarterly_financial_audit(
    project_id: int,
    year: int,
    quarter: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)
    if quarter not in [1, 2, 3, 4]:
        raise HTTPException(status_code=400, detail="Quarter must be between 1 and 4")
    quarter_map = {
        1: (1, 3),
        2: (4, 6),
        3: (7, 9),
        4: (10, 12),
    }
    start_month, end_month = quarter_map[quarter]
    total_expense = await db.scalar(
        select(func.sum(Expense.amount)).where(
            Expense.project_id == project_id,
            func.extract("year", Expense.expense_date) == year,
            func.extract("month", Expense.expense_date).between(start_month, end_month),
        )
    )
    total_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id,
            func.extract("year", Invoice.created_at) == year,
            func.extract("month", Invoice.created_at).between(start_month, end_month),
        )
    )
    completed_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(m.Task.project_id == project_id, m.Task.status == TaskStatus.COMPLETED)
    )
    delayed_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(
            m.Task.project_id == project_id,
            m.Task.end_date.isnot(None),
            m.Task.end_date < date.today(),
            m.Task.status != TaskStatus.COMPLETED,
        )
    )
    return {
        "project_id": project_id,
        "quarter": f"Q{quarter}",
        "year": year,
        "total_expense": round(float(total_expense or 0), 2),
        "total_invoice": round(float(total_invoice or 0), 2),
        "completed_tasks": int(completed_tasks or 0),
        "delayed_tasks": int(delayed_tasks or 0),
    }


# ===================== WORK SUMMARY =====================
@router.get("/work-summary")
async def work_summary(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)
    result = await db.execute(select(m.Task).where(m.Task.project_id == project_id))
    tasks = result.scalars().all()
    summary = []
    for task in tasks:
        actual = round(float(task.completion_percentage or 0), 2)
        # Future ready
        planned = 100
        if actual >= 90:
            efficiency = "HIGH"
        elif actual >= 60:
            efficiency = "MEDIUM"
        else:
            efficiency = "LOW"
        summary.append(
            {
                "task_id": task.id,
                "category": task.title,
                "plan_percentage": planned,
                "actual_percentage": actual,
                "efficiency": efficiency,
                "status": task.status.value if task.status else None,
            }
        )
    return {
        "project_id": project_id,
        "total_tasks": len(summary),
        "work_summary": summary,
    }


# ===================== AUDIT PDF =====================
@router.get("/audit-pdf")
async def audit_pdf(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)
    # ================= FINANCIAL =================
    total_expense = await db.scalar(
        select(func.sum(Expense.amount)).where(Expense.project_id == project_id)
    )
    total_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(Invoice.project_id == project_id)
    )
    paid_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PAID
        )
    )
    pending_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id == project_id, Invoice.status == InvoiceStatus.PENDING
        )
    )
    # ================= TASKS =================
    total_tasks = await db.scalar(
        select(func.count()).select_from(m.Task).where(m.Task.project_id == project_id)
    )
    completed_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(m.Task.project_id == project_id, m.Task.status == TaskStatus.COMPLETED)
    )
    in_progress_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(m.Task.project_id == project_id, m.Task.status == TaskStatus.IN_PROGRESS)
    )
    # ================= ISSUES =================
    open_issues = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(m.Issue.project_id == project_id, m.Issue.status == IssueStatus.OPEN)
    )
    closed_issues = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(m.Issue.project_id == project_id, m.Issue.status == IssueStatus.CLOSED)
    )
    # ================= PROGRESS =================
    progress = await db.scalar(
        select(func.avg(m.Task.completion_percentage)).where(
            m.Task.project_id == project_id
        )
    )

    b = PdfReportBuilder("Detailed Audit Report")
    b.add_info_table(
        [
            (
                "Project ID",
                str(project_id),
                "Generated On",
                datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
            )
        ]
    )
    b.add_summary_box(
        "FINANCIAL SUMMARY",
        [
            f"Total Expense: Rs. {round(float(total_expense or 0), 2)}",
            f"Total Invoice: Rs. {round(float(total_invoice or 0), 2)}",
            f"Paid Invoice: Rs. {round(float(paid_invoice or 0), 2)}",
            f"Pending Invoice: Rs. {round(float(pending_invoice or 0), 2)}",
            f"Profit: Rs. {round(float((total_invoice or 0) - (total_expense or 0)), 2)}",
        ],
    )
    b.add_summary_box(
        "WORK SUMMARY",
        [
            f"Total Tasks: {int(total_tasks or 0)}",
            f"Completed Tasks: {int(completed_tasks or 0)}",
            f"In Progress Tasks: {int(in_progress_tasks or 0)}",
        ],
    )
    b.add_summary_box(
        "ISSUE SUMMARY",
        [
            f"Open Issues: {int(open_issues or 0)}",
            f"Closed Issues: {int(closed_issues or 0)}",
        ],
    )
    b.add_summary_box(
        "PROJECT PROGRESS",
        [f"Overall Progress: {round(float(progress or 0), 2)} %"],
    )

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=audit_report_{project_id}.pdf"
        },
    )


# =========================================================
# UNIFIED PROJECT REPORT
# =========================================================


@router.get("/project")
async def project_report(
    type: str,
    project_id: Optional[int] = None,
    report_date: date | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    month: int | None = None,
    year: int | None = None,
    quarter: int | None = None,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    project_ids = []
    if project_id is not None:
        await assert_project_access(
            db, project_id=project_id, current_user=current_user
        )
        project_ids = [project_id]
    else:
        # Get all accessible projects
        if current_user.role in [
            UserRole.ADMIN.value,
            UserRole.OWNER.value,
            UserRole.ACCOUNTANT.value,
        ]:
            res = await db.execute(select(m.Project.id))
        else:
            res = await db.execute(
                select(m.Project.id).where(
                    m.Project.id.in_(
                        select(m.ProjectAssignment.project_id).where(
                            m.ProjectAssignment.user_id == current_user.id
                        )
                    )
                )
            )
        project_ids = res.scalars().all()
        if not project_ids:
            raise HTTPException(status_code=403, detail="No accessible projects found")

    # =====================================================
    # VALIDATION
    # =====================================================

    if type not in ["daily", "weekly", "monthly", "quarterly"]:
        raise HTTPException(status_code=400, detail="Invalid report type")

    if type == "daily" and not report_date:
        raise HTTPException(
            status_code=400, detail="report_date is required for daily report"
        )

    if type == "weekly" and (not start_date or not end_date):
        raise HTTPException(
            status_code=400, detail="start_date and end_date required for weekly report"
        )

    if type == "monthly":
        if not month or not year:
            raise HTTPException(
                status_code=400, detail="month and year required for monthly report"
            )

        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])

    # =====================================================
    # QUARTERLY
    # =====================================================

    if type == "quarterly":

        if not quarter or not year:
            raise HTTPException(
                status_code=400, detail="quarter and year required for quarterly report"
            )

        if quarter not in [1, 2, 3, 4]:
            raise HTTPException(
                status_code=400, detail="quarter must be between 1 and 4"
            )

        quarter_map = {
            1: (1, 3),
            2: (4, 6),
            3: (7, 9),
            4: (10, 12),
        }

        start_month, end_month = quarter_map[quarter]

        start_date = date(year, start_month, 1)

        end_date = date(year, end_month, monthrange(year, end_month)[1])

    if type == "daily":
        start_date = report_date
        end_date = report_date

    # =====================================================
    # PROJECT
    # =====================================================

    if project_id is not None:
        project = await db.scalar(select(m.Project).where(m.Project.id == project_id))
        project_data = {"id": project.id, "project_name": project.project_name}
    else:
        project_data = {"id": None, "project_name": "All Authorized Projects"}

    # =====================================================
    # TASK SUMMARY
    # =====================================================

    total_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(m.Task.project_id.in_(project_ids))
    )

    completed_tasks = await db.scalar(
        select(func.count())
        .select_from(m.Task)
        .where(
            m.Task.project_id.in_(project_ids),
            (
                m.Task.status == TaskStatus.COMPLETED.value
                if hasattr(TaskStatus.COMPLETED, "value")
                else TaskStatus.COMPLETED
            ),
        )
    )

    progress = await db.scalar(
        select(func.avg(m.Task.completion_percentage)).where(
            m.Task.project_id.in_(project_ids)
        )
    )

    # =====================================================
    # FINANCIALS
    # =====================================================

    total_invoice = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.project_id.in_(project_ids)
        )
    )

    total_expense = await db.scalar(
        select(func.sum(Expense.amount)).where(Expense.project_id.in_(project_ids))
    )

    # =====================================================
    # ISSUES
    # =====================================================

    open_issues = await db.scalar(
        select(func.count())
        .select_from(m.Issue)
        .where(
            m.Issue.project_id.in_(project_ids),
            (
                m.Issue.status == IssueStatus.OPEN.value
                if hasattr(IssueStatus.OPEN, "value")
                else IssueStatus.OPEN
            ),
        )
    )

    # =====================================================
    # DSR
    # =====================================================

    dsr_list = []
    if project_id is not None:
        dsr_result = await db.execute(
            select(m.DailySiteReport)
            .where(
                m.DailySiteReport.project_id == project_id,
                m.DailySiteReport.report_date >= start_date,
                m.DailySiteReport.report_date <= end_date,
            )
            .order_by(m.DailySiteReport.report_date.desc())
        )
        dsr_list = dsr_result.scalars().all()

    # =====================================================
    # RESPONSE
    # =====================================================

    return {
        "project": project_data,
        "report_type": type,
        "quarter": f"Q{quarter}" if type == "quarterly" else None,
        "date_range": {
            "start_date": start_date,
            "end_date": end_date,
        },
        "summary": {
            "total_tasks": int(total_tasks or 0),
            "completed_tasks": int(completed_tasks or 0),
            "open_issues": int(open_issues or 0),
            "overall_progress": round(float(progress or 0), 2),
        },
        "financials": {
            "total_invoice": round(float(total_invoice or 0), 2),
            "total_expense": round(float(total_expense or 0), 2),
            "profit": round(float((total_invoice or 0) - (total_expense or 0)), 2),
        },
        "daily_reports": [
            {
                "date": r.report_date,
                "work_done": r.work_done,
                "weather": r.weather,
                "remarks": r.remarks,
            }
            for r in dsr_list
        ],
        "generated_at": datetime.utcnow(),
    }


# =========================================================
# EXPORT PDF
# =========================================================


@router.get("/project/export/pdf")
async def export_project_report_pdf(
    type: str,
    project_id: Optional[int] = None,
    report_date: date | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    month: int | None = None,
    year: int | None = None,
    quarter: int | None = None,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    response = await project_report(
        project_id=project_id,
        type=type,
        report_date=report_date,
        start_date=start_date,
        end_date=end_date,
        month=month,
        year=year,
        quarter=quarter,
        current_user=current_user,
        db=db,
    )

    b = PdfReportBuilder(f"{type.title()} Project Report")
    b.add_info_table(
        [
            (
                "Project",
                response["project"]["project_name"],
                "Report Type",
                type.title(),
            )
        ]
    )
    b.add_summary_box(
        "SUMMARY",
        [
            f"<b>Progress:</b> {response['summary']['overall_progress']}%",
            f"Completed Tasks: {response['summary']['completed_tasks']}",
            f"Open Issues: {response['summary']['open_issues']}",
        ],
    )

    headers = ["Date", "Work Done"]
    rows = [[str(r["date"]), r["work_done"]] for r in response["daily_reports"]]
    b.add_section_table(
        "DAILY WORK LOGS", headers, rows, empty_text="No data available"
    )

    stream = b.build()

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={type}_project_report.pdf"
        },
    )


# =========================================================
# EXPORT EXCEL
# =========================================================


@router.get("/project/export/excel")
async def export_project_report_excel(
    type: str,
    project_id: Optional[int] = None,
    report_date: date | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    month: int | None = None,
    year: int | None = None,
    quarter: int | None = None,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    response = await project_report(
        project_id=project_id,
        type=type,
        report_date=report_date,
        start_date=start_date,
        end_date=end_date,
        month=month,
        year=year,
        quarter=quarter,
        current_user=current_user,
        db=db,
    )

    eb = ExcelReportBuilder(
        f"{type.title()} Project Report",
        project_line=response["project"]["project_name"],
    )
    eb.add_summary_row("Report Type", response["report_type"])
    eb.add_summary_row("Overall Progress (%)", response["summary"]["overall_progress"])
    eb.add_summary_row("Completed Tasks", response["summary"]["completed_tasks"])
    eb.add_summary_row("Open Issues", response["summary"]["open_issues"])
    eb.add_summary_row(
        "Total Invoice", response["financials"]["total_invoice"], is_currency=True
    )
    eb.add_summary_row(
        "Total Expense", response["financials"]["total_expense"], is_currency=True
    )
    eb.add_summary_row("Profit", response["financials"]["profit"], is_currency=True)
    eb.build_summary_sheet()

    headers = ["Date", "Work Done", "Weather", "Remarks"]
    rows = [
        [str(r["date"]), r["work_done"], r["weather"], r["remarks"]]
        for r in response["daily_reports"]
    ]
    eb.add_data_sheet("Daily Reports", headers, rows)

    stream = eb.build()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={type}_project_report.xlsx"
        },
    )


# ===================== BUSINESS INTELLIGENCE KPIs =====================
@router.get("/business-intelligence")
async def business_intelligence_kpis(
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    # Revenue (owner invoices)
    revenue = await db.scalar(
        select(func.sum(Invoice.total_amount)).where(
            Invoice.type == "owner", Invoice.status == InvoiceStatus.PAID
        )
    )

    # Expenditure (all expenses)
    expense = await db.scalar(select(func.sum(Expense.amount)))

    revenue_val = float(revenue or 0)
    expense_val = float(expense or 0)
    net_profit = revenue_val - expense_val

    # Activity Log
    documented_reports = await db.scalar(select(func.count(m.DailySiteReport.id)))

    # Efficiency (Active Sites)
    active_sites = await db.scalar(
        select(func.count(m.Project.id)).where(
            m.Project.status == ProjectStatus.ONGOING.value
        )
    )

    return {
        "revenue_focus": net_profit,
        "expenditure": expense_val,
        "activity_log": int(documented_reports or 0),
        "efficiency": f"Syncing from {active_sites or 0} sites",
    }


# ===================== WORK CATEGORY SUMMARY =====================
@router.get("/work-category")
async def work_category_summary(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    result = await db.execute(
        select(
            m.Task.discipline,
            func.count(m.Task.id).label("total_tasks"),
            func.sum(case((m.Task.status == TaskStatus.COMPLETED, 1), else_=0)).label(
                "completed_tasks"
            ),
            func.avg(m.Task.completion_percentage).label("avg_progress"),
        )
        .where(m.Task.project_id == project_id)
        .group_by(m.Task.discipline)
    )

    categories = []
    for row in result.all():
        discipline, total_tasks, completed_tasks, avg_progress = row
        categories.append(
            {
                "category": discipline or "General",
                "total_tasks": int(total_tasks or 0),
                "completed_tasks": int(completed_tasks or 0),
                "avg_progress": round(float(avg_progress or 0), 2),
            }
        )

    return {"work_categories": categories}


# ===================== QUARTERLY AUDIT SUMMARY =====================
@router.get("/audit-summary")
async def quarterly_audit_summary(
    project_id: int,
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    await assert_project_access(db, project_id=project_id, current_user=current_user)

    # Calculate current quarter bounds
    today = date.today()
    current_quarter = (today.month - 1) // 3 + 1
    start_month = 3 * current_quarter - 2
    start_date = date(today.year, start_month, 1)

    # High Priority Issues in Quarter
    critical_issues = await db.scalar(
        select(func.count(m.Issue.id)).where(
            m.Issue.project_id == project_id,
            m.Issue.priority == IssuePriority.HIGH.value,
            m.Issue.created_at >= start_date,
        )
    )

    # Audit trail activities
    audit_logs = await db.scalar(
        select(func.count(ActivityLog.id)).where(
            ActivityLog.entity == "project",
            ActivityLog.entity_id == project_id,
            ActivityLog.created_at >= start_date,
        )
    )

    # Expense Audits
    quarterly_expenses = await db.scalar(
        select(func.sum(Expense.amount)).where(
            Expense.project_id == project_id, Expense.expense_date >= start_date
        )
    )

    return {
        "quarter": f"Q{current_quarter} {today.year}",
        "critical_issues_found": int(critical_issues or 0),
        "audit_activities_logged": int(audit_logs or 0),
        "quarterly_expenses_audited": float(quarterly_expenses or 0),
        "compliance_status": (
            "Passed" if (critical_issues or 0) < 5 else "Review Needed"
        ),
    }


# ===================== COMMERCIAL & BOQ EXECUTION ======================


@router.get("/commercial-execution")
async def commercial_execution_analytics(
    project_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
):
    from app.models.boq import BOQ

    # BOQ Items
    boq_items = (
        (await db.execute(select(BOQ).where(BOQ.project_id == project_id)))
        .scalars()
        .all()
    )

    # Planned Cost from BOQ
    boq_total_planned_cost = sum(float(item.total_cost or 0) for item in boq_items)

    # Final Measurements
    measurements = (
        (
            await db.execute(
                select(FinalMeasurement).where(
                    FinalMeasurement.project_id == project_id,
                    FinalMeasurement.status.in_(["VERIFIED", "APPROVED", "BILLED"]),
                )
            )
        )
        .scalars()
        .all()
    )

    # Actual Certified Amount
    actual_certified_amount = sum(
        float(getattr(meas, "total_amount", 0) or 0) for meas in measurements
    )

    variance = boq_total_planned_cost - actual_certified_amount

    billing_efficiency = round(
        (
            (actual_certified_amount / boq_total_planned_cost * 100)
            if boq_total_planned_cost > 0
            else 0
        ),
        2,
    )

    return {
        "project_id": project_id,
        "boq_items_count": len(boq_items),
        "measurements_count": len(measurements),
        "boq_total_planned_cost": boq_total_planned_cost,
        "actual_certified_amount": actual_certified_amount,
        "variance": variance,
        "billing_efficiency": billing_efficiency,
    }


# ===================== CONTRACTOR EXECUTION =====================
@router.get("/contractor-execution")
async def contractor_execution_analytics(
    project_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
):
    from app.models.billing import RABill

    bills = (
        (await db.execute(select(RABill).where(RABill.project_id == project_id)))
        .scalars()
        .all()
    )

    contractor_stats = {}
    for bill in bills:
        cid = bill.contractor_id
        if cid not in contractor_stats:
            contractor_stats[cid] = {
                "total_billed": 0,
                "paid_amount": 0,
                "bill_count": 0,
            }

        contractor_stats[cid]["total_billed"] += float(bill.gross_amount)
        contractor_stats[cid]["bill_count"] += 1
        if bill.status == "Paid":
            contractor_stats[cid]["paid_amount"] += float(bill.net_amount)

    return {"project_id": project_id, "contractor_stats": contractor_stats}


# NOTE: The following email/WhatsApp sharing endpoints were commented out in the
# original implementation and relied on send_email / send_report_template /
# BackgroundTasks. They are preserved here as-is (still disabled) for reference.
# Re-enable by uncommenting and restoring the relevant imports if needed.

# @router.post("/combined/share/email")
# async def share_combined_report_email(
#     project_id: int,
#     start_date: date,
#     end_date: date,
#     email: str,
#     background_tasks: BackgroundTasks,
#     current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
#     db: AsyncSession = Depends(get_db_session),
# ):
#     await assert_project_access(db, project_id=project_id, current_user=current_user)
#
#     progress = await db.scalar(
#         select(func.avg(m.Task.completion_percentage)).where(
#             m.Task.project_id == project_id
#         )
#     )
#
#     total_paid = await db.scalar(
#         select(func.sum(Invoice.total_amount)).where(
#             Invoice.project_id == project_id,
#             Invoice.status == InvoiceStatus.PAID
#         )
#     )
#
#     total_pending = await db.scalar(
#         select(func.sum(Invoice.total_amount)).where(
#             Invoice.project_id == project_id,
#             Invoice.status == InvoiceStatus.PENDING
#         )
#     )
#
#     reports = await db.execute(
#         select(m.DailySiteReport).where(
#             m.DailySiteReport.project_id == project_id,
#             m.DailySiteReport.report_date >= start_date,
#             m.DailySiteReport.report_date <= end_date,
#         )
#     )
#     dsr_list = reports.scalars().all()
#
#     b = PdfReportBuilder("Combined Project Report")
#     b.add_info_table([("Project ID", str(project_id), "Period", f"{start_date} to {end_date}")])
#     b.add_summary_box("SUMMARY", [
#         f"Progress: {round(progress or 0, 2)}%",
#         f"Paid: {float(total_paid or 0)}",
#         f"Pending: {float(total_pending or 0)}",
#     ])
#     headers = ["Date", "Work Done"]
#     rows = [[str(r.report_date), r.work_done] for r in dsr_list]
#     b.add_section_table("WORK SUMMARY", headers, rows)
#     buffer = b.build()
#
#     background_tasks.add_task(
#         send_email,
#         to_email=email,
#         subject="Combined Project Report",
#         body=f"Report from {start_date} to {end_date}",
#         attachment=buffer.read(),
#         filename="combined_report.pdf",
#     )
#
#     return {"message": "Email queued successfully"}


# @router.post("/combined/share/whatsapp")
# async def share_combined_whatsapp(
#     project_id: int,
#     start_date: date,
#     end_date: date,
#     phone: str,
#     current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
#     db: AsyncSession = Depends(get_db_session),
# ):
#     await assert_project_access(db, project_id=project_id, current_user=current_user)
#
#     report_url = f"http://localhost:8000/reports/combined?project_id={project_id}&start_date={start_date}&end_date={end_date}"
#
#     result = await send_report_template(
#         to=phone,
#         name="Client",
#         report_url=report_url
#     )
#
#     return {
#         "message": "WhatsApp message sent",
#         "response": result
#     }


# @router.post("/daily/share/email")
# async def share_daily_email(
#     project_id: int,
#     report_date: date,
#     email: str,
#     background_tasks: BackgroundTasks,
#     current_user: User = Depends(require_roles(REPORT_READ_ROLES)),
#     db: AsyncSession = Depends(get_db_session),
# ):
#     await assert_project_access(db, project_id=project_id, current_user=current_user)
#
#     dsr = await db.scalar(
#         select(m.DailySiteReport).where(
#             m.DailySiteReport.project_id == project_id,
#             m.DailySiteReport.report_date == report_date,
#         )
#     )
#
#     b = PdfReportBuilder(f"Daily Report - {report_date}")
#     if dsr:
#         b.add_summary_box("SITE DETAILS", [
#             f"Work Done: {dsr.work_done}",
#             f"Weather: {dsr.weather}",
#         ])
#     else:
#         b.add_summary_box("SITE DETAILS", ["No data available"])
#     buffer = b.build()
#
#     body = f"""
#     <html>
#     <body style="font-family: Arial, sans-serif;">
#         <h2>Daily Site Report</h2>
#         <p><b>Date:</b> {report_date}</p>
#         <p>Please find the attached report.</p>
#         <hr>
#         <p style="font-size:12px;color:gray;">Construction Management System</p>
#     </body>
#     </html>
#     """
#
#     background_tasks.add_task(
#         send_email,
#         to_email=email,
#         subject="Daily Report",
#         body=body,
#         attachment=buffer.read(),
#         filename="daily_report.pdf",
#     )
#
#     return {"message": "Email queued successfully"}
