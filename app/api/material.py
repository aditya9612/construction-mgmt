from app.utils.helpers import NotFoundError
from datetime import date
from typing import Optional, List
from decimal import Decimal, InvalidOperation
import uuid
import os
from sqlalchemy.exc import IntegrityError
from fastapi import APIRouter, Depends, HTTPException, Query, Header
from sqlalchemy import exists, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from reportlab.platypus import SimpleDocTemplate, Table
from openpyxl import Workbook
from app.models.boq import BOQ
from app.models.user import User, UserRole, ActivityLog
from app.utils.boq_calc import recalculate_boq_actuals
from app.schemas.material import (
    InventoryAdjustResponse,
    MaterialReportResponse,
    MaterialReportSummary,
    TransferMaterial,
    TransferProject,
)
from app.cache.redis import (
    bump_cache_version,
    cache_get_json,
    cache_set_json,
    get_cache_version,
)

from app.core.dependencies import (
    get_current_active_user,
    get_request_redis,
    require_roles,
)
from app.db.session import get_db_session
from app.models.ai_prediction import AIPrediction
from app.schemas.material import (
    MaterialReport,
    PriceHistoryOut,
    AIMaterialRecommendationRequest,
    AIMaterialRecommendationResponse,
    SupplierRecommendationResponse,
    AIProcurementSummaryResponse,
    MaterialConsumptionTrendResponse,
    ReorderAlertsResponse,
    ProjectTransactionOut,
    TransferListResponse,
    MessageResponse,
    AIMaterialRecommendationRequest,
    AIMaterialRecommendationResponse,
    RecommendationSummary,
    MaterialRecommendationDetail,
    SupplierRecommendationDetail,
    SupplierRecommendationResponse,
    AIProcurementSummaryResponse,
    DailyConsumptionItem,
    ConsumptionDaySummary,
    MaterialConsumptionTrendResponse,
    ReorderAlertOut,
    ReorderAlertsResponse,
)

from app.utils.project_report_pdf import generate_procurement_report_pdf

from app.core.enums import IssueType, TransactionType, TransferStatus
from app.utils.common import generate_business_id
from app.utils.qr import generate_qr
from sqlalchemy.orm import selectinload
from fastapi.responses import FileResponse, StreamingResponse
from starlette.background import BackgroundTask
from app.services.notification_service import create_notification

from app.models import project as proj_model
import os
from app.models.project import Project
from app.models.material import (
    Material,
    MaterialLedger,
    MaterialTransaction,
    Supplier,
    PurchaseOrder,
    MaterialTransfer,
)
from reportlab.platypus import Image
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics import renderPM
from reportlab.lib.units import inch
from starlette.background import BackgroundTask
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models.material import MaterialUsage
from app.schemas.material import InventoryAdjustRequest
from app.models.project import Project
from sqlalchemy.orm import aliased
import re
from sqlalchemy import case
from decimal import Decimal
import tempfile, os, uuid
from datetime import datetime
from app.models.master_data import MaterialMaster, Unit
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle

from app.core.enums import TransactionType as DBTransactionType
from app.schemas.material import TransactionType as SchemaTransactionType
from app.schemas.material import (
    MaterialLogOut,
    MaterialCreate,
    MaterialOut,
    MaterialUpdate,
    SummaryOut,
    PurchaseMaterial,
    UsageMaterial,
    SupplierCreate,
    SupplierOut,
    PurchaseOrderCreate,
    PurchaseOrderOut,
    TransferCreate,
    TransferOut,
)

from app.core.logger import logger

MATERIAL_READ_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.ACCOUNTANT,
        UserRole.CLIENT,
    ]
]

MATERIAL_WRITE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
    ]
]

router = APIRouter(prefix="/materials", tags=["materials"])
VERSION_KEY = "cache_version:materials"


def safe_delete(file_path: str):
    try:
        os.remove(file_path)
    except OSError:
        pass


# ================= CENTRAL CALCULATION =================


def build_material_response(
    obj,
    supplier_name: str | None,
    unit_name=None,
):
    total_amount = float(obj.total_amount or 0)
    payment_given = float(obj.payment_given or 0)

    payment_pending = max(
        0,
        total_amount - payment_given,
    )

    extra_paid = max(
        0,
        payment_given - total_amount,
    )

    remaining = float(obj.remaining_stock or 0)

    min_level = float(obj.minimum_stock_level or 0)

    if remaining == 0:
        alert = "OUT_OF_STOCK"

    elif remaining <= min_level:
        alert = "LOW_STOCK"

    else:
        alert = "IN_STOCK"

    resolved_unit_name = unit_name if unit_name else (obj.unit.name if obj.unit else "")

    return MaterialOut(
        id=obj.id,
        material_code=obj.material_code,
        project_id=obj.project_id,
        material_master_id=obj.material_master_id,
        material_master_name=(
            obj.material_master.name if obj.material_master else None
        ),
        material_master_brand=(
            obj.material_master.brand if obj.material_master else None
        ),
        material_master_specification=(
            obj.material_master.specification if obj.material_master else None
        ),
        material_master_hsn_code=(
            obj.material_master.hsn_code if obj.material_master else None
        ),
        material_name=(obj.material_name or "").strip().title(),
        category=obj.category,
        unit_id=obj.unit_id,
        unit_name=resolved_unit_name,
        supplier_id=obj.supplier_id,
        supplier_name=(supplier_name if supplier_name else "N/A"),
        purchase_rate=round(
            float(obj.purchase_rate or 0),
            2,
        ),
        rate_type=obj.rate_type,
        quantity_purchased=round(
            float(obj.quantity_purchased or 0),
            2,
        ),
        quantity_used=round(
            float(obj.quantity_used or 0),
            2,
        ),
        remaining_stock=round(
            remaining,
            2,
        ),
        total_amount=round(
            total_amount,
            2,
        ),
        payment_given=round(
            payment_given,
            2,
        ),
        payment_pending=round(
            payment_pending,
            2,
        ),
        extra_paid=round(
            extra_paid,
            2,
        ),
        minimum_stock_level=round(
            min_level,
            2,
        ),
        alert_type=alert,
    )


def update_material_fields(obj: Material):
    qty_purchased = obj.quantity_purchased or Decimal("0")
    qty_used = obj.quantity_used or Decimal("0")
    payment_given = obj.payment_given or Decimal("0")
    total_amount = obj.total_amount or Decimal("0")

    obj.remaining_stock = max(qty_purchased - qty_used, Decimal("0"))
    obj.payment_pending = max(total_amount - payment_given, Decimal("0"))
    obj.advance_amount = max(payment_given - total_amount, Decimal("0"))


def build_po_response(po: PurchaseOrder) -> PurchaseOrderOut:
    return PurchaseOrderOut(
        id=po.id,
        supplier_id=po.supplier_id,
        project_id=po.project_id,
        material_id=po.material_id,
        material_name=(po.material_name or "").strip().title(),
        quantity=round(float(po.quantity or 0), 2),
        rate=round(float(po.rate or 0), 2),
        total_amount=round(float(po.total_amount or 0), 2),
        status=po.status,
    )


def get_signed_quantity(tx: MaterialTransaction) -> float:
    if tx.type in {DBTransactionType.USAGE, DBTransactionType.TRANSFER_OUT}:
        return -abs(float(tx.quantity or 0))
    return abs(float(tx.quantity or 0))


def build_transfer_response(obj, material, from_project, to_project):
    return TransferOut(
        id=obj.id,
        material=TransferMaterial(
            id=material.id, name=(material.material_name or "").title()
        ),
        from_project=TransferProject(
            id=from_project.id, name=from_project.project_name
        ),
        to_project=TransferProject(id=to_project.id, name=to_project.project_name),
        quantity=obj.quantity,
        status=obj.status,
        created_at=obj.created_at,
    )


def get_alert_type(obj):
    from decimal import Decimal

    remaining = obj.remaining_stock or Decimal("0")
    min_level = obj.minimum_stock_level or Decimal("0")

    if remaining == 0:
        return "OUT_OF_STOCK"
    elif remaining <= min_level:
        return "LOW_STOCK"
    else:
        return "IN_STOCK"


def calculate_fields(obj):
    from decimal import Decimal

    total_amount = obj.total_amount or Decimal("0")
    payment_given = obj.payment_given or Decimal("0")

    payment_pending = max(total_amount - payment_given, Decimal("0"))
    extra_paid = max(payment_given - total_amount, Decimal("0"))

    return (
        float(total_amount),
        float(payment_given),
        float(payment_pending),
        float(extra_paid),
    )


def calculate_payment(total, paid):
    return (
        float(total),
        float(paid),
        float(max(total - paid, 0)),
        float(max(paid - total, 0)),
    )


def calculate_avg_rate(material):
    """Weighted average purchase rate = total_amount / quantity_purchased."""
    qty = material.quantity_purchased or Decimal("0")
    total = material.total_amount or Decimal("0")

    return total / qty if qty > 0 else Decimal("0")


def generate_chart(rows, path):
    import matplotlib.pyplot as plt

    names = []
    stock = []

    for m, _ in rows[:10]:
        names.append((m.material_name or "")[:8])
        stock.append(float(m.remaining_stock or 0))

    plt.figure(figsize=(6, 3))
    plt.bar(names, stock)
    plt.xticks(rotation=30)
    plt.tight_layout()
    plt.savefig(path)
    plt.close()


from decimal import Decimal
from typing import Any, Dict


async def calculate_best_supplier_for_material(db, material) -> Dict[str, Any]:
    from sqlalchemy import select, or_
    from app.models.material import Supplier, PurchaseOrder, Material

    stmt = select(Supplier).where(Supplier.is_deleted == False)
    res = await db.execute(stmt)
    suppliers = res.scalars().all()

    if not suppliers:
        return {
            "supplier_id": material.supplier_id or 0,
            "supplier_name": "Default Supplier",
            "last_purchase_rate": round(float(material.purchase_rate or 0), 2),
            "average_delivery_days": 3.0,
            "supplier_score": 75.0,
            "recommendation_reason": "Default supplier assigned to material.",
        }

    supplier_ids = [s.id for s in suppliers]
    all_pos_stmt = (
        select(PurchaseOrder)
        .where(
            PurchaseOrder.supplier_id.in_(supplier_ids),
            PurchaseOrder.is_deleted == False,
        )
        .order_by(PurchaseOrder.id.desc())
    )
    all_pos_res = await db.execute(all_pos_stmt)
    all_pos = all_pos_res.scalars().all()

    po_by_supplier = {}
    for p in all_pos:
        po_by_supplier.setdefault(p.supplier_id, []).append(p)

    supp_mats_stmt = select(Material).where(
        Material.supplier_id.in_(supplier_ids), Material.is_deleted == False
    )
    supp_mats_res = await db.execute(supp_mats_stmt)
    supp_mats = supp_mats_res.scalars().all()
    mats_by_supplier = {}
    for m in supp_mats:
        mats_by_supplier.setdefault(m.supplier_id, []).append(m)

    supplier_evals = []
    rates = []

    for supp in suppliers:
        pos = po_by_supplier.get(supp.id, [])
        mat_pos = [
            p
            for p in pos
            if p.material_id == material.id
            or (
                material.material_name
                and material.material_name.lower() in (p.material_name or "").lower()
            )
        ]
        if mat_pos and float(mat_pos[0].rate or 0) > 0:
            rates.append(float(mat_pos[0].rate))
        elif supp.id == material.supplier_id and material.purchase_rate:
            rates.append(float(material.purchase_rate))

    min_rate = min(rates) if rates else float(material.purchase_rate or 1.0)
    if min_rate <= 0:
        min_rate = 1.0

    for supp in suppliers:
        pos = po_by_supplier.get(supp.id, [])
        mat_pos = [
            p
            for p in pos
            if p.material_id == material.id
            or (
                material.material_name
                and material.material_name.lower() in (p.material_name or "").lower()
            )
        ]

        if mat_pos and float(mat_pos[0].rate or 0) > 0:
            last_rate = float(mat_pos[0].rate)
        elif supp.id == material.supplier_id:
            last_rate = float(material.purchase_rate or 0)
        else:
            sm_list = mats_by_supplier.get(supp.id, [])
            last_rate = (
                float(sm_list[0].purchase_rate)
                if sm_list and sm_list[0].purchase_rate
                else float(material.purchase_rate or 0)
            )

        total_pos = len(pos)
        completed_pos = sum(
            1
            for p in pos
            if str(p.status).upper() in ["COMPLETED", "DELIVERED", "APPROVED"]
        )
        rejected_pos = sum(
            1 for p in pos if str(p.status).upper() in ["REJECTED", "CANCELLED"]
        )

        rejection_rate = (rejected_pos / total_pos) if total_pos > 0 else 0.0

        deliv_days_list = []
        for p in pos:
            if (
                str(p.status).upper() in ["COMPLETED", "DELIVERED", "APPROVED"]
                and p.created_at
                and p.updated_at
            ):
                diff = (p.updated_at - p.created_at).total_seconds() / 86400.0
                if diff >= 0:
                    deliv_days_list.append(diff)

        avg_delivery_days = (
            (sum(deliv_days_list) / len(deliv_days_list)) if deliv_days_list else 3.0
        )
        if avg_delivery_days <= 0:
            avg_delivery_days = 1.0

        price_ratio = (last_rate - min_rate) / min_rate if min_rate > 0 else 0
        price_score = max(0.0, 100.0 * (1.0 - price_ratio))
        delivery_score = max(0.0, 100.0 * (1.0 - (avg_delivery_days / 14.0)))
        completed_po_score = min(100.0, completed_pos * 20.0)
        rejection_score = 100.0 * (1.0 - rejection_rate)

        is_preferred = supp.id == material.supplier_id
        preferred_bonus = (
            100.0
            if (is_preferred and completed_pos > 0 and rejection_rate == 0)
            else (50.0 if is_preferred else 0.0)
        )

        score = round(
            0.35 * price_score
            + 0.25 * delivery_score
            + 0.20 * completed_po_score
            + 0.10 * rejection_score
            + 0.10 * preferred_bonus,
            2,
        )
        score = max(0.0, min(100.0, score))

        reasons = []
        if last_rate <= min_rate:
            reasons.append(f"lowest purchase rate (₹{last_rate:,.2f})")
        else:
            reasons.append(f"competitive rate (₹{last_rate:,.2f})")
        if completed_pos > 0:
            reasons.append(f"{completed_pos} completed order(s)")
        reasons.append(f"fast avg delivery ({round(avg_delivery_days, 1)} days)")
        if rejection_rate == 0 and total_pos > 0:
            reasons.append("0% rejection rate")
        if is_preferred:
            reasons.append("preferred supplier on record")

        rec_reason = f"Recommended based on {', '.join(reasons)}."

        supplier_evals.append(
            {
                "supplier_id": supp.id,
                "supplier_name": supp.supplier_name,
                "last_purchase_rate": round(last_rate, 2),
                "average_delivery_days": round(avg_delivery_days, 1),
                "supplier_score": score,
                "recommendation_reason": rec_reason,
            }
        )

    supplier_evals.sort(key=lambda x: x["supplier_score"], reverse=True)
    return supplier_evals[0]


async def generate_supplier_recommendations(db, project_id: int) -> Dict[str, Any]:
    from sqlalchemy import select
    from app.models.material import Material
    from app.models.project import Project

    project = await db.scalar(select(Project).where(Project.id == project_id))
    if not project:
        raise NotFoundError(f"Project with id {project_id} not found")

    materials = (
        (
            await db.execute(
                select(Material).where(
                    Material.project_id == project_id, Material.is_deleted == False
                )
            )
        )
        .scalars()
        .all()
    )

    supplier_recs = []
    for mat in materials:
        best_supp = await calculate_best_supplier_for_material(db, mat)
        supplier_recs.append(
            {
                "material_id": mat.id,
                "material_name": mat.material_name,
                "recommended_supplier": best_supp,
            }
        )

    return {
        "project_id": project_id,
        "supplier_recommendations": supplier_recs,
    }


async def generate_material_recommendations(
    db, project_id: int, target_days: int = 30
) -> Dict[str, Any]:
    from sqlalchemy import select
    from app.models.material import Material
    from app.models.project import Project

    project = await db.scalar(select(Project).where(Project.id == project_id))
    if not project:
        raise NotFoundError(f"Project with id {project_id} not found")

    target_days = max(1, target_days)

    stmt = (
        select(Material)
        .options(
            selectinload(Material.unit),
            selectinload(Material.supplier),
        )
        .where(
            Material.project_id == project_id,
            Material.is_deleted == False,
        )
    )
    result = await db.execute(stmt)
    materials = result.scalars().all()

    recommendations = []
    critical_items = 0
    warning_items = 0
    total_estimated_budget = Decimal("0.00")

    for mat in materials:
        qty_used = Decimal(str(mat.quantity_used or 0))
        remaining_stock = Decimal(str(mat.remaining_stock or 0))
        min_stock = Decimal(str(mat.minimum_stock_level or 0))
        rate = Decimal(str(mat.purchase_rate or 0))

        avg_daily_consumption = qty_used / Decimal(str(target_days))

        if avg_daily_consumption > Decimal("0"):
            remaining_days = float(remaining_stock / avg_daily_consumption)
        else:
            remaining_days = 999.0 if remaining_stock > Decimal("0") else 0.0

        buffer_qty = avg_daily_consumption * Decimal("7")
        needed_qty = (min_stock + buffer_qty) - remaining_stock
        if remaining_stock <= min_stock and needed_qty <= Decimal("0"):
            needed_qty = min_stock - remaining_stock
        recommended_purchase = max(Decimal("0.000"), needed_qty)

        estimated_cost = recommended_purchase * rate

        if remaining_days <= 3:
            priority = "High"
            critical_items += 1
        elif remaining_days <= 7:
            priority = "High"
            critical_items += 1
        elif remaining_days <= 15:
            priority = "Medium"
            warning_items += 1
        else:
            priority = "Low"

        if remaining_stock <= Decimal("0"):
            reason = "Material is completely out of stock."
            risk = "Active work dependent on this material is halted."
        elif remaining_stock <= min_stock:
            reason = f"Current stock ({round(float(remaining_stock), 2)}) is below minimum level ({round(float(min_stock), 2)})."
            risk = "Potential site delay due to material depletion."
        elif remaining_days <= 7:
            reason = f"Stock estimated to run out within {round(remaining_days, 1)} days based on consumption rate."
            risk = "Work interruption expected within a week if not reordered."
        else:
            reason = "Stock level is adequate for target period."
            risk = "Low operational risk."

        total_estimated_budget += estimated_cost

        # Calculate AI Best Supplier Recommendation (FEATURE 1)
        best_supplier = await calculate_best_supplier_for_material(db, mat)

        recommendations.append(
            {
                "material_id": mat.id,
                "material_name": (mat.material_name or "Material").strip().title(),
                "unit_name": (mat.unit.name if getattr(mat, "unit", None) else None),
                "supplier_name": (
                    mat.supplier.supplier_name
                    if getattr(mat, "supplier", None)
                    else None
                ),
                "current_stock": round(float(remaining_stock), 2),
                "minimum_stock": round(float(min_stock), 2),
                "remaining_days": round(remaining_days, 1),
                "recommended_purchase": round(float(recommended_purchase), 2),
                "estimated_cost": round(float(estimated_cost), 2),
                "priority": priority,
                "reason": reason,
                "risk": risk,
                "recommended_supplier": best_supplier,
            }
        )

    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    recommendations.sort(
        key=lambda x: (priority_order.get(x["priority"], 3), x["remaining_days"])
    )

    return {
        "project_id": project_id,
        "summary": {
            "critical_items": critical_items,
            "warning_items": warning_items,
            "estimated_budget": round(float(total_estimated_budget), 2),
        },
        "recommendations": recommendations,
    }


async def generate_procurement_summary(db, project_id: int) -> Dict[str, Any]:
    from sqlalchemy import select
    from app.models.project import Project

    project = await db.scalar(select(Project).where(Project.id == project_id))
    if not project:
        raise NotFoundError(f"Project with id {project_id} not found")

    rec_data = await generate_material_recommendations(db, project_id, target_days=30)
    recommendations = rec_data.get("recommendations", [])
    summary = rec_data.get("summary", {})

    critical_items = [
        item
        for item in recommendations
        if item.get("remaining_days", 999) <= 7
        or item.get("priority") in ["High", "Medium"]
    ]

    critical_names = [item["material_name"] for item in critical_items]
    total_budget = float(summary.get("estimated_budget", 0.0))

    if summary.get("critical_items", 0) > 0:
        overall_risk = "HIGH"
    elif summary.get("warning_items", 0) > 0:
        overall_risk = "MEDIUM"
    else:
        overall_risk = "LOW"

    if critical_items:
        first_mat = critical_items[0]
        rem_days = first_mat.get("remaining_days", 0)
        p_text = f"{first_mat['material_name']} inventory is projected to finish within {rem_days} days. Immediate procurement is recommended."
        if len(critical_items) > 1:
            second_mat = critical_items[1]
            p_text += f" {second_mat['material_name']} stock is low and requires reordering soon."
        healthy = [
            item["material_name"]
            for item in recommendations
            if item["material_name"] not in critical_names
        ]
        if healthy:
            p_text += f" {healthy[0]} inventory is healthy for the next week."
    else:
        p_text = "All material inventory levels are healthy and within safe operational limits for the target period."

    action_text = (
        f"Initiate purchase orders for {len(critical_names)} critical material(s): {', '.join(critical_names)}. Estimated budget required: ₹{total_budget:,.2f}."
        if critical_names
        else "Maintain routine inventory monitoring; no immediate procurement required."
    )

    return {
        "project_id": project_id,
        "procurement_summary": p_text,
        "overall_risk": overall_risk,
        "recommended_action": action_text,
        "estimated_budget": round(total_budget, 2),
        "critical_materials": critical_names,
    }


async def get_material_consumption_trend(
    db, project_id: int, days: int = 30
) -> Dict[str, Any]:
    from datetime import datetime, timedelta
    from sqlalchemy import select
    from app.models.material import MaterialUsage, Material, MaterialTransaction
    from app.models.project import Project
    from app.core.enums import TransactionType

    project = await db.scalar(select(Project).where(Project.id == project_id))
    if not project:
        raise NotFoundError(f"Project with id {project_id} not found")

    cutoff_date = datetime.utcnow() - timedelta(days=days)

    stmt = (
        select(MaterialUsage, Material.material_name)
        .join(Material, Material.id == MaterialUsage.material_id)
        .where(
            MaterialUsage.project_id == project_id,
            MaterialUsage.usage_date >= cutoff_date,
        )
        .order_by(MaterialUsage.usage_date.asc())
    )
    result = await db.execute(stmt)
    usage_rows = result.all()

    daily_map = {}
    day_total_map = {}

    for usage, mat_name in usage_rows:
        d_str = (
            usage.usage_date.strftime("%Y-%m-%d")
            if usage.usage_date
            else datetime.utcnow().strftime("%Y-%m-%d")
        )
        qty = float(usage.quantity_used or 0)
        key = (d_str, mat_name or "Material")
        daily_map[key] = daily_map.get(key, 0.0) + qty
        day_total_map[d_str] = day_total_map.get(d_str, 0.0) + qty

    if not daily_map:
        tx_stmt = (
            select(MaterialTransaction, Material.material_name)
            .join(Material, Material.id == MaterialTransaction.material_id)
            .where(
                MaterialTransaction.project_id == project_id,
                MaterialTransaction.created_at >= cutoff_date,
                MaterialTransaction.type.in_(
                    [TransactionType.USAGE, TransactionType.ISSUE]
                ),
            )
            .order_by(MaterialTransaction.created_at.asc())
        )
        tx_rows = (await db.execute(tx_stmt)).all()
        for tx, mat_name in tx_rows:
            d_str = (
                tx.created_at.strftime("%Y-%m-%d")
                if tx.created_at
                else datetime.utcnow().strftime("%Y-%m-%d")
            )
            qty = float(tx.quantity or 0)
            key = (d_str, mat_name or "Material")
            daily_map[key] = daily_map.get(key, 0.0) + qty
            day_total_map[d_str] = day_total_map.get(d_str, 0.0) + qty

    daily_trends = [
        {
            "date": d_str,
            "material_name": mat_name,
            "quantity_used": round(qty, 2),
        }
        for (d_str, mat_name), qty in sorted(daily_map.items(), key=lambda x: x[0][0])
    ]

    total_consumed = sum(item["quantity_used"] for item in daily_trends)
    avg_daily = round(total_consumed / float(days), 2)

    highest_day = None
    lowest_day = None

    if day_total_map:
        sorted_days = sorted(day_total_map.items(), key=lambda x: x[1], reverse=True)
        max_d, max_q = sorted_days[0]
        min_d, min_q = sorted_days[-1]

        max_mats = [m for (d, m), q in daily_map.items() if d == max_d]
        min_mats = [m for (d, m), q in daily_map.items() if d == min_d]

        highest_day = {
            "date": max_d,
            "quantity_used": round(max_q, 2),
            "material_name": max_mats[0] if max_mats else None,
        }
        lowest_day = {
            "date": min_d,
            "quantity_used": round(min_q, 2),
            "material_name": min_mats[0] if min_mats else None,
        }

    return {
        "project_id": project_id,
        "period_days": days,
        "daily_trends": daily_trends,
        "average_daily_consumption": avg_daily,
        "highest_consumption_day": highest_day,
        "lowest_consumption_day": lowest_day,
    }


async def check_and_trigger_reorder_alerts(
    db, project_id: int, current_user_id: Optional[int] = None
) -> Dict[str, Any]:
    from datetime import datetime, timedelta
    from sqlalchemy import select
    from app.models.material import Material
    from app.models.notification import Notification
    from app.models.alert import Alert
    from app.models.user import User, UserRole
    from app.services.notification_service import create_notification

    stmt = (
        select(Material)
        .options(
            selectinload(Material.unit),
            selectinload(Material.supplier),
        )
        .where(
            Material.project_id == project_id,
            Material.is_deleted == False,
        )
    )
    materials = (await db.execute(stmt)).scalars().all()

    target_user_id = current_user_id
    if not target_user_id:
        user_stmt = select(User.id).where(User.role == UserRole.ADMIN.value).limit(1)
        target_user_id = await db.scalar(user_stmt) or 1

    alerts_list = []
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    cutoff_time = datetime.utcnow() - timedelta(hours=24)

    for mat in materials:
        rem_stock = float(mat.remaining_stock or 0)
        min_stock = float(mat.minimum_stock_level or 0)
        qty_used = float(mat.quantity_used or 0)

        avg_daily = qty_used / 30.0
        buffer_qty = avg_daily * 7.0
        rec_qty = max(1.0, (min_stock + buffer_qty) - rem_stock)

        priority = "HIGH" if rem_stock <= 0 else "MEDIUM"
        msg = (
            f"AUTOMATIC REORDER ALERT: Material '{mat.material_name}' (Code: {mat.material_code}) "
            f"in Project ID {project_id} has remaining stock of {round(rem_stock, 2)} "
            f"which is <= minimum stock ({round(min_stock, 2)}). "
            f"Recommended reorder quantity: {round(rec_qty, 2)} units. Priority: {priority}."
        )

        dup_stmt = select(Notification).where(
            Notification.user_id == target_user_id,
            Notification.title.ilike(f"%REORDER ALERT: {mat.material_name}%"),
            Notification.created_at >= cutoff_time,
        )
        existing_notif = await db.scalar(dup_stmt)

        notif_id = None
        if not existing_notif:
            notif = await create_notification(
                db=db,
                user_id=target_user_id,
                title=f"AUTOMATIC REORDER ALERT: {mat.material_name}",
                message=msg,
                type="alert" if priority == "HIGH" else "warning",
                link=f"/materials/{mat.id}",
            )
            await db.flush()
            notif_id = notif.id

            try:
                alert_obj = Alert(
                    project_id=project_id,
                    alert_type="REORDER_ALERT",
                    message=msg,
                    user_id=target_user_id,
                    status="active",
                )
                db.add(alert_obj)
                await db.flush()
            except Exception as e:
                pass
        else:
            notif_id = existing_notif.id

        alerts_list.append(
            {
                "id": notif_id,
                "material_id": mat.id,
                "material_name": mat.material_name,
                "project_id": project_id,
                "remaining_stock": round(rem_stock, 2),
                "minimum_stock": round(min_stock, 2),
                "recommended_quantity": round(rec_qty, 2),
                "priority": priority,
                "timestamp": now_str,
                "message": msg,
            }
        )

    return {
        "project_id": project_id,
        "alerts": alerts_list,
        "total_alerts": len(alerts_list),
    }


# ================= SUMMARY =================


@router.get("/summary", response_model=SummaryOut)
async def material_summary(
    project_id: Optional[int] = Query(
        None,
        description="Filter by project ID",
    ),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    from decimal import Decimal
    from fastapi import HTTPException

    try:
        # ==========================================================
        # BASE QUERY
        # ==========================================================

        material_filter = [Material.is_deleted == False]

        if project_id is not None:
            material_filter.append(Material.project_id == project_id)

        # ==========================================================
        # STEP 1: TOTAL MATERIALS
        # ==========================================================

        total_materials = await db.scalar(
            select(func.count(Material.id)).where(*material_filter)
        )

        # ==========================================================
        # STEP 2: FETCH MATERIALS
        # ==========================================================

        result = await db.execute(select(Material).where(*material_filter))

        rows = result.scalars().all()

        # ==========================================================
        # STEP 3: STOCK VALUE
        # ==========================================================

        total_stock = Decimal("0")

        for m in rows:

            purchased = m.quantity_purchased or Decimal("0")
            total_amt = m.total_amount or Decimal("0")
            remaining = m.remaining_stock or Decimal("0")

            avg_rate = total_amt / purchased if purchased > 0 else Decimal("0")

            total_stock += remaining * avg_rate

        # ==========================================================
        # STEP 4: TOTAL PENDING
        # ==========================================================

        total_pending = await db.scalar(
            select(func.sum(Material.payment_pending)).where(*material_filter)
        )

        # ==========================================================
        # RESPONSE
        # ==========================================================

        return {
            "total_materials": total_materials or 0,
            "total_stock_value": round(float(total_stock or 0), 2),
            "total_pending_payments": round(float(total_pending or 0), 2),
        }

    except Exception as e:
        logger.exception("Material summary failed")

        raise HTTPException(
            status_code=500,
            detail=f"Material summary failed: {type(e).__name__}: {str(e)}",
        )


# ================= SUPPLIERS =================


@router.get("/suppliers", response_model=List[SupplierOut])
async def list_suppliers(
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):

    rows = (
        (
            await db.execute(
                select(Supplier)
                .where(Supplier.is_deleted == False)
                .order_by(Supplier.id.desc())
                .offset(skip)
                .limit(limit)
            )
        )
        .scalars()
        .all()
    )

    return [SupplierOut.model_validate(r) for r in rows]


# ================Get_supplier===============

@router.get("/suppliers/{supplier_id}/qr", response_class=StreamingResponse)
async def generate_supplier_qr(
    supplier_id: int,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    supplier = await db.get(Supplier, supplier_id)

    if not supplier or supplier.is_deleted:
        raise HTTPException(
            status_code=404,
            detail="Supplier not found",
        )

    qr_buf = generate_qr(entity_type="VEN", entity_id=supplier.id)

    headers = {
        "Cache-Control": "no-store",
        "Content-Disposition": f'inline; filename="supplier_{supplier.id}.png"'
    }

    return StreamingResponse(
        qr_buf,
        media_type="image/png",
        headers=headers
    )

@router.get("/suppliers/{supplier_id}", response_model=SupplierOut)
async def get_supplier(
    supplier_id: int,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    supplier = await db.get(Supplier, supplier_id)

    if not supplier or supplier.is_deleted:
        raise HTTPException(
            status_code=404,
            detail="Supplier not found",
        )

    return SupplierOut.model_validate(supplier)


# ================Create_supplier===============


@router.post("/suppliers", response_model=SupplierOut, status_code=201)
async def create_supplier(
    payload: SupplierCreate,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    import re

    # Name validation
    supplier_name = payload.supplier_name.strip().title()
    if len(supplier_name) < 3:
        raise HTTPException(400, "Supplier name must be at least 3 characters")

    # Contact person validation
    contact_person = (
        payload.contact_person.strip().title() if payload.contact_person else None
    )

    if contact_person and not re.fullmatch(r"[A-Za-z ]{3,}", contact_person):
        raise HTTPException(400, "Invalid contact person name")

    # Phone / Email validation
    phone_email = payload.phone_email.strip() if payload.phone_email else None

    if phone_email:
        cleaned = re.sub(r"[^\d]", "", phone_email)

        if cleaned.startswith("91") and len(cleaned) == 12:
            cleaned = cleaned[2:]

        if cleaned.isdigit():
            if not re.fullmatch(r"[6-9]\d{9}", cleaned):
                raise HTTPException(400, "Invalid Indian mobile number")

            phone_email = cleaned

        elif "@" in phone_email:
            if not re.fullmatch(r"[^@]+@[^@]+\.[^@]+", phone_email):
                raise HTTPException(400, "Invalid email format")

        else:
            raise HTTPException(400, "Enter valid phone number or email")

    # GST validation
    gst_number = payload.gst_number.strip().upper() if payload.gst_number else None

    if gst_number and not re.fullmatch(
        r"\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d]",
        gst_number,
    ):
        raise HTTPException(400, "Invalid GST number format")

    # Address validation
    address = payload.address.strip() if payload.address else None

    if address and len(address) < 3:
        raise HTTPException(400, "Address too short")

    # ==================================================
    # Duplicate Check
    # ==================================================
    conditions = []

    if gst_number:
        conditions.append(Supplier.gst_number == gst_number)

    if phone_email:
        conditions.append(Supplier.phone_email == phone_email)

    existing = None

    if conditions:
        existing = await db.scalar(
            select(Supplier).where(
                Supplier.is_deleted.is_(False),
                or_(*conditions),
            )
        )

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Supplier with same GST or phone/email already exists",
        )

    # Create supplier
    supplier = Supplier(
        supplier_name=supplier_name,
        contact_person=contact_person,
        phone_email=phone_email,
        gst_number=gst_number,
        address=address,
    )

    try:
        db.add(supplier)
        await db.commit()
        await db.refresh(supplier)

    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Duplicate supplier",
        )

    return SupplierOut.model_validate(supplier)


# ================update_supplier===============


@router.put("/suppliers/{supplier_id}", response_model=SupplierOut)
async def update_supplier(
    supplier_id: int,
    payload: SupplierCreate,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    import re

    supplier = await db.get(Supplier, supplier_id)

    if not supplier or supplier.is_deleted:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # ================= NORMALIZE =================
    new_name = payload.supplier_name.strip().title()

    new_contact_person = (
        payload.contact_person.strip().title() if payload.contact_person else None
    )

    new_phone_email = payload.phone_email.strip() if payload.phone_email else None

    new_gst = payload.gst_number.strip().upper() if payload.gst_number else None

    new_address = payload.address.strip() if payload.address else None

    # ================= VALIDATIONS =================

    if len(new_name) < 3:
        raise HTTPException(
            status_code=400,
            detail="Supplier name must be at least 3 characters",
        )

    if new_contact_person:
        if not re.fullmatch(r"[A-Za-z ]{3,}", new_contact_person):
            raise HTTPException(
                status_code=400,
                detail="Invalid contact person name",
            )

    # Phone / Email validation
    if new_phone_email:

        cleaned = re.sub(r"[^\d]", "", new_phone_email)

        if cleaned.startswith("91") and len(cleaned) == 12:
            cleaned = cleaned[2:]

        if cleaned.isdigit():

            if not re.fullmatch(r"[6-9]\d{9}", cleaned):
                raise HTTPException(
                    status_code=400,
                    detail="Invalid Indian mobile number",
                )

            new_phone_email = cleaned

        elif "@" in new_phone_email:

            if not re.fullmatch(
                r"[^@]+@[^@]+\.[^@]+",
                new_phone_email,
            ):
                raise HTTPException(
                    status_code=400,
                    detail="Invalid email format",
                )

        else:
            raise HTTPException(
                status_code=400,
                detail="Enter valid phone number or email",
            )

    # GST Validation
    if new_gst:
        if not re.fullmatch(
            r"\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d]",
            new_gst,
        ):
            raise HTTPException(
                status_code=400,
                detail="Invalid GST number format",
            )

    # ================= DUPLICATE CHECK =================

    duplicate_conditions = []

    if new_gst:
        duplicate_conditions.append(Supplier.gst_number == new_gst)

    if new_phone_email:
        duplicate_conditions.append(Supplier.phone_email == new_phone_email)

    if duplicate_conditions:

        existing = await db.scalar(
            select(Supplier).where(
                Supplier.id != supplier_id,
                Supplier.is_deleted == False,
                or_(*duplicate_conditions),
            )
        )

        if existing:
            raise HTTPException(
                status_code=400,
                detail="GST number or phone/email already exists",
            )

    # ================= UPDATE =================

    supplier.supplier_name = new_name
    supplier.contact_person = new_contact_person
    supplier.phone_email = new_phone_email
    supplier.gst_number = new_gst
    supplier.address = new_address

    try:
        await db.commit()
        await db.refresh(supplier)

    except IntegrityError:
        await db.rollback()

        raise HTTPException(
            status_code=400,
            detail="Duplicate supplier",
        )

    return SupplierOut.model_validate(supplier)


# =============delete_supplier============


@router.delete(
    "/suppliers/{id}",
    response_model=MessageResponse,
    summary="Delete Supplier",
    description="Soft deletes a supplier if not currently assigned to active materials.",
)
async def delete_supplier(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
):
    obj = await db.get(Supplier, id)

    if not obj or obj.is_deleted:
        raise HTTPException(status_code=404, detail="Supplier not found")

    in_use = await db.scalar(
        select(func.count()).where(
            Material.supplier_id == id, Material.is_deleted == False
        )
    )

    if in_use > 0:
        raise HTTPException(status_code=400, detail="Supplier is used in materials")

    obj.is_deleted = True
    await db.commit()

    return MessageResponse(
        success=True,
        message="Deleted successfully",
        resource_id=id,
    )


# ================= supplier materials =================


@router.get(
    "/suppliers/{supplier_id}/materials",
    response_model=list[MaterialOut],
)
async def get_supplier_materials(
    supplier_id: int,
    project_id: Optional[int] = Query(
        None,
        description="Filter by project ID",
    ),
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    limit = min(max(limit, 1), 100)
    skip = max(skip, 0)

    query = (
        select(
            Material,
            Supplier.supplier_name,
        )
        .options(
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .join(
            Supplier,
            Supplier.id == Material.supplier_id,
            isouter=True,
        )
        .where(
            Material.supplier_id == supplier_id,
            Material.is_deleted == False,
        )
    )

    # Optional Project Filter
    if project_id is not None:
        query = query.where(Material.project_id == project_id)

    query = query.order_by(Material.id.desc()).offset(skip).limit(limit)

    rows = (await db.execute(query)).all()

    return [
        build_material_response(
            m,
            supplier_name=supplier_name,
        )
        for m, supplier_name in rows
    ]


# ================= material_alerts =================


@router.get("/alerts", response_model=list[MaterialOut])
async def get_material_alerts(
    project_id: Optional[int] = Query(
        None,
        description="Filter by project ID",
    ),
    threshold: Optional[float] = Query(
        None,
        description="Custom low stock threshold",
    ),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    query = (
        select(
            Material,
            Supplier.supplier_name,
            Unit.name.label("unit_name"),
        )
        .options(
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .join(
            Supplier,
            Supplier.id == Material.supplier_id,
            isouter=True,
        )
        .join(
            Unit,
            Unit.id == Material.unit_id,
            isouter=True,
        )
        .where(Material.is_deleted == False)
    )

    # Project Filter
    if project_id is not None:
        query = query.where(Material.project_id == project_id)

    # Low Stock Filter
    if threshold is not None:
        query = query.where(Material.remaining_stock <= threshold)
    else:
        query = query.where(Material.remaining_stock <= Material.minimum_stock_level)

    query = query.order_by(Material.remaining_stock.asc())

    result = await db.execute(query)
    rows = result.all()

    data = []

    for material, supplier_name, unit_name in rows:

        response = build_material_response(
            material,
            supplier_name=supplier_name,
            unit_name=unit_name or "",
        )

        if (
            threshold is not None
            and response.alert_type == "IN_STOCK"
            and response.remaining_stock <= threshold
        ):
            response.alert_type = "NEAR_LOW"

        data.append(MaterialOut.model_validate(response))

    return data


# ================= PURCHASE ORDERS =================


@router.post("/purchase-orders", response_model=PurchaseOrderOut, status_code=201)
async def create_po(
    payload: PurchaseOrderCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
):

    if payload.quantity <= 0 or payload.rate <= 0:
        raise HTTPException(400, "Quantity and rate must be greater than 0")

    if current_user.role != UserRole.ADMIN.value and payload.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    material = await db.get(Material, payload.material_id)
    if not material:
        raise HTTPException(404, "Material not found")

    if material.supplier_id != payload.supplier_id:
        raise HTTPException(400, "Material does not belong to supplier")

    if material.project_id != payload.project_id:
        raise HTTPException(400, "Material does not belong to project")

    total_amount = payload.quantity * payload.rate

    po = PurchaseOrder(
        supplier_id=payload.supplier_id,
        boq_item_id=payload.boq_item_id,
        project_id=payload.project_id,
        material_id=payload.material_id,
        material_name=material.material_name,
        quantity=payload.quantity,
        rate=payload.rate,
        total_amount=total_amount,
        status="CREATED",
    )

    db.add(po)
    await db.flush()
    db.add(
        ActivityLog(
            action="CREATE_PO",
            entity="project",
            entity_id=payload.project_id,
            performed_by=current_user.id,
            details={
                "message": f"Purchase Order created for {payload.quantity} {material.material_name}"
            },
        )
    )
    await db.commit()
    await db.refresh(po)

    return build_po_response(po)


# ==============================================================


@router.get("/purchase-orders/{id}", response_model=PurchaseOrderOut)
async def get_po(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):

    po = await db.get(PurchaseOrder, id)

    if not po or po.is_deleted:
        raise HTTPException(404, "PO not found")

    if current_user.role != UserRole.ADMIN.value and po.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    return build_po_response(po)


# ==============================================================


@router.get("/purchase-orders", response_model=List[PurchaseOrderOut])
async def list_po(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    limit = min(max(limit, 1), 100)

    if (
        project_id is not None
        and current_user.role != UserRole.ADMIN.value
        and (project_id not in (current_user.allowed_projects or []))
    ):
        raise HTTPException(403, "Access denied")

    query = (
        select(PurchaseOrder)
        .where(PurchaseOrder.is_deleted == False)
        .order_by(PurchaseOrder.id.desc())
    )

    if project_id is not None:
        query = query.where(PurchaseOrder.project_id == project_id)
    elif current_user.role != UserRole.ADMIN.value:
        query = query.where(
            PurchaseOrder.project_id.in_(current_user.allowed_projects or [])
        )

    query = query.offset(skip).limit(limit)

    rows = (await db.execute(query)).scalars().all()

    return [build_po_response(r) for r in rows]


# ==============================================================


@router.put("/purchase-orders/{id}", response_model=PurchaseOrderOut)
async def update_po(
    id: int,
    payload: PurchaseOrderCreate,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.get(PurchaseOrder, id)

    if not obj or obj.is_deleted:
        raise HTTPException(404, "PO not found")

    if current_user.role != UserRole.ADMIN.value and obj.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    if obj.status in ["PENDING", "APPROVED"]:
        raise HTTPException(
            400, "Cannot edit a Purchase Order that is pending or approved."
        )

    if payload.quantity <= 0 or payload.rate <= 0:
        raise HTTPException(400, "Invalid quantity or rate")

    material = await db.get(Material, payload.material_id)
    if not material:
        raise HTTPException(404, "Material not found")

    if material.supplier_id != payload.supplier_id:
        raise HTTPException(400, "Material does not belong to supplier")

    if material.project_id != payload.project_id:
        raise HTTPException(400, "Material does not belong to project")

    obj.supplier_id = payload.supplier_id
    obj.project_id = payload.project_id
    obj.material_id = payload.material_id
    obj.material_name = material.material_name
    obj.quantity = payload.quantity
    obj.rate = payload.rate
    obj.total_amount = payload.quantity * payload.rate

    await db.commit()
    await db.refresh(obj)

    return build_po_response(obj)


# ==============================================================


@router.delete(
    "/purchase-orders/{id}",
    response_model=MessageResponse,
    summary="Delete Purchase Order",
    description="Soft deletes a purchase order if not pending or approved.",
)
async def delete_po(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
):

    obj = await db.get(PurchaseOrder, id)

    if not obj or obj.is_deleted:
        raise HTTPException(status_code=404, detail="PO not found")

    if current_user.role != UserRole.ADMIN.value and obj.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(status_code=403, detail="Access denied")

    if obj.status in ["PENDING", "APPROVED"]:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete a Purchase Order that is pending or approved.",
        )

    obj.is_deleted = True

    await db.commit()

    return MessageResponse(
        success=True,
        message="Purchase order deleted successfully",
        resource_id=id,
    )


# =========================project_transactions=============================


@router.get(
    "/projects/{project_id}/transactions",
    response_model=List[ProjectTransactionOut],
)
async def project_transactions(
    project_id: int,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)
    if current_user.role != UserRole.ADMIN.value and project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    query = (
        select(MaterialTransaction, Material.material_name, Supplier.supplier_name)
        .join(Material, Material.id == MaterialTransaction.material_id)
        .join(Supplier, Supplier.id == Material.supplier_id, isouter=True)
        .where(MaterialTransaction.project_id == project_id)
        .order_by(MaterialTransaction.created_at.desc())
        .offset(offset)
        .limit(limit)
    )

    rows = (await db.execute(query)).all()

    return [
        {
            "id": tx.id,
            "type": tx.type.value,
            "material_id": tx.material_id,
            "material_name": material_name,
            "supplier_name": supplier_name or "N/A",
            "quantity": get_signed_quantity(tx),
            "total_amount": round(float(tx.total_amount or 0), 2),
            "project_id": tx.project_id,
            "created_at": tx.created_at,
        }
        for tx, material_name, supplier_name in rows
    ]


# ================= material_transactions =================


@router.get("/{material_id}/transactions", response_model=List[MaterialLogOut])
async def get_material_transactions(
    material_id: int,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    limit = min(max(limit, 1), 100)
    offset = max(offset, 0)

    # validation
    material = await db.get(Material, material_id)
    if not material or material.is_deleted:
        raise HTTPException(404, "Material not found")

    if current_user.role != UserRole.ADMIN.value and material.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    result = await db.execute(
        select(MaterialTransaction)
        .where(MaterialTransaction.material_id == material_id)
        .order_by(MaterialTransaction.created_at.desc())
        .offset(offset)
        .limit(limit)
    )

    rows = result.scalars().all()

    data = []

    for r in rows:
        quantity = float(r.quantity or 0)
        total_amount = float(r.total_amount or 0)

        avg_rate = abs(total_amount / quantity) if quantity != 0 else 0

        data.append(
            MaterialLogOut(
                id=r.id,
                boq_item_id=r.boq_item_id,
                material_id=r.material_id,
                type=r.type,
                quantity=round(quantity, 3),
                rate=round(float(r.rate or 0), 2),
                avg_rate=round(avg_rate, 2),
                total_amount=round(total_amount, 2),
                amount_paid=float(r.amount_paid or 0),
                payment_pending=float(r.payment_pending or 0),
                issue_type=r.issue_type,
                project_id=r.project_id,
                created_at=r.created_at,
            )
        )

    return data


# ================= TRANSFERS =================


@router.post("/transfers", response_model=TransferOut)
async def create_transfer(
    payload: TransferCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    redis=Depends(get_request_redis),
):
    try:

        if payload.quantity <= 0:
            raise HTTPException(
                status_code=400,
                detail="Quantity must be greater than zero",
            )

        if payload.from_project_id == payload.to_project_id:
            raise HTTPException(
                status_code=400,
                detail="Source and destination project cannot be same",
            )

        if current_user.role != UserRole.ADMIN.value and (
            payload.from_project_id not in (current_user.allowed_projects or [])
            or payload.to_project_id not in (current_user.allowed_projects or [])
        ):
            raise HTTPException(
                status_code=403,
                detail="Access denied",
            )

        from_project = await db.get(
            Project,
            payload.from_project_id,
        )

        to_project = await db.get(
            Project,
            payload.to_project_id,
        )

        if not from_project:
            raise HTTPException(
                status_code=404,
                detail="Source project not found",
            )

        if not to_project:
            raise HTTPException(
                status_code=404,
                detail="Destination project not found",
            )

        reference_id = f"TRF-{uuid.uuid4().hex[:8].upper()}"

        # =====================================================
        # SOURCE MATERIAL
        # =====================================================
        # NOTE: to reduce deadlock risk when two transfers cross-reference
        # projects (A->B and B->A concurrently), always lock the lower
        # material id first would require knowing destination id up front,
        # which isn't possible before this point (it may not exist yet).
        # Kept as-is; flagged as a known limitation.

        material = await db.scalar(
            select(Material)
            .options(
                selectinload(Material.unit),
                selectinload(Material.supplier),
                selectinload(Material.material_master),
            )
            .where(
                Material.id == payload.material_id,
                Material.is_deleted == False,
            )
            .with_for_update()
        )

        if not material:
            raise HTTPException(
                status_code=404,
                detail="Material not found",
            )

        if material.project_id != payload.from_project_id:
            raise HTTPException(
                status_code=400,
                detail="Material does not belong to source project",
            )

        if material.remaining_stock < payload.quantity:
            raise HTTPException(
                status_code=400,
                detail="Insufficient stock available",
            )

        qty_purchased = material.quantity_purchased or Decimal("0")
        total_amount = material.total_amount or Decimal("0")

        avg_rate = total_amount / qty_purchased if qty_purchased > 0 else Decimal("0")

        transfer_amount = avg_rate * payload.quantity

        # =====================================================
        # REDUCE STOCK FROM SOURCE
        # =====================================================

        material.quantity_used += payload.quantity

        update_material_fields(material)

        # =====================================================
        # FIND DESTINATION MATERIAL
        # =====================================================

        destination_material = await db.scalar(
            select(Material)
            .where(
                Material.project_id == payload.to_project_id,
                Material.material_master_id == material.material_master_id,
                Material.supplier_id == material.supplier_id,
                Material.is_deleted == False,
            )
            .with_for_update()
        )

        # =====================================================
        # CREATE DESTINATION MATERIAL IF NOT EXISTS
        # =====================================================

        if not destination_material:

            material_code = await generate_business_id(
                db=db,
                model=Material,
                column_name="material_code",
                prefix="MAT",
            )

            destination_material = Material(
                material_code=material_code,
                project_id=payload.to_project_id,
                material_master_id=material.material_master_id,
                material_name=material.material_name,
                category=material.category,
                unit_id=material.unit_id,
                supplier_id=material.supplier_id,
                purchase_rate=avg_rate,
                rate_type=material.rate_type,
                quantity_purchased=payload.quantity,
                quantity_used=Decimal("0"),
                total_amount=transfer_amount,
                payment_given=Decimal("0"),
                payment_pending=Decimal("0"),
                minimum_stock_level=material.minimum_stock_level,
            )

            update_material_fields(destination_material)

            db.add(destination_material)

            await db.flush()

        else:

            destination_material.quantity_purchased += payload.quantity

            destination_material.total_amount += transfer_amount

            update_material_fields(destination_material)

        # =====================================================
        # MATERIAL TRANSACTION
        # =====================================================

        tx_out = MaterialTransaction(
            material_id=material.id,
            project_id=payload.from_project_id,
            type=TransactionType.TRANSFER_OUT,
            quantity=-payload.quantity,
            rate=avg_rate,
            total_amount=transfer_amount,
            issue_type=IssueType.TRANSFER,
            reference_id=reference_id,
        )

        tx_in = MaterialTransaction(
            material_id=destination_material.id,
            project_id=payload.to_project_id,
            type=TransactionType.TRANSFER_IN,
            quantity=payload.quantity,
            rate=avg_rate,
            total_amount=transfer_amount,
            issue_type=IssueType.TRANSFER,
            reference_id=reference_id,
        )

        db.add(tx_out)
        db.add(tx_in)

        # =====================================================
        # MATERIAL LEDGER
        # =====================================================

        ledger_out = MaterialLedger(
            material_id=material.id,
            project_id=payload.from_project_id,
            type=TransactionType.TRANSFER_OUT,
            quantity=-payload.quantity,
            rate=avg_rate,
            total_amount=transfer_amount,
            issue_type=IssueType.TRANSFER,
            reference_id=reference_id,
        )

        ledger_in = MaterialLedger(
            material_id=destination_material.id,
            project_id=payload.to_project_id,
            type=TransactionType.TRANSFER_IN,
            quantity=payload.quantity,
            rate=avg_rate,
            total_amount=transfer_amount,
            issue_type=IssueType.TRANSFER,
            reference_id=reference_id,
        )

        db.add(ledger_out)
        db.add(ledger_in)

        # =====================================================
        # TRANSFER RECORD
        # =====================================================

        transfer = MaterialTransfer(
            material_id=material.id,
            from_project_id=payload.from_project_id,
            to_project_id=payload.to_project_id,
            quantity=payload.quantity,
            status="COMPLETED",
            reference_id=reference_id,
        )

        db.add(transfer)

        await db.commit()

        await db.refresh(transfer)

        await bump_cache_version(
            redis,
            VERSION_KEY,
        )

        return build_transfer_response(
            transfer,
            material,
            from_project,
            to_project,
        )

    except HTTPException:
        await db.rollback()
        raise

    except Exception:
        await db.rollback()
        raise


# ================= LIST TRANSFERS =================


@router.get("/transfers", response_model=TransferListResponse)
async def list_transfers(
    project_id: Optional[int] = Query(None, description="Filter by project ID"),
    skip: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    skip = max(skip, 0)
    limit = min(max(limit, 1), 100)

    if (
        project_id is not None
        and current_user.role != UserRole.ADMIN.value
        and (project_id not in (current_user.allowed_projects or []))
    ):
        raise HTTPException(403, "Access denied")

    FromProject = aliased(Project)
    ToProject = aliased(Project)

    query = (
        select(
            MaterialTransfer,
            Material.material_name,
            FromProject.project_name,
            ToProject.project_name,
        )
        .select_from(MaterialTransfer)
        .join(Material)
        .join(FromProject, FromProject.id == MaterialTransfer.from_project_id)
        .join(ToProject, ToProject.id == MaterialTransfer.to_project_id)
    )

    if project_id is not None:
        query = query.where(
            or_(
                MaterialTransfer.from_project_id == project_id,
                MaterialTransfer.to_project_id == project_id,
            )
        )

    total = await db.scalar(
        select(func.count())
        .select_from(MaterialTransfer)
        .where(
            or_(
                MaterialTransfer.from_project_id == project_id,
                MaterialTransfer.to_project_id == project_id,
            )
        )
        if project_id is not None
        else select(func.count()).select_from(MaterialTransfer)
    )

    rows = (
        await db.execute(
            query.order_by(MaterialTransfer.id.desc()).offset(skip).limit(limit)
        )
    ).all()

    data = [
        {
            "id": t.id,
            "material": {"id": t.material_id, "name": material_name},
            "from_project": {"id": t.from_project_id, "name": from_name},
            "to_project": {"id": t.to_project_id, "name": to_name},
            "quantity": float(t.quantity),
            "status": t.status,
        }
        for t, material_name, from_name, to_name in rows
    ]

    return {
        "total": total or 0,
        "skip": skip,
        "limit": limit,
        "data": data,
    }


# ================= GET SINGLE TRANSFER =================


@router.get("/transfers/{id}", response_model=TransferOut)
async def get_transfer(
    id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):

    FromProject = aliased(Project)
    ToProject = aliased(Project)

    result = await db.execute(
        select(
            MaterialTransfer,
            Material,
            FromProject,
            ToProject,
        )
        .join(Material, Material.id == MaterialTransfer.material_id)
        .join(FromProject, FromProject.id == MaterialTransfer.from_project_id)
        .join(ToProject, ToProject.id == MaterialTransfer.to_project_id)
        .where(MaterialTransfer.id == id)
    )

    row = result.first()

    if not row:
        raise HTTPException(404, "Transfer not found")

    obj, material, from_project, to_project = row

    if current_user.role != UserRole.ADMIN.value and (
        obj.from_project_id not in (current_user.allowed_projects or [])
        and obj.to_project_id not in (current_user.allowed_projects or [])
    ):
        raise HTTPException(403, "Access denied")

    return build_transfer_response(obj, material, from_project, to_project)


# =================update_transfer_status=========


VALID_STATUS = {"PENDING", "COMPLETED", "CANCELLED"}


@router.put("/transfers/{id}", response_model=TransferOut)
async def update_transfer_status(
    id: int,
    status: str,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    status = status.upper().strip()

    if status not in VALID_STATUS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid status. Allowed: {', '.join(VALID_STATUS)}",
        )

    obj = await db.get(MaterialTransfer, id)

    if not obj:
        raise HTTPException(404, "Transfer not found")

    # NOTE: create_transfer() applies the stock/ledger movement immediately
    # and marks the transfer COMPLETED. This endpoint only flips the status
    # label — it does NOT reverse quantity_used/quantity_purchased or the
    # ledger entries. Changing status away from COMPLETED here will leave
    # inventory numbers out of sync with the displayed status. Block that
    # until a proper reversal flow exists.
    if obj.status == "COMPLETED" and status != "COMPLETED":
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot change status of a completed transfer: stock has "
                "already moved. Create a reverse transfer instead."
            ),
        )

    obj.status = status

    await db.commit()
    await db.refresh(obj)

    material = await db.get(Material, obj.material_id)
    from_project = await db.get(Project, obj.from_project_id)
    to_project = await db.get(Project, obj.to_project_id)

    return TransferOut(
        id=obj.id,
        material=(
            TransferMaterial(id=material.id, name=material.material_name)
            if material
            else None
        ),
        from_project=(
            TransferProject(id=from_project.id, name=from_project.project_name)
            if from_project
            else None
        ),
        to_project=(
            TransferProject(id=to_project.id, name=to_project.project_name)
            if to_project
            else None
        ),
        quantity=obj.quantity,
        status=obj.status,
        created_at=obj.created_at,
    )


# ================= USAGE =================


async def _build_material_out_from_id(db, material_id: int):
    obj = await db.scalar(
        select(Material)
        .options(
            selectinload(Material.supplier),
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .where(Material.id == material_id)
    )
    if not obj:
        raise HTTPException(404, "Material not found")

    supplier = obj.supplier
    unit = obj.unit
    master = obj.material_master

    total_amount = float(obj.total_amount or 0)
    payment_given = float(obj.payment_given or 0)

    payment_pending = max(
        0,
        total_amount - payment_given,
    )

    extra_paid = max(
        0,
        payment_given - total_amount,
    )

    if (obj.remaining_stock or 0) <= 0:
        alert_type = "OUT_OF_STOCK"
    elif (obj.remaining_stock or 0) <= (obj.minimum_stock_level or 0):
        alert_type = "LOW_STOCK"
    else:
        alert_type = "IN_STOCK"

    return MaterialOut(
        id=obj.id,
        material_code=obj.material_code,
        project_id=obj.project_id,
        material_master_id=obj.material_master_id,
        material_master_name=master.name if master else None,
        material_master_brand=master.brand if master else None,
        material_master_specification=master.specification if master else None,
        material_master_hsn_code=master.hsn_code if master else None,
        material_name=obj.material_name.strip().title() if obj.material_name else "",
        category=obj.category,
        unit_id=obj.unit_id,
        unit_name=unit.name if unit else "",
        supplier_id=obj.supplier_id,
        supplier_name=supplier.supplier_name if supplier else None,
        purchase_rate=float(obj.purchase_rate or 0),
        rate_type=obj.rate_type,
        quantity_purchased=float(obj.quantity_purchased or 0),
        quantity_used=float(obj.quantity_used or 0),
        remaining_stock=float(obj.remaining_stock or 0),
        total_amount=round(total_amount, 2),
        payment_given=round(payment_given, 2),
        payment_pending=round(payment_pending, 2),
        extra_paid=round(extra_paid, 2),
        minimum_stock_level=float(obj.minimum_stock_level or 0),
        alert_type=alert_type,
    )

@router.post("/{material_id}/usage", response_model=MaterialOut)
async def usage(
    material_id: int,
    data: UsageMaterial,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
):

    import json
    import hashlib

    request_hash = None
    if idempotency_key:
        payload_dict = data.model_dump(mode="json")
        payload_dict["material_id"] = material_id
        payload_str = json.dumps(payload_dict, sort_keys=True)
        request_hash = hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

        existing = await db.scalar(
            select(MaterialTransaction)
            .where(MaterialTransaction.idempotency_key == idempotency_key)
        )
        if existing:
            if existing.request_hash != request_hash:
                raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
            return await _build_material_out_from_id(db, existing.material_id)


    obj = await db.scalar(
        select(Material)
        .options(
            selectinload(Material.supplier),
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .where(
            Material.id == material_id,
            Material.is_deleted == False,
        )
        .with_for_update()
    )

    if not obj:
        raise HTTPException(404, "Material not found")

    qty = Decimal(str(data.quantity))

    if qty <= 0:
        raise HTTPException(400, "Quantity must be greater than 0")

    purchased = obj.quantity_purchased or Decimal("0")
    used = obj.quantity_used or Decimal("0")

    current_stock = purchased - used

    if current_stock <= 0:
        raise HTTPException(400, "Stock exhausted")

    if qty > current_stock:
        raise HTTPException(
            400,
            f"Insufficient stock. Available: {current_stock}",
        )

    total_amount_current = obj.total_amount or Decimal("0")

    avg_rate = total_amount_current / purchased if purchased > 0 else Decimal("0")

    used_value = qty * avg_rate

    # =====================================
    # BOQ VALIDATION (OPTIONAL)
    # =====================================

    boq = None

    if data.boq_item_id:

        boq = await db.scalar(
            select(BOQ)
            .where(BOQ.id == data.boq_item_id)
            .with_for_update()
        )

        if not boq:
            raise HTTPException(
                status_code=404,
                detail="BOQ item not found",
            )

        if boq.project_id != data.project_id:
            raise HTTPException(
                status_code=400,
                detail="BOQ does not belong to project",
            )

        if str(boq.status).upper() == "DELETED":
            raise HTTPException(
                status_code=400,
                detail="BOQ item is deleted",
            )

    reference = f"USE-{uuid.uuid4().hex[:8]}"
    issue_type = data.issue_type or IssueType.SYSTEM

    try:

        transaction = MaterialTransaction(
            material_id=obj.id,
                idempotency_key=idempotency_key,
                request_hash=request_hash,
            boq_item_id=data.boq_item_id,
            type=DBTransactionType.USAGE,
            quantity=-qty,
            rate=avg_rate,
            total_amount=used_value,
            amount_paid=Decimal("0"),
            payment_pending=Decimal("0"),
            issue_type=issue_type,
            project_id=data.project_id,
            task_id=data.task_id,
            remarks="Material used",
            reference_id=reference,
        )

        ledger = MaterialLedger(
            material_id=obj.id,
            boq_item_id=data.boq_item_id,
            type=DBTransactionType.USAGE,
            quantity=-qty,
            rate=avg_rate,
            total_amount=used_value,
            amount_paid=Decimal("0"),
            payment_pending=Decimal("0"),
            issue_type=issue_type,
            project_id=data.project_id,
            remarks="Material used",
            reference_id=reference,
        )

        usage_entry = MaterialUsage(
            material_id=obj.id,
            boq_item_id=data.boq_item_id,
            project_id=data.project_id,
            task_id=data.task_id,
            quantity_used=qty,
            usage_date=datetime.utcnow(),
        )

        db.add(transaction)
        db.add(ledger)
        db.add(usage_entry)

        # =====================================
        # UPDATE MATERIAL STOCK
        # =====================================

        obj.quantity_used = used + qty

        # =====================================
        # UPDATE BOQ ACTUALS
        # =====================================

        if boq:

            # Replace inline mutations with centralized recalculation
            # Note: We don't need to lock the BOQ again since it's already locked above.
            # But recalculate_boq_actuals takes `lock=True` by default, which is fine (reentrant).
            await db.flush()
            await recalculate_boq_actuals(db, boq.id)

        # =====================================
        # RECALCULATE MATERIAL
        # =====================================

        update_material_fields(obj)

        await db.commit()

        # Re-fetch with relationships eagerly loaded instead of db.refresh(),
        # which can leave relationship attributes expired and trigger a
        # MissingGreenlet lazy-load error under AsyncSession.
        obj = await db.scalar(
            select(Material)
            .options(
                selectinload(Material.supplier),
                selectinload(Material.unit),
                selectinload(Material.material_master),
            )
            .where(Material.id == material_id)
        )


    except IntegrityError:
        await db.rollback()
        if idempotency_key:
            existing = await db.scalar(
                select(MaterialTransaction)
                .where(MaterialTransaction.idempotency_key == idempotency_key)
            )
            if existing:
                if existing.request_hash != request_hash:
                    raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
                return await _build_material_out_from_id(db, existing.material_id)
        raise
    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(redis, VERSION_KEY)

    supplier = obj.supplier
    master = obj.material_master

    total_amount = float(obj.total_amount or 0)
    payment_given = float(obj.payment_given or 0)

    payment_pending = max(
        0,
        total_amount - payment_given,
    )

    extra_paid = max(
        0,
        payment_given - total_amount,
    )

    if obj.remaining_stock <= 0:
        alert_type = "OUT_OF_STOCK"
    elif obj.remaining_stock <= obj.minimum_stock_level:
        alert_type = "LOW_STOCK"
    else:
        alert_type = "IN_STOCK"

    if alert_type in ["LOW_STOCK", "OUT_OF_STOCK"]:

        pm = await db.scalar(
            select(proj_model.ProjectMember.user_id)
            .join(
                User,
                User.id == proj_model.ProjectMember.user_id,
            )
            .where(
                proj_model.ProjectMember.project_id == data.project_id,
                User.role == UserRole.PROJECT_MANAGER.value,
            )
            .limit(1)
        )

        if pm:
            await create_notification(
                db=db,
                user_id=pm,
                title=f"Material Alert: {alert_type.replace('_', ' ')}",
                message=(
                    f"Stock for {obj.material_name} "
                    f"is now {obj.remaining_stock} "
                    f"{obj.unit.name if obj.unit else ''}"
                ),
                type="alert",
            )

            await db.commit()

    return MaterialOut(
        id=obj.id,
        material_code=obj.material_code,
        project_id=obj.project_id,
        material_master_id=obj.material_master_id,
        material_master_name=master.name if master else None,
        material_master_brand=master.brand if master else None,
        material_master_specification=master.specification if master else None,
        material_master_hsn_code=master.hsn_code if master else None,
        material_name=obj.material_name,
        category=obj.category,
        unit_id=obj.unit_id,
        unit_name=obj.unit.name if obj.unit else "",
        supplier_id=obj.supplier_id,
        supplier_name=supplier.supplier_name if supplier else None,
        purchase_rate=float(obj.purchase_rate),
        rate_type=obj.rate_type,
        quantity_purchased=float(obj.quantity_purchased),
        quantity_used=float(obj.quantity_used),
        remaining_stock=float(obj.remaining_stock),
        total_amount=round(total_amount, 2),
        payment_given=round(payment_given, 2),
        payment_pending=round(payment_pending, 2),
        extra_paid=round(extra_paid, 2),
        minimum_stock_level=float(obj.minimum_stock_level),
        alert_type=alert_type,
    )


# ================= PURCHASE =================


@router.post("/{material_id}/purchase", response_model=MaterialOut)
async def purchase(
    material_id: int,
    data: PurchaseMaterial,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
):
    import uuid

    import json
    import hashlib

    request_hash = None
    if idempotency_key:
        payload_dict = data.model_dump(mode="json")
        payload_dict["material_id"] = material_id
        payload_str = json.dumps(payload_dict, sort_keys=True)
        request_hash = hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

        existing = await db.scalar(
            select(MaterialTransaction)
            .where(MaterialTransaction.idempotency_key == idempotency_key)
        )
        if existing:
            if existing.request_hash != request_hash:
                raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
            return await _build_material_out_from_id(db, existing.material_id)

    from decimal import Decimal, ROUND_HALF_UP

    try:

        obj = await db.scalar(
            select(Material)
            .options(
                selectinload(Material.supplier),
                selectinload(Material.unit),
                selectinload(Material.material_master),
            )
            .where(
                Material.id == material_id,
                Material.is_deleted == False,
            )
            .with_for_update()
        )

        if not obj:
            raise HTTPException(
                status_code=404,
                detail="Material not found",
            )

        qty = Decimal(str(data.quantity))
        rate = Decimal(str(data.rate))
        paid = Decimal(str(data.amount_paid or 0))

        if qty <= 0:
            raise HTTPException(400, "Quantity must be > 0")

        if rate <= 0:
            raise HTTPException(400, "Rate must be > 0")

        if paid < 0:
            raise HTTPException(400, "Payment cannot be negative")

        old_qty = obj.quantity_purchased or Decimal("0")
        old_rate = obj.purchase_rate or Decimal("0")

        if old_qty > 0:
            new_rate = ((old_qty * old_rate) + (qty * rate)) / (old_qty + qty)
        else:
            new_rate = rate

        obj.purchase_rate = new_rate.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        total = (qty * rate).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        payment_pending = max(
            Decimal("0"),
            total - paid,
        )

        # =====================================
        # BOQ VALIDATION (OPTIONAL)
        # =====================================

        if data.boq_item_id:

            boq = await db.get(
                BOQ,
                data.boq_item_id,
            )

            if not boq:
                raise HTTPException(
                    status_code=404,
                    detail="BOQ item not found",
                )

            if boq.project_id != data.project_id:
                raise HTTPException(
                    status_code=400,
                    detail="BOQ does not belong to project",
                )

            if str(boq.status).upper() == "DELETED":
                raise HTTPException(
                    status_code=400,
                    detail="BOQ item is deleted",
                )

            # Removed BOQ actual_cost mutation per Phase 1C rules

        reference = f"PUR-{uuid.uuid4().hex[:8]}"

        # =====================================
        # MATERIAL TRANSACTION
        # =====================================

        db.add(
            MaterialTransaction(
                material_id=obj.id,
                idempotency_key=idempotency_key,
                request_hash=request_hash,
                boq_item_id=data.boq_item_id,
                type=DBTransactionType.PURCHASE,
                quantity=qty,
                rate=rate,
                total_amount=total,
                amount_paid=paid,
                payment_pending=payment_pending,
                issue_type=IssueType.PURCHASE,
                project_id=data.project_id,
                remarks="Material purchased",
                reference_id=reference,
            )
        )

        # =====================================
        # MATERIAL LEDGER
        # =====================================

        db.add(
            MaterialLedger(
                material_id=obj.id,
                boq_item_id=data.boq_item_id,
                type=DBTransactionType.PURCHASE,
                quantity=qty,
                rate=rate,
                total_amount=total,
                amount_paid=paid,
                payment_pending=payment_pending,
                issue_type=IssueType.PURCHASE,
                project_id=data.project_id,
                remarks="Material purchased",
                reference_id=reference,
            )
        )

        # =====================================
        # UPDATE MATERIAL
        # =====================================

        obj.quantity_purchased = (obj.quantity_purchased or Decimal("0")) + qty

        obj.remaining_stock = (obj.remaining_stock or Decimal("0")) + qty

        obj.payment_given = (obj.payment_given or Decimal("0")) + paid

        obj.total_amount = (obj.total_amount or Decimal("0")) + total

        obj.payment_pending = max(
            Decimal("0"),
            obj.total_amount - obj.payment_given,
        )

        update_material_fields(obj)

        db.add(
            ActivityLog(
                action="RECEIVE_MATERIAL",
                entity="project",
                entity_id=data.project_id,
                performed_by=current_user.id,
                details={"message": f"Received {qty} {obj.material_name}"},
            )
        )

        await db.commit()

        result = await db.execute(
            select(Material)
            .options(
                selectinload(Material.supplier),
                selectinload(Material.unit),
                selectinload(Material.material_master),
            )
            .where(Material.id == obj.id)
        )

        obj = result.scalar_one()


    except IntegrityError:
        await db.rollback()
        if idempotency_key:
            existing = await db.scalar(
                select(MaterialTransaction)
                .where(MaterialTransaction.idempotency_key == idempotency_key)
            )
            if existing:
                if existing.request_hash != request_hash:
                    raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
                return await _build_material_out_from_id(db, existing.material_id)
        raise
    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    supplier = obj.supplier
    unit = obj.unit
    master = obj.material_master

    total_amount = float(obj.total_amount or 0)
    payment_given = float(obj.payment_given or 0)

    payment_pending = max(
        0,
        total_amount - payment_given,
    )

    extra_paid = max(
        0,
        payment_given - total_amount,
    )

    if obj.remaining_stock == 0:
        alert_type = "OUT_OF_STOCK"
    elif obj.remaining_stock <= obj.minimum_stock_level:
        alert_type = "LOW_STOCK"
    else:
        alert_type = "IN_STOCK"

    return MaterialOut(
        id=obj.id,
        material_code=obj.material_code,
        project_id=obj.project_id,
        material_master_id=obj.material_master_id,
        material_master_name=master.name if master else None,
        material_master_brand=master.brand if master else None,
        material_master_specification=master.specification if master else None,
        material_master_hsn_code=master.hsn_code if master else None,
        material_name=obj.material_name.strip().title(),
        category=obj.category,
        unit_id=obj.unit_id,
        unit_name=unit.name if unit else "",
        supplier_id=obj.supplier_id,
        supplier_name=supplier.supplier_name if supplier else None,
        purchase_rate=float(obj.purchase_rate or 0),
        rate_type=obj.rate_type,
        quantity_purchased=float(obj.quantity_purchased or 0),
        quantity_used=float(obj.quantity_used or 0),
        remaining_stock=float(obj.remaining_stock or 0),
        total_amount=round(total_amount, 2),
        payment_given=round(payment_given, 2),
        payment_pending=round(payment_pending, 2),
        extra_paid=round(extra_paid, 2),
        minimum_stock_level=float(obj.minimum_stock_level or 0),
        alert_type=alert_type,
    )


# ================= ADD INVENTORY =================


@router.post(
    "/inventory",
    response_model=InventoryAdjustResponse,
)
async def adjust_inventory(
    payload: InventoryAdjustRequest,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    redis=Depends(get_request_redis),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
):

    import json
    import hashlib

    request_hash = None
    if idempotency_key:
        payload_dict = payload.model_dump(mode="json")
        payload_dict["material_id"] = payload.material_id
        payload_str = json.dumps(payload_dict, sort_keys=True)
        request_hash = hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

        existing = await db.scalar(
            select(MaterialTransaction)
            .where(MaterialTransaction.idempotency_key == idempotency_key)
        )
        if existing:
            if existing.request_hash != request_hash:
                raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
            adj_mat = await db.scalar(select(Material).where(Material.id == existing.material_id))
            return InventoryAdjustResponse(
                material_id=adj_mat.id,
                material_name=adj_mat.material_name,
                old_stock=float(adj_mat.remaining_stock - existing.quantity),
                new_stock=float(adj_mat.remaining_stock),
                difference=float(existing.quantity),
                avg_rate=float(existing.rate),
                reason=existing.remarks.split('|')[-1].strip() if existing.remarks else "Idempotent response",
                reference_id=existing.reference_id,
                message="Inventory adjusted successfully"
            )


    material_id = payload.material_id

    reason = " ".join((payload.reason or "").strip().split())

    try:
        new_stock = Decimal(str(payload.new_stock)).quantize(Decimal("0.001"))

    except (InvalidOperation, ValueError):
        raise HTTPException(
            status_code=400,
            detail="Invalid stock value",
        )

    # ================= VALIDATIONS =================

    if new_stock < 0:
        raise HTTPException(
            status_code=400,
            detail="Stock cannot be negative",
        )

    if new_stock > Decimal("999999999"):
        raise HTTPException(
            status_code=400,
            detail="Stock value too large",
        )

    if not reason:
        raise HTTPException(
            status_code=400,
            detail="Reason is required",
        )

    if len(reason) > 500:
        raise HTTPException(
            status_code=400,
            detail="Reason too long",
        )

    try:
        material = await db.scalar(
            select(Material)
            .where(
                Material.id == material_id,
                Material.is_deleted == False,
            )
            .with_for_update()
        )

        if not material:
            raise HTTPException(
                status_code=404,
                detail="Material not found",
            )

        old_stock = material.remaining_stock or Decimal("0")

        diff = new_stock - old_stock

        if diff == 0:
            raise HTTPException(
                status_code=400,
                detail="No stock change detected",
            )

        reference = f"ADJ-{uuid.uuid4().hex[:8]}"

        qty_purchased = material.quantity_purchased or Decimal("0")

        total_amt = material.total_amount or Decimal("0")

        avg_rate = (
            total_amt / qty_purchased
            if qty_purchased > 0
            else (material.purchase_rate or Decimal("0"))
        )

        # ================= STOCK RECONCILIATION =================

        if diff > 0:
            # Physical stock found extra

            material.quantity_purchased += diff

            # Increase inventory valuation
            material.total_amount += diff * avg_rate

        else:
            # Physical stock less than system stock

            material.quantity_used += abs(diff)

            # DO NOT reduce total_amount
            # total_amount = historical purchase cost

        # ================= SET ACTUAL STOCK =================

        material.remaining_stock = new_stock

        # ================= PAYMENT FIELDS =================

        material.payment_pending = max(
            (material.total_amount or Decimal("0"))
            - (material.payment_given or Decimal("0")),
            Decimal("0"),
        )

        material.advance_amount = max(
            (material.payment_given or Decimal("0"))
            - (material.total_amount or Decimal("0")),
            Decimal("0"),
        )

        audit_remark = f"Stock adjusted: {old_stock} -> {new_stock} | {reason}"

        adjustment_total = abs(diff) * avg_rate

        # ================= TRANSACTION =================

        db.add(
            MaterialTransaction(
                material_id=material.id,
                idempotency_key=idempotency_key,
                request_hash=request_hash,
                type=DBTransactionType.ADJUSTMENT,
                quantity=diff,
                rate=avg_rate,
                total_amount=adjustment_total,
                amount_paid=0,
                payment_pending=0,
                issue_type=IssueType.SYSTEM,
                project_id=material.project_id,
                remarks=audit_remark,
                reference_id=reference,
            )
        )

        # ================= LEDGER =================

        db.add(
            MaterialLedger(
                material_id=material.id,
                type=DBTransactionType.ADJUSTMENT,
                quantity=diff,
                rate=avg_rate,
                total_amount=adjustment_total,
                amount_paid=0,
                payment_pending=0,
                project_id=material.project_id,
                remarks=audit_remark,
                reference_id=reference,
            )
        )

        await db.commit()
        await db.refresh(material)

    except HTTPException:
        await db.rollback()
        raise


    except IntegrityError:
        await db.rollback()
        if idempotency_key:
            existing = await db.scalar(
                select(MaterialTransaction)
                .where(MaterialTransaction.idempotency_key == idempotency_key)
            )
            if existing:
                if existing.request_hash != request_hash:
                    raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
                adj_mat = await db.scalar(select(Material).where(Material.id == existing.material_id))
            return InventoryAdjustResponse(
                material_id=adj_mat.id,
                material_name=adj_mat.material_name,
                old_stock=float(adj_mat.remaining_stock - existing.quantity),
                new_stock=float(adj_mat.remaining_stock),
                difference=float(existing.quantity),
                avg_rate=float(existing.rate),
                reason=existing.remarks.split('|')[-1].strip() if existing.remarks else "Idempotent response",
                reference_id=existing.reference_id,
                message="Inventory adjusted successfully"
            )
        raise
    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return InventoryAdjustResponse(
        material_id=material.id,
                idempotency_key=idempotency_key,
                request_hash=request_hash,
        material_name=material.material_name,
        old_stock=float(old_stock),
        new_stock=float(material.remaining_stock),
        difference=float(diff),
        avg_rate=float(avg_rate),
        reason=reason,
        reference_id=reference,
        message="Inventory adjusted successfully",
    )


# ===============get_all_inventory===========================


@router.get("/inventory")
async def get_all_inventory(
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    # NOTE: `Material.unit` is a relationship, not a column - selecting it
    # directly used to be a bug. Fixed by selecting `Material.unit_id` and
    # joining Unit for the human-readable name.
    query = (
        select(
            Material.id,
            Material.material_name,
            Material.remaining_stock,
            Material.unit_id,
            Unit.name.label("unit_name"),
            Material.project_id,
            Material.total_amount,
            Material.quantity_purchased,
        )
        .outerjoin(Unit, Unit.id == Material.unit_id)
        .where(Material.is_deleted == False)
    )

    if current_user.role != UserRole.ADMIN.value:
        query = query.where(
            Material.project_id.in_(current_user.allowed_projects or [])
        )

    result = await db.execute(query)

    rows = result.all()

    data = []

    for r in rows:
        qty_purchased = r.quantity_purchased or Decimal("0")
        remaining = r.remaining_stock or Decimal("0")
        total_amount = r.total_amount or Decimal("0")

        avg_rate = total_amount / qty_purchased if qty_purchased > 0 else Decimal("0")

        total_value = remaining * avg_rate

        # rounding at final stage
        avg_rate = avg_rate.quantize(Decimal("0.01"))
        total_value = total_value.quantize(Decimal("0.01"))

        data.append(
            {
                "material_id": r.id,
                "material_name": (r.material_name or "").strip().title(),
                "remaining_stock": float(remaining),
                "unit_id": r.unit_id,
                "unit_name": r.unit_name or "",
                "avg_rate": float(avg_rate),
                "total_value": float(total_value),
                "project_id": r.project_id,
            }
        )

    return data


# ==================get_inventory_valuation=======================


@router.get("/inventory/valuation")
async def get_inventory_valuation(
    project_id: Optional[int] = Query(
        None,
        description="Filter by project ID",
    ),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    query = select(
        Material.quantity_purchased,
        Material.remaining_stock,
        Material.total_amount,
    ).where(Material.is_deleted == False)

    # Optional Project Filter
    if project_id is not None:
        query = query.where(Material.project_id == project_id)

    result = await db.execute(query)
    rows = result.all()

    total_value = Decimal("0")

    for r in rows:
        purchased = r.quantity_purchased or Decimal("0")
        remaining = r.remaining_stock or Decimal("0")
        total_amount = r.total_amount or Decimal("0")

        avg_rate = total_amount / purchased if purchased > 0 else Decimal("0")

        total_value += remaining * avg_rate

    return {
        "project_id": project_id,
        "total_value": float(total_value.quantize(Decimal("0.01"))),
    }


# ======================================================


@router.get("/inventory/{project_id}")
async def get_project_inventory(
    project_id: int,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(
            Material.id,
            Material.material_name,
            Material.remaining_stock,
            Material.total_amount,
            Material.quantity_purchased,
        )
        .where(Material.project_id == project_id, Material.is_deleted == False)
        .offset(skip)
        .limit(limit)
    )

    rows = result.all()

    data = []

    for r in rows:
        purchased = r.quantity_purchased or Decimal("0")
        remaining = r.remaining_stock or Decimal("0")
        total_amt = r.total_amount or Decimal("0")

        avg_rate = total_amt / purchased if purchased > 0 else Decimal("0")

        total_value = remaining * avg_rate

        data.append(
            {
                "material_id": r.id,
                "material_name": (r.material_name or "").strip().title(),
                "remaining_stock": float(remaining),
                "avg_rate": float(avg_rate.quantize(Decimal("0.01"))),
                "total_value": float(total_value.quantize(Decimal("0.01"))),
            }
        )

    return data


# ================= FILTERED LOGS =================


@router.get("/logs", response_model=List[MaterialLogOut])
async def logs(
    limit: int = 50,
    offset: int = 0,
    material_id: Optional[int] = None,
    project_id: Optional[int] = None,
    type: Optional[SchemaTransactionType] = None,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    limit = min(max(limit, 1), 100)

    if (
        project_id is not None
        and current_user.role != UserRole.ADMIN.value
        and (project_id not in (current_user.allowed_projects or []))
    ):
        raise HTTPException(403, "Access denied")

    query = select(MaterialTransaction)

    if material_id is not None:
        query = query.where(MaterialTransaction.material_id == material_id)

    if project_id is not None:
        query = query.where(MaterialTransaction.project_id == project_id)
    elif current_user.role != UserRole.ADMIN.value:
        query = query.where(
            MaterialTransaction.project_id.in_(current_user.allowed_projects or [])
        )

    if type is not None:
        query = query.where(MaterialTransaction.type == type.value)

    query = (
        query.order_by(MaterialTransaction.created_at.desc())
        .offset(offset)
        .limit(limit)
    )

    result = await db.execute(query)
    rows = result.scalars().all()

    logs = []

    for r in rows:
        quantity = r.quantity or Decimal("0")
        total_amount = r.total_amount or Decimal("0")
        rate = r.rate or Decimal("0")

        logs.append(
            MaterialLogOut(
                id=r.id,
                material_id=r.material_id,
                boq_item_id=r.boq_item_id,
                type=r.type.value,
                quantity=float(round(quantity, 3)),
                rate=float(round(rate, 2)),
                avg_rate=float(round(rate, 2)),
                total_amount=float(round(total_amount, 2)),
                amount_paid=float(r.amount_paid or 0),
                payment_pending=float(r.payment_pending or 0),
                issue_type=r.issue_type,
                project_id=r.project_id,
                created_at=r.created_at,
            )
        )

    return logs


# ================= REPORTS =================


@router.get(
    "/reports",
    response_model=MaterialReportResponse,
)
async def material_report(
    project_id: int = Query(...),
    supplier_id: Optional[int] = None,
    material_id: Optional[int] = None,
    category: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    limit = min(max(limit, 1), 100)
    skip = max(skip, 0)

    if current_user.role != UserRole.ADMIN.value and project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    query = (
        select(Material)
        .options(
            selectinload(Material.unit),
            selectinload(Material.supplier),
            selectinload(Material.material_master),
        )
        .where(Material.is_deleted == False)
    )

    query = query.where(Material.project_id == project_id)

    if supplier_id:
        query = query.where(Material.supplier_id == supplier_id)

    if category:
        query = query.where(func.lower(Material.category) == category.lower())

    if material_id:
        query = query.where(Material.id == material_id)

    query = query.order_by(Material.id.desc()).offset(skip).limit(limit)

    materials = (await db.execute(query)).scalars().all()

    report_rows = []

    total_purchased = Decimal("0")
    total_used = Decimal("0")
    total_remaining = Decimal("0")
    total_stock_value = Decimal("0")
    total_payment_given = Decimal("0")
    total_payment_pending = Decimal("0")

    in_stock_count = 0
    low_stock_count = 0
    out_of_stock_count = 0

    for m in materials:

        purchased = m.quantity_purchased or Decimal("0")
        used = m.quantity_used or Decimal("0")
        remaining = m.remaining_stock or Decimal("0")
        total_amount = m.total_amount or Decimal("0")

        avg_rate = total_amount / purchased if purchased > 0 else Decimal("0")

        stock_value = remaining * avg_rate

        if remaining == 0:
            alert_type = "OUT_OF_STOCK"
            out_of_stock_count += 1

        elif remaining <= (m.minimum_stock_level or Decimal("0")):
            alert_type = "LOW_STOCK"
            low_stock_count += 1

        else:
            alert_type = "IN_STOCK"
            in_stock_count += 1

        total_purchased += purchased
        total_used += used
        total_remaining += remaining
        total_stock_value += stock_value

        total_payment_given += m.payment_given or Decimal("0")

        total_payment_pending += m.payment_pending or Decimal("0")

        report_rows.append(
            MaterialReport(
                material_id=m.id,
                material_code=m.material_code,
                material_master_id=m.material_master_id,
                material_master_name=(
                    m.material_master.name if m.material_master else None
                ),
                material_master_brand=(
                    m.material_master.brand if m.material_master else None
                ),
                material_master_specification=(
                    m.material_master.specification if m.material_master else None
                ),
                material_master_hsn_code=(
                    m.material_master.hsn_code if m.material_master else None
                ),
                material_name=m.material_name,
                category=m.category,
                unit_id=m.unit_id,
                unit_name=(m.unit.name if m.unit else None),
                supplier_id=m.supplier_id,
                supplier_name=(m.supplier.supplier_name if m.supplier else None),
                project_id=m.project_id,
                total_purchased=float(purchased),
                total_used=float(used),
                remaining_stock=float(remaining),
                avg_rate=float(avg_rate),
                stock_value=float(stock_value),
                payment_given=float(m.payment_given or 0),
                payment_pending=float(m.payment_pending or 0),
                minimum_stock_level=float(m.minimum_stock_level or 0),
                alert_type=alert_type,
            )
        )

    return MaterialReportResponse(
        summary=MaterialReportSummary(
            total_materials=len(report_rows),
            total_purchased=float(total_purchased),
            total_used=float(total_used),
            total_remaining=float(total_remaining),
            total_stock_value=float(total_stock_value),
            total_payment_given=float(total_payment_given),
            total_payment_pending=float(total_payment_pending),
            in_stock_count=in_stock_count,
            low_stock_count=low_stock_count,
            out_of_stock_count=out_of_stock_count,
        ),
        materials=report_rows,
    )


# ======================================================================
#  MATERIAL INVENTORY REPORT  —  PDF + EXCEL
# ======================================================================

from reportlab.lib.units import inch, cm
from openpyxl.utils import get_column_letter

# ─────────────────────────── COLOR PALETTE (matches sample) ────────────────
NAVY_BLUE = colors.HexColor("#0B2B5C")
LIGHT_BLUE = colors.HexColor("#3498DB")
GREEN = colors.HexColor("#27AE60")
RED = colors.HexColor("#E74C3C")
ORANGE = colors.HexColor("#F39C12")
LIGHT_GRAY = colors.HexColor("#F8F9FA")
BORDER_GRAY = colors.HexColor("#E2E8F0")

NAVY_HEX = "0B2B5C"
LIGHT_GRAY_HEX = "F8F9FA"
BORDER_GRAY_HEX = "E2E8F0"
GREEN_HEX = "27AE60"
RED_HEX = "E74C3C"
ORANGE_HEX = "F39C12"


# ─────────────────────────── FORMATTING HELPERS ─────────────────────────────
def fmt(val, dec: int = 2) -> str:
    try:
        val = val or Decimal("0")
        return f"{float(val):,.{dec}f}"
    except Exception:
        return "0"


def rs(val, dec: int = 2) -> str:
    return f"Rs. {fmt(val, dec)}"


def alert_to_status(alert_type: str) -> str:
    mapping = {"OUT_OF_STOCK": "OUT", "LOW_STOCK": "LOW", "IN_STOCK": "OK"}
    return mapping.get((alert_type or "").upper(), "OK")


def compute_avg_rate(total_amount, quantity_purchased) -> Decimal:
    """Weighted average cost per unit = total purchase cost / total qty purchased."""
    total_amount = total_amount or Decimal("0")
    quantity_purchased = quantity_purchased or Decimal("0")
    if quantity_purchased > 0:
        return total_amount / quantity_purchased
    return Decimal("0")


# ═══════════════════════════ PDF BUILDER ════════════════════════════════════
def _build_pdf(
    file_path: str,
    rows: list,
    project_name: Optional[str] = None,
    project_code: Optional[str] = None,
):
    doc = SimpleDocTemplate(
        file_path,
        pagesize=A4,
        rightMargin=cm,
        leftMargin=cm,
        topMargin=cm,
        bottomMargin=cm,
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "MatTitle",
        parent=styles["Heading1"],
        fontSize=20,
        textColor=NAVY_BLUE,
        alignment=0,
        spaceAfter=15,
        fontName="Helvetica-Bold",
    )
    heading2_style = ParagraphStyle(
        "MatH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=NAVY_BLUE,
        spaceBefore=6,
        spaceAfter=10,
        fontName="Helvetica-Bold",
    )
    normal_style = ParagraphStyle(
        "MatNormal",
        fontSize=9,
        textColor=colors.black,
        fontName="Helvetica",
        leading=11,
    )
    bold_style = ParagraphStyle(
        "MatBold",
        fontSize=9,
        textColor=colors.black,
        fontName="Helvetica-Bold",
        leading=11,
    )
    small_style = ParagraphStyle(
        "MatSmall", fontSize=8, textColor=colors.black, fontName="Helvetica", leading=10
    )

    # ── 1. HEADER (logo left, title right — same as sample) ────────────────
    logo_path = "static/logo.png"
    if os.path.exists(logo_path):
        logo_img = Image(logo_path, width=2 * inch, height=0.75 * inch)
    else:
        logo_img = Paragraph("<b>INFRA PILOT</b>", title_style)

    header_data = [
        [logo_img, Paragraph("<b>MATERIAL INVENTORY REPORT</b>", title_style)]
    ]
    header_table = Table(header_data, colWidths=[2.5 * inch, 4.5 * inch])
    header_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    elements.append(header_table)
    elements.append(Spacer(1, 0.15 * inch))

    # ── 2. PROJECT INFORMATION ──────────────────────────────────────────────
    pi_data = [
        [
            Paragraph("<b>Project Name</b>", bold_style),
            project_name or "N/A",
            Paragraph("<b>Report Date</b>", bold_style),
            datetime.now().strftime("%Y-%m-%d"),
        ],
        [
            Paragraph("<b>Project Code</b>", bold_style),
            project_code or "N/A",
            Paragraph("<b>Report Type</b>", bold_style),
            "Material Inventory Report",
        ],
    ]
    pi_table = Table(pi_data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 2 * inch])
    pi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                ("BACKGROUND", (2, 0), (2, -1), LIGHT_GRAY),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(Paragraph("1. PROJECT INFORMATION", heading2_style))
    elements.append(pi_table)
    elements.append(Spacer(1, 0.2 * inch))

    # ── PROCESS DATA (single source of truth for avg_rate / stock value) ───
    totals = {
        k: Decimal("0")
        for k in (
            "purchased",
            "used",
            "remaining",
            "value",
            "pending",
            "given",
            "advance",
            "purchase_cost",
        )
    }
    processed = []
    alerts_out, alerts_low = [], []

    for m, sup, unit_name in rows:
        purchased = m.quantity_purchased or Decimal("0")
        used = m.quantity_used or Decimal("0")
        remaining = m.remaining_stock or Decimal("0")
        total_amt = m.total_amount or Decimal("0")
        pending = m.payment_pending or Decimal("0")
        given = m.payment_given or Decimal("0")
        advance = m.advance_amount or Decimal("0")

        avg_rate = compute_avg_rate(total_amt, purchased)
        value = remaining * avg_rate

        for key, val in [
            ("purchased", purchased),
            ("used", used),
            ("remaining", remaining),
            ("value", value),
            ("pending", pending),
            ("given", given),
            ("advance", advance),
            ("purchase_cost", total_amt),
        ]:
            totals[key] += val

        status = alert_to_status(m.alert_type)
        if status == "OUT":
            alerts_out.append((m.material_name or "").title())
        elif status == "LOW":
            alerts_low.append((m.material_name or "").title())

        processed.append(
            (m, sup, unit_name, purchased, used, remaining, avg_rate, value, status)
        )

    # ── 3. INVENTORY SUMMARY ────────────────────────────────────────────────
    overall = "HIGH" if alerts_out else ("MEDIUM" if alerts_low else "LOW")
    elements.append(Paragraph("2. INVENTORY SUMMARY", heading2_style))
    summary_box_data = [
        [
            Paragraph(
                f"<b>Overall Stock Risk:</b> {overall} | "
                f"<b>Total Materials:</b> {len(rows)} | "
                f"<b>Total Stock Value:</b> {rs(totals['value'], 0)}",
                bold_style,
            )
        ],
        [
            Paragraph(
                f"<b>Summary:</b> {len(alerts_out)} material(s) out of stock, "
                f"{len(alerts_low)} material(s) running low on stock.",
                normal_style,
            )
        ],
        [
            Paragraph(
                f"<b>Pending Payment:</b> {rs(totals['pending'], 2)} across all suppliers.",
                normal_style,
            )
        ],
    ]
    summary_box = Table(summary_box_data, colWidths=[7 * inch])
    summary_box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
                ("BOX", (0, 0), (-1, -1), 1, NAVY_BLUE),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(summary_box)
    elements.append(Spacer(1, 0.2 * inch))

    # ── 4. MATERIAL DETAILS ──────────────────────────────────────────────
    elements.append(Paragraph("3. MATERIAL DETAILS", heading2_style))
    det_headers = [
        Paragraph("<b>Material</b>", bold_style),
        Paragraph("<b>Unit</b>", bold_style),
        Paragraph("<b>Supplier</b>", bold_style),
        Paragraph("<b>Purchased</b>", bold_style),
        Paragraph("<b>Used</b>", bold_style),
        Paragraph("<b>Remaining</b>", bold_style),
        Paragraph("<b>Avg Rate</b>", bold_style),
        Paragraph("<b>Stock Value</b>", bold_style),
    ]
    det_data = [det_headers]
    for item in processed:
        m, sup, unit_name, purchased, used, remaining, avg_rate, value, status = item
        det_data.append(
            [
                Paragraph((m.material_name or "").title(), small_style),
                unit_name or "-",
                Paragraph(sup or "N/A", small_style),
                fmt(purchased, 1),
                fmt(used, 1),
                fmt(remaining, 1),
                rs(avg_rate, 2),
                rs(value, 2),
            ]
        )
    det_data.append(
        [
            Paragraph("<b>TOTAL</b>", bold_style),
            "",
            "",
            fmt(totals["purchased"], 1),
            fmt(totals["used"], 1),
            fmt(totals["remaining"], 1),
            "",
            rs(totals["value"], 2),
        ]
    )
    det_table = Table(
        det_data,
        colWidths=[
            1.6 * inch,
            0.6 * inch,
            1.2 * inch,
            0.8 * inch,
            0.7 * inch,
            0.8 * inch,
            0.9 * inch,
            1.0 * inch,
        ],
        repeatRows=1,
    )
    total_row_idx = len(det_data) - 1
    det_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                ("BACKGROUND", (0, total_row_idx), (-1, total_row_idx), LIGHT_GRAY),
                ("LINEABOVE", (0, total_row_idx), (-1, total_row_idx), 1, NAVY_BLUE),
            ]
        )
    )
    elements.append(det_table)
    elements.append(Spacer(1, 0.2 * inch))

    # ── 5. CRITICAL STOCK ALERTS ─────────────────────────────────────────
    elements.append(Paragraph("4. CRITICAL STOCK ALERTS", heading2_style))
    crit_headers = [
        Paragraph("<b>Material</b>", bold_style),
        Paragraph("<b>Status</b>", bold_style),
        Paragraph("<b>Action</b>", bold_style),
    ]
    crit_data = [crit_headers]
    for name in alerts_out:
        crit_data.append(
            [
                Paragraph(name, small_style),
                Paragraph(
                    '<font color="#E74C3C"><b>OUT OF STOCK</b></font>', small_style
                ),
                Paragraph("Immediate replenishment required", small_style),
            ]
        )
    for name in alerts_low:
        crit_data.append(
            [
                Paragraph(name, small_style),
                Paragraph('<font color="#F39C12"><b>LOW STOCK</b></font>', small_style),
                Paragraph("Schedule reorder soon", small_style),
            ]
        )
    if len(crit_data) == 1:
        crit_data.append([Paragraph("No critical materials.", small_style), "-", "-"])
    crit_table = Table(crit_data, colWidths=[2.5 * inch, 1.8 * inch, 2.7 * inch])
    crit_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(crit_table)
    elements.append(Spacer(1, 0.2 * inch))

    # ── 6. PAYMENT SUMMARY ────────────────────────────────────────────────
    elements.append(Paragraph("5. PAYMENT SUMMARY", heading2_style))
    pay_data = [
        [
            Paragraph("<b>Total Purchase Cost</b>", bold_style),
            rs(totals["purchase_cost"], 2),
        ],
        [Paragraph("<b>Total Paid</b>", bold_style), rs(totals["given"], 2)],
        [Paragraph("<b>Pending Payment</b>", bold_style), rs(totals["pending"], 2)],
        [Paragraph("<b>Advance Paid</b>", bold_style), rs(totals["advance"], 2)],
    ]
    pay_table = Table(pay_data, colWidths=[3.5 * inch, 3.5 * inch])
    pay_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), LIGHT_GRAY),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(pay_table)

    # ── FOOTER (signature lines — identical to sample) ────────────────────
    def add_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setStrokeColor(NAVY_BLUE)
        canvas.setLineWidth(1)
        canvas.line(cm, 1.5 * cm, A4[0] - cm, 1.5 * cm)

        canvas.drawString(cm, 1.2 * cm, "Prepared By: ______________")
        canvas.drawString(A4[0] / 2 - 1.5 * cm, 1.2 * cm, "Reviewed By: ______________")
        canvas.drawString(A4[0] - 5 * cm, 1.2 * cm, "Approved By: ______________")

        canvas.drawString(
            cm, 0.8 * cm, "Generated by InfraPilot Construction Management System"
        )
        canvas.drawRightString(A4[0] - cm, 0.8 * cm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)


# ═══════════════════════════ EXCEL BUILDER (navy theme, same palette) ═══════
def _build_excel(project, rows) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    border = Border(
        left=Side(style="thin", color=BORDER_GRAY_HEX),
        right=Side(style="thin", color=BORDER_GRAY_HEX),
        top=Side(style="thin", color=BORDER_GRAY_HEX),
        bottom=Side(style="thin", color=BORDER_GRAY_HEX),
    )
    header_fill = PatternFill(
        start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=10)
    accent_fill = PatternFill(
        start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid"
    )
    title_font = Font(bold=True, size=14, color=NAVY_HEX)
    subtitle_font = Font(size=10, color="6B7280")

    ws.merge_cells("A1:N1")
    ws["A1"] = "MATERIAL INVENTORY REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:N2")
    ws["A2"] = (
        f"Project: {project.project_name if project else 'N/A'}  |  "
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    row = 4

    totals = {
        k: Decimal("0")
        for k in [
            "purchased",
            "used",
            "remaining",
            "value",
            "paid",
            "pending",
            "advance",
            "purchase_cost",
        ]
    }
    status_counts = {"OK": 0, "LOW": 0, "OUT": 0}
    computed_rows = []

    for material, supplier_name, unit_name in rows:
        purchased = material.quantity_purchased or Decimal("0")
        used = material.quantity_used or Decimal("0")
        remaining = material.remaining_stock or Decimal("0")
        total_amt = material.total_amount or Decimal("0")
        avg_rate = compute_avg_rate(total_amt, purchased)
        value = remaining * avg_rate

        totals["purchased"] += purchased
        totals["used"] += used
        totals["remaining"] += remaining
        totals["value"] += value
        totals["paid"] += material.payment_given or Decimal("0")
        totals["pending"] += material.payment_pending or Decimal("0")
        totals["advance"] += material.advance_amount or Decimal("0")
        totals["purchase_cost"] += total_amt

        status = alert_to_status(material.alert_type)
        status_counts[status] += 1
        computed_rows.append((material, supplier_name, unit_name, avg_rate, value))

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "INVENTORY SUMMARY"
    ws[f"A{row}"].font = Font(bold=True, size=10, color=NAVY_HEX)
    ws[f"A{row}"].fill = PatternFill(
        start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"
    )
    row += 1

    summary_data = [
        ("Total Materials", len(rows), NAVY_HEX),
        ("Total Stock Value", f"Rs. {fmt(totals['value'], 2)}", "3498DB"),
        ("In Stock", status_counts["OK"], GREEN_HEX),
        ("Low Stock", status_counts["LOW"], ORANGE_HEX),
        ("Out of Stock", status_counts["OUT"], RED_HEX),
    ]
    for col, (label, value, color) in enumerate(summary_data, 1):
        ws.cell(row, col, label).font = Font(size=9, color="6B7280")
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        val_cell = ws.cell(row + 1, col, value)
        val_cell.font = Font(bold=True, size=11, color=color)
        val_cell.alignment = Alignment(horizontal="center")

    row += 3

    headers = [
        "ID",
        "Material",
        "Code",
        "Unit",
        "Supplier",
        "Purchased",
        "Used",
        "Remaining",
        "Avg Rate",
        "Stock Value",
        "Paid",
        "Pending",
        "Advance",
        "Alert",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row, col_num, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[row].height = 18
    header_row = row
    row += 1

    for idx, (material, supplier_name, unit_name, avg_rate, value) in enumerate(
        computed_rows
    ):
        values = [
            material.id,
            material.material_name or "-",
            material.material_code or "-",
            unit_name or "-",
            supplier_name or "-",
            float(material.quantity_purchased or 0),
            float(material.quantity_used or 0),
            float(material.remaining_stock or 0),
            float(avg_rate),
            float(value),
            float(material.payment_given or 0),
            float(material.payment_pending or 0),
            float(material.advance_amount or 0),
            material.alert_type or "-",
        ]
        fill = accent_fill if idx % 2 == 1 else None
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row, col_num, val)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if col_num > 5 else "left", vertical="center"
            )
            if fill:
                cell.fill = fill
        row += 1

    total_row = row
    ws.cell(total_row, 2, "TOTAL").font = Font(bold=True, color=NAVY_HEX)
    ws.cell(total_row, 6, float(totals["purchased"])).font = Font(bold=True)
    ws.cell(total_row, 7, float(totals["used"])).font = Font(bold=True)
    ws.cell(total_row, 8, float(totals["remaining"])).font = Font(bold=True)
    ws.cell(total_row, 10, float(totals["value"])).font = Font(bold=True)
    ws.cell(total_row, 11, float(totals["paid"])).font = Font(bold=True)
    ws.cell(total_row, 12, float(totals["pending"])).font = Font(bold=True)
    ws.cell(total_row, 13, float(totals["advance"])).font = Font(bold=True)

    total_fill = PatternFill(
        start_color=LIGHT_GRAY_HEX, end_color=LIGHT_GRAY_HEX, fill_type="solid"
    )
    for col in range(1, len(headers) + 1):
        cell = ws.cell(total_row, col)
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right" if col > 5 else "left")

    row = total_row + 3
    ws.cell(row, 1, "Prepared By: ______________").font = Font(size=9)
    ws.cell(row, 6, "Reviewed By: ______________").font = Font(size=9)
    ws.cell(row, 11, "Approved By: ______________").font = Font(size=9)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:N{total_row - 1}"

    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_width = 12
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    max_width = max(max_width, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width + 2, 25)

    file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb.save(file_path)
    return file_path


# ═══════════════════════════ PDF ENDPOINT — ACTUAL ROUTE ═════════════════════
@router.get("/reports/pdf", response_class=FileResponse)
async def export_pdf(
    project_id: int = Query(...),
    category: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    material_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    try:
        if current_user.role != UserRole.ADMIN.value and project_id not in (
            current_user.allowed_projects or []
        ):
            raise HTTPException(403, "Access denied")

        query = (
            select(Material, Supplier.supplier_name, Unit.name)
            .join(Supplier, Supplier.id == Material.supplier_id, isouter=True)
            .join(Unit, Unit.id == Material.unit_id, isouter=True)
            .where(Material.is_deleted == False, Material.project_id == project_id)
        )
        if category:
            query = query.where(func.lower(Material.category) == category.lower())
        if supplier_id:
            query = query.where(Material.supplier_id == supplier_id)
        if material_id:
            query = query.where(Material.id == material_id)

        rows = (await db.execute(query.order_by(Material.id.desc()))).all()
        if not rows:
            raise HTTPException(status_code=404, detail="No material data found")

        project = await db.get(Project, project_id)
        project_name = project.project_name if project else None
        project_code = getattr(project, "project_code", None) if project else None

        file_path = os.path.join(
            tempfile.gettempdir(), f"material_report_{uuid.uuid4()}.pdf"
        )
        try:
            _build_pdf(
                file_path=file_path,
                rows=rows,
                project_name=project_name,
                project_code=project_code,
            )
        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"PDF generation failed: {str(e)}"
            )

        return FileResponse(
            path=file_path,
            filename="material_report.pdf",
            media_type="application/pdf",
            background=BackgroundTask(safe_delete, file_path),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Material report error: {str(e)}")


# ═══════════════════════════ EXCEL ENDPOINT — ACTUAL ROUTE ═══════════════════
@router.get("/reports/excel", response_class=FileResponse)
async def export_excel(
    project_id: int = Query(...),
    category: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    material_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    try:
        if current_user.role != UserRole.ADMIN.value and project_id not in (
            current_user.allowed_projects or []
        ):
            raise HTTPException(403, "Access denied")

        project = await db.get(Project, project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

        query = (
            select(Material, Supplier.supplier_name, Unit.name)
            .outerjoin(Supplier, Supplier.id == Material.supplier_id)
            .outerjoin(Unit, Unit.id == Material.unit_id)
            .where(Material.is_deleted == False, Material.project_id == project_id)
            .order_by(Material.id.desc())
        )
        if category:
            query = query.where(func.lower(Material.category) == category.lower())
        if supplier_id:
            query = query.where(Material.supplier_id == supplier_id)
        if material_id:
            query = query.where(Material.id == material_id)

        rows = (await db.execute(query)).all()
        if not rows:
            raise HTTPException(status_code=404, detail="No material data found")

        file_path = _build_excel(project, rows)

        return FileResponse(
            path=file_path,
            filename=f"material_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(safe_delete, file_path),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel report error: {str(e)}")


# ===============price-history==========================


@router.get("/price-history/{material_id}", response_model=list[PriceHistoryOut])
async def price_history(
    material_id: int,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    result = await db.execute(
        select(MaterialTransaction.rate, MaterialTransaction.created_at)
        .where(
            MaterialTransaction.material_id == material_id,
            MaterialTransaction.type == DBTransactionType.PURCHASE,
        )
        .order_by(MaterialTransaction.created_at.asc())
    )

    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="No price history found")

    history = []
    last_rate: Decimal | None = None

    for rate, created_at in rows:
        rate = rate or Decimal("0")

        # avoid float precision issue
        if last_rate is None or rate != last_rate:
            history.append(
                {
                    "rate": float(round(rate, 2)),
                    "date": created_at.strftime("%Y-%m-%d %H:%M"),
                }
            )
            last_rate = rate

    return history


# ================= MATERIALS - DYNAMIC ROUTES =================


@router.post("", response_model=MaterialOut)
async def create_material(
    payload: MaterialCreate,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    redis=Depends(get_request_redis),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
):
    from decimal import Decimal

    import json
    import hashlib

    request_hash = None
    if idempotency_key:
        payload_dict = payload.model_dump(mode="json")

        payload_str = json.dumps(payload_dict, sort_keys=True)
        request_hash = hashlib.sha256(payload_str.encode("utf-8")).hexdigest()

        existing = await db.scalar(
            select(MaterialTransaction)
            .where(MaterialTransaction.idempotency_key == idempotency_key)
        )
        if existing:
            if existing.request_hash != request_hash:
                raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
            return await _build_material_out_from_id(db, existing.material_id)

    import uuid

    data = payload.model_dump()

    # ================= PROJECT VALIDATION =================

    project = await db.get(Project, payload.project_id)

    if not project:
        raise HTTPException(
            status_code=404,
            detail="Project not found",
        )

    if current_user.role != UserRole.ADMIN.value and payload.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(403, "Access denied")

    # ================= SUPPLIER VALIDATION =================

    supplier = await db.get(
        Supplier,
        payload.supplier_id,
    )

    if not supplier or supplier.is_deleted:
        raise HTTPException(
            status_code=404,
            detail="Supplier not found",
        )

    # ================= MATERIAL MASTER VALIDATION =================

    material_master = await db.get(
        MaterialMaster,
        payload.material_master_id,
    )

    if not material_master:
        raise HTTPException(
            status_code=404,
            detail="Material master not found",
        )

    if getattr(material_master, "is_deleted", False):
        raise HTTPException(
            status_code=400,
            detail="Material master is deleted",
        )

    # ================= INPUT VALIDATION =================

    qty = Decimal(str(payload.quantity_purchased or 0))
    rate = Decimal(str(payload.purchase_rate or 0))
    payment = Decimal(str(payload.payment_given or 0))

    if qty < 0:
        raise HTTPException(
            status_code=400,
            detail="Quantity purchased cannot be negative",
        )

    if rate <= 0:
        raise HTTPException(
            status_code=400,
            detail="Purchase rate must be greater than 0",
        )

    if payment < 0:
        raise HTTPException(
            status_code=400,
            detail="Payment cannot be negative",
        )

    # ================= DUPLICATE CHECK =================

    existing = await db.scalar(
        select(Material).where(
            Material.project_id == payload.project_id,
            Material.material_master_id == payload.material_master_id,
            Material.supplier_id == payload.supplier_id,
            Material.is_deleted == False,
        )
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Material already exists for this project & supplier",
        )

    # ================= AUTO FILL FROM MASTER =================

    data["material_name"] = material_master.name
    data["category"] = material_master.category or "GENERAL"
    data["unit_id"] = material_master.unit_id
    data["purchase_rate"] = (
        rate  # keep the quantized/validated Decimal, not raw payload value
    )

    # ================= GENERATE MATERIAL CODE =================

    material_code = await generate_business_id(
        db=db,
        model=Material,
        column_name="material_code",
        prefix="MAT",
    )

    # ================= CREATE MATERIAL =================

    obj = Material(
        **data,
        material_code=material_code,
    )

    obj.quantity_purchased = qty
    obj.quantity_used = Decimal("0")
    obj.payment_given = payment

    try:

        db.add(obj)
        await db.flush()

        # ================= CALCULATIONS =================

        obj.total_amount = (qty * rate).quantize(Decimal("0.01"))

        if obj.payment_given > obj.total_amount:
            raise HTTPException(
                status_code=400,
                detail="Payment cannot exceed total amount",
            )

        update_material_fields(obj)

        alert_type = get_alert_type(obj)

        # ================= REFERENCE =================

        reference = f"INIT-{uuid.uuid4().hex[:8]}"

        # ================= OPENING STOCK TRANSACTION =================

        if qty > 0:

            transaction = MaterialTransaction(
                material_id=obj.id,
                idempotency_key=idempotency_key,
                request_hash=request_hash,
                type=DBTransactionType.PURCHASE,
                quantity=qty,
                rate=rate,
                total_amount=obj.total_amount,
                amount_paid=obj.payment_given,
                payment_pending=obj.payment_pending,
                issue_type=IssueType.PURCHASE,
                project_id=obj.project_id,
                remarks="Initial material entry",
                reference_id=reference,
            )

            ledger = MaterialLedger(
                material_id=obj.id,
                type=DBTransactionType.PURCHASE,
                quantity=qty,
                rate=rate,
                total_amount=obj.total_amount,
                amount_paid=obj.payment_given,
                payment_pending=obj.payment_pending,
                project_id=obj.project_id,
                remarks="Initial material entry",
                reference_id=reference,
            )

            db.add(transaction)
            db.add(ledger)

        await db.commit()

    except HTTPException:
        await db.rollback()
        raise


    except IntegrityError as e:
        await db.rollback()

        if idempotency_key:
            existing = await db.scalar(
                select(MaterialTransaction)
                .where(MaterialTransaction.idempotency_key == idempotency_key)
            )
            if existing:
                if existing.request_hash != request_hash:
                    raise HTTPException(409, "Idempotency-Key already used with a different request payload.")
                return await _build_material_out_from_id(db, existing.material_id)

        # If it wasn't the idempotency key, it might be the material unique constraint
        raise HTTPException(
            status_code=400,
            detail="Material already exists or integrity constraint violated"
        )


    except Exception:
        await db.rollback()
        raise

    await db.refresh(obj)

    # ================= CACHE =================

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    # ================= RESPONSE =================

    unit = await db.get(
        Unit,
        obj.unit_id,
    )

    response = MaterialOut.model_validate(
        {
            **obj.__dict__,
            "unit_name": unit.name if unit else "",
            "alert_type": alert_type,
        }
    )

    response.supplier_name = supplier.supplier_name
    response.material_master_name = material_master.name
    response.material_master_brand = material_master.brand
    response.material_master_specification = material_master.specification
    response.material_master_hsn_code = material_master.hsn_code

    return response


# =================list_materials=========================


@router.get("", response_model=list[MaterialOut])
async def list_materials(
    project_id: int | None = None,
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    skip = max(skip, 0)
    limit = min(max(limit, 1), 100)

    if (
        project_id is not None
        and current_user.role != UserRole.ADMIN.value
        and project_id not in (current_user.allowed_projects or [])
    ):
        raise HTTPException(
            status_code=403,
            detail="Access denied",
        )

    query = (
        select(
            Material,
            Supplier.supplier_name,
        )
        .options(
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .join(
            Supplier,
            Supplier.id == Material.supplier_id,
            isouter=True,
        )
        .where(
            Material.is_deleted == False,
        )
    )

    if project_id is not None:
        query = query.where(
            Material.project_id == project_id,
        )
    elif current_user.role != UserRole.ADMIN.value:
        query = query.where(
            Material.project_id.in_(current_user.allowed_projects or [])
        )

    query = query.order_by(Material.id.desc()).offset(skip).limit(limit)

    rows = (await db.execute(query)).all()

    return [
        build_material_response(
            obj,
            supplier_name=supplier_name,
            unit_name=(obj.unit.name if obj.unit else ""),
        )
        for obj, supplier_name in rows
    ]


# =====================================================
# FEATURE 1: WEEKLY PROCUREMENT REPORT (PDF)
# =====================================================
@router.get(
    "/procurement-report/{project_id}",
    summary="Weekly Procurement Report (PDF)",
    description="Generates a downloadable Weekly Procurement Report PDF including Project Info, Current Inventory, Critical Materials, Supplier Recommendations, Procurement Budget, Consumption Summary, and AI Procurement Summary.",
)
async def download_procurement_report_pdf(
    project_id: int,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    if (
        current_user.role != UserRole.ADMIN.value
        and current_user.allowed_projects
        and project_id not in current_user.allowed_projects
    ):
        raise HTTPException(status_code=403, detail="Access denied")

    project = await db.scalar(select(Project).where(Project.id == project_id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    mats = (
        (
            await db.execute(
                select(Material).where(
                    Material.project_id == project_id, Material.is_deleted == False
                )
            )
        )
        .scalars()
        .all()
    )

    inv_data = []
    for m in mats:
        inv_data.append(
            {
                "material_name": m.material_name,
                "category": m.category,
                "remaining_stock": round(float(m.remaining_stock or 0), 2),
                "minimum_stock_level": round(float(m.minimum_stock_level or 0), 2),
                "purchase_rate": float(m.purchase_rate or 0),
                "total_amount": float(m.total_amount or 0),
            }
        )

    rec_data = await generate_material_recommendations(db=db, project_id=project_id)
    supplier_rec_data = await generate_supplier_recommendations(
        db=db, project_id=project_id
    )
    ai_summary = await generate_procurement_summary(db=db, project_id=project_id)
    consumption_summary = await get_material_consumption_trend(
        db=db, project_id=project_id, days=30
    )

    pdf_payload = {
        "project": {
            "name": project.project_name,
            "code": getattr(project, "business_id", f"PRJ-{project_id}"),
            "location": getattr(project, "site_address", "N/A") or "N/A",
        },
        "report_date": datetime.utcnow().strftime("%Y-%m-%d"),
        "current_inventory": inv_data,
        "critical_materials": rec_data.get("recommendations", []),
        "supplier_recommendations": supplier_rec_data.get(
            "supplier_recommendations", []
        ),
        "procurement_recommendations": rec_data.get("recommendations", []),
        "ai_procurement_summary": ai_summary,
        "consumption_summary": consumption_summary,
    }

    pdf_buffer = generate_procurement_report_pdf(pdf_payload)

    headers = {
        "Content-Disposition": f"attachment; filename=procurement_report_{project_id}.pdf"
    }
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers=headers,
    )


# ==============get_material=================

async def _get_active_material_or_404(db: AsyncSession, material_id: int):
    obj = await db.scalar(
        select(Material).where(
            Material.id == material_id,
            Material.is_deleted == False
        )
    )
    if not obj:
        raise HTTPException(status_code=404, detail="Material not found")
    return obj


@router.get("/{material_id}/qr", response_class=StreamingResponse)
async def generate_material_qr(
    material_id: int,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    material = await _get_active_material_or_404(db, material_id)

    qr_buf = generate_qr(entity_type="MAT", entity_id=material.id)

    headers = {
        "Cache-Control": "no-store",
        "Content-Disposition": f'inline; filename="material_{material.id}.png"'
    }

    return StreamingResponse(
        qr_buf,
        media_type="image/png",
        headers=headers
    )


@router.get("/{material_id}", response_model=MaterialOut)
async def get_material(
    material_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
):
    result = await db.execute(
        select(
            Material,
            Supplier.supplier_name,
        )
        .options(
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .join(
            Supplier,
            Supplier.id == Material.supplier_id,
            isouter=True,
        )
        .where(
            Material.id == material_id,
            Material.is_deleted == False,
        )
    )

    row = result.first()

    if not row:
        raise HTTPException(
            status_code=404,
            detail="Material not found",
        )

    obj, supplier_name = row

    if current_user.role != UserRole.ADMIN.value and obj.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(
            status_code=403,
            detail="Access denied",
        )

    return build_material_response(
        obj,
        supplier_name=supplier_name,
        unit_name=(obj.unit.name if obj.unit else ""),
    )


# =============update_material==================


@router.put("/{material_id}", response_model=MaterialOut)
async def update_material(
    material_id: int,
    payload: MaterialUpdate,
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):
    obj = await db.scalar(
        select(Material)
        .where(
            Material.id == material_id,
            Material.is_deleted == False,
        )
        .with_for_update()
    )

    if not obj:
        raise HTTPException(
            status_code=404,
            detail="Material not found",
        )

    # =====================================================
    # PROJECT ACCESS VALIDATION
    # =====================================================

    if (
        hasattr(current_user, "allowed_projects")
        and current_user.role != UserRole.ADMIN.value
        and current_user.allowed_projects
        and obj.project_id not in current_user.allowed_projects
    ):
        raise HTTPException(
            status_code=403,
            detail="Access denied",
        )

    update_data = payload.model_dump(exclude_unset=True)

    # =====================================================
    # BLOCK SYSTEM MANAGED FIELDS
    # =====================================================

    blocked_fields = {
        "quantity_purchased",
        "quantity_used",
        "remaining_stock",
        "total_amount",
        "payment_given",
        "payment_pending",
        "advance_amount",
    }

    invalid_fields = blocked_fields.intersection(update_data.keys())

    if invalid_fields:
        raise HTTPException(
            status_code=400,
            detail=(
                "Direct update not allowed for: "
                + ", ".join(sorted(invalid_fields))
                + ". Use Purchase / Usage APIs."
            ),
        )

    # =====================================================
    # VALIDATE SUPPLIER
    # =====================================================

    if "supplier_id" in update_data:

        supplier = await db.get(
            Supplier,
            update_data["supplier_id"],
        )

        if not supplier:
            raise HTTPException(
                status_code=404,
                detail="Supplier not found",
            )

    # =====================================================
    # VALIDATE MATERIAL MASTER
    # =====================================================

    if "material_master_id" in update_data:

        material_master = await db.get(
            MaterialMaster,
            update_data["material_master_id"],
        )

        if not material_master:
            raise HTTPException(
                status_code=404,
                detail="Material master not found",
            )

        update_data["material_name"] = material_master.name
        update_data["category"] = (
            material_master.category if material_master.category else "GENERAL"
        )
        update_data["unit_id"] = material_master.unit_id

    # =====================================================
    # DUPLICATE CHECK
    # =====================================================

    new_supplier_id = update_data.get(
        "supplier_id",
        obj.supplier_id,
    )

    new_material_master_id = update_data.get(
        "material_master_id",
        obj.material_master_id,
    )

    existing = await db.scalar(
        select(Material).where(
            Material.project_id == obj.project_id,
            Material.material_master_id == new_material_master_id,
            Material.supplier_id == new_supplier_id,
            Material.id != obj.id,
            Material.is_deleted == False,
        )
    )

    if existing:
        raise HTTPException(
            status_code=400,
            detail="Material already exists for this project & supplier",
        )

    try:

        # =====================================================
        # APPLY UPDATES
        # =====================================================

        for key, value in update_data.items():
            setattr(obj, key, value)

        update_material_fields(obj)

        await db.commit()
        await db.refresh(obj)

    except IntegrityError:

        await db.rollback()

        raise HTTPException(
            status_code=400,
            detail="Material update failed",
        )

    except Exception:

        await db.rollback()
        raise

    # =====================================================
    # RELOAD WITH RELATIONSHIPS + SUPPLIER
    # =====================================================

    result = await db.execute(
        select(
            Material,
            Supplier.supplier_name,
        )
        .options(
            selectinload(Material.unit),
            selectinload(Material.material_master),
        )
        .join(
            Supplier,
            Supplier.id == Material.supplier_id,
            isouter=True,
        )
        .where(
            Material.id == material_id,
            Material.is_deleted == False,
        )
    )

    row = result.first()

    if not row:
        raise HTTPException(
            status_code=404,
            detail="Material not found",
        )

    obj, supplier_name = row

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    total_amount, payment_given, payment_pending, extra_paid = calculate_fields(obj)

    if obj.remaining_stock == 0:
        alert_type = "OUT_OF_STOCK"
    elif obj.remaining_stock <= obj.minimum_stock_level:
        alert_type = "LOW_STOCK"
    else:
        alert_type = "IN_STOCK"

    return MaterialOut(
        id=obj.id,
        material_code=obj.material_code,
        project_id=obj.project_id,
        material_master_id=obj.material_master_id,
        material_master_name=(
            obj.material_master.name if obj.material_master else None
        ),
        material_master_brand=(
            obj.material_master.brand if obj.material_master else None
        ),
        material_master_specification=(
            obj.material_master.specification if obj.material_master else None
        ),
        material_master_hsn_code=(
            obj.material_master.hsn_code if obj.material_master else None
        ),
        material_name=obj.material_name,
        category=obj.category,
        unit_id=obj.unit_id,
        unit_name=(obj.unit.name if obj.unit else ""),
        supplier_id=obj.supplier_id,
        supplier_name=supplier_name,
        purchase_rate=float(obj.purchase_rate or 0),
        rate_type=obj.rate_type,
        quantity_purchased=float(obj.quantity_purchased or 0),
        quantity_used=float(obj.quantity_used or 0),
        remaining_stock=float(obj.remaining_stock or 0),
        total_amount=round(total_amount, 2),
        payment_given=round(payment_given, 2),
        payment_pending=round(payment_pending, 2),
        extra_paid=round(extra_paid, 2),
        minimum_stock_level=float(obj.minimum_stock_level or 0),
        alert_type=alert_type,
    )


# ============delete_material===========


@router.delete(
    "/{material_id}",
    response_model=MessageResponse,
    summary="Delete Material",
    description="Soft deletes a material record.",
)
async def delete_material(
    material_id: int,
    db: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_roles(MATERIAL_WRITE_ROLES)),
    redis=Depends(get_request_redis),
):

    obj = await db.scalar(
        select(Material).where(
            Material.id == material_id,
            Material.is_deleted == False,
        )
    )

    if not obj:
        raise HTTPException(
            status_code=404,
            detail="Material not found",
        )

    if current_user.role != UserRole.ADMIN.value and obj.project_id not in (
        current_user.allowed_projects or []
    ):
        raise HTTPException(
            status_code=403,
            detail="Access denied",
        )

    try:
        obj.is_deleted = True
        await db.commit()

    except Exception:
        await db.rollback()
        raise

    await bump_cache_version(
        redis,
        VERSION_KEY,
    )

    return MessageResponse(
        success=True,
        message="Material deleted successfully",
        resource_id=material_id,
    )


# =====================================================
# FEATURE 2: AI MATERIAL RECOMMENDATION ENGINE
# =====================================================
@router.post(
    "/ai-recommendation",
    response_model=AIMaterialRecommendationResponse,
)
async def get_ai_material_recommendation(
    payload: AIMaterialRecommendationRequest,
    current_user: User = Depends(require_roles(MATERIAL_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
    redis=Depends(get_request_redis),
):

    version = await get_cache_version(redis, VERSION_KEY)
    cache_key = (
        f"cache:materials:ai_rec:{version}:{payload.project_id}:{payload.target_days}"
    )
    cached = await cache_get_json(redis, cache_key)
    if cached is not None:
        return AIMaterialRecommendationResponse.model_validate(cached)

    result = await generate_material_recommendations(
        db=db,
        project_id=payload.project_id,
        target_days=payload.target_days,
    )

    try:
        ai_obj = AIPrediction(
            module_name="material_recommendation",
            prompt=f"project_id={payload.project_id}, target_days={payload.target_days}",
            prediction=result,
            created_by_user_id=current_user.id,
        )
        db.add(ai_obj)
        await db.flush()
    except Exception as e:
        logger.warning(f"Failed to store AIPrediction log: {e}")

    await cache_set_json(redis, cache_key, result)
    return AIMaterialRecommendationResponse.model_validate(result)
