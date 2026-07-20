from datetime import date, datetime, time
from io import BytesIO
import os
from typing import Annotated, List

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from starlette.concurrency import run_in_threadpool
from fastapi import UploadFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core import dependencies as d
from app.core.logger import logger
from app.core.validators import validate_and_save_image
from app.db.session import get_db_session

from app.models.master_data import ActivityType
from app.models.project import (
    Project,
    ProjectMember,
    Task,
    WorkActivity,
)

from app.models.user import User, UserRole

from app.models.work_update import (
    WorkUpdate,
    WorkUpdateImage,
    WorkUpdateImageType,
    WorkUpdateStatus,
)

from app.schemas import work_update as s
from app.schemas.work_update import ExportFormat

from app.utils.common import (
    assert_project_access,
    assert_task_project,
    generate_business_id,
)

from app.utils.helpers import (
    NotFoundError,
    ValidationError,
)

from collections import defaultdict

WORK_UPDATE_READ_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.CONTRACTOR,
        UserRole.CLIENT,
        UserRole.LABOUR,
    ]
]

WORK_UPDATE_WRITE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
        UserRole.CONTRACTOR,
        UserRole.LABOUR,
    ]
]

WORK_UPDATE_DELETE_ROLES = [
    r.value
    for r in [
        UserRole.ADMIN,
        UserRole.PROJECT_MANAGER,
        UserRole.SITE_ENGINEER,
    ]
]

router = APIRouter(
    prefix="/work-updates",
    tags=["Work Updates"],
)

MAX_PAGE_SIZE = 100


# =====================================================
# PHASE 1 — CREATE / UPDATE VALIDATION HELPERS
# (router-only: no changes to the WorkUpdate model. Project/Task status
# blocking relies on whatever `status` attribute already exists on those
# models — see the ASSUMPTION notes below and adjust the string sets if
# your actual enum values differ.)
# =====================================================

# ASSUMPTION: Project.status is a string/enum whose value is one of these
# (case-insensitive). Adjust this set to match your real ProjectStatus enum.
BLOCKED_PROJECT_STATUSES = {"closed", "completed", "inactive", "cancelled"}

# ASSUMPTION: Task.status is a string/enum whose value is one of these
# (case-insensitive). Adjust this set to match your real TaskStatus enum.
BLOCKED_TASK_STATUSES = {"completed", "closed", "cancelled"}


def _status_value(status_obj) -> str | None:
    """Normalizes a plain string or an Enum member into a lowercase string."""
    if status_obj is None:
        return None
    raw = getattr(status_obj, "value", status_obj)
    return str(raw).strip().lower()


def _assert_project_open(project: Project) -> None:
    """
    Blocks work-update create/update against a project that is
    Closed / Completed / Inactive / Cancelled.
    If `Project` has no `status` attribute in your actual model, this is a
    silent no-op (getattr default), so it fails safe rather than crashing.
    """
    value = _status_value(getattr(project, "status", None))
    if value is not None and value in BLOCKED_PROJECT_STATUSES:
        raise ValidationError(
            f"Cannot create or update a work update — project status is "
            f"'{getattr(project, 'status', value)}'."
        )


def _assert_task_open(task: Task) -> None:
    """
    Blocks work-update create/update against a task that is
    Completed / Closed / Cancelled.
    Same fail-safe behavior as _assert_project_open if `status` isn't
    present on your Task model.
    """
    value = _status_value(getattr(task, "status", None))
    if value is not None and value in BLOCKED_TASK_STATUSES:
        raise ValidationError(
            f"Cannot create or update a work update — task status is "
            f"'{getattr(task, 'status', value)}'."
        )


def _assert_activity_belongs_to_task(task: Task, activity_type_id: int) -> None:
    """
    Verifies the selected activity_type is valid for the given task —
    previously only `db.get(ActivityType, id)` (existence only) was
    checked, with no link back to the task at all.

    ASSUMPTION: Task exposes a single `activity_type_id` FK (a task is
    restricted to one activity type). If your schema instead has a
    many-to-many (e.g. `task.activity_types`), replace the check below
    with:
        allowed_ids = {a.id for a in task.activity_types}
        if activity_type_id not in allowed_ids: raise ValidationError(...)
    If a task has NO activity restriction at all, delete this function and
    its call sites — existence-only validation (already in place) is
    sufficient.
    """
    allowed_activity_type_id = getattr(task, "activity_type_id", None)
    if (
        allowed_activity_type_id is not None
        and allowed_activity_type_id != activity_type_id
    ):
        raise ValidationError(
            "Selected activity_type does not belong to the selected task."
        )


async def _assert_no_duplicate_draft(
    db: AsyncSession,
    created_by_id: int,
    task_id: int,
    work_date: date,
    exclude_id: int | None = None,
) -> None:
    """
    Duplicate = same created_by_id + task_id + work_date, status=Draft.
    Excludes `exclude_id` so an update can "duplicate-check against itself"
    without false-positiving.
    """
    query = select(WorkUpdate.id).where(
        WorkUpdate.created_by_id == created_by_id,
        WorkUpdate.task_id == task_id,
        WorkUpdate.work_date == work_date,
        WorkUpdate.status == WorkUpdateStatus.DRAFT.value,
    )
    if exclude_id is not None:
        query = query.where(WorkUpdate.id != exclude_id)

    existing_id = await db.scalar(query)
    if existing_id is not None:
        raise ValidationError(
            "A Draft work update already exists for this task and date. "
            "Edit the existing draft instead of creating a new one."
        )


# =====================================================
# PHASE 2 — LISTING / PAGINATION / SORT HELPERS
# =====================================================

# Maps the public `sort_by` enum values to actual ORM columns.
SORT_COLUMN_MAP = {
    s.WorkUpdateSortBy.CREATED_AT.value: WorkUpdate.created_at,
    s.WorkUpdateSortBy.UPDATED_AT.value: WorkUpdate.updated_at,
    s.WorkUpdateSortBy.WORK_DATE.value: WorkUpdate.work_date,
    s.WorkUpdateSortBy.TOTAL_HOURS.value: WorkUpdate.total_hours,
    s.WorkUpdateSortBy.BUSINESS_ID.value: WorkUpdate.business_id,
}


def _apply_sort(query, sort_by: "s.WorkUpdateSortBy", sort_order: "s.SortOrder"):
    """
    Applies ORDER BY <sort_by> <sort_order>, with WorkUpdate.id as a
    stable tiebreaker so pagination doesn't reshuffle rows that tie on the
    sort column (e.g. two rows with the same work_date).
    """
    column = SORT_COLUMN_MAP[sort_by.value]
    if sort_order == s.SortOrder.ASC:
        return query.order_by(column.asc(), WorkUpdate.id.asc())
    return query.order_by(column.desc(), WorkUpdate.id.desc())


def _build_pagination_meta(total: int, limit: int, offset: int) -> dict:
    """
    Derives page/page_size/total_pages/has_next/has_previous from the
    existing limit/offset query contract — the request shape (limit,
    offset) is unchanged so no existing caller breaks; this only adds
    richer metadata to the response.
    """
    page_size = max(limit, 1)
    page = (offset // page_size) + 1
    total_pages = (total + page_size - 1) // page_size if page_size else 0
    return {
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_next": (offset + page_size) < total,
        "has_previous": offset > 0,
    }


# =====================================================
# STATUS -> COLOR MAP (used by both Excel & PDF)
# =====================================================

STATUS_COLORS_EXCEL = {
    WorkUpdateStatus.DRAFT.value: "FFF3CD",
    WorkUpdateStatus.SUBMITTED.value: "D4EDDA",
}

STATUS_COLORS_PDF = {
    WorkUpdateStatus.DRAFT.value: colors.HexColor("#FFF3CD"),
    WorkUpdateStatus.SUBMITTED.value: colors.HexColor("#D4EDDA"),
}

BRAND_DARK = "1F4E78"
BRAND_DARK_HEX = colors.HexColor("#1F4E78")
BRAND_LIGHT_ROW = "F2F6FA"


# =====================================================
# EXCEL EXPORT (unchanged)
# =====================================================


def export_work_updates_excel(
    work_updates: list[WorkUpdate],
) -> StreamingResponse:

    wb = Workbook()

    ws = wb.active
    ws.title = "Work Updates"

    header_fill = PatternFill(fill_type="solid", fgColor=BRAND_DARK)
    header_font = Font(bold=True, color="FFFFFF", size=11)

    title_font = Font(bold=True, size=16, color=BRAND_DARK)
    subtitle_font = Font(italic=True, size=10, color="666666")

    thin_border = Border(
        left=Side(style="thin", color="B7C4CF"),
        right=Side(style="thin", color="B7C4CF"),
        top=Side(style="thin", color="B7C4CF"),
        bottom=Side(style="thin", color="B7C4CF"),
    )

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    alt_fill = PatternFill(fill_type="solid", fgColor=BRAND_LIGHT_ROW)

    headers = [
        "Business ID",
        "Date",
        "Project",
        "Task",
        "Activity",
        "Engineer",
        "Description",
        "Start Time",
        "End Time",
        "Hours",
        "Status",
    ]
    num_cols = len(headers)
    last_col_letter = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = "WORK UPDATE REPORT"
    cell.font = title_font
    cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    )
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center

    header_row = 4

    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[header_row].height = 22

    current_row = header_row + 1
    total_hours = 0.0

    for idx, item in enumerate(work_updates):

        hours = float(item.total_hours or 0)
        total_hours += hours

        row_values = [
            item.business_id,
            item.work_date,
            item.project.project_name if item.project else "",
            item.task.title if item.task else "",
            item.activity_type.name if item.activity_type else "",
            item.created_by.full_name if item.created_by else "",
            item.work_description,
            str(item.start_time or ""),
            str(item.end_time or ""),
            hours,
            (item.status or "").upper(),
        ]

        row_fill = alt_fill if idx % 2 == 1 else None

        for col, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center if col in (1, 2, 8, 9, 10, 11) else left

            if col == 2 and isinstance(value, date):
                cell.number_format = "dd-mmm-yyyy"

            if col == 10:
                cell.number_format = "0.00"

            if col == 11:
                status_hex = STATUS_COLORS_EXCEL.get(item.status, "FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor=status_hex)
                cell.font = Font(bold=True, size=10)
            elif row_fill:
                cell.fill = row_fill

        current_row += 1

    fixed_widths = {
        1: 16,
        2: 13,
        3: 20,
        4: 20,
        5: 18,
        6: 18,
        7: 40,
        8: 12,
        9: 12,
        10: 10,
        11: 14,
    }
    for col, width in fixed_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    current_row += 1
    summary_start = current_row

    ws.merge_cells(f"A{current_row}:{last_col_letter}{current_row}")
    ws.cell(current_row, 1, "SUMMARY").font = Font(bold=True, size=12, color=BRAND_DARK)
    current_row += 1

    label_font = Font(bold=True)

    ws.cell(current_row, 1, "Total Work Updates").font = label_font
    ws.cell(current_row, 2, len(work_updates))
    current_row += 1

    ws.cell(current_row, 1, "Total Hours").font = label_font
    total_cell = ws.cell(current_row, 2, round(total_hours, 2))
    total_cell.number_format = "0.00"
    current_row += 1

    draft_count = sum(
        1 for w in work_updates if w.status == WorkUpdateStatus.DRAFT.value
    )
    submitted_count = sum(
        1 for w in work_updates if w.status == WorkUpdateStatus.SUBMITTED.value
    )

    ws.cell(current_row, 1, "Draft").font = label_font
    ws.cell(current_row, 2, draft_count)
    current_row += 1

    ws.cell(current_row, 1, "Submitted").font = label_font
    ws.cell(current_row, 2, submitted_count)

    for r in range(summary_start, current_row + 1):
        ws.cell(r, 1).border = thin_border
        ws.cell(r, 2).border = thin_border

    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_view.showGridLines = False

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=work_updates.xlsx"},
    )


# =====================================================
# PDF EXPORT (unchanged)
# =====================================================


def export_work_updates_pdf(
    work_updates: list[WorkUpdate],
) -> StreamingResponse:

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]
    title_style.alignment = TA_CENTER
    title_style.textColor = BRAND_DARK_HEX

    subtitle_style = styles["Normal"]
    subtitle_style.alignment = TA_CENTER
    subtitle_style.textColor = colors.grey

    normal_style = styles["BodyText"]

    footer_style = styles["Normal"]
    footer_style.alignment = TA_CENTER
    footer_style.textColor = colors.grey
    footer_style.fontSize = 8

    elements = []

    elements.append(Paragraph("Work Update Report", title_style))
    elements.append(
        Paragraph(
            f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
            subtitle_style,
        )
    )
    elements.append(Spacer(1, 0.2 * inch))

    total_hours = sum(float(item.total_hours or 0) for item in work_updates)
    draft_count = sum(
        1 for w in work_updates if w.status == WorkUpdateStatus.DRAFT.value
    )
    submitted_count = sum(
        1 for w in work_updates if w.status == WorkUpdateStatus.SUBMITTED.value
    )

    summary = (
        f"<b>Total Records:</b> {len(work_updates)} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Total Hours:</b> {round(total_hours, 2)} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Draft:</b> {draft_count} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Submitted:</b> {submitted_count}"
    )

    elements.append(Paragraph(summary, normal_style))
    elements.append(Spacer(1, 0.25 * inch))

    header_row = [
        "Business ID",
        "Date",
        "Project",
        "Task",
        "Engineer",
        "Description",
        "Hours",
        "Status",
    ]

    table_data = [header_row]

    for item in work_updates:
        table_data.append(
            [
                item.business_id or "",
                str(item.work_date) if item.work_date else "",
                item.project.project_name if item.project else "",
                item.task.title if item.task else "",
                item.created_by.full_name if item.created_by else "",
                (item.work_description or "")[:60],
                str(round(float(item.total_hours or 0), 2)),
                (item.status or "").upper(),
            ]
        )

    col_widths = [
        0.9 * inch,
        0.8 * inch,
        1.3 * inch,
        1.3 * inch,
        1.3 * inch,
        3.4 * inch,
        0.6 * inch,
        1.0 * inch,
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK_HEX),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7C4CF")),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (4, -1), "LEFT"),
        ("ALIGN", (6, 1), (7, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
    ]

    for i, item in enumerate(work_updates, start=1):
        if i % 2 == 0:
            style_commands.append(
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F2F6FA"))
            )
        status_color = STATUS_COLORS_PDF.get(item.status)
        if status_color:
            style_commands.append(("BACKGROUND", (7, i), (7, i), status_color))
            style_commands.append(("FONTNAME", (7, i), (7, i), "Helvetica-Bold"))

    table.setStyle(TableStyle(style_commands))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    elements.append(Paragraph("Generated by Construction ERP", footer_style))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=work_updates.pdf"},
    )


# ========================create_work_update=======================================
@router.post(
    "",
    response_model=s.WorkUpdateOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_work_update(
    payload: s.WorkUpdateCreate,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Creating Work Update | project={payload.project_id}")

    await assert_project_access(
        db=db,
        project_id=payload.project_id,
        current_user=current_user,
    )

    project = await db.get(Project, payload.project_id)
    if not project:
        raise NotFoundError("Project not found")

    # FIX (Phase 1 / Section 1.2): reject work updates against a project
    # that's Closed / Completed / Inactive / Cancelled.
    _assert_project_open(project)

    await assert_task_project(
        db=db,
        task_id=payload.task_id,
        project_id=payload.project_id,
    )

    # FIX (Phase 1 / Section 1.3): assert_task_project only confirmed the
    # task exists and belongs to the project — it never checked the task's
    # own status. Fetch the task and reject Completed / Closed / Cancelled.
    task = await db.get(Task, payload.task_id)
    if not task:
        raise NotFoundError("Task not found")
    _assert_task_open(task)

    activity = await db.get(ActivityType, payload.activity_type_id)
    if not activity:
        raise ValidationError("Invalid activity_type_id")

    # FIX (Phase 1 / Section 1.4): previously activity_type_id was only
    # checked for existence, with no link back to the task at all — any
    # activity type could be attached to any task. See the ASSUMPTION note
    # on _assert_activity_belongs_to_task if your Task model's field name
    # differs from `activity_type_id`.
    _assert_activity_belongs_to_task(task, payload.activity_type_id)

    # FIX (Phase 1 / Section 1.1): reject a second Draft for the same
    # user + task + date instead of silently allowing duplicates.
    await _assert_no_duplicate_draft(
        db=db,
        created_by_id=current_user.id,
        task_id=payload.task_id,
        work_date=payload.work_date,
    )

    for _ in range(3):
        try:
            business_id = await generate_business_id(
                db=db,
                model=WorkUpdate,
                column_name="business_id",
                prefix="WU",
            )

            obj = WorkUpdate(
                business_id=business_id,
                project_id=payload.project_id,
                task_id=payload.task_id,
                activity_type_id=payload.activity_type_id,
                work_description=payload.work_description,
                before_remarks=payload.before_remarks,
                work_date=payload.work_date,
                start_time=payload.start_time,
                location=payload.location,
                status=WorkUpdateStatus.DRAFT.value,
                created_by_id=current_user.id,
            )

            db.add(obj)

            await db.flush()
            await db.commit()
            await db.refresh(obj)

            await db.refresh(
                obj,
                attribute_names=[
                    "project",
                    "task",
                    "activity_type",
                    "created_by",
                    "images",
                ],
            )

            logger.info(f"Work Update Created | id={obj.id}")

            return s.WorkUpdateOut.model_validate(
                obj,
                from_attributes=True,
            )

        except IntegrityError as e:
            await db.rollback()
            logger.warning(f"Retry creating Work Update because of IntegrityError: {e}")
            continue

        # FIX (Medium #7): previously only IntegrityError was caught inside
        # this loop. Any other exception (e.g. from generate_business_id,
        # or a driver-level error during flush/commit) propagated without
        # rolling back, potentially leaving the session dirty for whatever
        # ran next on it. Roll back and re-raise so the caller/dependency
        # sees a clean failure either way.
        except Exception:
            await db.rollback()
            logger.exception("Unexpected error while creating Work Update")
            raise

    raise ValidationError("Unable to generate unique Work Update ID.")


# =================get_my_work_updates=================
# FIX (Critical #1): moved this — and export_work_updates / get_project_timeline
# below — ABOVE get_work_update. "/{work_update_id}" matches ANY single path
# segment as a string before FastAPI even tries to coerce it to int, so with
# the old ordering "GET /work-updates/my" and "GET /work-updates/export"
# were being intercepted by get_work_update and failing int-conversion with
# a 422 — those two endpoints were effectively dead code. Declaration order
# matters for path routing; static single-segment routes must come before
# the dynamic "/{work_update_id}" route.


@router.get(
    "/my",
    response_model=s.WorkUpdateListOut,
    summary="Get my work updates",
    description=(
        "Lists the current user's work updates. `search` matches against "
        "Business ID, Project name, Task name, Engineer name, Description, "
        "Before/After remarks. Supports sorting via `sort_by`/`sort_order` "
        "and standard limit/offset pagination (response includes derived "
        "page/page_size/total_pages/has_next/has_previous)."
    ),
)
async def get_my_work_updates(
    project_id: int | None = None,
    # FIX (Phase 2 / Section 2): status_filter is now the real
    # WorkUpdateStatus enum instead of a free string — Swagger renders it
    # as a dropdown and invalid values 422 instead of silently matching
    # zero rows.
    status_filter: WorkUpdateStatus | None = Query(
        None, description="Filter by status (Draft / Submitted)."
    ),
    search: str | None = Query(
        None,
        description=(
            "Matches Business ID, Project name, Task name, Engineer name, "
            "Description, Before remarks, After remarks."
        ),
    ),
    date_from: date | None = Query(None, description="work_date >= this date"),
    date_to: date | None = Query(None, description="work_date <= this date"),
    sort_by: s.WorkUpdateSortBy = Query(s.WorkUpdateSortBy.WORK_DATE),
    sort_order: s.SortOrder = Query(s.SortOrder.DESC),
    limit: int = Query(20, ge=1, le=MAX_PAGE_SIZE),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"My Work Updates | user_id={current_user.id}")

    query = (
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
            selectinload(WorkUpdate.images),
        )
        .where(WorkUpdate.created_by_id == current_user.id)
    )

    if project_id:
        query = query.where(WorkUpdate.project_id == project_id)

    if status_filter:
        query = query.where(WorkUpdate.status == status_filter.value)

    if date_from:
        query = query.where(WorkUpdate.work_date >= date_from)

    if date_to:
        query = query.where(WorkUpdate.work_date <= date_to)

    # FIX (Phase 2 / Section 2): `search` previously only matched
    # work_description. Project/Task/Engineer names live on related
    # tables, so `.has(...)` (an EXISTS subquery against the relationship)
    # is used instead of a join — this avoids duplicate WorkUpdate rows
    # that a plain JOIN would introduce when combined with the `images`
    # eager-load, and keeps the query index-friendly.
    if search:
        like = f"%{search}%"
        query = query.where(
            or_(
                WorkUpdate.business_id.ilike(like),
                WorkUpdate.work_description.ilike(like),
                WorkUpdate.before_remarks.ilike(like),
                WorkUpdate.after_remarks.ilike(like),
                WorkUpdate.project.has(Project.name.ilike(like)),
                WorkUpdate.task.has(Task.name.ilike(like)),
                WorkUpdate.created_by.has(User.full_name.ilike(like)),
            )
        )

    total = await db.scalar(select(func.count()).select_from(query.subquery()))
    total = total or 0

    query = _apply_sort(query, sort_by, sort_order)

    result = await db.scalars(query.offset(offset).limit(limit))
    items = result.unique().all()

    logger.info(f"My Work Updates fetched | total={total}")

    return s.WorkUpdateListOut(
        total=total,
        items=[
            s.WorkUpdateOut.model_validate(item, from_attributes=True) for item in items
        ],
        **_build_pagination_meta(total, limit, offset),
    )


# ================get_project_timeline===========================


@router.get(
    "/project/{project_id}/timeline",
    response_model=s.WorkUpdateListOut,
    summary="Get project work-update timeline",
    description=(
        "Lists work updates for a project with filters for engineer, task, "
        "activity, status, hours range, and date range, plus sorting and "
        "pagination metadata."
    ),
)
async def get_project_timeline(
    project_id: int,
    search: str | None = Query(
        None,
        description=(
            "Matches Business ID, Description, Before/After remarks, "
            "Engineer name, Task name."
        ),
    ),
    # FIX (Phase 2 / Section 3): enum instead of free string, same
    # reasoning as get_my_work_updates.
    status_filter: WorkUpdateStatus | None = Query(None),
    # FIX (Phase 2 / Section 3): "Engineer" / "Created By" filters — both
    # map to created_by_id since a work update has a single creator/engineer.
    created_by_id: int | None = Query(
        None, description="Filter by the engineer who created the entry."
    ),
    task_id: int | None = Query(None, description="Filter by task."),
    activity_type_id: int | None = Query(None, description="Filter by activity type."),
    min_hours: float | None = Query(
        None, ge=0, description="total_hours >= this value"
    ),
    max_hours: float | None = Query(
        None, le=24, description="total_hours <= this value"
    ),
    start_date: date | None = None,
    end_date: date | None = None,
    sort_by: s.WorkUpdateSortBy = Query(s.WorkUpdateSortBy.WORK_DATE),
    sort_order: s.SortOrder = Query(s.SortOrder.DESC),
    limit: int = Query(20, ge=1, le=MAX_PAGE_SIZE),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Fetching Work Timeline | project={project_id}")

    await assert_project_access(
        db=db,
        project_id=project_id,
        current_user=current_user,
    )

    project = await db.get(Project, project_id)
    if not project:
        raise NotFoundError("Project not found.")

    if min_hours is not None and max_hours is not None and min_hours > max_hours:
        raise ValidationError("min_hours cannot be greater than max_hours.")

    query = (
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
            selectinload(WorkUpdate.images),
        )
        .where(WorkUpdate.project_id == project_id)
    )

    if status_filter:
        query = query.where(WorkUpdate.status == status_filter.value)

    if created_by_id:
        query = query.where(WorkUpdate.created_by_id == created_by_id)

    if task_id:
        query = query.where(WorkUpdate.task_id == task_id)

    if activity_type_id:
        query = query.where(WorkUpdate.activity_type_id == activity_type_id)

    if min_hours is not None:
        query = query.where(WorkUpdate.total_hours >= min_hours)

    if max_hours is not None:
        query = query.where(WorkUpdate.total_hours <= max_hours)

    if start_date:
        query = query.where(WorkUpdate.work_date >= start_date)

    if end_date:
        query = query.where(WorkUpdate.work_date <= end_date)

    if search:
        like = f"%{search}%"
        query = query.where(
            or_(
                WorkUpdate.business_id.ilike(like),
                WorkUpdate.work_description.ilike(like),
                WorkUpdate.before_remarks.ilike(like),
                WorkUpdate.after_remarks.ilike(like),
                WorkUpdate.task.has(Task.name.ilike(like)),
                WorkUpdate.created_by.has(User.full_name.ilike(like)),
            )
        )

    total = await db.scalar(select(func.count()).select_from(query.subquery()))
    total = total or 0

    query = _apply_sort(query, sort_by, sort_order)

    result = await db.scalars(query.offset(offset).limit(limit))
    items = result.unique().all()

    logger.info(f"Timeline fetched | total={total}")

    return s.WorkUpdateListOut(
        total=total,
        items=[
            s.WorkUpdateOut.model_validate(item, from_attributes=True) for item in items
        ],
        **_build_pagination_meta(total, limit, offset),
    )


# =====================Export Work Updates=============================


@router.get(
    "/export",
    summary="Export Work Updates",
)
async def export_work_updates(
    project_id: int | None = Query(None),
    user_id: int | None = Query(None),
    status_filter: str | None = Query(None),
    from_date: date | None = Query(None),
    to_date: date | None = Query(None),
    format: ExportFormat = Query(ExportFormat.EXCEL),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    """
    Export Work Updates into Excel or PDF.
    """

    if project_id:
        project = await db.get(Project, project_id)
        if not project:
            raise ValidationError("Project not found.")

        await assert_project_access(
            db=db,
            project_id=project_id,
            current_user=current_user,
        )

    # FIX (Critical #3): removed `WorkUpdate.is_deleted == False`. The
    # WorkUpdate model (unchanged, per request) has no `is_deleted` column,
    # so this filter would raise on every call to this endpoint unless it's
    # inherited from Base/TimestampMixin. Since this pass doesn't touch the
    # model, the filter is removed rather than assumed. If soft-delete is
    # actually needed, add the column on the model side and reinstate this
    # filter (and switch delete_work_update to set the flag instead of a
    # hard delete) in a follow-up model change.
    query = select(WorkUpdate).options(
        selectinload(WorkUpdate.project),
        selectinload(WorkUpdate.task),
        selectinload(WorkUpdate.activity_type),
        selectinload(WorkUpdate.created_by),
    )

    if project_id:
        query = query.where(WorkUpdate.project_id == project_id)

    if user_id:
        query = query.where(WorkUpdate.created_by_id == user_id)

    if status_filter:
        query = query.where(WorkUpdate.status == status_filter)

    if from_date:
        query = query.where(WorkUpdate.work_date >= from_date)

    if to_date:
        query = query.where(WorkUpdate.work_date <= to_date)

    query = query.order_by(WorkUpdate.work_date.desc(), WorkUpdate.id.desc())

    result = await db.scalars(query)
    work_updates = result.unique().all()

    if not work_updates:
        raise ValidationError("No work updates found.")

    if format == ExportFormat.EXCEL:
        return await run_in_threadpool(export_work_updates_excel, work_updates)

    if format == ExportFormat.PDF:
        return await run_in_threadpool(export_work_updates_pdf, work_updates)

    raise ValidationError("Invalid export format.")


# ==================get_work_update=============


@router.get(
    "/{work_update_id:int}",
    response_model=s.WorkUpdateOut,
)
async def get_work_update(
    work_update_id: int,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_READ_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Fetching Work Update | id={work_update_id}")

    obj = await db.scalar(
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
            selectinload(WorkUpdate.images),
        )
        .where(WorkUpdate.id == work_update_id)
    )

    if not obj:
        raise NotFoundError("Work update not found")

    await assert_project_access(
        db=db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    return s.WorkUpdateOut.model_validate(obj, from_attributes=True)


# ==================upload_before_images=============================


@router.post(
    "/{work_update_id}/before-image",
    response_model=s.WorkUpdateOut,
    status_code=status.HTTP_200_OK,
)
async def upload_before_image(
    work_update_id: int,
    image: UploadFile = File(...),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Uploading BEFORE image | work_update_id={work_update_id}")

    work_update = await db.scalar(
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.images),
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
        )
        .where(WorkUpdate.id == work_update_id)
    )

    if not work_update:
        raise NotFoundError("Work update not found.")

    await assert_project_access(
        db=db,
        project_id=work_update.project_id,
        current_user=current_user,
    )

    if work_update.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update cannot be modified.")

    # Check existing BEFORE images
    before_count = (
        await db.scalar(
            select(func.count())
            .select_from(WorkUpdateImage)
            .where(
                WorkUpdateImage.work_update_id == work_update_id,
                WorkUpdateImage.image_type == WorkUpdateImageType.BEFORE,
            )
        )
        or 0
    )

    if before_count >= 10:
        raise ValidationError("Maximum 10 BEFORE images are allowed.")

    display_order = (
        await db.scalar(
            select(func.max(WorkUpdateImage.display_order)).where(
                WorkUpdateImage.work_update_id == work_update_id
            )
        )
        or 0
    )

    try:
        image_path = await validate_and_save_image(
            file=image,
            upload_dir="uploads/work_updates/before",
            prefix="before",
        )

        db.add(
            WorkUpdateImage(
                work_update_id=work_update.id,
                image_type=WorkUpdateImageType.BEFORE,
                image_url=image_path,
                display_order=display_order + 1,
            )
        )

        await db.commit()

        work_update = await db.scalar(
            select(WorkUpdate)
            .options(
                selectinload(WorkUpdate.images),
                selectinload(WorkUpdate.project),
                selectinload(WorkUpdate.task),
                selectinload(WorkUpdate.activity_type),
                selectinload(WorkUpdate.created_by),
            )
            .where(WorkUpdate.id == work_update_id)
        )

        logger.info(
            f"BEFORE image uploaded successfully | work_update_id={work_update.id}"
        )

        return s.WorkUpdateOut.model_validate(
            work_update,
            from_attributes=True,
        )

    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to upload BEFORE image: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to upload BEFORE image.",
        )


# ===============update_work_update (PARTIAL update)====================


@router.put(
    "/{work_update_id:int}",
    response_model=s.WorkUpdateOut,
)
async def update_work_update(
    work_update_id: int,
    payload: s.WorkUpdateUpdate,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Updating Work Update | id={work_update_id}")

    obj = await db.scalar(
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.images),
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
        )
        .where(WorkUpdate.id == work_update_id)
    )

    if not obj:
        raise NotFoundError("Work update not found")

    await assert_project_access(
        db=db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update cannot be updated.")

    # FIX (Phase 1 / Section 7): "only creator can edit draft, Admin can
    # override" — previously ANY user with a write role (including a
    # different Contractor/Engineer on the same project) could edit
    # someone else's draft. assert_project_access only confirms project
    # membership, not ownership of this specific record.
    is_owner = current_user.id == obj.created_by_id
    is_admin = current_user.role == UserRole.ADMIN.value
    if not is_owner and not is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only the creator or an Admin can edit this work update.",
        )

    update_data = payload.model_dump(exclude_unset=True)

    if not update_data:
        raise ValidationError("No fields provided to update.")

    # FIX (Phase 1 / Section 1.2 / 1.3, applied to update): re-validate
    # project/task status on every update too — e.g. a task could have
    # moved to Closed after the draft was created but before it was edited.
    _assert_project_open(obj.project)

    new_task_id = update_data.get("task_id", obj.task_id)

    if "task_id" in update_data:
        await assert_task_project(
            db=db,
            task_id=update_data["task_id"],
            project_id=obj.project_id,
        )
        task = await db.get(Task, new_task_id)
        if not task:
            raise NotFoundError("Task not found")
    else:
        task = obj.task

    _assert_task_open(task)

    new_activity_type_id = update_data.get("activity_type_id", obj.activity_type_id)

    if "activity_type_id" in update_data:
        activity = await db.get(ActivityType, update_data["activity_type_id"])
        if not activity:
            raise ValidationError("Invalid activity_type_id")

    # FIX (Phase 1 / Section 1.4, applied to update): validate the
    # activity/task link whenever either side of that pair changes.
    if "activity_type_id" in update_data or "task_id" in update_data:
        _assert_activity_belongs_to_task(task, new_activity_type_id)

    # FIX (Phase 1 / Section 1.1, applied to update): only re-check
    # duplicates if the fields that define a "duplicate" are actually
    # changing — no need to duplicate-check a description-only edit.
    new_work_date = update_data.get("work_date", obj.work_date)
    if "task_id" in update_data or "work_date" in update_data:
        await _assert_no_duplicate_draft(
            db=db,
            created_by_id=obj.created_by_id,
            task_id=new_task_id,
            work_date=new_work_date,
            exclude_id=obj.id,
        )

    for field, value in update_data.items():
        setattr(obj, field, value)

    try:
        await db.flush()

        await db.commit()

        # Reload complete object with all relationships
        obj = await db.scalar(
            select(WorkUpdate)
            .options(
                selectinload(WorkUpdate.images),
                selectinload(WorkUpdate.project),
                selectinload(WorkUpdate.task),
                selectinload(WorkUpdate.activity_type),
                selectinload(WorkUpdate.created_by),
            )
            .where(WorkUpdate.id == work_update_id)
        )

        logger.info(f"Work Update Updated | id={obj.id}")

        return s.WorkUpdateOut.model_validate(
            obj,
            from_attributes=True,
        )

    except IntegrityError:
        await db.rollback()
        raise ValidationError("Unable to update work update.")


# =========upload_after_images============


@router.post(
    "/{work_update_id:int}/after-image",
    response_model=s.WorkUpdateOut,
    status_code=status.HTTP_200_OK,
)
async def upload_after_image(
    work_update_id: int,
    image: UploadFile = File(...),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Uploading AFTER image | work_update_id={work_update_id}")

    work_update = await db.scalar(
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.images),
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
        )
        .where(WorkUpdate.id == work_update_id)
    )

    if not work_update:
        raise NotFoundError("Work update not found.")

    await assert_project_access(
        db=db,
        project_id=work_update.project_id,
        current_user=current_user,
    )

    if work_update.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update cannot be modified.")

    # BEFORE image validation
    before_count = (
        await db.scalar(
            select(func.count())
            .select_from(WorkUpdateImage)
            .where(
                WorkUpdateImage.work_update_id == work_update_id,
                WorkUpdateImage.image_type == WorkUpdateImageType.BEFORE,
            )
        )
        or 0
    )

    if before_count == 0:
        raise ValidationError("Upload BEFORE image first.")

    # Maximum 10 AFTER images
    after_count = (
        await db.scalar(
            select(func.count())
            .select_from(WorkUpdateImage)
            .where(
                WorkUpdateImage.work_update_id == work_update_id,
                WorkUpdateImage.image_type == WorkUpdateImageType.AFTER,
            )
        )
        or 0
    )

    if after_count >= 10:
        raise ValidationError("Maximum 10 AFTER images are allowed.")

    display_order = (
        await db.scalar(
            select(func.max(WorkUpdateImage.display_order)).where(
                WorkUpdateImage.work_update_id == work_update_id
            )
        )
        or 0
    )

    try:
        image_path = await validate_and_save_image(
            file=image,
            upload_dir="uploads/work_updates/after",
            prefix="after",
        )

        db.add(
            WorkUpdateImage(
                work_update_id=work_update.id,
                image_type=WorkUpdateImageType.AFTER,
                image_url=image_path,
                display_order=display_order + 1,
            )
        )

        await db.commit()

        work_update = await db.scalar(
            select(WorkUpdate)
            .options(
                selectinload(WorkUpdate.images),
                selectinload(WorkUpdate.project),
                selectinload(WorkUpdate.task),
                selectinload(WorkUpdate.activity_type),
                selectinload(WorkUpdate.created_by),
            )
            .where(WorkUpdate.id == work_update_id)
        )

        logger.info(
            f"AFTER image uploaded successfully | work_update_id={work_update.id}"
        )

        return s.WorkUpdateOut.model_validate(
            work_update,
            from_attributes=True,
        )

    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to upload AFTER image: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Unable to upload AFTER image.",
        )


# =========submit_work_update============


@router.post(
    "/{work_update_id:int}/submit",
    response_model=s.WorkUpdateOut,
)
async def submit_work_update(
    work_update_id: int,
    payload: s.WorkUpdateSubmit,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Submitting Work Update | id={work_update_id}")

    obj = await db.scalar(
        select(WorkUpdate)
        .options(
            selectinload(WorkUpdate.images),
            selectinload(WorkUpdate.project),
            selectinload(WorkUpdate.task),
            selectinload(WorkUpdate.activity_type),
            selectinload(WorkUpdate.created_by),
        )
        .where(WorkUpdate.id == work_update_id)
    )

    if not obj:
        raise NotFoundError("Work update not found")

    await assert_project_access(
        db=db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Work update already submitted.")

    before_count = await db.scalar(
        select(func.count()).where(
            WorkUpdateImage.work_update_id == work_update_id,
            WorkUpdateImage.image_type == WorkUpdateImageType.BEFORE.value,
        )
    )

    if before_count == 0:
        raise ValidationError("Please upload BEFORE images.")

    after_count = await db.scalar(
        select(func.count()).where(
            WorkUpdateImage.work_update_id == work_update_id,
            WorkUpdateImage.image_type == WorkUpdateImageType.AFTER.value,
        )
    )

    if after_count == 0:
        raise ValidationError("Please upload AFTER images.")

    if payload.end_time <= obj.start_time:
        raise ValidationError("End time must be greater than start time.")

    if payload.total_hours <= 0:
        raise ValidationError("Total hours must be greater than zero.")

    computed_hours = (
        datetime.combine(date.min, payload.end_time)
        - datetime.combine(date.min, obj.start_time)
    ).total_seconds() / 3600

    if abs(computed_hours - float(payload.total_hours)) > 0.01:
        raise ValidationError(
            "total_hours does not match the elapsed time between "
            "start_time and end_time."
        )

    obj.end_time = payload.end_time
    obj.total_hours = payload.total_hours
    obj.after_remarks = payload.after_remarks
    obj.status = WorkUpdateStatus.SUBMITTED.value

    try:
        await db.flush()
        await db.commit()

        # Reload fresh object with all relationships after commit
        result = await db.execute(
            select(WorkUpdate)
            .options(
                selectinload(WorkUpdate.images),
                selectinload(WorkUpdate.project),
                selectinload(WorkUpdate.task),
                selectinload(WorkUpdate.activity_type),
                selectinload(WorkUpdate.created_by),
            )
            .where(WorkUpdate.id == work_update_id)
        )

        obj = result.scalar_one()

        logger.info(f"Work Update Submitted | id={obj.id}")

        return s.WorkUpdateOut.model_validate(
            obj,
            from_attributes=True,
        )

    except IntegrityError:
        await db.rollback()
        raise ValidationError("Unable to submit work update.")


# =============delete_work_update==================


@router.delete(
    "/{work_update_id:int}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_work_update(
    work_update_id: int,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_DELETE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    obj = await db.scalar(
        select(WorkUpdate)
        .options(selectinload(WorkUpdate.images))
        .where(WorkUpdate.id == work_update_id)
    )

    if not obj:
        raise NotFoundError("Work update not found")

    await assert_project_access(
        db=db,
        project_id=obj.project_id,
        current_user=current_user,
    )

    if obj.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update cannot be deleted.")

    image_paths = [img.image_url for img in obj.images if img.image_url]

    try:
        await db.delete(obj)
        await db.flush()
        # FIX (Critical #2): added the missing commit. Without it, the
        # delete would be flushed within the transaction but never
        # durably committed, while the file-cleanup loop below runs
        # unconditionally — leaving files deleted on disk even if the DB
        # row silently survives.
        await db.commit()
    except Exception:
        await db.rollback()
        raise

    for path in image_paths:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            logger.warning(f"Unable to delete image file on work update delete: {e}")


# ==================delete_work_update_image===========================


@router.delete(
    "/images/{image_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_work_update_image(
    image_id: int,
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Deleting Work Update Image | image_id={image_id}")

    image = await db.scalar(
        select(WorkUpdateImage)
        .options(selectinload(WorkUpdateImage.work_update))
        .where(WorkUpdateImage.id == image_id)
    )

    if not image:
        raise NotFoundError("Image not found.")

    work_update = image.work_update

    await assert_project_access(
        db=db,
        project_id=work_update.project_id,
        current_user=current_user,
    )

    if work_update.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update images cannot be deleted.")

    if image.image_type == WorkUpdateImageType.BEFORE.value:
        before_count = await db.scalar(
            select(func.count()).where(
                WorkUpdateImage.work_update_id == work_update.id,
                WorkUpdateImage.image_type == WorkUpdateImageType.BEFORE.value,
            )
        )
        if before_count <= 1:
            raise ValidationError("At least one BEFORE image is required.")

    elif image.image_type == WorkUpdateImageType.AFTER.value:
        after_count = await db.scalar(
            select(func.count()).where(
                WorkUpdateImage.work_update_id == work_update.id,
                WorkUpdateImage.image_type == WorkUpdateImageType.AFTER.value,
            )
        )
        if after_count <= 1:
            raise ValidationError("At least one AFTER image is required.")

    try:
        await db.delete(image)
        await db.flush()
        # FIX (Critical #2): added the missing commit — previously flush()
        # only, so the image row deletion could be silently lost.
        await db.commit()
        logger.info(f"Image deleted successfully | image_id={image_id}")
    except Exception:
        await db.rollback()
        raise

    # FIX: file removal moved to AFTER the DB commit succeeds (mirrors the
    # pattern used in delete_work_update). Previously the file was removed
    # from disk BEFORE the DB delete was attempted — if the DB delete then
    # failed/rolled back, the row would still reference a file that no
    # longer exists on disk.
    try:
        if image.image_url and os.path.exists(image.image_url):
            os.remove(image.image_url)
    except Exception as e:
        logger.warning(f"Unable to delete image file: {e}")


# ==============replace_work_update_image==================


@router.put(
    "/images/{image_id}",
    response_model=s.WorkUpdateImageOut,
)
async def replace_work_update_image(
    image_id: int,
    new_image: UploadFile = File(...),
    current_user: User = Depends(d.require_roles(WORK_UPDATE_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db_session),
):
    logger.info(f"Replacing Work Update Image | image_id={image_id}")

    image = await db.scalar(
        select(WorkUpdateImage)
        .options(selectinload(WorkUpdateImage.work_update))
        .where(WorkUpdateImage.id == image_id)
    )

    if not image:
        raise NotFoundError("Image not found.")

    work_update = image.work_update

    await assert_project_access(
        db=db,
        project_id=work_update.project_id,
        current_user=current_user,
    )

    if work_update.status == WorkUpdateStatus.SUBMITTED.value:
        raise ValidationError("Submitted work update images cannot be replaced.")

    folder = (
        "uploads/work_updates/before"
        if image.image_type == WorkUpdateImageType.BEFORE.value
        else "uploads/work_updates/after"
    )
    prefix = (
        "before" if image.image_type == WorkUpdateImageType.BEFORE.value else "after"
    )

    new_image_path = await validate_and_save_image(
        file=new_image,
        upload_dir=folder,
        prefix=prefix,
    )

    old_image_path = image.image_url
    image.image_url = new_image_path

    try:
        await db.flush()
        # FIX (Critical #2): added the missing commit.
        await db.commit()
        await db.refresh(image)

        logger.info(f"Image replaced successfully | image_id={image.id}")

    except Exception:
        await db.rollback()
        try:
            if os.path.exists(new_image_path):
                os.remove(new_image_path)
        except Exception:
            pass
        raise

    # FIX: old file removal moved to AFTER the commit succeeds — previously
    # it ran before the DB write was attempted, so a failed/rolled-back
    # update would leave the DB still pointing at a file that had already
    # been deleted from disk.
    try:
        if old_image_path and os.path.exists(old_image_path):
            os.remove(old_image_path)
    except Exception as e:
        logger.warning(f"Unable to delete old image: {e}")

    return s.WorkUpdateImageOut.model_validate(image, from_attributes=True)
